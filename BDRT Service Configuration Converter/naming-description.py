
import sys
import os
import re
import datetime
import time
import openpyxl
from openpyxl.styles import PatternFill
from netaddr import IPNetwork
from string import ascii_lowercase

t = datetime.datetime.now()
inputname = 'migration-input.xlsx'
sdpname = 'SDP number.xlsm'
wbmap = 'nodemapping.xlsx'
naming = 'naming'
namingname = 'migration-naming-%s.xlsx'%t.strftime("%Y%m%d-%H%M")
cfgpath = '.\iptn-cfg\\'
extractpath = '.\extraction\\'
inputpath = '.\input\\'
attrpath = 'attr-cfg\\'
migratepath = '.\migration\\'
wsindex = 'index'
wsport = 'port-lag'
wsvprn = 'vprn'
wsvprnint = 'vprn-int'
wsl2vpn = 'l2vpn'
wsvpls = 'vpls'
wsbridge = 'bridge-domain'
wsvsw = 'virtual-switch'


def createinputfile():
    iptnnode = ['ASAY_IPTNJ_MX2020_PE03','ASAY_IPTNJ_MX2020_PE04','ASAY_IPTNJ_MX960_PE01','ASAY_IPTNJ_MX960_PE02','BPLNB_IPTNJ_MX960_PE01','BPLNB_IPTNJ_MX960_PE02','BPLNB_IPTNJ_MX960_PE05','BPLNB_IPTNJ_MX960_PE06','CMBP_IPTNJ_MX2020_PE03','CMBP_IPTNJ_MX2020_PE04','CMBP_IPTNJ_MX960_PE01','CMBP_IPTNJ_MX960_PE02','CWDC_IPTNJ_MX2020_PE07','CWDC_IPTNJ_MX2020_PE08','CWDC_IPTNJ_MX960_PE01','CWDC_IPTNJ_MX960_PE02','CWDC_IPTNJ_MX960_PE05','CWDC_IPTNJ_MX960_PE06','ERHQ_IPTNJ_MX2020_PE03','ERHQ_IPTNJ_MX2020_PE04','ERHQ_IPTNJ_MX960_PE01','ERHQ_IPTNJ_MX960_PE02','KNKON_IPTNJ_MX2020_PE03','KNKON_IPTNJ_MX2020_PE04','MSCS_IPTNJ_MX2020_PE03','MSCS_IPTNJ_MX2020_PE04','MSCS_IPTNJ_MX960_PE01','MSCS_IPTNJ_MX960_PE02','NKY2_IPTNJ_MX2020_PE03','NKY2_IPTNJ_MX2020_PE04','NKY2_IPTNJ_MX960_PE01','NKY2_IPTNJ_MX960_PE02','PYOF_IPTNJ_MX2020_PE03','PYOF_IPTNJ_MX2020_PE04','PYOF_IPTNJ_MX960_PE01','PYOF_IPTNJ_MX960_PE02','RONESUB_IPTNJ_MX960_PE01','RONESUB_IPTNJ_MX960_PE02','RONE_IPTNJ_MX2020_PE03','RONE_IPTNJ_MX2020_PE04','RONE_IPTNJ_MX960_PE01','RONE_IPTNJ_MX960_PE02','SILA1_IPTNJ_MX2020_PE01','SILA1_IPTNJ_MX2020_PE02','SINKO_IPTNJ_MX2020_PE03','SINKO_IPTNJ_MX2020_PE04','SUK_IPTNJ_MX960_PE01','SUK_IPTNJ_MX960_PE02','SUK_IPTNJ_MX960_PE05','SUK_IPTNJ_MX960_PE06','TLC_IPTNJ_MX960_PE01','TLC_IPTNJ_MX960_PE02','TLC_IPTNJ_MX960_PE05','TLC_IPTNJ_MX960_PE06','TLS1_IPTNJ_MX2020_PE03','TLS1_IPTNJ_MX2020_PE04','TLS1_IPTNJ_MX960_PE01','TLS1_IPTNJ_MX960_PE02','TLS2_IPTNJ_MX2020_PE01','TLS2_IPTNJ_MX2020_PE02','WCLMNB_IPTNJ_MX2020_PE03','WCLMNB_IPTNJ_MX2020_PE04','WCLMNB_IPTNJ_MX960_PE01','WCLMNB_IPTNJ_MX960_PE02','SCT2_IPTNJ_MX960_PE03','SCT2_IPTNJ_MX960_PE04','SUK_IPTNJ_PTX_P01','TLS1_IPTNJ_PTX_P01','CWDC_IPTNJ_PTX_P01']
    iptnlo0 = ['10.129.116.15','10.129.116.16','10.129.116.9','10.129.116.10','10.129.112.50','10.129.112.51','10.129.112.54','10.129.112.55','10.129.118.53','10.129.118.54','10.129.118.49','10.129.118.50','10.129.112.83','10.129.112.84','10.129.112.34','10.129.112.35','10.129.112.38','10.129.112.39','10.129.120.3','10.129.120.4','10.129.120.1','10.129.120.2','10.129.122.15','10.129.122.16','10.129.124.20','10.129.124.21','10.129.124.17','10.129.124.18','10.129.116.5','10.129.116.6','10.129.116.1','10.129.116.2','10.129.120.51','10.129.120.52','10.129.120.49','10.129.120.50','10.129.122.9','10.129.122.10','10.129.122.29','10.129.122.30','10.129.122.8','10.129.122.28','10.129.113.157','10.129.113.158','10.129.124.11','10.129.124.12','10.129.112.18','10.129.112.19','10.129.112.22','10.129.112.23','10.129.112.2','10.129.112.3','10.129.112.6','10.129.112.7','10.129.113.155','10.129.113.156','10.129.113.153','10.129.113.154','10.129.112.103','10.129.112.104','10.129.118.37','10.129.118.38','10.129.118.33','10.129.118.34','10.129.113.172','10.129.113.173','10.129.112.16','10.129.113.152','10.129.112.32']
    bdrnode = ['BDRT_CWDC_P01','BDRT_SUKE_P01','BDRT_TLS1_P01','BDRT_CWDC_PE01','BDRT_CWDC_PE02','BDRT_SUKE_PE01','BDRT_SUKE_PE02','BDRT_TLS1_PE01','BDRT_TLS1_PE02','BDRT_TLS2_PE01','BDRT_TLS2_PE02','BDRT_SILA1_PE01','BDRT_SILA1_PE02','BDRT_BPL2_PE01','BDRT_BPL2_PE02','BDRT_TWAE_PE01','BDRT_TWAE_PE02','BDRT_ASAY_PE01','BDRT_ASAY_PE02','BDRT_NKY2_PE01','BDRT_NKY2_PE02','BDRT_WCLM_PE01','BDRT_WCLM_PE02','BDRT_CMBP_PE01','BDRT_CMBP_PE02','BDRT_ERHQ_PE01','BDRT_ERHQ_PE02','BDRT_PYOF_PE01','BDRT_PYOF_PE02','BDRT_RONE_PE01','BDRT_RONE_PE02','BDRT_KNKON_PE01','BDRT_KNKON_PE02','BDRT_SINKO_PE01','BDRT_SINKO_PE02','BDRT_MSCS_PE01','BDRT_MSCS_PE02']
    bdrsystem = ['10.129.147.1','10.129.147.2','10.129.147.3','10.129.147.4','10.129.147.5','10.129.147.6','10.129.147.7','10.129.147.8','10.129.147.9','10.129.147.10','10.129.147.11','10.129.147.12','10.129.147.13','10.129.147.14','10.129.147.15','10.129.147.16','10.129.147.17','10.129.151.1','10.129.151.2','10.129.151.3','10.129.151.4','10.129.153.1','10.129.153.2','10.129.153.3','10.129.153.4','10.129.155.1','10.129.155.2','10.129.155.3','10.129.155.4','10.129.157.1','10.129.157.2','10.129.157.3','10.129.157.4','10.129.159.1','10.129.159.2','10.129.159.3','10.129.159.4']
    if os.path.exists(migratepath+inputpath+inputname):
        existing = openpyxl.load_workbook(filename=migratepath+inputpath+inputname)
        ws = existing['input']
        ws['B1'] = t.strftime("%Y%m%d-%H%M")
        existing.save(migratepath + inputpath + inputname)
        existing.close()
        os.rename(migratepath+inputpath+inputname, migratepath+t.strftime("%Y%m%d-%H%M")+'\\'+'migration-input_'+ t.strftime("%Y%m%d-%H%M") +'.xlsx')


    row = 3
    input = openpyxl.Workbook()
    input.remove(input['Sheet'])
    inputws = input.create_sheet('input')
    inputws['A1'] = row
    inputws['A2'] = 'iptn-node'
    inputws['B2'] = 'iptn-port'
    inputws['C2'] = 'nokia-bdr-node'
    inputws['D2'] = 'nokia-bdr-port'
    inputws['F2'] = '## iptn-node and iptn-port are mandatory required fields. (please make sure the physical port is exact match with the configuration on each node)'
    inputws['F3'] = '## * if not specify nokia-bdr node and port. this migration file will be use default node mapping (odd iptn node to odd bdr node and even iptn node to even bdr node)'
    inputws['F4'] = '## * and will be use temporary port instead : e.g., 1/a/a, 1/a/b'
    inputws['F7'] = 'iptn node'
    inputws['G7'] = 'iptn loopback0'
    inputws['H7'] = 'bdr node'
    inputws['I7'] = 'bdr system-ip'

    ri = 8
    rb = 8
    for n in range(len(iptnnode)):
        inputws['F%s' % ri] = iptnnode[n]
        inputws['G%s' % ri] = iptnlo0[n]
        ri += 1
    for n in range(len(bdrnode)):
        inputws['H%s' % rb] = bdrnode[n]
        inputws['I%s' % rb] = bdrsystem[n]
        rb += 1

    inputws.auto_filter.ref = 'A2:D2'
    inputws.freeze_panes = inputws['A3']
    inputws.column_dimensions['A'].width = 30.0
    inputws.column_dimensions['B'].width = 20.0
    inputws.column_dimensions['C'].width = 30.0
    inputws.column_dimensions['D'].width = 20.0
    inputws.column_dimensions['F'].width = 30.0
    inputws.column_dimensions['G'].width = 20.0
    inputws.column_dimensions['H'].width = 30.0
    inputws.column_dimensions['I'].width = 20.0
    input.save(migratepath+inputpath+inputname)

def createnokianamingfile():
    row = 3
    namingwb = openpyxl.Workbook()
    namingwb.remove(namingwb['Sheet'])
    namingws = namingwb.create_sheet(naming)
    namingws['A2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['B2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['C2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['D2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['H2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['L2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['O2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['S2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['X2'].fill = PatternFill(start_color='FFd3d3d3',end_color='FFd3d3d3',fill_type='solid')
    namingws['A1'] = row
    namingws['A2'] = 'iptn-node'
    namingws['B2'] = 'iptn-loopback'
    namingws['C2'] = 'iptn-interface.unit'
    namingws['D2'] = 'iptn-ae'
    namingws['E2'] = 'nokia-node'
    namingws['F2'] = 'system-ip'
    namingws['G2'] = 'nokia-bdr-port'
    namingws['H2'] = 'iptn-port-description'
    namingws['I2'] = 'nokia-bdr-port-description'
    namingws['J2'] = 'length (max 160)'
    namingws['K2'] = 'nokia-bdr-LAG'
    namingws['L2'] = 'iptn-port-encapsulation'
    namingws['M2'] = 'nokia-bdr-port-encapsulation'
    namingws['N2'] = 'length (max 160)'
    namingws['O2'] = 'iptn-service-name'
    namingws['P2'] = 'bdr-service-name'
    namingws['Q2'] = 'length (max 64)'
    namingws['R2'] = 'nokia-bdr-VLAN'
    namingws['S2'] = 'iptn-unit-description'
    namingws['T2'] = 'nokia-bdr-SAP-description(for L2 service)'
    namingws['U2'] = 'length (max 160)'
    namingws['V2'] = 'L3-interface-name'
    namingws['W2'] = 'length (max 32)'
    namingws['X2'] = 'iptn-L3-unit-description'
    namingws['Y2'] = 'nokia-bdr-L3-int-description'
    namingws['Z2'] = 'length (max 160)'
    namingws['AA2'] = 'vrrp_via_switch(yes)'
    namingws.auto_filter.ref = 'A2:AA2'
    namingws.freeze_panes = namingws['A3']
    namingws.column_dimensions['A'].width = 25.0
    namingws.column_dimensions['B'].width = 15.0
    namingws.column_dimensions['C'].width = 20.0
    namingws.column_dimensions['D'].width = 10.0
    namingws.column_dimensions['E'].width = 20.0
    namingws.column_dimensions['F'].width = 15.0
    namingws.column_dimensions['G'].width = 16.0
    namingws.column_dimensions['H'].width = 50.0
    namingws.column_dimensions['I'].width = 50.0
    namingws.column_dimensions['J'].width = 16.0
    namingws.column_dimensions['K'].width = 16.0
    namingws.column_dimensions['L'].width = 25.0
    namingws.column_dimensions['M'].width = 30.0
    namingws.column_dimensions['N'].width = 16.0
    namingws.column_dimensions['O'].width = 70.0
    namingws.column_dimensions['P'].width = 70.0
    namingws.column_dimensions['Q'].width = 16.0
    namingws.column_dimensions['R'].width = 16.0
    namingws.column_dimensions['S'].width = 50.0
    namingws.column_dimensions['T'].width = 50.0
    namingws.column_dimensions['U'].width = 16.0
    namingws.column_dimensions['V'].width = 35.0
    namingws.column_dimensions['W'].width = 16.0
    namingws.column_dimensions['X'].width = 50.0
    namingws.column_dimensions['Y'].width = 50.0
    namingws.column_dimensions['Z'].width = 16.0
    namingws.column_dimensions['AA'].width = 22.0
    return namingwb,namingws


def writeinfo(export,exvsw,exbd, namingws,iptnnode,iptnport,bdrnode,bdrport,pr,namingrow,startportrow, endportrow,exvpls):
    nodenum = iptnnode[-1]
    iptnservicename = export['V%s' % pr].value
    iptnunitdesc = export['N%s' % pr].value
    iptnportdesc = export['F%s' % pr].value
    namingws['A%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['B%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['C%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['D%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['H%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['L%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['O%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['S%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['X%s' % namingrow].fill = PatternFill(start_color='FFe0e0e0',end_color='FFe0e0e0',fill_type='solid')
    namingws['G%s' % namingrow].fill = PatternFill(start_color='FFffffe0',end_color='FFffffe0',fill_type='solid')
    namingws['K%s' % namingrow].fill = PatternFill(start_color='FFffffe0',end_color='FFffffe0',fill_type='solid')
    namingws['P%s' % namingrow].fill = PatternFill(start_color='FFffffe0',end_color='FFffffe0',fill_type='solid')
    namingws['V%s' % namingrow].fill = PatternFill(start_color='FFffffe0',end_color='FFffffe0',fill_type='solid')
    namingws['A%s' % namingrow] = iptnnode  # 'iptn-node'
    namingws['B%s' % namingrow] = export['B%s' % pr].value  # 'iptn-loopback'
    if export['M%s' % pr].value == None: # if don't have unit information. that means this port might be member of lag
        namingws['C%s' % namingrow] = export['C%s' % pr].value  # 'iptn-interface.unit'
        namingws['L%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['M%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['O%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['P%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['R%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['S%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['T%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['V%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['V%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['X%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        namingws['Y%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
    else:
        namingws['C%s' % namingrow] = '%s.%s' % (export['C%s' % pr].value, export['M%s' % pr].value)  # 'iptn-interface.unit'
        namingws['R%s' % namingrow] = export['M%s' % pr].value  # 'nokia-bdr-VLAN'
    namingws['D%s' % namingrow] = export['G%s' % pr].value  # 'iptn-ae'
    namingws['E%s' % namingrow] = bdrnode #'nokia-node'
    namingws['F%s' % namingrow] = export['Y%s' % pr].value #'system-ip'
    namingws['G%s' % namingrow] = bdrport  # 'nokia-bdr-port'
    namingws['H%s' % namingrow] = iptnportdesc #'iptn-port-description'
    if iptnportdesc != None:
        bdrdesc = iptnportdesc.strip('"')
        bdrdesc = bdrdesc.replace('_PE03','_PE01').replace('_PE05','_PE01').replace('_PE07','_PE01').replace('_PE09','_PE01')
        bdrdesc = bdrdesc.replace('_PE04','_PE02').replace('_PE06','_PE02').replace('_PE08','_PE02').replace('_PE10','_PE02')
        bdrdesc = bdrdesc.replace('_pe03','_pe01').replace('_pe05','_pe01').replace('_pe07','_pe01').replace('_pe09','_pe01')
        bdrdesc = bdrdesc.replace('_pe04','_pe02').replace('_pe06','_pe02').replace('_pe08','_pe02').replace('_pe10','_pe02')
        if '_SW' not in bdrdesc:
            bdrdesc = bdrdesc.replace('IPTNJ','BDRT')
            if 'IPTN_NMS' not in bdrdesc:
                bdrdesc = bdrdesc.replace('IPTN','BDRT')
        namingws['I%s' % namingrow] = bdrdesc #'nokia-bdr-port-description'
    namingws['J%s' % namingrow] = '=len(I%s)' % namingrow
    if export['G%s' % pr].value != None: # if existing have lag id
        if 'ae' in export['G%s' % pr].value:
            #print(export['G%s' % pr].value.split('ae')[1])
            if nodenum == '1' or nodenum == '2':
                if iptnnode == 'RONE_IPTNJ_MX960_PE01' or iptnnode == 'RONE_IPTNJ_MX960_PE02':
                    namingws['K%s' % namingrow] = 'lag-%s' % (600 + int(export['G%s' % pr].value.split('ae')[1]))  # 'nokia-bdr-LAG'
                else:
                    namingws['K%s' % namingrow] = 'lag-%s' % (300 + int(export['G%s' % pr].value.split('ae')[1]))  # 'nokia-bdr-LAG'
            if nodenum == '3' or nodenum == '4':
                namingws['K%s' % namingrow] = 'lag-%s' % (400 + int(export['G%s' % pr].value.split('ae')[1]))  # 'nokia-bdr-LAG'
            if nodenum == '5' or nodenum == '6':
                namingws['K%s' % namingrow] = 'lag-%s' % (500 + int(export['G%s' % pr].value.split('ae')[1]))  # 'nokia-bdr-LAG'
            if nodenum == '7' or nodenum == '8':
                namingws['K%s' % namingrow] = 'lag-%s' % (200 + int(export['G%s' % pr].value.split('ae')[1]))  # 'nokia-bdr-LAG'
    namingws['L%s' % namingrow] = export['E%s' % pr].value #'iptn-port-encapsulation'
    if export['W%s' % pr].value == 'VPRN':
        namingws['M%s' % namingrow] = 'dot1q'  # 'nokia-bdr-port-encapsulation'
    elif export['E%s' % pr].value != None:
        if re.search('ethernet-ccc',export['E%s' % pr].value):
            namingws['M%s' % namingrow] = 'null'  # 'nokia-bdr-port-encapsulation'
        elif re.search('ethernet-bridge', export['E%s' % pr].value):
            namingws['M%s' % namingrow] = 'null'  # 'nokia-bdr-port-encapsulation'
        elif re.search('ethernet-vpls', export['E%s' % pr].value):
            namingws['M%s' % namingrow] = 'null'  # 'nokia-bdr-port-encapsulation'
        elif re.search('access', export['E%s' % pr].value):
            namingws['M%s' % namingrow] = 'null'  # 'nokia-bdr-port-encapsulation'
        else:
            namingws['M%s' % namingrow] = 'dot1q'  # 'nokia-bdr-port-encapsulation'
    namingws['N%s' % namingrow] = '=len(M%s)' % namingrow
    namingws['O%s' % namingrow] = iptnservicename #'iptn-service-name'
    namingws['Q%s' % namingrow] = '=len(P%s)' % namingrow
    if export['W%s' % pr].value != 'VPRN':
        if iptnunitdesc != None: # if existing port have unit description
            bdrsapdesc = iptnunitdesc.strip('"')
            bdrsapdesc = bdrsapdesc.replace('_PE03','_PE01').replace('_PE05','_PE01').replace('_PE07','_PE01').replace('_PE09','_PE01')
            bdrsapdesc = bdrsapdesc.replace('_PE04','_PE02').replace('_PE06','_PE02').replace('_PE08','_PE02').replace('_PE10','_PE02')
            bdrsapdesc = bdrsapdesc.replace('_pe03','_pe01').replace('_pe05','_pe01').replace('_pe07','_pe01').replace('_pe09','_pe01')
            bdrsapdesc = bdrsapdesc.replace('_pe04','_pe02').replace('_pe06','_pe02').replace('_pe08','_pe02').replace('_pe10','_pe02')
            if '_SW' not in bdrsapdesc:
                bdrsapdesc = bdrsapdesc.replace('IPTNJ','BDRT')
                if 'IPTN_NMS' not in bdrsapdesc:
                    bdrsapdesc = bdrsapdesc.replace('IPTN','BDRT')
            namingws['S%s' % namingrow] = iptnunitdesc  # 'iptn-unit-description'
            namingws['T%s' % namingrow] = bdrsapdesc #'nokia-bdr-SAP-description(for L2 service)'
        elif export['M%s' % pr].value != None: # if existing port don't have unit description. will be use physical port description instead
            bdrsapdesc = iptnportdesc.strip('"')
            bdrsapdesc = bdrsapdesc.replace('_PE03','_PE01').replace('_PE05','_PE01').replace('_PE07','_PE01').replace('_PE09','_PE01')
            bdrsapdesc = bdrsapdesc.replace('_PE04','_PE02').replace('_PE06','_PE02').replace('_PE08','_PE02').replace('_PE10','_PE02')
            bdrsapdesc = bdrsapdesc.replace('_pe03','_pe01').replace('_pe05','_pe01').replace('_pe07','_pe01').replace('_pe09','_pe01')
            bdrsapdesc = bdrsapdesc.replace('_pe04','_pe02').replace('_pe06','_pe02').replace('_pe08','_pe02').replace('_pe10','_pe02')
            if '_SW' not in bdrsapdesc:
                bdrsapdesc = bdrsapdesc.replace('IPTNJ','BDRT')
                if 'IPTN_NMS' not in bdrsapdesc:
                    bdrsapdesc = bdrsapdesc.replace('IPTN','BDRT')
            namingws['S%s' % namingrow] = iptnportdesc  # 'iptn-unit-description'
            namingws['T%s' % namingrow] = bdrsapdesc #'nokia-bdr-SAP-description(for L2 service)'
        if export['W%s' % pr].value == 'EPIPE':
            iptnservicename = iptnservicename.replace('_PE03','_PE01').replace('_PE05','_PE01').replace('_PE07','_PE01').replace('_PE09','_PE01')
            iptnservicename = iptnservicename.replace('_PE04','_PE02').replace('_PE06','_PE02').replace('_PE08','_PE02').replace('_PE10','_PE02')
            iptnservicename = iptnservicename.replace('_pe03','_pe01').replace('_pe05','_pe01').replace('_pe07','_pe01').replace('_pe09','_pe01')
            iptnservicename = iptnservicename.replace('_pe04','_pe02').replace('_pe06','_pe02').replace('_pe08','_pe02').replace('_pe10','_pe02')
            iptnservicename = iptnservicename.replace('IPTNJ','BDRT')
            iptnservicename = iptnservicename.replace('IPTN','BDRT')
            if re.match('L2Circuit',iptnservicename):
                namingws['P%s' % namingrow] = 'L2_%s' % iptnservicename.split('L2Circuit_')[-1]
            elif re.match('L2circuit',iptnservicename):
                namingws['P%s' % namingrow] = 'L2_%s' % iptnservicename.split('L2circuit_')[-1]
            elif re.match('L2_Circuit',iptnservicename):
                namingws['P%s' % namingrow] = 'L2_%s' % iptnservicename.split('L2_Circuit_')[-1]
            elif re.match('L2_4Circuit',iptnservicename):
                namingws['P%s' % namingrow] = 'L2_%s' % iptnservicename.split('L2_4Circuit_')[-1]
            elif re.match('L2_4PWE',iptnservicename):
                namingws['P%s' % namingrow] = 'L2_%s' % iptnservicename.split('L2_4PWE_')[-1]
            elif re.match('pwe',iptnservicename):
                namingws['P%s' % namingrow] = 'L2_%s' % iptnservicename.split('pwe-')[-1]
            else:
                namingws['P%s' % namingrow] = 'L2_%s' % iptnservicename
                #print(namingws['P%s' % namingrow].value)
            namingws['V%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
            namingws['X%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
            namingws['Y%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
        elif export['W%s' % pr].value == 'VPLS':
            for vplsrow in range(3,exvpls.max_row + 1):
                if iptnnode == exvpls['A%s' %vplsrow].value:
                    if (export['C%s' % pr].value == exvpls['I%s'%vplsrow].value) and (export['M%s' % pr].value == exvpls['J%s'%vplsrow].value):
                        if exvpls['P%s' % vplsrow].value == None:
                            if re.match('VPLS_',iptnservicename):
                                iptnservicename.replace('IPTNJ','BDRT')
                                namingws['P%s' % namingrow] = iptnservicename
                            else:
                                namingws['P%s' % namingrow] = 'VPLS_%s' % iptnservicename
                            namingws['V%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
                            namingws['X%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
                            namingws['Y%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
                        else:
                            if re.match('VPLS_',iptnservicename):
                                iptnservicename.replace('IPTNJ', 'BDRT')
                                namingws['P%s' % namingrow] = 'r%s' % iptnservicename
                            else:
                                namingws['P%s' % namingrow] = 'rVPLS_%s' % iptnservicename
                            for i in range(startportrow, endportrow):
                                namingws['X%s' % namingrow] = export['N%s' % i].value  # 'iptn-L3-unit-description'
                                if (export['C%s' % i].value == exvpls['P%s' % vplsrow].value.split('.')[0]) and (export['M%s' % i].value == exvpls['P%s' % vplsrow].value.split('.')[-1]):
                                    if export['N%s' % i].value != None:  # if unit description have value
                                        namingws['V%s' % namingrow] = 'Intf_%s' % export['N%s' % i].value.strip('"').replace(' ', '_').replace('Vlan','V')  # 'L3-interface-name'
                                        namingws['Y%s' % namingrow] = export['N%s' % i].value.strip('"')  # 'nokia-bdr-L3-int-description'
                                    elif iptnportdesc != None:  # if existing port don't have unit description. will be use unit description of port instead
                                        namingws['V%s' % namingrow] = 'Intf_%s' % iptnportdesc.strip('"').replace(' ', '_').replace('Vlan','V')  # 'L3-interface-name'
                                        namingws['Y%s' % namingrow] = iptnportdesc.strip('"')  # 'nokia-bdr-L3-int-description'
        elif export['W%s' % pr].value == 'Bridge-Domain':
            for bdrow in range(3,exbd.max_row + 1):
                if iptnnode == exbd['A%s' %bdrow].value:
                    if (export['C%s' % pr].value == exbd['H%s'%bdrow].value) and (export['M%s' % pr].value == exbd['I%s'%bdrow].value):
                        if exbd['F%s'%bdrow].value == None:
                            namingws['P%s' % namingrow] = 'VPLS_%s' % iptnservicename
                            namingws['V%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
                            namingws['X%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
                            namingws['Y%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
                        elif re.search('irb',exbd['F%s'%bdrow].value):
                            namingws['P%s' % namingrow] = 'rVPLS_%s' % iptnservicename
                            for i in range(startportrow, endportrow):
                                if (export['C%s' % i].value == 'irb') and (export['M%s' % i].value == exbd['F%s'%bdrow].value.split('.')[-1]):
                                    namingws['X%s' % namingrow] = export['N%s' % i].value  # 'iptn-L3-unit-description'
                                    if export['N%s' % i].value != None:  # if unit description have value
                                        namingws['V%s' % namingrow] = 'Intf_%s' % export['N%s' % i].value.strip('"').replace(' ', '_').replace('Vlan','V')  # 'L3-interface-name'
                                        namingws['Y%s' % namingrow] = export['N%s' % i].value.strip('"')  # 'nokia-bdr-L3-int-description'
                                    elif iptnportdesc != None:  # if existing port don't have unit description. will be use unit description of port instead
                                        namingws['V%s' % namingrow] = 'Intf_%s' % iptnportdesc.strip('"').replace(' ', '_').replace('Vlan','V')  # 'L3-interface-name'
                                        namingws['Y%s' % namingrow] = iptnportdesc.strip('"')  # 'nokia-bdr-L3-int-description'
                        #print(namingws['P%s' % namingrow].value)
                        #print(exbd['H%s'%bdrow].value)
        elif export['W%s' % pr].value == 'Virtual-Switch':
            for vswrow in range(3,exvsw.max_row+1):
                if iptnnode == exvsw['A%s' % vswrow].value:
                    #print('port %s : vsw %s' % (export['C%s' % pr].value, exvsw['K%s' % vswrow].value))
                    if export['C%s' % pr].value == exvsw['K%s' % vswrow].value:
                        if export['M%s' % pr].value == exvsw['L%s' % vswrow].value:
                            namingws['P%s' % namingrow] = 'rVPLS_%s' % exvsw['I%s' % vswrow].value
                            namingws['R%s' % namingrow] = exvsw['O%s' % vswrow].value  # 'nokia-bdr-VLAN'
                            for i in range(startportrow, endportrow):
                                if (export['C%s' % i].value == 'irb') and (export['M%s' % i].value == exvsw['L%s' % vswrow].value):
                                    namingws['X%s' % namingrow] = export['N%s' % i].value  # 'iptn-L3-unit-description'
                                    if export['N%s' % i].value != None:  # if unit description have value
                                        namingws['V%s' % namingrow] = 'Intf_%s' % export['N%s' % i].value.strip('"').replace(' ', '_').replace('Vlan','V')  # 'L3-interface-name'
                                        namingws['Y%s' % namingrow] = export['N%s' % i].value.strip('"')  # 'nokia-bdr-L3-int-description'
                                    elif iptnportdesc != None:  # if existing port don't have unit description. will be use unit description of port instead
                                        namingws['V%s' % namingrow] = 'Intf_%s' % iptnportdesc.strip('"').replace(' ', '_').replace('Vlan','V')  # 'L3-interface-name'
                                        namingws['Y%s' % namingrow] = iptnportdesc.strip('"')  # 'nokia-bdr-L3-int-description'
                                elif (export['C%s' % i].value == exvsw['Q%s' % vswrow].value.split('.')[0]) and (export['M%s' % i].value == exvsw['Q%s' % vswrow].value.split('.')[1]):
                                    namingws['X%s' % namingrow] = export['N%s' % i].value  # 'iptn-L3-unit-description'
                                    if export['N%s' % i].value != None:  # if unit description have value
                                        namingws['V%s' % namingrow] = 'Intf_%s' % export['N%s' % i].value.strip('"').replace(' ', '_').replace('Vlan','V')  # 'L3-interface-name'
                                        namingws['Y%s' % namingrow] = export['N%s' % i].value.strip('"')  # 'nokia-bdr-L3-int-description'
                                    elif iptnportdesc != None:  # if existing port don't have unit description. will be use unit description of port instead
                                        namingws['V%s' % namingrow] = 'Intf_%s' % iptnportdesc.strip('"').replace(' ', '_').replace('Vlan','V')  # 'L3-interface-name'
                                        namingws['Y%s' % namingrow] = iptnportdesc.strip('"')  # 'nokia-bdr-L3-int-description'
    else: # if it not L2 service. will be VPRN service
        namingws['P%s' % namingrow] = 'L3_%s' % iptnservicename  # 'bdr-service-name'
        namingws['X%s' % namingrow] = iptnunitdesc  # 'iptn-L3-unit-description'
        if iptnunitdesc != None: # if unit description have value
            namingws['V%s' % namingrow] = 'Intf_%s' % iptnunitdesc.strip('"').replace(' ','_').replace('Vlan','V')  # 'L3-interface-name'
            namingws['Y%s' % namingrow] = iptnunitdesc.strip('"')  # 'nokia-bdr-L3-int-description'
            namingws['S%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
            namingws['T%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
        elif export['M%s' % pr].value != None:  # if existing port don't have unit description. will be use physical port description instead
            namingws['V%s' % namingrow] = 'Intf_%s' % iptnportdesc.strip('"').replace(' ','_').replace('Vlan','V')  # 'L3-interface-name'
            namingws['Y%s' % namingrow] = iptnportdesc.strip('"')  # 'nokia-bdr-L3-int-description'
            namingws['S%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
            namingws['T%s' % namingrow].fill = PatternFill(start_color='FF565656', end_color='FF565656',fill_type='solid')
    namingws['U%s' % namingrow] = '=len(T%s)' % namingrow
    namingws['W%s' % namingrow] = '=len(V%s)' % namingrow
    namingws['Z%s' % namingrow] = '=len(Y%s)' % namingrow



def readinputfile():
    mapsheet = openpyxl.load_workbook(inputpath + wbmap)['node-map']
    namingwb,namingws = createnokianamingfile()
    dummyport = []
    laglist = []
    lastiptnnode = ''
    if not os.path.exists(migratepath + inputpath + inputname):
        createinputfile()
    input = openpyxl.load_workbook(filename=migratepath + inputpath + inputname)
    inputws = input['input']
    max_port_row = 1
    for cell in inputws['B']:
        if cell.value != None:
            max_port_row += 1


    le = open(extractpath+'latest.txt','r')
    latestextract = le.readlines()[0]
    #print(latestextract)
    extr = openpyxl.load_workbook(filename=extractpath + latestextract)
    exindex = extr[wsindex]
    export = extr[wsport]
    exbd = extr[wsbridge]
    exvsw = extr[wsvsw]
    exvpls = extr[wsvpls]


    for a in ascii_lowercase: # create dummy port id
        if a == 'c':
            a = '@'
        for b in ascii_lowercase:
            if b == 'c':
                b = '@'
            dummyport.append('1/%s/%s' %(a,b))

    p = 0
    namingrow = namingws['A1'].value
    for row in range(3,max_port_row + 2):
        startportrow = 0
        endportrow = 0
        iptnnode = inputws['A%s' % row].value
        iptnport = inputws['B%s' % row].value
        if inputws['A%s' % row].value == None:
            iptnnode = 'blank'
            iptnport = 'blank'
            bdrnode = 'blank'
            bdrport = 'blank'

        if inputws['C%s' % row].value == None: # if not specify bdr-node. will be use bdr-node from nodemapping file
            for x in range(3, (mapsheet.max_row + 1)):
                if iptnnode:
                    if re.match(iptnnode, mapsheet['A%s' % x].value):
                        bdrnode = mapsheet['C%s' % x].value

        else:
            bdrnode = inputws['C%s' % row].value
        if inputws['D%s' % row].value == None: # if not specify bdr-port. will be use dummy port instead
            bdrport = dummyport[p]
            p += 1
        else:
            bdrport = inputws['D%s' % row].value

        for i in range(3, (exindex.max_row + 1)): # search iptnnode in index sheet of extraction file to get start and end row of port sheet
            if iptnnode == exindex['B%s'%i].value:
                startportrow = exindex['C%s'%i].value
                endportrow = exindex['D%s'%i].value+1
        if iptnnode != lastiptnnode and lastiptnnode != '':
            #print('new')
            #print(laglist)
            lagnum = 0
            for iptnae in laglist:
                #print(iptnae)
                for lpr in range(laststartportrow,lastendportrow):
                    if lastiptnnode == export['A%s' % lpr].value:
                        nodenum = lastiptnnode[-1]
                        if iptnae == export['C%s' % lpr].value:
                            if nodenum == '1' or nodenum == '2':
                                if lastiptnnode == 'RONE_IPTNJ_MX960_PE01' or lastiptnnode == 'RONE_IPTNJ_MX960_PE02':
                                    bdrlag = 'lag-%s' % (600 + int(iptnae.split('ae')[1]))
                                else:
                                    bdrlag = 'lag-%s' % (300 + int(iptnae.split('ae')[1]))
                            if nodenum == '3' or nodenum == '4':
                                bdrlag = 'lag-%s' % (400 + int(iptnae.split('ae')[1]))
                            if nodenum == '5' or nodenum == '6':
                                bdrlag = 'lag-%s' % (500 + int(iptnae.split('ae')[1]))
                            if nodenum == '7' or nodenum == '8':
                                bdrlag = 'lag-%s' % (200 + int(iptnae.split('ae')[1]))
                            writeinfo(export,exvsw,exbd, namingws, lastiptnnode, iptnae, lastbdrnode, bdrlag, lpr, namingrow,laststartportrow, lastendportrow,exvpls)
                            namingrow += 1
            laglist = []
        for pr in range(startportrow,endportrow): # search iptnnode and iptnport in port sheet
            if iptnnode == export['A%s' % pr].value:
                if iptnport == export['C%s' % pr].value:
                    #print('iptn-node : %s # iptn-port : %s' % (export['A%s' % pr].value, export['C%s' % pr].value))
                    #print('nokia-node : %s # nokia-port : %s' %(bdrnode,bdrport))
                    writeinfo(export,exvsw,exbd,namingws,iptnnode,iptnport,bdrnode,bdrport,pr,namingrow,startportrow, endportrow,exvpls)
                    namingrow += 1
                    lastiptnnode = iptnnode
                    laststartportrow = startportrow
                    lastendportrow = endportrow
                    lastbdrnode = bdrnode
                    if export['M%s' % pr].value == None and export['G%s' % pr].value:
                        #print(export['G%s' % pr].value)
                        if export['G%s' % pr].value not in laglist:
                            laglist.append(export['G%s' % pr].value)
                            #print(laglist)

        namingws['B1'] = t.strftime("%Y%m%d-%H%M")
        namingws['A1'] = namingrow
        namingwb.save(migratepath + inputpath + namingname)
        if os.path.exists(migratepath + inputpath + 'latestnaming.txt'):  # create text file for specific latest workbook
            os.remove(migratepath + inputpath + 'latestnaming.txt')
        f = open(migratepath + inputpath + "latestnaming.txt", "w")
        f.write(namingname)
        f.close()
        #print('iptn-node : %s # iptn-port : %s # bdr-node : %s # bdr-port : %s' %(iptnnode,iptnport,bdrnode,bdrport))

    #createinputfile()

def main():
    if not os.path.exists(migratepath):
        os.mkdir(migratepath)
    if not os.path.exists(migratepath+inputpath):
        os.mkdir(migratepath+inputpath)
    #if not os.path.exists(migratepath+t.strftime("%Y%m%d-%H%M")+'\\'):
        #os.mkdir(migratepath+t.strftime("%Y%m%d-%H%M"))
    readinputfile()


if __name__ == "__main__":
    main()