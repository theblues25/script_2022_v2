
import sys
import os
import re
import datetime
import time
import openpyxl
from openpyxl.styles import PatternFill
from string import ascii_lowercase
from collections import Counter
from netaddr import IPNetwork, IPAddress

t = datetime.datetime.now()
inputname = 'migration-input.xlsx'
sdpname = 'SDP number.xlsm'
wbmap = 'nodemapping.xlsx'
summary = 'summary'
misumname = 'migration-summary_%s.xlsx' %t.strftime("%Y%m%d-%H%M")
cfgpath = '.\iptn-cfg\\'
extractpath = '.\extraction\\'
inputpath = '.\input\\'
attrpath = 'attr-cfg\\'
migratepath = '.\migration\\'
wsindex = 'index'
wsport = 'port-lag'
wsvprn = 'vprn'
wsvprnint = 'vprn-int'
wsl2vpn = 'l2vpn'
wsvpls = 'vpls'
wsbridge = 'bridge-domain'
wsvsw = 'virtual-switch'
# open sdp xls file
sdpmapwb = openpyxl.load_workbook(filename=inputpath + sdpname)
sdpmapws = sdpmapwb['Nokia NE Names']
# open extraction xls file
le = open(extractpath + 'latest.txt', 'r')
latestextract = le.readlines()[0]
extr = openpyxl.load_workbook(filename=extractpath + latestextract)
exindex = extr[wsindex]
export = extr[wsport]
exvprn = extr[wsvprn]
exvprnint = extr[wsvprnint]
exl2vpn = extr[wsl2vpn]
exvpls = extr[wsvpls]
exbridge = extr[wsbridge]
exvsw = extr[wsvsw]
l2mapsheet = openpyxl.load_workbook(inputpath + wbmap)['node-l2-peer']
asbrmapws = openpyxl.load_workbook(inputpath + wbmap)['inter-as']
communitysheet = openpyxl.load_workbook(inputpath + wbmap)['route-target']
customerws = openpyxl.load_workbook(inputpath + wbmap)['customer']
policyprefixduplist = []
policyprefixdupdict = {}
for filename in os.listdir(inputpath):
    if 'AIS BDRT_ASBR-IP&VLAN' in filename:
        asbrvplsvlansheet = openpyxl.load_workbook(inputpath + filename)['VPLS']
        asbrepipevlansheet = openpyxl.load_workbook(inputpath + filename)['Epipe']


def createnokiasummaryfile():
    row = 3
    misum = openpyxl.Workbook()
    misum.remove(misum['Sheet'])
    sumws = misum.create_sheet(summary)
    sumws['A1'] = row
    sumws['A2'] = 'nokia-node'
    sumws['B2'] = 'system-ip'
    sumws['C2'] = 'service-name'
    sumws['D2'] = 'service-id'
    sumws['E2'] = 'service-type'
    sumws['F2'] = 'port'
    sumws['G2'] = 'port-state'
    sumws['H2'] = 'LAG'
    sumws['I2'] = 'SAP'
    sumws['J2'] = 'SAP-state'
    sumws['K2'] = 'vprn-interface-name'
    sumws['L2'] = 'iptn-node'
    sumws['M2'] = 'iptn-loopback'
    sumws['N2'] = 'iptn-int'
    sumws['O2'] = 'iptn-service-type'
    sumws['P2'] = 'iptn-service-name'
    sumws['Q2'] = 'iptn-service-rd'
    sumws.auto_filter.ref = 'A2:Q2'
    sumws.freeze_panes = sumws['A3']
    sumws.column_dimensions['A'].width = 17.0
    sumws.column_dimensions['B'].width = 12.0
    sumws.column_dimensions['C'].width = 64.0
    sumws.column_dimensions['D'].width = 12.0
    sumws.column_dimensions['E'].width = 14.0
    sumws.column_dimensions['F'].width = 10.0
    sumws.column_dimensions['G'].width = 12.0
    sumws.column_dimensions['H'].width = 10.0
    sumws.column_dimensions['I'].width = 12.0
    sumws.column_dimensions['J'].width = 12.0
    sumws.column_dimensions['K'].width = 35.0
    sumws.column_dimensions['L'].width = 25.0
    sumws.column_dimensions['M'].width = 15.0
    sumws.column_dimensions['N'].width = 12.0
    sumws.column_dimensions['O'].width = 18.0
    sumws.column_dimensions['P'].width = 60.0
    sumws.column_dimensions['Q'].width = 14.0

    filternamews = misum.create_sheet('filter-naming')
    filternamews['A1'] = row
    filternamews['A2'] = 'nokia-node'
    filternamews['B2'] = 'service-type'
    filternamews['C2'] = 'service-name'
    filternamews['D2'] = 'iptn-filter-name'
    filternamews['E2'] = 'bdr-filter-name'
    filternamews['F2'] = 'length(max 64)'
    filternamews['G2'] = 'entry'
    filternamews['H2'] = 'existing-prefix-name'
    filternamews['I2'] = 'bdr-prefix-name'
    filternamews['J2'] = 'length(max 32)'
    filternamews['K2'] = 'existing-port-name'
    filternamews['L2'] = 'bdr-port-name'
    filternamews['M2'] = 'length(max 32)'
    filternamews.auto_filter.ref = 'A2:M2'
    filternamews.freeze_panes = filternamews['A3']
    filternamews.column_dimensions['A'].width = 17.0
    filternamews.column_dimensions['B'].width = 14.0
    filternamews.column_dimensions['C'].width = 56.0
    filternamews.column_dimensions['D'].width = 64.0
    filternamews.column_dimensions['E'].width = 64.0
    filternamews.column_dimensions['F'].width = 16.0
    filternamews.column_dimensions['G'].width = 16.0
    filternamews.column_dimensions['H'].width = 32.0
    filternamews.column_dimensions['I'].width = 32.0
    filternamews.column_dimensions['J'].width = 16.0
    filternamews.column_dimensions['K'].width = 32.0
    filternamews.column_dimensions['L'].width = 32.0
    filternamews.column_dimensions['M'].width = 16.0


    policynamews = misum.create_sheet('policy-naming')
    policynamews['A1'] = row
    policynamews['A2'] = 'nokia-node'
    policynamews['B2'] = 'service-type'
    policynamews['C2'] = 'service-name'
    policynamews['D2'] = 'iptn-policy-name'
    policynamews['E2'] = 'bdr-policy-name'
    policynamews['F2'] = 'length(max 64)'
    policynamews['G2'] = 'iptn-term-name'
    policynamews['H2'] = 'bdr-entry-name'
    policynamews['I2'] = 'length(max 255)'
    policynamews['J2'] = 'existing-prefix-list-name'
    policynamews['K2'] = 'bdr-prefix-list-name'
    policynamews['L2'] = 'length(max 32)'
    policynamews.auto_filter.ref = 'A2:L2'
    policynamews.freeze_panes = policynamews['A3']
    policynamews.column_dimensions['A'].width = 17.0
    policynamews.column_dimensions['B'].width = 14.0
    policynamews.column_dimensions['C'].width = 56.0
    policynamews.column_dimensions['D'].width = 64.0
    policynamews.column_dimensions['E'].width = 64.0
    policynamews.column_dimensions['F'].width = 16.0
    policynamews.column_dimensions['G'].width = 32.0
    policynamews.column_dimensions['H'].width = 32.0
    policynamews.column_dimensions['I'].width = 16.0
    policynamews.column_dimensions['J'].width = 32.0
    policynamews.column_dimensions['K'].width = 32.0
    policynamews.column_dimensions['L'].width = 16.0

    nportws = misum.create_sheet('port-lag')
    nportws['A1'] = row
    nportws['A2'] = 'nokia-node'
    nportws['B2'] = 'system-ip'
    nportws['C2'] = 'service-type'
    nportws['D2'] = 'service-name'
    nportws['E2'] = 'service-id'
    nportws['F2'] = 'port/lag'
    nportws['G2'] = 'port-encapsulation'
    nportws['H2'] = 'admin-state'
    nportws['I2'] = 'physical-description'
    nportws['J2'] = 'speed'
    nportws['K2'] = 'auto-negotiation'
    nportws['L2'] = 'LAG'
    nportws['M2'] = 'LAG-protocol'
    nportws['N2'] = 'LAG-port-threshold'
    nportws['O2'] = 'SAP'
    nportws['P2'] = 'SAP-state'
    nportws['Q2'] = 'SAP-description'
    nportws['R2'] = 'input-filter-policy-name'
    nportws['S2'] = 'input-filter-configuration'
    nportws['T2'] = 'output-filter-policy-name'
    nportws['U2'] = 'output-filter-configuration'
    nportws['V2'] = 'sap-ingress qos'
    nportws['W2'] = 'sap-egress qos'
    nportws['X2'] = 'iptn-node'
    nportws['Y2'] = 'iptn-service-type'
    nportws['Z2'] = 'iptn-service'
    nportws['AA2'] = 'iptn-interface'
    nportws.auto_filter.ref = 'A2:AA2'
    nportws.freeze_panes = nportws['A3']
    nportws.column_dimensions['A'].width = 17.0
    nportws.column_dimensions['B'].width = 12.0
    nportws.column_dimensions['C'].width = 14.0
    nportws.column_dimensions['D'].width = 56.0
    nportws.column_dimensions['E'].width = 12.0
    nportws.column_dimensions['F'].width = 10.0
    nportws.column_dimensions['G'].width = 16.0
    nportws.column_dimensions['H'].width = 14.0
    nportws.column_dimensions['I'].width = 60.0
    nportws.column_dimensions['J'].width = 8.0
    nportws.column_dimensions['K'].width = 19.0
    nportws.column_dimensions['L'].width = 7.0
    nportws.column_dimensions['M'].width = 14.0
    nportws.column_dimensions['N'].width = 20.0
    nportws.column_dimensions['O'].width = 12.0
    nportws.column_dimensions['P'].width = 12.0
    nportws.column_dimensions['Q'].width = 60.0
    nportws.column_dimensions['R'].width = 40.0
    nportws.column_dimensions['S'].width = 60.0
    nportws.column_dimensions['T'].width = 40.0
    nportws.column_dimensions['U'].width = 60.0
    nportws.column_dimensions['V'].width = 17.0
    nportws.column_dimensions['W'].width = 17.0
    nportws.column_dimensions['X'].width = 25.0
    nportws.column_dimensions['Y'].width = 18.0
    nportws.column_dimensions['Z'].width = 14.0
    nportws.column_dimensions['AA'].width = 15.0

    l2ws = misum.create_sheet('l2service-related-pe')
    l2ws['A1'] = row
    l2ws['A2'] = 'iptn-node'
    l2ws['B2'] = 'iptn-service-type'
    l2ws['C2'] = 'iptn-service'
    l2ws['D2'] = 'iptn-route-target'
    l2ws['E2'] = 'iptn-site-preference'
    l2ws['F2'] = 'iptn-site-id'
    l2ws['G2'] = 'nokia-node'
    l2ws['H2'] = 'nokia-service-type'
    l2ws['I2'] = 'nokia-service-name'
    l2ws['J2'] = 'nokia-service-id'
    l2ws['K2'] = 'nokia-customer-id'
    l2ws.auto_filter.ref = 'A2:K2'
    l2ws.freeze_panes = l2ws['A3']
    l2ws.column_dimensions['A'].width = 25.0
    l2ws.column_dimensions['B'].width = 18.0
    l2ws.column_dimensions['C'].width = 60.0
    l2ws.column_dimensions['D'].width = 18.0
    l2ws.column_dimensions['E'].width = 20.0
    l2ws.column_dimensions['F'].width = 10.0
    l2ws.column_dimensions['G'].width = 17.0
    l2ws.column_dimensions['H'].width = 19.0
    l2ws.column_dimensions['I'].width = 60.0
    l2ws.column_dimensions['J'].width = 15.0
    l2ws.column_dimensions['K'].width = 20.0


    sdpws = misum.create_sheet('l2service-sdp')
    sdpws['A1'] = row
    sdpws['A2'] = 'nokia-node'
    sdpws['B2'] = 'service-type'
    sdpws['C2'] = 'service-name'
    sdpws['D2'] = 'service-id'
    sdpws['E2'] = 'farend-node'
    sdpws['F2'] = 'farend-ip'
    sdpws['G2'] = 'sdp-number'
    sdpws['H2'] = 'sdp-description'
    sdpws['I2'] = 'vc-id'
    sdpws['J2'] = 'customer-id'
    sdpws['K2'] = 'entropy'
    sdpws.auto_filter.ref = 'A2:K2'
    sdpws.freeze_panes = sdpws['A3']
    sdpws.column_dimensions['A'].width = 17.0
    sdpws.column_dimensions['B'].width = 14.0
    sdpws.column_dimensions['C'].width = 56.0
    sdpws.column_dimensions['D'].width = 12.0
    sdpws.column_dimensions['E'].width = 17.0
    sdpws.column_dimensions['F'].width = 13.0
    sdpws.column_dimensions['G'].width = 14.0
    sdpws.column_dimensions['H'].width = 25.0
    sdpws.column_dimensions['I'].width = 10.0
    sdpws.column_dimensions['J'].width = 14.0
    sdpws.column_dimensions['K'].width = 14.0


    l2evpnws = misum.create_sheet('l2-evpn')
    l2evpnws['A1'] = row
    l2evpnws['A2'] = 'nokia-node'
    l2evpnws['B2'] = 'system-ip'
    l2evpnws['C2'] = 'service-name'
    l2evpnws['D2'] = 'service-id'
    l2evpnws['E2'] = 'ethernet-segment'
    l2evpnws['F2'] = 'esi'
    l2evpnws['G2'] = 'preference'
    l2evpnws['H2'] = 'port/lag'
    l2evpnws['I2'] = 'vlan'
    l2evpnws['J2'] = 'local-ac-name'
    l2evpnws['K2'] = 'local-eth-tag'
    l2evpnws['L2'] = 'remote-ac-name'
    l2evpnws['M2'] = 'remote-eth-tag'
    l2evpnws['N2'] = 'evi'
    l2evpnws['O2'] = 'customer-id'
    l2evpnws['P2'] = 'entropy'
    l2evpnws.auto_filter.ref = 'A2:P2'
    l2evpnws.freeze_panes = l2evpnws['A3']
    l2evpnws.column_dimensions['A'].width = 18.0
    l2evpnws.column_dimensions['B'].width = 12.0
    l2evpnws.column_dimensions['C'].width = 56.0
    l2evpnws.column_dimensions['D'].width = 12.0
    l2evpnws.column_dimensions['E'].width = 32.0
    l2evpnws.column_dimensions['F'].width = 22.0
    l2evpnws.column_dimensions['G'].width = 13.0
    l2evpnws.column_dimensions['H'].width = 10.0
    l2evpnws.column_dimensions['I'].width = 7.0
    l2evpnws.column_dimensions['J'].width = 17.0
    l2evpnws.column_dimensions['K'].width = 14.0
    l2evpnws.column_dimensions['L'].width = 18.0
    l2evpnws.column_dimensions['M'].width = 17.0
    l2evpnws.column_dimensions['N'].width = 8.0
    l2evpnws.column_dimensions['O'].width = 14.0
    l2evpnws.column_dimensions['P'].width = 14.0


    l2interasws = misum.create_sheet('l2inter-as')
    l2interasws['A1'] = row
    l2interasws['A2'] = 'PE-node'
    l2interasws['B2'] = 'service-type'
    l2interasws['C2'] = 'service-name'
    l2interasws['D2'] = 'service-id'
    l2interasws['E2'] = 'PE-sdp-number'
    l2interasws['F2'] = 'PE-sdp-description'
    l2interasws['G2'] = 'asbr-node'
    l2interasws['H2'] = 'asbr-ip'
    l2interasws['I2'] = 'asbr-sdp-number'
    l2interasws['J2'] = 'asbr-sdp-description'
    l2interasws['K2'] = 'vc-id'
    l2interasws['L2'] = 'sdp-precedence'
    l2interasws['M2'] = 'endpoint-name'
    l2interasws['N2'] = 'inter-as-lag'
    l2interasws['O2'] = 'md-index'
    l2interasws['P2'] = 'md-name'
    l2interasws['Q2'] = 'ma-index'
    l2interasws['R2'] = 'bridge-id'
    l2interasws['S2'] = 'bridge-vlan'
    l2interasws['T2'] = 'local-mep-id'
    l2interasws['U2'] = 'remote-mep-id'
    l2interasws['V2'] = 'customer-id'
    l2interasws['W2'] = 'entropy'
    l2interasws.auto_filter.ref = 'A2:W2'
    l2interasws.freeze_panes = sdpws['A3']
    l2interasws.column_dimensions['A'].width = 17.0
    l2interasws.column_dimensions['B'].width = 14.0
    l2interasws.column_dimensions['C'].width = 64.0
    l2interasws.column_dimensions['D'].width = 12.0
    l2interasws.column_dimensions['E'].width = 17.0
    l2interasws.column_dimensions['F'].width = 25.0
    l2interasws.column_dimensions['G'].width = 17.0
    l2interasws.column_dimensions['H'].width = 12.0
    l2interasws.column_dimensions['I'].width = 18.0
    l2interasws.column_dimensions['J'].width = 24.0
    l2interasws.column_dimensions['K'].width = 8.0
    l2interasws.column_dimensions['L'].width = 17.0
    l2interasws.column_dimensions['M'].width = 17.0
    l2interasws.column_dimensions['N'].width = 13.0
    l2interasws.column_dimensions['O'].width = 12.0
    l2interasws.column_dimensions['P'].width = 12.0
    l2interasws.column_dimensions['Q'].width = 12.0
    l2interasws.column_dimensions['R'].width = 60.0
    l2interasws.column_dimensions['S'].width = 13.0
    l2interasws.column_dimensions['T'].width = 14.0
    l2interasws.column_dimensions['U'].width = 17.0
    l2interasws.column_dimensions['V'].width = 14.0
    l2interasws.column_dimensions['W'].width = 14.0

    l2evpninterws = misum.create_sheet('l2-evpn-inter-as')
    l2evpninterws['A1'] = row
    l2evpninterws['A2'] = 'nokia-node'
    l2evpninterws['B2'] = 'system-ip'
    l2evpninterws['C2'] = 'service-name'
    l2evpninterws['D2'] = 'service-id'
    l2evpninterws['E2'] = 'ethernet-segment'
    l2evpninterws['F2'] = 'esi'
    l2evpninterws['G2'] = 'preference'
    l2evpninterws['H2'] = 'port/lag'
    l2evpninterws['I2'] = 'vlan'
    l2evpninterws['J2'] = 'pe-ac-name'
    l2evpninterws['K2'] = 'pe-eth-tag'
    l2evpninterws['L2'] = 'asbr-ac-name'
    l2evpninterws['M2'] = 'asbr-eth-tag'
    l2evpninterws['N2'] = 'evi'
    l2evpninterws['O2'] = 'asbr-node'
    l2evpninterws['P2'] = 'asbr-ethernet-segment'
    l2evpninterws['Q2'] = 'asbr-esi'
    l2evpninterws['R2'] = 'asbr-preference'
    l2evpninterws['S2'] = 'inter-as-lag'
    l2evpninterws['T2'] = 'md-index'
    l2evpninterws['U2'] = 'md-name'
    l2evpninterws['V2'] = 'ma-index'
    l2evpninterws['W2'] = 'bridge-id'
    l2evpninterws['X2'] = 'bridge-vlan'
    l2evpninterws['Y2'] = 'local-mep-id'
    l2evpninterws['Z2'] = 'remote-mep-id'
    l2evpninterws['AA2'] = 'customer-id'
    l2evpninterws['AB2'] = 'entropy'
    l2evpninterws.auto_filter.ref = 'A2:AB2'
    l2evpninterws.freeze_panes = l2evpninterws['A3']
    l2evpninterws.column_dimensions['A'].width = 17.0
    l2evpninterws.column_dimensions['B'].width = 12.0
    l2evpninterws.column_dimensions['C'].width = 64.0
    l2evpninterws.column_dimensions['D'].width = 12.0
    l2evpninterws.column_dimensions['E'].width = 28.0
    l2evpninterws.column_dimensions['F'].width = 21.0
    l2evpninterws.column_dimensions['G'].width = 13.0
    l2evpninterws.column_dimensions['H'].width = 10.0
    l2evpninterws.column_dimensions['I'].width = 7.0
    l2evpninterws.column_dimensions['J'].width = 17.0
    l2evpninterws.column_dimensions['K'].width = 12.0
    l2evpninterws.column_dimensions['L'].width = 15.0
    l2evpninterws.column_dimensions['M'].width = 14.0
    l2evpninterws.column_dimensions['N'].width = 6.0
    l2evpninterws.column_dimensions['O'].width = 16.0
    l2evpninterws.column_dimensions['P'].width = 26.0
    l2evpninterws.column_dimensions['Q'].width = 21.0
    l2evpninterws.column_dimensions['R'].width = 17.0
    l2evpninterws.column_dimensions['S'].width = 13.0
    l2evpninterws.column_dimensions['T'].width = 12.0
    l2evpninterws.column_dimensions['U'].width = 12.0
    l2evpninterws.column_dimensions['V'].width = 12.0
    l2evpninterws.column_dimensions['W'].width = 60.0
    l2evpninterws.column_dimensions['X'].width = 13.0
    l2evpninterws.column_dimensions['Y'].width = 14.0
    l2evpninterws.column_dimensions['Z'].width = 17.0
    l2evpninterws.column_dimensions['AA'].width = 14.0
    l2evpninterws.column_dimensions['AB'].width = 14.0



    vprnws = misum.create_sheet('global-vprn')
    vprnws['A1'] = row
    vprnws['A2'] = 'nokia-node'
    vprnws['B2'] = 'system-ip'
    vprnws['C2'] = 'service-name'
    vprnws['D2'] = 'service-id'
    vprnws['E2'] = 'admin-state'
    vprnws['F2'] = 'rd'
    vprnws['G2'] = 'vrf-import-policy-name'
    vprnws['H2'] = 'vrf-import-policy'
    vprnws['I2'] = 'vrf-import-prefix'
    vprnws['J2'] = 'vrf-import-commu'
    vprnws['K2'] = 'vrf-export-policy-name'
    vprnws['L2'] = 'vrf-export-policy'
    vprnws['M2'] = 'vrf-export-prefix'
    vprnws['N2'] = 'vrf-export-commu'
    vprnws['O2'] = 'vrf-target'
    vprnws['P2'] = 'aggregate-route'
    vprnws['Q2'] = 'customer-id'
    vprnws['R2'] = 'iptn-node'
    vprnws['S2'] = 'iptn-service'
    vprnws.auto_filter.ref = 'A2:S2'
    vprnws.freeze_panes = vprnws['A3']
    vprnws.column_dimensions['A'].width = 18.0
    vprnws.column_dimensions['B'].width = 12.0
    vprnws.column_dimensions['C'].width = 40.0
    vprnws.column_dimensions['D'].width = 12.0
    vprnws.column_dimensions['E'].width = 14.0
    vprnws.column_dimensions['F'].width = 20.0
    vprnws.column_dimensions['G'].width = 25.0
    vprnws.column_dimensions['H'].width = 70.0
    vprnws.column_dimensions['I'].width = 70.0
    vprnws.column_dimensions['J'].width = 20.0
    vprnws.column_dimensions['K'].width = 25.0
    vprnws.column_dimensions['L'].width = 70.0
    vprnws.column_dimensions['M'].width = 70.0
    vprnws.column_dimensions['N'].width = 20.0
    vprnws.column_dimensions['O'].width = 20.0
    vprnws.column_dimensions['P'].width = 17.0
    vprnws.column_dimensions['Q'].width = 14.0
    vprnws.column_dimensions['R'].width = 27.0
    vprnws.column_dimensions['S'].width = 14.0

    vprnintws = misum.create_sheet('vprn-interface')
    vprnintws['A1']  = row
    vprnintws['A2']  = 'nokia-node'
    vprnintws['B2']  = 'system-ip'
    vprnintws['C2']  = 'service-name'
    vprnintws['D2']  = 'service-id'
    vprnintws['E2']  = 'interface-name'
    vprnintws['F2']  = 'interface description'
    vprnintws['G2']  = 'sap'
    vprnintws['H2']  = 'sap-ingress-filter'
    vprnintws['I2']  = 'sap-egress-filter'
    vprnintws['J2']  = 'sap-in/egress qos'
    vprnintws['K2']  = 'ip/ipv6 address'
    vprnintws['L2']  = 'secondary ip/ipv6'
    vprnintws['M2']  = 'family'
    vprnintws['N2']  = 'vrrp id'
    vprnintws['O2']  = 'vrrp vip'
    vprnintws['P2']  = 'vrrp priority'
    vprnintws['Q2']  = 'vrrp interval'
    vprnintws['R2']  = 'static-route'
    vprnintws['S2']  = 'bgp-features'
    vprnintws['T2']  = 'group'
    vprnintws['U2']  = 'peer-as'
    vprnintws['V2']  = 'neighbor'
    vprnintws['W2']  = 'neighbor-description'
    vprnintws['X2']  = 'bgp-imp-policy-name'
    vprnintws['Y2']  = 'bgp-imp-policy'
    vprnintws['Z2']  = 'bgp-imp-policy-prefix'
    vprnintws['AA2'] = 'bgp-imp-policy-commu'
    vprnintws['AB2'] = 'bgp-exp-policy-name'
    vprnintws['AC2'] = 'bgp-exp-policy'
    vprnintws['AD2'] = 'bgp-exp-policy-prefix'
    vprnintws['AE2'] = 'bgp-exp-policy-commu'
    vprnintws['AF2'] = 'bgp-bfd-interval'
    vprnintws['AG2'] = 'bgp-bfd-multiply'
    vprnintws['AH2'] = 'iptn-node'
    vprnintws['AI2'] = 'iptn-service'
    vprnintws['AJ2'] = 'iptn-interface'
    vprnintws.auto_filter.ref = 'A2:AJ2'
    vprnintws.freeze_panes = vprnintws['A3']
    vprnintws.column_dimensions['A'].width = 17.0
    vprnintws.column_dimensions['B'].width = 12.0
    vprnintws.column_dimensions['C'].width = 40.0
    vprnintws.column_dimensions['D'].width = 12.0
    vprnintws.column_dimensions['E'].width = 40.0
    vprnintws.column_dimensions['F'].width = 50.0
    vprnintws.column_dimensions['G'].width = 12.0
    vprnintws.column_dimensions['H'].width = 40.0
    vprnintws.column_dimensions['I'].width = 40.0
    vprnintws.column_dimensions['J'].width = 20.0
    vprnintws.column_dimensions['K'].width = 24.0
    vprnintws.column_dimensions['L'].width = 24.0
    vprnintws.column_dimensions['M'].width = 9.0
    vprnintws.column_dimensions['N'].width = 9.0
    vprnintws.column_dimensions['O'].width = 24.0
    vprnintws.column_dimensions['P'].width = 14.0
    vprnintws.column_dimensions['Q'].width = 14.0
    vprnintws.column_dimensions['R'].width = 90.0
    vprnintws.column_dimensions['S'].width = 10.0
    vprnintws.column_dimensions['T'].width = 16.0
    vprnintws.column_dimensions['U'].width = 10.0
    vprnintws.column_dimensions['V'].width = 14.0
    vprnintws.column_dimensions['W'].width = 40.0
    vprnintws.column_dimensions['X'].width = 27.0
    vprnintws.column_dimensions['Y'].width = 80.0
    vprnintws.column_dimensions['Z'].width = 80.0
    vprnintws.column_dimensions['AA'].width = 24.0
    vprnintws.column_dimensions['AB'].width = 27.0
    vprnintws.column_dimensions['AC'].width = 80.0
    vprnintws.column_dimensions['AD'].width = 80.0
    vprnintws.column_dimensions['AE'].width = 24.0
    vprnintws.column_dimensions['AF'].width = 18.0
    vprnintws.column_dimensions['AG'].width = 18.0
    vprnintws.column_dimensions['AH'].width = 24.0
    vprnintws.column_dimensions['AI'].width = 14.0
    vprnintws.column_dimensions['AJ'].width = 15.0


    epipews = misum.create_sheet('epipe-sap')
    epipews['A1'] = row
    epipews['A2'] = 'nokia-node'
    epipews['B2'] = 'system-ip'
    epipews['C2'] = 'service-name'
    epipews['D2'] = 'service-id'
    epipews['E2'] = 'sap'
    epipews['F2'] = 'sap-description'
    epipews['G2'] = 'sap-ingress-filter'
    epipews['H2'] = 'sap-egress-filter'
    epipews['I2'] = 'sap-in/egress qos'
    epipews['J2'] = 'ingress-rate'
    epipews['K2'] = 'egress-rate'
    epipews['L2'] = 'iptn-node'
    epipews['M2'] = 'iptn-service'
    epipews['N2'] = 'iptn-interface'
    epipews.auto_filter.ref = 'A2:N2'
    epipews.freeze_panes = epipews['A3']
    epipews.column_dimensions['A'].width = 17.0
    epipews.column_dimensions['B'].width = 12.0
    epipews.column_dimensions['C'].width = 65.0
    epipews.column_dimensions['D'].width = 12.0
    epipews.column_dimensions['E'].width = 12.0
    epipews.column_dimensions['F'].width = 64.0
    epipews.column_dimensions['G'].width = 30.0
    epipews.column_dimensions['H'].width = 30.0
    epipews.column_dimensions['I'].width = 19.0
    epipews.column_dimensions['J'].width = 14.0
    epipews.column_dimensions['K'].width = 14.0
    epipews.column_dimensions['L'].width = 26.0
    epipews.column_dimensions['M'].width = 64.0
    epipews.column_dimensions['N'].width = 15.0

    vplsws = misum.create_sheet('vpls-sap')
    vplsws['A1'] = row
    vplsws['A2'] = 'nokia-node'
    vplsws['B2'] = 'system-ip'
    vplsws['C2'] = 'service-name'
    vplsws['D2'] = 'service-id'
    vplsws['E2'] = 'sap'
    vplsws['F2'] = 'sap-description'
    vplsws['G2'] = 'sap-ingress-filter'
    vplsws['H2'] = 'sap-egress-filter'
    vplsws['I2'] = 'sap-in/egress qos'
    vplsws['J2'] = 'iptn-node'
    vplsws['K2'] = 'iptn-service'
    vplsws['L2'] = 'iptn-interface'
    vplsws.auto_filter.ref = 'A2:L2'
    vplsws.freeze_panes = vplsws['A3']
    vplsws.column_dimensions['A'].width = 17.0
    vplsws.column_dimensions['B'].width = 12.0
    vplsws.column_dimensions['C'].width = 65.0
    vplsws.column_dimensions['D'].width = 12.0
    vplsws.column_dimensions['E'].width = 12.0
    vplsws.column_dimensions['F'].width = 64.0
    vplsws.column_dimensions['G'].width = 30.0
    vplsws.column_dimensions['H'].width = 30.0
    vplsws.column_dimensions['I'].width = 19.0
    vplsws.column_dimensions['J'].width = 26.0
    vplsws.column_dimensions['K'].width = 64.0
    vplsws.column_dimensions['L'].width = 15.0





    rvplsws = misum.create_sheet('rvpls-sap')
    rvplsws['A1'] = row
    rvplsws['A2'] = 'nokia-node'
    rvplsws['B2'] = 'system-ip'
    rvplsws['C2'] = 'service-name'
    rvplsws['D2'] = 'service-id'
    rvplsws['E2'] = 'sap'
    rvplsws['F2'] = 'sap-description'
    rvplsws['G2'] = 'vpls-ingress-filter'
    rvplsws['H2'] = 'vpls-egress-filter'
    rvplsws['I2'] = 'sap-in/egress qos'
    rvplsws['J2'] = 'iptn-node'
    rvplsws['K2'] = 'iptn-service'
    rvplsws['L2'] = 'iptn-interface'
    rvplsws.auto_filter.ref = 'A2:L2'
    rvplsws.freeze_panes = rvplsws['A3']
    rvplsws.column_dimensions['A'].width = 17.0
    rvplsws.column_dimensions['B'].width = 12.0
    rvplsws.column_dimensions['C'].width = 65.0
    rvplsws.column_dimensions['D'].width = 12.0
    rvplsws.column_dimensions['E'].width = 12.0
    rvplsws.column_dimensions['F'].width = 64.0
    rvplsws.column_dimensions['G'].width = 30.0
    rvplsws.column_dimensions['H'].width = 30.0
    rvplsws.column_dimensions['I'].width = 19.0
    rvplsws.column_dimensions['J'].width = 26.0
    rvplsws.column_dimensions['K'].width = 64.0
    rvplsws.column_dimensions['L'].width = 15.0

    rvplsintws = misum.create_sheet('rvpls-interface')
    rvplsintws['A1'] = row
    rvplsintws['A2'] = 'nokia-node'
    rvplsintws['B2'] = 'system-ip'
    rvplsintws['C2'] = 'service-name'
    rvplsintws['D2'] = 'service-id'
    rvplsintws['E2'] = 'interface-name'
    rvplsintws['F2'] = 'interface description'
    rvplsintws['G2'] = 'vpls'
    rvplsintws['H2'] = 'vpls-ingress-filter'
    rvplsintws['I2'] = 'rvpls-ingress-filter-policy'
    rvplsintws['J2'] = 'vpls-egress-filter'
    rvplsintws['K2'] = 'rvpls-egress-filter-policy'
    rvplsintws['L2'] = 'ip/ipv6 address'
    rvplsintws['M2'] = 'secondary ip/ipv6'
    rvplsintws['N2'] = 'family'
    rvplsintws['O2'] = 'vrrp id'
    rvplsintws['P2'] = 'vrrp vip'
    rvplsintws['Q2'] = 'vrrp priority'
    rvplsintws['R2'] = 'vrrp interval'
    rvplsintws['S2'] = 'static-route'
    rvplsintws['T2'] = 'iptn-node'
    rvplsintws['U2'] = 'iptn-service'
    rvplsintws['V2'] = 'iptn-interface'
    rvplsintws.auto_filter.ref = 'A2:V2'
    rvplsintws.freeze_panes = rvplsintws['A3']
    rvplsintws.column_dimensions['A'].width = 18.0
    rvplsintws.column_dimensions['B'].width = 12.0
    rvplsintws.column_dimensions['C'].width = 30.0
    rvplsintws.column_dimensions['D'].width = 12.0
    rvplsintws.column_dimensions['E'].width = 32.0
    rvplsintws.column_dimensions['F'].width = 50.0
    rvplsintws.column_dimensions['G'].width = 32.0
    rvplsintws.column_dimensions['H'].width = 25.0
    rvplsintws.column_dimensions['I'].width = 64.0
    rvplsintws.column_dimensions['J'].width = 25.0
    rvplsintws.column_dimensions['K'].width = 64.0
    rvplsintws.column_dimensions['L'].width = 18.0
    rvplsintws.column_dimensions['M'].width = 18.0
    rvplsintws.column_dimensions['N'].width = 9.0
    rvplsintws.column_dimensions['O'].width = 9.0
    rvplsintws.column_dimensions['P'].width = 14.0
    rvplsintws.column_dimensions['Q'].width = 14.0
    rvplsintws.column_dimensions['R'].width = 14.0
    rvplsintws.column_dimensions['S'].width = 87.0
    rvplsintws.column_dimensions['T'].width = 27.0
    rvplsintws.column_dimensions['U'].width = 27.0
    rvplsintws.column_dimensions['V'].width = 16.0

    misum.save(migratepath+t.strftime("%Y%m%d-%H%M")+'\\'+misumname)
    return misum


def findnamingobject(objectlist,objectline, worksheet, objectrow, column):
    rowlist = []
    for row in objectrow:
        object = worksheet['%s%s' % (column,row)].value
        if object not in objectlist:
            objectlist.append(object)
    for ob in objectlist:
        for row in objectrow:
            if ob == worksheet['%s%s' % (column,row)].value:
                objectline.setdefault(ob,[]).append(row)
                #print('node %s row %s' % (node, nrow))

    #print(objectlist, objectline)
    return objectlist,objectline

def createnamingdict(naming,n,row):
    namingdict = {}
    namingdict['iptnnode'] = naming['A%s' % row].value
    namingdict['iptnloopback'] = naming['B%s' % row].value
    namingdict['iptnintunit'] = naming['C%s' % row].value
    namingdict['iptnae'] = naming['D%s' % row].value
    namingdict['bdrnode'] = n
    namingdict['bdrsystem'] = naming['F%s' % row].value
    namingdict['bdrport'] = naming['G%s' % row].value
    namingdict['bdrportdesc'] = naming['I%s' % row].value
    namingdict['bdrlag'] = naming['K%s' % row].value
    namingdict['bdrportencap'] = naming['M%s' % row].value
    namingdict['iptnservice'] = naming['O%s' % row].value
    namingdict['bdrservice'] = naming['P%s' % row].value
    namingdict['bdrvlan'] = naming['R%s' % row].value
    namingdict['bdrsapdesc'] = naming['T%s' % row].value
    namingdict['bdrl3int'] = naming['V%s' % row].value
    namingdict['bdrl3desc'] = naming['Y%s' % row].value
    namingdict['vrrpviasw'] = naming['AA%s' % row].value
    return namingdict

def createextractionportdict(export,node,row):
    exportdict = {}
    exportdict['iptnnode'] = export['A%s' % row].value #'node'
    exportdict['iptnloopback'] = export['B%s' % row].value #'system-ip'
    exportdict['iptnport'] = export['C%s' % row].value #'physical-port'
    exportdict['iptnportadminstate'] = export['D%s' % row].value #'admin-state'
    exportdict['iptnportencap'] = export['E%s' % row].value #'port-encap'
    exportdict['iptnportdesc'] = export['F%s' % row].value #'physical-description'
    exportdict['iptnae'] = export['G%s' % row].value #'LAG'
    exportdict['iptnaedesc'] = export['H%s' % row].value #'LAG-protocol'
    exportdict['iptnaeminlink'] = export['I%s' % row].value #'LAG-minimum-link'
    exportdict['iptnportspeed'] = export['J%s' % row].value #'speed'
    exportdict['iptnportmtu'] = export['K%s' % row].value #'mtu'
    exportdict['iptnportautonego'] = export['L%s' % row].value #'auto-negotiation'
    exportdict['iptnunit'] = export['M%s' % row].value #'unit'
    exportdict['iptnunitdesc'] = export['N%s' % row].value #'unit-description'
    exportdict['iptnunitstate'] = export['O%s' % row].value #'unit-state'
    exportdict['iptnunitvlan'] = export['P%s' % row].value #'vlan'
    exportdict['iptninputfiltername'] = export['Q%s' % row].value #'input-filter-policy-name'
    exportdict['iptninputfilter'] = export['R%s' % row].value #'input-filter-configuration'
    exportdict['iptnoutputfiltername'] = export['S%s' % row].value #'output-filter-policy-name'
    exportdict['iptnoutputfilter'] = export['T%s' % row].value #'output-filter-configuration'
    exportdict['iptnfc'] = export['U%s' % row].value #'forwarding-class'
    exportdict['iptnservice'] = export['V%s' % row].value #'service'
    exportdict['iptnservicetype'] = export['W%s' % row].value #'service-type'
    exportdict['bdrnode'] = export['X%s' % row].value #'nokia-node'
    exportdict['bdrsystem'] = export['Y%s' % row].value #'nokia-system-ip'
    exportdict['bdrsap'] = export['Z%s' % row].value #'nokia-sap'
    return exportdict

def createvprnglobaldict(vprnglobal, node, row):
    exvprnglobaldict = {}
    exvprnglobaldict['iptnnode'] = vprnglobal['A%s' % row].value #'node'
    exvprnglobaldict['iptnloopback'] = vprnglobal['B%s' % row].value #'system-ip'
    exvprnglobaldict['iptnservice'] = vprnglobal['C%s' % row].value #'vpn-name'
    exvprnglobaldict['iptnservicestate'] = vprnglobal['D%s' % row].value #'admin-state'
    exvprnglobaldict['iptnrd'] = vprnglobal['E%s' % row].value #'rd'
    exvprnglobaldict['iptnrtimppolname'] = vprnglobal['F%s' % row].value #'rt-import-policy-name'
    exvprnglobaldict['iptnrtimppol'] = vprnglobal['G%s' % row].value #'rt-import-policy'
    exvprnglobaldict['iptnrtimpprefix'] = vprnglobal['H%s' % row].value #'rt-import-prefix'
    exvprnglobaldict['iptnrtimpcommu'] = vprnglobal['I%s' % row].value #'rt-import-commu'
    exvprnglobaldict['iptnrtexppolname'] = vprnglobal['J%s' % row].value #'rt-export-policy-name'
    exvprnglobaldict['iptnrtexppol'] = vprnglobal['K%s' % row].value #'rt-export-policy'
    exvprnglobaldict['iptnrtexpprefix'] = vprnglobal['L%s' % row].value #'rt-export-prefix'
    exvprnglobaldict['iptnrtexpcommu'] = vprnglobal['M%s' % row].value #'rt-export-commu'
    exvprnglobaldict['iptnstatic'] = vprnglobal['N%s' % row].value #'static-route'
    exvprnglobaldict['iptnaggregate'] = vprnglobal['O%s' % row].value #'aggregate-route'
    exvprnglobaldict['iptnbgp'] = vprnglobal['P%s' % row].value #'bgp'
    exvprnglobaldict['iptnbgpimp'] = vprnglobal['Q%s' % row].value #'bgp-imp-policy'
    exvprnglobaldict['iptnbgpimpprefix'] = vprnglobal['R%s' % row].value #'bgp-imp-policy-prefix'
    exvprnglobaldict['iptnbgpimpcom'] = vprnglobal['S%s' % row].value #'bgp-imp-policy-commu'
    exvprnglobaldict['iptnbgpexp'] = vprnglobal['T%s' % row].value #'bgp-exp-policy'
    exvprnglobaldict['iptnbgpexpprefix'] = vprnglobal['U%s' % row].value #'bgp-exp-policy-prefix'
    exvprnglobaldict['iptnbgpexpcom'] = vprnglobal['V%s' % row].value #'bgp-exp-policy-commu'
    return exvprnglobaldict

def createvprnintdict(vprnint, node, row):
    exvprnintdict = {}
    exvprnintdict['iptnnode'] = vprnint['A%s' % row].value  #'node'
    exvprnintdict['iptnloopback'] = vprnint['B%s' % row].value  #'system-ip'
    exvprnintdict['iptnservice'] = vprnint['C%s' % row].value  #'vpn-name'
    exvprnintdict['iptnrd'] = vprnint['D%s' % row].value  #'rd'
    exvprnintdict['iptnl3int'] = vprnint['E%s' % row].value  #'layer 3 interface'
    exvprnintdict['iptnunit'] = vprnint['F%s' % row].value  #'unit'
    exvprnintdict['iptnunitstate'] = vprnint['G%s' % row].value  #'unit-state'
    exvprnintdict['iptnvlan'] = vprnint['H%s' % row].value  #'vlan'
    exvprnintdict['iptnintdesc'] = vprnint['I%s' % row].value  #'interface description'
    exvprnintdict['iptnintfc'] = vprnint['J%s' % row].value  #'forwarding-class'
    exvprnintdict['iptnip'] = vprnint['K%s' % row].value  #'ip/ipv6 address'
    exvprnintdict['iptnsecip'] = vprnint['L%s' % row].value  #'secondary ip/ipv6'
    exvprnintdict['iptnfam'] = vprnint['M%s' % row].value  #'family'
    exvprnintdict['iptnvrrpid'] = vprnint['N%s' % row].value  #'vrrp id'
    exvprnintdict['iptnvrrpvip'] = vprnint['O%s' % row].value  #'vrrp vip'
    exvprnintdict['iptnvrrppri'] = vprnint['P%s' % row].value  #'vrrp priority'
    exvprnintdict['iptnvrrpinterval'] = vprnint['Q%s' % row].value  #'vrrp interval'
    exvprnintdict['iptnstatic'] = vprnint['R%s' % row].value  #'static-route'
    exvprnintdict['iptnbgp'] = vprnint['S%s' % row].value  #'bgp'
    exvprnintdict['iptnbgpgroup'] = vprnint['T%s' % row].value  #'group'
    exvprnintdict['iptnbgppeeras'] = vprnint['U%s' % row].value  #'peer-as'
    exvprnintdict['iptnbgppeer'] = vprnint['V%s' % row].value  #'neighbor'
    exvprnintdict['iptnbgppeerdesc'] = vprnint['W%s' % row].value  #'neighbor-description'
    exvprnintdict['iptnbgpimpname'] = vprnint['X%s' % row].value  #'bgp-imp-policy-name'
    exvprnintdict['iptnbgpimp'] = vprnint['Y%s' % row].value  #'bgp-imp-policy'
    exvprnintdict['iptnbgpimppre'] = vprnint['Z%s' % row].value  #'bgp-imp-policy-prefix'
    exvprnintdict['iptnbgpimpcom'] = vprnint['AA%s' % row].value #'bgp-imp-policy-commu'
    exvprnintdict['iptnbgpexpname'] = vprnint['AB%s' % row].value #'bgp-exp-policy-name'
    exvprnintdict['iptnbgpexp'] = vprnint['AC%s' % row].value #'bgp-exp-policy'
    exvprnintdict['iptnbgpexppre'] = vprnint['AD%s' % row].value #'bgp-exp-policy-prefix'
    exvprnintdict['iptnbgpexpcom'] = vprnint['AE%s' % row].value #'bgp-exp-policy-commu'
    exvprnintdict['iptnbgpbfdinterval'] = vprnint['AF%s' % row].value #'bgp-bfd-interval'
    exvprnintdict['iptnbgpbfdmul'] = vprnint['AG%s' % row].value #'bgp-bfd-multiply'
    exvprnintdict['bdrnode'] = vprnint['AH%s' % row].value #'nokia-node'
    exvprnintdict['bdrsystem'] = vprnint['AI%s' % row].value #'nokia-system-ip'
    exvprnintdict['bdrsap'] = vprnint['AJ%s' % row].value #'nokia-sap'
    return exvprnintdict


def createl2dict(l2vpn, node, row):
    exl2vpndict = {}
    exl2vpndict['iptnnode'] = l2vpn['A%s' % row].value #'node'
    exl2vpndict['iptnloopback'] = l2vpn['B%s' % row].value #'system-ip'
    exl2vpndict['iptnservice'] = l2vpn['C%s' % row].value #'l2vpn-name'
    exl2vpndict['iptnrd'] = l2vpn['D%s' % row].value #'rd'
    exl2vpndict['iptnrt'] = l2vpn['E%s' % row].value #'vrf-target'
    exl2vpndict['iptnsite'] = l2vpn['F%s' % row].value #'site'
    exl2vpndict['iptnsiteid'] = l2vpn['G%s' % row].value #'site-id'
    exl2vpndict['iptnremotesiteid'] = l2vpn['H%s' % row].value #'remote-site-id'
    exl2vpndict['iptnsitepreference'] = l2vpn['I%s' % row].value #'site-preference'
    exl2vpndict['iptnint'] = l2vpn['J%s' % row].value #'interface'
    exl2vpndict['iptnunit'] = l2vpn['K%s' % row].value #'unit'
    exl2vpndict['iptnunitstate'] = l2vpn['L%s' % row].value #'unit-state'
    exl2vpndict['iptnunitdesc'] = l2vpn['M%s' % row].value #'unit-description'
    exl2vpndict['iptnvlan'] = l2vpn['N%s' % row].value #'vlan'
    exl2vpndict['iptninvlanmap'] = l2vpn['O%s' % row].value #'input-vlan-map'
    exl2vpndict['iptninvlanmapid'] = l2vpn['P%s' % row].value #'input-map-vlan-id'
    exl2vpndict['iptnoutvlanmap'] = l2vpn['Q%s' % row].value #'output-vlan-map'
    exl2vpndict['iptnoutvlanmapid'] = l2vpn['R%s' % row].value #'output-map-vlan-id'
    exl2vpndict['iptnpolicerin'] = l2vpn['S%s' % row].value #'policer-input'
    exl2vpndict['iptnpolicerout'] = l2vpn['T%s' % row].value #'policer-output'
    exl2vpndict['iptnintfc'] = l2vpn['U%s' % row].value #'forwarding-class'
    exl2vpndict['bdrnode'] = l2vpn['V%s' % row].value #'nokia-node'
    exl2vpndict['bdrsystem'] = l2vpn['W%s' % row].value #'nokia-system-ip'
    exl2vpndict['bdrsap'] = l2vpn['X%s' % row].value #'nokia-sap'
    return exl2vpndict

def createvplsdict(vpls, node, row):
    exvplsdict = {}
    exvplsdict['iptnnode'] = vpls['A%s' % row].value #'node'
    exvplsdict['iptnloopback'] = vpls['B%s' % row].value #'system-ip'
    exvplsdict['iptnservice'] = vpls['C%s' % row].value #'vpls-name'
    exvplsdict['iptnrd'] = vpls['D%s' % row].value #'rd'
    exvplsdict['iptnrt'] = vpls['E%s' % row].value #'vrf-target'
    exvplsdict['iptnsite'] = vpls['F%s' % row].value #'site'
    exvplsdict['iptnsiteid'] = vpls['G%s' % row].value #'site-id'
    exvplsdict['iptnsitepreference'] = vpls['H%s' % row].value #'site-preference'
    exvplsdict['iptnint'] = vpls['I%s' % row].value #'interface'
    exvplsdict['iptnunit'] = vpls['J%s' % row].value #'unit'
    exvplsdict['iptnunitstate'] = vpls['K%s' % row].value #'unit-state'
    exvplsdict['iptnunitdesc'] = vpls['L%s' % row].value #'unit-description'
    exvplsdict['iptnvlan'] = vpls['M%s' % row].value #'vlan'
    exvplsdict['iptnphydesc'] = vpls['N%s' % row].value #'input-vlan-map'
    exvplsdict['iptnl3vrf'] = vpls['O%s' % row].value #'input-map-vlan-id'
    exvplsdict['iptnl3int'] = vpls['P%s' % row].value #'output-vlan-map'
    exvplsdict['iptnl3intdesc'] = vpls['Q%s' % row].value #'output-map-vlan-id'
    exvplsdict['iptnpolicerin'] = vpls['R%s' % row].value #'policer-input'
    exvplsdict['iptnpolicerout'] = vpls['S%s' % row].value #'policer-output'
    exvplsdict['iptnintfc'] = vpls['T%s' % row].value #'forwarding-class'
    exvplsdict['bdrnode'] = vpls['U%s' % row].value #'nokia-node'
    exvplsdict['bdrsystem'] = vpls['V%s' % row].value #'nokia-system-ip'
    exvplsdict['bdrsap'] = vpls['W%s' % row].value #'nokia-sap'
    return exvplsdict

def createbddict(bd, node, row):
    exbddict = {}
    exbddict['iptnnode'] = bd['A%s' % row].value #'node'
    exbddict['iptnloopback'] = bd['B%s' % row].value #'system-ip'
    exbddict['iptnservice'] = bd['C%s' % row].value #'bridge-domain'
    exbddict['iptnbdvlan'] = bd['D%s' % row].value #'bd-vlan'
    exbddict['iptnbdfilter'] = bd['E%s' % row].value #'filter'
    exbddict['iptnirb'] = bd['F%s' % row].value #'routing-interface'
    exbddict['iptnirbvrf'] = bd['G%s' % row].value #'irb-vprn'
    exbddict['iptnint'] = bd['H%s' % row].value #'interface'
    exbddict['iptnunit'] = bd['I%s' % row].value #'unit'
    exbddict['iptnunitstate'] = bd['J%s' % row].value #'unit-state'
    exbddict['iptnunitdesc'] = bd['K%s' % row].value #'unit-description'
    exbddict['iptnintvlan'] = bd['L%s' % row].value #'int-vlan'
    exbddict['iptnintfc'] = bd['M%s' % row].value #'forwarding-class'
    exbddict['bdrnode'] = bd['N%s' % row].value #'nokia-node'
    exbddict['bdrsystem'] = bd['O%s' % row].value #'nokia-system-ip'
    exbddict['bdrsap'] = bd['P%s' % row].value #'nokia-sap'
    return exbddict

def createvswdict(vsw, node, row):
    exvswdict = {}
    exvswdict['iptnnode'] = vsw['A%s' % row].value #'node'
    exvswdict['iptnloopback'] = vsw['B%s' % row].value #'system-ip'
    exvswdict['iptnservice'] = vsw['C%s' % row].value #'virtual-switch-name'
    exvswdict['iptnrd'] = vsw['D%s' % row].value #'rd'
    exvswdict['iptnrt'] = vsw['E%s' % row].value #'vrf-target'
    exvswdict['iptnsite'] = vsw['F%s' % row].value #'site'
    exvswdict['iptnsiteid'] = vsw['G%s' % row].value #'site-id'
    exvswdict['iptnintunit'] = vsw['H%s' % row].value #'interfaceunit'
    exvswdict['iptnbd'] = vsw['I%s' % row].value #'bridge-domain'
    exvswdict['iptnbdvlan'] = vsw['J%s' % row].value #'bridge-vlan'
    exvswdict['iptnint'] = vsw['K%s' % row].value #'interface'
    exvswdict['iptnunit'] = vsw['L%s' % row].value #'unit'
    exvswdict['iptnunitstate'] = vsw['M%s' % row].value #'unit-state'
    exvswdict['iptnunitdesc'] = vsw['N%s' % row].value #'unit-description'
    exvswdict['iptnunitvlan'] = vsw['O%s' % row].value #'vlan'
    exvswdict['iptnintfc'] = vsw['P%s' % row].value #'forwarding-class'
    exvswdict['iptnirb'] = vsw['Q%s' % row].value #'routing-interface'
    exvswdict['iptnirbvrf'] = vsw['R%s' % row].value #'irb-vprn'
    exvswdict['iptnbdfilter'] = vsw['S%s' % row].value #'filter'
    exvswdict['bdrnode'] = vsw['T%s' % row].value #'nokia-node'
    exvswdict['bdrsystem'] = vsw['U%s' % row].value #'nokia-system-ip'
    exvswdict['bdrsap'] = vsw['V%s' % row].value #'nokia-sap'
    return exvswdict

def checkvprnint(namingdict,exportdict, iptnp, iptnu, noderow):
    exvprnintdict = {}
    bdrservicetype = 'VPRN'
    bdrserviceid = ''
    for i in range(noderow['exvprnintstartrow'], noderow['exvprnintendrow']+1):
        iptnvprnint = createvprnintdict(exvprnint, namingdict['iptnnode'], i)
        if '%s.%s' %(iptnp,iptnu) == '%s.%s' % (iptnvprnint['iptnl3int'], iptnvprnint['iptnunit']):
            if iptnvprnint['iptnfam'] == 'ipv4':
                exvprnintdict = iptnvprnint
                bdrserviceid = 3000000 + int(iptnvprnint['iptnrd'].split(':')[-1])
                return exvprnintdict, bdrservicetype, bdrserviceid
                break
            elif iptnvprnint['iptnfam'] == 'ipv6':
                exvprnintdict = iptnvprnint
                bdrserviceid = 3000000 + int(iptnvprnint['iptnrd'].split(':')[-1])
                return exvprnintdict, bdrservicetype, bdrserviceid
                break

def checkl2vpn(namingdict,exportdict, iptnp, iptnu, noderow):
    exl2dict = {}
    bdrservicetype = 'EPIPE'
    bdrserviceid = ''
    for i in range(noderow['exl2startrow'], noderow['exl2endrow']+1):
        iptnl2vpn = createl2dict(exl2vpn, namingdict['iptnnode'], i)
        if '%s.%s' %(iptnp,iptnu) == '%s.%s' % (iptnl2vpn['iptnint'], iptnl2vpn['iptnunit']):
            if iptnl2vpn['iptnsitepreference'] == None:
                exl2dict = iptnl2vpn
                bdrserviceid = 1000000 + int(iptnl2vpn['iptnrd'].split(':')[-1])
                return exl2dict, bdrservicetype, bdrserviceid
                break
            else:
                bdrservicetype = 'EPIPE-EVPN'
                exl2dict = iptnl2vpn
                bdrserviceid = 1000000 + int(iptnl2vpn['iptnrd'].split(':')[-1])
                return exl2dict, bdrservicetype, bdrserviceid
                break

def checkvpls(namingdict,exportdict, iptnp, iptnu, noderow):
    exvplsd = {}
    if 'rVPLS_' in namingdict['bdrservice']:
        bdrservicetype = 'rVPLS'
        bdrserviceid = ''
        for i in range(noderow['exvplsstartrow'], noderow['exvplsendrow']+1):
            iptnvpls = createvplsdict(exvpls, namingdict['iptnnode'], i)
            if '%s.%s' %(iptnp,iptnu) == '%s.%s' % (iptnvpls['iptnint'], iptnvpls['iptnunit']):
                #print('iptnport : %s vplsint %s' % (iptnp, iptnvpls['iptnint']))
                exvplsd = iptnvpls
                bdrserviceid = 4000000 + int(iptnvpls['iptnrd'].split(':')[-1])
                return exvplsd, bdrservicetype, bdrserviceid
                break
    else:
        bdrservicetype = 'VPLS'
        bdrserviceid = ''
        for i in range(noderow['exvplsstartrow'], noderow['exvplsendrow']+1):
            iptnvpls = createvplsdict(exvpls, namingdict['iptnnode'], i)
            if '%s.%s' %(iptnp,iptnu) == '%s.%s' % (iptnvpls['iptnint'], iptnvpls['iptnunit']):
                #print('iptnport : %s vplsint %s' % (iptnp, iptnvpls['iptnint']))
                exvplsd = iptnvpls
                bdrserviceid = 2000000 + int(iptnvpls['iptnrd'].split(':')[-1])
                return exvplsd, bdrservicetype, bdrserviceid
                break

def checkbd(namingdict,exportdict, iptnp, iptnu, noderow):
    exbdd = {}
    l2map = {}
    bdrservicetype = 'rVPLS'
    bdrserviceid = ''
    for i in range(noderow['exbdstartrow'], noderow['exbdendrow']+1):
        iptnbd = createbddict(exbridge, namingdict['iptnnode'], i)
        if '%s.%s' %(iptnp,iptnu) == '%s.%s' % (iptnbd['iptnint'], iptnbd['iptnunit']):
            if iptnbd['iptnirb'] == None:
                for p in range(3,l2mapsheet.max_row +1):
                    if iptnbd['iptnnode'] == l2mapsheet['A%s' % p].value:
                        l2map['peerid'] = l2mapsheet['C%s' % p].value
                        l2map['bdrnodeingress'] = l2mapsheet['D%s' % p].value
                        l2map['bdrsystemingress'] = l2mapsheet['E%s' % p].value
                        l2map['bdrnodeegress'] = l2mapsheet['F%s' % p].value
                        l2map['bdrsystemegress'] = l2mapsheet['G%s' % p].value
                        exbdd = iptnbd
                        bdrservicetype = 'VPLS'
                        bdrserviceid = 2000000 + (int(l2map['peerid'])*10000) + int(iptnbd['iptnbdvlan'])
                        return exbdd, l2map, bdrservicetype, bdrserviceid
                        break
            else:
                for p in range(3,l2mapsheet.max_row +1):
                    if iptnbd['iptnnode'] == l2mapsheet['A%s' % p].value:
                        l2map['peerid'] = l2mapsheet['C%s' % p].value
                        l2map['bdrnodeingress'] = l2mapsheet['D%s' % p].value
                        l2map['bdrsystemingress'] = l2mapsheet['E%s' % p].value
                        l2map['bdrnodeegress'] = l2mapsheet['F%s' % p].value
                        l2map['bdrsystemegress'] = l2mapsheet['G%s' % p].value
                        exbdd = iptnbd
                        #print(iptnbd['iptnbdvlan'])
                        #print(l2map['peerid'])
                        bdrserviceid = 4000000 + (int(l2map['peerid'])*10000) + int(iptnbd['iptnbdvlan'])
                        #print(bdrserviceid)
                        return exbdd, l2map, bdrservicetype, bdrserviceid
                        break

def checkvsw(namingdict,exportdict, iptnp, iptnu, noderow):
    exvswd = {}
    l2map = {}
    bdrservicetype = 'rVPLS'
    bdrserviceid = ''
    for i in range(noderow['exvswstartrow'], noderow['exvswendrow']+1):
        iptnvsw = createvswdict(exvsw, namingdict['iptnnode'], i)
        if '%s.%s' %(iptnp,iptnu) == '%s.%s' % (iptnvsw['iptnint'], iptnvsw['iptnunit']):

            for p in range(3,l2mapsheet.max_row +1):
                if namingdict['bdrvlan'] == iptnvsw['iptnbdvlan']:
                    if iptnvsw['iptnnode'] == l2mapsheet['A%s' % p].value:
                        l2map['peerid'] = l2mapsheet['C%s' % p].value
                        l2map['bdrnodeingress'] = l2mapsheet['D%s' % p].value
                        l2map['bdrsystemingress'] = l2mapsheet['E%s' % p].value
                        l2map['bdrnodeegress'] = l2mapsheet['F%s' % p].value
                        l2map['bdrsystemegress'] = l2mapsheet['G%s' % p].value
                        if iptnvsw['iptnservice'] == 'vsw-vpls-for-vrrp-cwdc05-suk05':
                            l2map['peerid'] = 80
                            if l2map['bdrnodeingress'] == 'BDRT_SUKE_PE01':
                                l2map['bdrnodeegress'] = 'BDRT_CWDC_PE01'
                                l2map['bdrsystemegress'] = '10.129.147.4'
                            elif l2map['bdrnodeingress'] == 'BDRT_CWDC_PE01':
                                l2map['bdrnodeegress'] = 'BDRT_SUKE_PE01'
                                l2map['bdrsystemegress'] = '10.129.147.6'
                        exvswd = iptnvsw
                        bdrserviceid = 4000000 + (int(l2map['peerid'])*10000) + int(iptnvsw['iptnbdvlan'])
                        #print('%s.%s' % (iptnp, iptnu))
                        #print(exvswd['iptnservice'])
                        #print(exvswd['iptnbdvlan'])
                        #print(bdrserviceid)
                        return exvswd, l2map, bdrservicetype, bdrserviceid
                        break

def findextractnodeindex(exindex,iptnnode,nodecolumn,startcolumn,endcolumn):
    startportrow = 0
    endportrow = 0
    try:
        for i in range(3, (exindex.max_row + 1)):  # search iptnnode in index sheet of extraction file to get start and end row of port sheet
            if iptnnode == exindex['%s%s' % (nodecolumn,i)].value:
                startportrow = exindex['%s%s' % (startcolumn,i)].value
                endportrow = exindex['%s%s' % (endcolumn,i)].value
        #print('iptnnode : %s start %s end %s' %(iptnnode,startportrow,endportrow))
    except:
        pass
    return startportrow,endportrow

def creatervplsinterfaceworksheet(rvplsintws,rvplssap,namingdict,iptnvprnint,exportdict,qos,filternamews):
    row = rvplsintws['A1'].value
    filterinname = ''
    filterinfilepath = ''
    filteroutname = ''
    filteroutfilepath = ''
    bdrstatic = ''
    bdrvprnservice = 'L3_%s' %iptnvprnint['iptnservice']
    bdrvprnserviceid = int(iptnvprnint['iptnrd'].split(':')[-1]) + 3000000
    allvprn = namingdict['bdrnode']+'$%'+namingdict['bdrsystem']+'$%'+bdrvprnservice+'$%'+str(bdrvprnserviceid)+'$%'+namingdict['iptnnode']+'$%'+iptnvprnint['iptnrd']+'$%'+qos
    f = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'allvprn.txt', "a")
    f.write(allvprn + '\n')
    f.close()
    #print(bdrvprnserviceid)

    if iptnvprnint['iptnstatic'] != None:
        s = open(extractpath + iptnvprnint['iptnstatic'], 'r')
        bdrstatic = createbdrstaticconfig(s ,namingdict,bdrvprnservice,iptnvprnint['iptnfam'])
    if exportdict['iptninputfilter'] != None:
        i = open(extractpath + exportdict['iptninputfilter'], 'r')
        filterinfilepath, filterinname = createbdrfilter(i,namingdict,filternamews,'input')
    if exportdict['iptnoutputfilter'] != None:
        o = open(extractpath + exportdict['iptnoutputfilter'], 'r')
        filteroutfilepath, filteroutname = createbdrfilter(o,namingdict,filternamews,'output')

    rvplsintws['A%s' % row] = namingdict['bdrnode'] #'nokia-node'
    rvplsintws['B%s' % row] = namingdict['bdrsystem'] #'system-ip'
    rvplsintws['C%s' % row] = bdrvprnservice #'service-name'
    rvplsintws['D%s' % row] = bdrvprnserviceid #'service-id'
    rvplsintws['E%s' % row] = namingdict['bdrl3int'] #'interface-name'
    rvplsintws['F%s' % row] = namingdict['bdrl3desc'] #'interface description'
    if namingdict['vrrpviasw'] == 'yes':
        rvplsintws['G%s' % row] = 'SAP:$'+rvplssap[4]+':$'+rvplssap[5]+':$'+rvplssap[8]
    else:
        rvplsintws['G%s' % row] = namingdict['bdrservice']  # 'vpls'
    rvplsintws['H%s' % row] = filterinname #'vpls-ingress-filter'
    rvplsintws['I%s' % row].hyperlink = filterinfilepath #'rvpls-ingress-filter-policy'
    rvplsintws['J%s' % row] = filteroutname #'vpls-egress-filter'
    rvplsintws['K%s' % row].hyperlink = filteroutfilepath #'rvpls-egress-filter-policy'
    rvplsintws['L%s' % row] = iptnvprnint['iptnip'] #'ip/ipv6 address'
    rvplsintws['M%s' % row] = iptnvprnint['iptnsecip'] #'secondary ip/ipv6'
    rvplsintws['N%s' % row] = iptnvprnint['iptnfam'] #'family'
    rvplsintws['O%s' % row] = iptnvprnint['iptnvrrpid'] #'vrrp id'
    rvplsintws['P%s' % row] = iptnvprnint['iptnvrrpvip'] #'vrrp vip'
    rvplsintws['Q%s' % row] = iptnvprnint['iptnvrrppri'] #'vrrp priority'
    rvplsintws['R%s' % row] = iptnvprnint['iptnvrrpinterval'] #'vrrp interval'
    rvplsintws['S%s' % row].hyperlink = bdrstatic #'static-route'
    rvplsintws['T%s' % row] = namingdict['iptnnode'] #'iptn-node'
    rvplsintws['U%s' % row] = namingdict['iptnservice'] #'iptn-service'
    rvplsintws['V%s' % row] = namingdict['iptnintunit'] #'iptn-interface'
    row += 1
    rvplsintws['A1'] = row

def createepipesapworksheet(epipesapws,epipesap):
    row = epipesapws['A1'].value
    epipesapws['A%s' % row] = epipesap[0] #'nokia-node'
    epipesapws['B%s' % row] = epipesap[1] #'system-ip'
    epipesapws['C%s' % row] = epipesap[2] #'service-name'
    epipesapws['D%s' % row] = epipesap[3] #'service-id'
    epipesapws['E%s' % row] = epipesap[4] #'sap'
    epipesapws['F%s' % row] = epipesap[5] #'sap-description'
    epipesapws['G%s' % row] = epipesap[6] #'sap-ingress-filter'
    epipesapws['H%s' % row] = epipesap[7] #'sap-egress-filter'
    epipesapws['I%s' % row] = epipesap[8] #'sap-in/egress qos'
    epipesapws['J%s' % row] = epipesap[9] #'ingress-rate'
    epipesapws['K%s' % row] = epipesap[10] #'egress-rate'
    epipesapws['L%s' % row] = epipesap[11] #'iptn-node'
    epipesapws['M%s' % row] = epipesap[12] #'iptn-service'
    epipesapws['N%s' % row] = epipesap[13] #'iptn-interface'
    row += 1
    epipesapws['A1'] = row

def creatervplssapworksheet(rvplssapws,rvplssap):
    row = rvplssapws['A1'].value
    rvplssapws['A%s' % row] = rvplssap[0] #'nokia-node'
    rvplssapws['B%s' % row] = rvplssap[1] #'system-ip'
    rvplssapws['C%s' % row] = rvplssap[2] #'service-name'
    rvplssapws['D%s' % row] = rvplssap[3] #'service-id'
    rvplssapws['E%s' % row] = rvplssap[4] #'sap'
    rvplssapws['F%s' % row] = rvplssap[5] #'sap-description'
    rvplssapws['G%s' % row] = rvplssap[6] #'sap-ingress-filter'
    rvplssapws['H%s' % row] = rvplssap[7] #'sap-egress-filter'
    rvplssapws['I%s' % row] = rvplssap[8] #'sap-in/egress qos'
    rvplssapws['J%s' % row] = rvplssap[9] #'iptn-node'
    rvplssapws['K%s' % row] = rvplssap[10] #'iptn-service'
    rvplssapws['L%s' % row] = rvplssap[11] #'iptn-interface'
    row += 1
    rvplssapws['A1'] = row

def createvplssapworksheet(vplssapws,vplssap):
    row = vplssapws['A1'].value
    vplssapws['A%s' % row] = vplssap[0] #'nokia-node'
    vplssapws['B%s' % row] = vplssap[1] #'system-ip'
    vplssapws['C%s' % row] = vplssap[2] #'service-name'
    vplssapws['D%s' % row] = vplssap[3] #'service-id'
    vplssapws['E%s' % row] = vplssap[4] #'sap'
    vplssapws['F%s' % row] = vplssap[5] #'sap-description'
    vplssapws['G%s' % row] = vplssap[6] #'sap-ingress-filter'
    vplssapws['H%s' % row] = vplssap[7] #'sap-egress-filter'
    vplssapws['I%s' % row] = vplssap[8] #'sap-in/egress qos'
    vplssapws['J%s' % row] = vplssap[9] #'iptn-node'
    vplssapws['K%s' % row] = vplssap[10] #'iptn-service'
    vplssapws['L%s' % row] = vplssap[11] #'iptn-interface'
    row += 1
    vplssapws['A1'] = row

def createbdrvprnglobalworksheet(vprnglobalws,bdrvprn,exvprnglobal,qos, policynamews):
    row = vprnglobalws['A1'].value
    bdrrd = bdrvprn[1]+':'+bdrvprn[5].split(':')[-1]
    impcom = None
    expcom = None
    bdrpolicyimppath = ''
    bdrprefiximppath = ''
    bdrpolicyexppath = ''
    bdrprefixexppath = ''
    aggregatelist = []
    aggregateprefix = []
    bdraggcfg = []
    bdraggfilepath = ''

    if exvprnglobal['iptnrtimppol'] != None:
        impprefix = exvprnglobal['iptnrtimpprefix']
        impcom = exvprnglobal['iptnrtimpcommu']
        iptnstatic = exvprnglobal['iptnstatic']
        imp = open(extractpath + exvprnglobal['iptnrtimppol'], 'r')
        bdrpolicyimppath,bdrprefiximppath = createbdrpolicy(imp, bdrvprn[0],bdrvprn[2],'vrf_L3_%s' % exvprnglobal['iptnrtimppolname'], impprefix, impcom,policynamews,iptnstatic,'vrfimp')
    if exvprnglobal['iptnrtexppol'] != None:
        expprefix = exvprnglobal['iptnrtexpprefix']
        expcom = exvprnglobal['iptnrtexpcommu']
        exp = open(extractpath + exvprnglobal['iptnrtexppol'], 'r')
        bdrpolicyexppath,bdrprefixexppath = createvrfexportpolicy(exp, bdrvprn[0],bdrvprn[2],'vrf_L3_%s' % exvprnglobal['iptnrtexppolname'], expprefix, expcom,policynamews,bdrvprn[4])
    if exvprnglobal['iptnaggregate'] != None:
        ag = open(extractpath + exvprnglobal['iptnaggregate'], 'r')
        bdraggfilename = '%s_%s_bdraggregate.txt' % (bdrvprn[0],bdrvprn[2])
        bdraggfilepath = '.\cfg\\'+ bdraggfilename
        g = ag.readlines()
        for a in range(len(g)):
            g[a] = g[a].rstrip('\n')  # remove newline('\n') from end of line
            if g[a] not in aggregatelist:
                aggregatelist.append(g[a])
        for aggregate in aggregatelist:
            if 'set routing-instances %s routing-options aggregate route' % exvprnglobal['iptnservice'] in aggregate:
                #print(aggregate)
                aggprefix = aggregate.split(' ')[6]
                aggregateprefix.append(aggprefix)
                bdraggcfg.append('/configure service vprn "%s" aggregates aggregate %s blackhole' %(bdrvprn[2],aggprefix))
                #print(bdraggcfg)

        for staticsumprefix in aggregateprefix:
            for aggreprefix in aggregateprefix:
                if IPNetwork(aggreprefix) in IPNetwork(staticsumprefix):
                    if staticsumprefix != aggreprefix:
                        bdraggcfg.append('')
                        bdraggcfg.append('/configure policy-options prefix-list "Sum_%s" prefix %s type exact' % (staticsumprefix,aggreprefix))
                        bdraggcfg.append('/configure service vprn "%s" static-routes route %s route-type unicast blackhole admin-state enable' % (bdrvprn[2],staticsumprefix))
                        bdraggcfg.append('/configure service vprn "%s" static-routes route %s route-type unicast blackhole preference 131' % (bdrvprn[2],staticsumprefix))
                        bdraggcfg.append('/configure service vprn "%s" static-routes route %s route-type unicast blackhole prefix-list name "Sum_%s"' % (bdrvprn[2],staticsumprefix,staticsumprefix))

        f = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + bdraggfilename, "a")
        for cfg in bdraggcfg:
            f.write(cfg + '\n')
        f.close()

    if impcom != None:
        bdrimpcom = impcom.replace('65000','65051')
    else:
        bdrimpcom = ''
    if expcom != None:
        bdrexpcom = expcom.replace('65000','65051')
    else:
        bdrexpcom = ''
    vprnglobalws['A%s' % row] = bdrvprn[0] #'nokia-node'
    vprnglobalws['B%s' % row] = bdrvprn[1] #'system-ip'
    vprnglobalws['C%s' % row] = bdrvprn[2] #'service-name'
    vprnglobalws['D%s' % row] = bdrvprn[3] #'service-id'
    vprnglobalws['E%s' % row] = exvprnglobal['iptnservicestate'] #'admin-state'
    vprnglobalws['F%s' % row] = bdrrd #'rd'
    vprnglobalws['G%s' % row] = 'vrf_L3_%s' % exvprnglobal['iptnrtimppolname'] #'vrf-import-policy-name'
    vprnglobalws['H%s' % row].hyperlink = bdrpolicyimppath #'vrf-import-policy'
    vprnglobalws['I%s' % row].hyperlink = bdrprefiximppath #'vrf-import-prefix'
    vprnglobalws['J%s' % row] = bdrimpcom #'vrf-import-commu'
    vprnglobalws['K%s' % row] = 'vrf_L3_%s' % exvprnglobal['iptnrtexppolname'] #'vrf-export-policy-name'
    vprnglobalws['L%s' % row].hyperlink = bdrpolicyexppath #'vrf-export-policy'
    vprnglobalws['M%s' % row].hyperlink = bdrprefixexppath #'vrf-export-prefix'
    vprnglobalws['N%s' % row] = bdrexpcom #'vrf-export-commu'
    vprnglobalws['O%s' % row] = bdrexpcom #'vrf-target'
    vprnglobalws['P%s' % row].hyperlink = bdraggfilepath #'aggregate-route'
    vprnglobalws['Q%s' % row] = qos #'customer-id'
    vprnglobalws['R%s' % row] = bdrvprn[4] #'iptn-node'
    vprnglobalws['S%s' % row] = exvprnglobal['iptnservice'] #'iptn-service'
    row += 1
    vprnglobalws['A1'] = row

def createvprnintworksheet(vprnintws,vprnint):
    row = vprnintws['A1'].value
    vprnintws['A%s' % row]  = vprnint[0] # 'nokia-node'
    vprnintws['B%s' % row]  = vprnint[1] # 'system-ip'
    vprnintws['C%s' % row]  = vprnint[2] # 'service-name'
    vprnintws['D%s' % row]  = vprnint[3] # 'service-id'
    vprnintws['E%s' % row]  = vprnint[4] # 'interface-name'
    vprnintws['F%s' % row]  = vprnint[5] # 'interface description'
    vprnintws['G%s' % row]  = vprnint[6] # 'sap'
    vprnintws['H%s' % row]  = vprnint[7] # 'sap-ingress-filter'
    vprnintws['I%s' % row]  = vprnint[8] # 'sap-egress-filter'
    vprnintws['J%s' % row]  = vprnint[9] # 'sap-in/egress qos'
    vprnintws['K%s' % row]  = vprnint[10] # 'ip/ipv6 address'
    vprnintws['L%s' % row]  = vprnint[11] # 'secondary ip/ipv6'
    vprnintws['M%s' % row]  = vprnint[12] # 'family'
    vprnintws['N%s' % row]  = vprnint[13] # 'vrrp id'
    vprnintws['O%s' % row]  = vprnint[14] # 'vrrp vip'
    vprnintws['P%s' % row]  = vprnint[15] # 'vrrp priority'
    vprnintws['Q%s' % row]  = vprnint[16] # 'vrrp interval'
    vprnintws['R%s' % row].hyperlink  = vprnint[17] # 'static-route'
    vprnintws['S%s' % row]  = vprnint[18] # 'bgp'
    vprnintws['T%s' % row]  = vprnint[19] # 'group'
    vprnintws['U%s' % row]  = vprnint[20] # 'peer-as'
    vprnintws['V%s' % row]  = vprnint[21] # 'neighbor'
    vprnintws['W%s' % row]  = vprnint[22] # 'neighbor-description'
    vprnintws['X%s' % row]  = vprnint[23] # 'bgp-imp-policy-name'
    vprnintws['Y%s' % row].hyperlink  = vprnint[24] # 'bgp-imp-policy'
    vprnintws['Z%s' % row].hyperlink  = vprnint[25] # 'bgp-imp-policy-prefix'
    vprnintws['AA%s' % row] = vprnint[26] # 'bgp-imp-policy-commu'
    vprnintws['AB%s' % row] = vprnint[27] # 'bgp-exp-policy-name'
    vprnintws['AC%s' % row].hyperlink = vprnint[28] # 'bgp-exp-policy'
    vprnintws['AD%s' % row].hyperlink = vprnint[29] # 'bgp-exp-policy-prefix'
    vprnintws['AE%s' % row] = vprnint[30] # 'bgp-exp-policy-commu'
    vprnintws['AF%s' % row] = vprnint[31] # 'bgp-bfd-interval'
    vprnintws['AG%s' % row] = vprnint[32] # 'bgp-bfd-multiply'
    vprnintws['AH%s' % row] = vprnint[33] # 'iptn-node'
    vprnintws['AI%s' % row] = vprnint[34] # 'iptn-service'
    vprnintws['AJ%s' % row] = vprnint[35] # 'iptn-interface'
    row += 1
    vprnintws['A1'] = row

def createvprnintinfo(vprnintws, sapqos, namingdict, filterin, filterout, bdrservicetype,bdrserviceid, exvprnintdict,policynamews):
    vprnintlist = []
    bdrsap = ''
    bdrstatic = ''
    bdrfeature = ''
    vprnintdata = ''
    impcom = None
    expcom = None
    bdrpolicyimppath = ''
    bdrprefiximppath = ''
    bdrpolicyexppath = ''
    bdrprefixexppath = ''
    if namingdict['bdrportencap'] == 'null':
        bdrsap = namingdict['bdrport']  # 'sap'
    elif namingdict['bdrportencap'] == 'dot1q':
        bdrsap = '%s:%s' % (namingdict['bdrport'], namingdict['bdrvlan'])  # 'sap'

    allvprn = namingdict['bdrnode']+'$%'+namingdict['bdrsystem']+'$%'+namingdict['bdrservice']+'$%'+str(bdrserviceid)+'$%'+namingdict['iptnnode']+'$%'+exvprnintdict['iptnrd']+'$%'+sapqos
    f = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'allvprn.txt', "a")
    f.write(allvprn + '\n')
    f.close()

    if exvprnintdict['iptnstatic'] != None:
        s = open(extractpath + exvprnintdict['iptnstatic'], 'r')
        bdrstatic = createbdrstaticconfig(s ,namingdict,namingdict['bdrservice'],exvprnintdict['iptnfam'])
    if exvprnintdict['iptnbgp'] != None:
        b = open(extractpath + exvprnintdict['iptnbgp'], 'r')
        bgplines = b.readlines()
        b.close()
        for a in bgplines:
            #print (a)
            if 'as-override' in a:
                bdrfeature += ' as-override'
            if 'remove-private' in a:
                bdrfeature += ' remove-private'
            if 'authentication-key' in a:
                bdrfeature += ' authentication-key'
            if 'prefix-limit maximum' in a:
                bdrfeature += ' prefix-limit-maximum'+str(a.split()[-1])
            if 'prefix-limit teardown' in a:
                bdrfeature += ' prefix-limit-teardown'+str(a.split()[-1])
    if exvprnintdict['iptnbgpimp'] != None:
        impprefix = exvprnintdict['iptnbgpimppre']
        impcom = exvprnintdict['iptnbgpimpcom']
        imp = open(extractpath + exvprnintdict['iptnbgpimp'], 'r')
        bdrpolicyimppath,bdrprefiximppath = createbdrpolicy(imp, namingdict['bdrnode'],namingdict['bdrservice'],exvprnintdict['iptnbgpimpname'], impprefix, impcom,policynamews,None,'ebgpimp')
    if exvprnintdict['iptnbgpexp'] != None:
        expprefix = exvprnintdict['iptnbgpexppre']
        expcom = exvprnintdict['iptnbgpexpcom']
        exp = open(extractpath + exvprnintdict['iptnbgpexp'], 'r')
        bdrpolicyexppath,bdrprefixexppath = createbdrpolicy(exp, namingdict['bdrnode'],namingdict['bdrservice'],exvprnintdict['iptnbgpexpname'], expprefix, expcom,policynamews,None,'ebgpexp')

    vprnintlist.append(namingdict['bdrnode'])
    vprnintlist.append(namingdict['bdrsystem'])
    vprnintlist.append(namingdict['bdrservice'])
    vprnintlist.append(str(bdrserviceid))
    vprnintlist.append(namingdict['bdrl3int'])
    vprnintlist.append(namingdict['bdrl3desc'])
    vprnintlist.append(bdrsap)
    vprnintlist.append(filterin)
    vprnintlist.append(filterout)
    vprnintlist.append(sapqos)
    vprnintlist.append(exvprnintdict['iptnip'])
    vprnintlist.append(exvprnintdict['iptnsecip'])
    vprnintlist.append(exvprnintdict['iptnfam'])
    vprnintlist.append(exvprnintdict['iptnvrrpid'])
    vprnintlist.append(exvprnintdict['iptnvrrpvip'])
    vprnintlist.append(exvprnintdict['iptnvrrppri'])
    vprnintlist.append(exvprnintdict['iptnvrrpinterval'])
    vprnintlist.append(bdrstatic) # 'static-route'
    vprnintlist.append(bdrfeature) # 'bgp'
    bdrbgpgroup = exvprnintdict['iptnbgpgroup']
    if exvprnintdict['iptnbgpgroup'] != None:
        if 'ebgp' in exvprnintdict['iptnbgpgroup']:
            bdrbgpgroup = bdrbgpgroup.replace('ebgp', 'eBGP')
        if 'EBGP' in exvprnintdict['iptnbgpgroup']:
            bdrbgpgroup = bdrbgpgroup.replace('EBGP', 'eBGP')
        if 'eBgp' in exvprnintdict['iptnbgpgroup']:
            bdrbgpgroup = bdrbgpgroup.replace('eBgp', 'eBGP')
        if 'HUAWEI' in exvprnintdict['iptnbgpgroup']:
            bdrbgpgroup = bdrbgpgroup.replace('HUAWEI','H')
        if 'Huawei' in exvprnintdict['iptnbgpgroup']:
            bdrbgpgroup = bdrbgpgroup.replace('Huawei', 'H')
        if 'VLAN' in exvprnintdict['iptnbgpgroup']:
            bdrbgpgroup = bdrbgpgroup.replace('VLAN', 'V')
        if 'Vlan' in exvprnintdict['iptnbgpgroup']:
            bdrbgpgroup = bdrbgpgroup.replace('Vlan', 'V')
    vprnintlist.append(bdrbgpgroup)
    vprnintlist.append(exvprnintdict['iptnbgppeeras'])
    vprnintlist.append(exvprnintdict['iptnbgppeer'])
    if exvprnintdict['iptnbgppeerdesc'] != None:
        vprnintlist.append(exvprnintdict['iptnbgppeerdesc'].strip('"'))
    else:
        vprnintlist.append(exvprnintdict['iptnbgppeerdesc'])
    vprnintlist.append(exvprnintdict['iptnbgpimpname'])
    vprnintlist.append(bdrpolicyimppath) # 'bgp-imp-policy'
    vprnintlist.append(bdrprefiximppath) # 'bgp-imp-policy-prefix'
    if impcom != None:
        impcom = impcom.replace('65000','65051')
        vprnintlist.append(impcom)  # 'bgp-imp-policy-commu'
    else:
        vprnintlist.append('')  # 'bgp-imp-policy-commu'
    vprnintlist.append(exvprnintdict['iptnbgpexpname'])
    vprnintlist.append(bdrpolicyexppath) # 'bgp-exp-policy'
    vprnintlist.append(bdrprefixexppath) # 'bgp-exp-policy-prefix'
    if expcom != None:
        expcom = expcom.replace('65000','65051')
        vprnintlist.append(expcom) # 'bgp-exp-policy-commu'
    else:
        vprnintlist.append('')  # 'bgp-exp-policy-commu'
    vprnintlist.append(exvprnintdict['iptnbgpbfdinterval'])
    vprnintlist.append(exvprnintdict['iptnbgpbfdmul'])
    vprnintlist.append(namingdict['iptnnode'])
    vprnintlist.append(namingdict['iptnservice'])
    vprnintlist.append(namingdict['iptnintunit'])

    for vprnint in vprnintlist:
        if vprnint == None:
            vprnint = ''
        vprnintdata = vprnintdata+vprnint+'$%'

    #print(vprnintdata)
    eg = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'servicevprnintinfo.txt', "a")
    eg.write(vprnintdata + '\n')

    relatedbgpcheck = []
    iptnbgprelatedpelist = []
    relatedbgplastservice = ''
    if os.path.exists(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'related-bgp-pe.txt'):
        rs = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'related-bgp-pe.txt', "r")
        for lrs in rs.readlines():
            lrs = lrs.rstrip('\n')
            relatedbgpcheck.append(lrs)
    for i in range(3, exvprnint.max_row + 1):
        if namingdict['iptnservice'] == exvprnint['C%s' % i].value:
            if exvprnintdict['iptnbgppeeras'] != None:
                if exvprnintdict['iptnbgppeeras'] == exvprnint['U%s' % i].value:
                    if exvprnint['C%s' % i].value != relatedbgplastservice:
                        iptnbgprelatedpelist.append('#' * 120)
                        relatedbgplastservice = exvprnint['C%s' % i].value
                    iptnbgpint = exvprnint['E%s' % i].value + '.' + exvprnint['F%s' % i].value
                    if 'BGP Peer to AS %s %s %s %s' %(exvprnint['U%s' % i].value, exvprnint['A%s' % i].value,exvprnint['C%s' % i].value,iptnbgpint) not in relatedbgpcheck:
                        iptnbgprelatedpelist.append('BGP Peer to AS %s %s %s %s' %(exvprnint['U%s' % i].value, exvprnint['A%s' % i].value,exvprnint['C%s' % i].value,iptnbgpint))


    bf = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'related-bgp-pe.txt', "a")
    for bs in iptnbgprelatedpelist:
        bf.write(bs + '\n')
    bf.close()

def createepipesapinfo(epipesapws, sapqos, namingdict, filterin, filterout, bdrservicetype, bdrserviceid, exl2vpndict):
    ingressrate = ''
    egressrate = ''
    bdrsap = ''
    if exl2vpndict['iptnpolicerin'] != None:
        ingressrate = exl2vpndict['iptnpolicerin'].split('_')[1]
    if exl2vpndict['iptnpolicerout'] != None:
        egressrate = exl2vpndict['iptnpolicerout'].split('_')[1]
    if namingdict['bdrportencap'] == 'null':
        bdrsap = namingdict['bdrport'] # 'sap'
    elif namingdict['bdrportencap'] == 'dot1q':
        bdrsap = '%s:%s' %(namingdict['bdrport'],namingdict['bdrvlan']) #'sap'
    epipesapdata = namingdict['bdrnode']+'$%'+namingdict['bdrsystem']+'$%'+namingdict['bdrservice']+'$%'+str(bdrserviceid)+'$%'+bdrsap+'$%'+namingdict['bdrsapdesc']+'$%'+filterin+'$%'+filterout+'$%'+sapqos+'$%'+ingressrate+'$%'+egressrate+'$%'+namingdict['iptnnode']+'$%'+namingdict['iptnservice']+'$%'+namingdict['iptnintunit']
    eg = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceepipesapinfo.txt', "a")
    eg.write(epipesapdata + '\n')


def createvplssapinfo(vplssapws, sapqos, namingdict, filterin, filterout, bdrservicetype, bdrserviceid):
    bdrsap = ''
    if namingdict['bdrportencap'] == 'null':
        bdrsap = namingdict['bdrport'] # 'sap'
    elif namingdict['bdrportencap'] == 'dot1q':
        bdrsap = '%s:%s' %(namingdict['bdrport'],namingdict['bdrvlan']) #'sap'
    vplssapdata = namingdict['bdrnode']+'$%'+namingdict['bdrsystem']+'$%'+namingdict['bdrservice']+'$%'+str(bdrserviceid)+'$%'+bdrsap+'$%'+namingdict['bdrsapdesc']+'$%'+filterin+'$%'+filterout+'$%'+sapqos+'$%'+namingdict['iptnnode']+'$%'+namingdict['iptnservice']+'$%'+namingdict['iptnintunit']
    eg = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'servicevplssapinfo.txt', "a")
    eg.write(vplssapdata + '\n')

def creatervplssapinfo(rvplssapws, sapqos, namingdict, filterin, filterout, bdrservicetype, bdrserviceid,iptnservicetype):
    bdrsap = ''
    if namingdict['bdrportencap'] == 'null':
        bdrsap = namingdict['bdrport'] # 'sap'
    elif namingdict['bdrportencap'] == 'dot1q':
        bdrsap = '%s:%s' %(namingdict['bdrport'],namingdict['bdrvlan']) #'sap'
    rvplssapdata = namingdict['bdrnode']+'$%'+namingdict['bdrsystem']+'$%'+namingdict['bdrservice']+'$%'+str(bdrserviceid)+'$%'+bdrsap+'$%'+namingdict['bdrsapdesc']+'$%'+filterin+'$%'+filterout+'$%'+sapqos+'$%'+namingdict['iptnnode']+'$%'+namingdict['iptnservice']+'$%'+namingdict['iptnintunit']+'$%'+iptnservicetype
    eg = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'servicervplssapinfo.txt', "a")
    eg.write(rvplssapdata + '\n')


def createl2servicerelatedworksheet(relatedpews, l2service, namingdict, routetarget,iptnservicetype,bdrservicetype,bdrserviceid, iptnfccolumn, bdrnodecolumn,bdrsystemcolumn):
    row = relatedpews['A1'].value
    servicecheck = []
    custid = ''
    lastcustid = ''
    lastiptnservice = ''
    f = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'cfg'+'\\'+ 'l2servicecheck.txt', "a")
    g = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'cfg'+'\\'+ 'l2servicecheck.txt', 'r').readlines()
    #print('g : %s' % g)
    for a in range(len(g)):
        g[a] = g[a].rstrip('\n')  # remove newline('\n') from end of line
        if g[a] not in servicecheck:
            servicecheck.append(g[a])


    #print(servicecheck)
    for i in range(3,l2service.max_row + 1):
        #print(l2service['E%s'%i].value)
        if routetarget == l2service['E%s'%i].value:
            iptnnode = l2service['A%s' % i].value
            iptnservice = l2service['C%s' % i].value
            bdrnode = l2service['%s%s' %(bdrnodecolumn,i)].value
            bdrsystem = l2service['%s%s' % (bdrsystemcolumn, i)].value
            if l2service['%s%s'%(iptnfccolumn,i)].value == 'mobile':
                custid = 'mobile'
            elif l2service['%s%s'%(iptnfccolumn,i)].value == 'corporate':
                custid = 'corporate'
            elif l2service['%s%s'%(iptnfccolumn,i)].value == 'internal':
                custid = 'internal'
            elif l2service['%s%s'%(iptnfccolumn,i)].value == 'network-control':
                custid = 'network-control'
            elif l2service['%s%s'%(iptnfccolumn,i)].value == 'best-effort':
                custid = 'best-effort'
            elif l2service['%s%s'%(iptnfccolumn,i)].value == 'exp_classifier_ipcbb':
                custid = 'ipcbb'

            s = '%s%s%s'%(iptnnode,iptnservicetype,iptnservice)
            if s not in servicecheck:
                servicecheck.append(s)
                relatedpews['A%s' % row] = iptnnode
                relatedpews['B%s' % row] = iptnservicetype
                relatedpews['C%s' % row] = iptnservice
                relatedpews['D%s' % row] = routetarget
                if iptnservicetype == 'EPIPE':
                    relatedpews['E%s' % row] = l2service['I%s' % i].value
                    relatedpews['F%s' % row] = l2service['G%s' % i].value
                elif iptnservicetype == 'VPLS':
                    relatedpews['E%s' % row] = l2service['H%s' % i].value
                    relatedpews['F%s' % row] = l2service['G%s' % i].value
                else:
                    relatedpews['E%s' % row] = ''
                    relatedpews['F%s' % row] = ''
                relatedpews['G%s' % row] = bdrnode
                relatedpews['H%s' % row] = bdrservicetype
                relatedpews['I%s' % row] = namingdict['bdrservice']
                relatedpews['J%s' % row] = bdrserviceid
                relatedpews['K%s' % row] = custid

                if lastiptnservice == iptnservice:
                    if lastcustid != custid:
                        lastrow = int(row)-1
                        relatedpews['K%s' % lastrow].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                        relatedpews['K%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                lastiptnservice = iptnservice
                lastcustid = custid

                row += 1
                relatedpews['A1'] = row
                f.write(s + '\n')
            #print(servicecheck)
    f.close()


def createbdservicerelateworksheet(relatedpews,exbridge,namingdict,iptnservice,iptnservicetype,bdrservicetype,bdrserviceid):
    row = relatedpews['A1'].value
    servicecheck = []
    custid = ''
    lastcustid = ''
    lastiptnservice = ''
    peerid = ''
    peerlist = []
    #print(iptnservice)
    f = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'cfg'+'\\'+ 'l2servicecheck.txt', "a")
    g = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'cfg'+'\\'+ 'l2servicecheck.txt', 'r').readlines()
    #print('g : %s' % g)

    for n in range(3, l2mapsheet.max_row + 1):
        if namingdict['iptnnode'] == l2mapsheet['A%s' % n].value:
            peerid = l2mapsheet['C%s' % n].value

    for m in range(3, l2mapsheet.max_row + 1):
        if peerid == l2mapsheet['C%s' % m].value:
            peerlist.append(l2mapsheet['A%s' % m].value)

    for a in range(len(g)):
        g[a] = g[a].rstrip('\n')  # remove newline('\n') from end of line
        if g[a] not in servicecheck:
            servicecheck.append(g[a])

    for i in range(3, exbridge.max_row + 1):
        if iptnservice == exbridge['C%s' % i].value:
            if exbridge['A%s' % i].value in peerlist:
                #if 'ae0' == exbridge['H%s' % i].value:
                iptnnode = exbridge['A%s' % i].value
                bdrnode = exbridge['N%s' % i].value
                if exbridge['M%s' % i].value == 'mobile':
                    custid = 'mobile'
                elif exbridge['M%s' % i].value == 'corporate':
                    custid = 'corporate'
                elif exbridge['M%s' % i].value == 'internal':
                    custid = 'internal'
                elif exbridge['M%s' % i].value == 'network-control':
                    custid = 'network-control'
                elif exbridge['M%s' % i].value == 'best-effort':
                    custid = 'best-effort'
                elif exbridge['M%s' % i].value == 'exp_classifier_ipcbb':
                    custid = 'ipcbb'
                s = '%s%s%s' % (iptnnode, iptnservicetype, iptnservice)
                if s not in servicecheck:
                    if namingdict['vrrpviasw'] != 'yes':
                        servicecheck.append(s)
                        relatedpews['A%s' % row] = iptnnode
                        relatedpews['B%s' % row] = iptnservicetype
                        relatedpews['C%s' % row] = iptnservice
                        relatedpews['D%s' % row] = 'N/A'
                        relatedpews['E%s' % row] = ''
                        relatedpews['F%s' % row] = ''
                        relatedpews['G%s' % row] = bdrnode
                        relatedpews['H%s' % row] = bdrservicetype
                        relatedpews['I%s' % row] = namingdict['bdrservice']
                        relatedpews['J%s' % row] = bdrserviceid
                        relatedpews['K%s' % row] = custid
                        if lastiptnservice == iptnservice:
                            if lastcustid != custid:
                                lastrow = int(row)-1
                                relatedpews['K%s' % lastrow].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                                relatedpews['K%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                        lastiptnservice = iptnservice
                        lastcustid = custid
                        row += 1
                        relatedpews['A1'] = row
                        f.write(s + '\n')

def createvswservicerelatedworksheet(relatedpews, exvsw, namingdict, routetarget,iptnservice, iptnservicetype,bdrservicetype,bdrserviceid, iptnfccolumn, bdrnodecolumn,bdrsystemcolumn):
    row = relatedpews['A1'].value
    servicecheck = []
    custid = ''
    lastcustid = ''
    lastiptnservice = ''
    #print(bdrserviceid)
    f = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'cfg'+'\\'+ 'l2servicecheck.txt', "a")
    g = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'cfg'+'\\'+ 'l2servicecheck.txt', 'r').readlines()
    #print('g : %s' % g)
    for a in range(len(g)):
        g[a] = g[a].rstrip('\n')  # remove newline('\n') from end of line
        if g[a] not in servicecheck:
            servicecheck.append(g[a])


    #print(servicecheck)
    for i in range(3,exvsw.max_row + 1):
        #print(exvsw['E%s'%i].value)
        if routetarget == exvsw['E%s'%i].value:
            if iptnservice == exvsw['I%s'%i].value:
                iptnnode = exvsw['A%s' % i].value
                bdrnode = exvsw['%s%s' %(bdrnodecolumn,i)].value
                bdrsystem = exvsw['%s%s' % (bdrsystemcolumn, i)].value
                if exvsw['%s%s'%(iptnfccolumn,i)].value == 'mobile':
                    custid = 'mobile'
                elif exvsw['%s%s'%(iptnfccolumn,i)].value == 'corporate':
                    custid = 'corporate'
                elif exvsw['%s%s'%(iptnfccolumn,i)].value == 'internal':
                    custid = 'internal'
                elif exvsw['%s%s'%(iptnfccolumn,i)].value == 'network-control':
                    custid = 'network-control'
                elif exvsw['%s%s'%(iptnfccolumn,i)].value == 'best-effort':
                    custid = 'best-effort'
                elif exvsw['%s%s'%(iptnfccolumn,i)].value == 'exp_classifier_ipcbb':
                    custid = 'ipcbb'

                s = '%s%s%s'%(iptnnode,iptnservicetype,iptnservice)
                if s not in servicecheck:
                    if namingdict['vrrpviasw'] != 'yes':
                        servicecheck.append(s)
                        relatedpews['A%s' % row] = iptnnode
                        relatedpews['B%s' % row] = iptnservicetype
                        relatedpews['C%s' % row] = iptnservice
                        relatedpews['D%s' % row] = routetarget
                        relatedpews['E%s' % row] = ''
                        relatedpews['F%s' % row] = ''
                        relatedpews['G%s' % row] = bdrnode
                        relatedpews['H%s' % row] = bdrservicetype
                        relatedpews['I%s' % row] = namingdict['bdrservice']
                        relatedpews['J%s' % row] = bdrserviceid
                        relatedpews['K%s' % row] = custid

                        if lastiptnservice == iptnservice:
                            if lastcustid != custid:
                                lastrow = int(row)-1
                                relatedpews['K%s' % lastrow].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                                relatedpews['K%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                        lastiptnservice = iptnservice
                        lastcustid = custid

                        row += 1
                        relatedpews['A1'] = row
                        f.write(s + '\n')
                #print(servicecheck)
    f.close()

def createbdrevpnworksheet(misum,i,n,bdrlist,noentropyws):
    l2relatedws = misum['l2service-related-pe']
    miportws = misum['port-lag']
    l2evpnws = misum['l2-evpn']
    oddlocalpe = ''
    evenlocalpe = ''
    oddremotepe = ''
    evenremotepe = ''
    lastremotepe = ''
    peerid = ''
    sap = ''
    vlanid = ''
    bdrnode = l2relatedws['G%s' % i].value
    row = l2evpnws['A1'].value
    miportsystem = ''
    preference = ''
    port = ''
    localethtag = ''
    remoteethtag = ''
    entropy = 'yes'
    for bdrn in bdrlist:
        if l2relatedws['G%s' % i].value.split('_')[1] == bdrn.split('_')[1]:
            if (int(bdrn[-2:]) %2) == 0 and (int(l2relatedws['G%s' % i].value[-2:]) %2) == 1:
                oddlocalpe = l2relatedws['G%s' % i].value.split('_')[1] + l2relatedws['G%s' % i].value[-2:]
                evenlocalpe = bdrn.split('_')[1]+bdrn[-2:]
            elif (int(l2relatedws['G%s' % i].value[-2:]) %2) ==  0 and (int(bdrn[-2:]) %2) == 1:
                oddlocalpe = bdrn.split('_')[1]+bdrn[-2:]
                evenlocalpe = l2relatedws['G%s' % i].value.split('_')[1] + l2relatedws['G%s' % i].value[-2:]
        else:
            if lastremotepe:
                if (int(bdrn[-2:]) % 2) == 0 and (int(lastremotepe[-2:]) % 2) == 1:
                    oddremotepe = lastremotepe.split('_')[1] + lastremotepe[-2:]
                    evenremotepe = bdrn.split('_')[1] + bdrn[-2:]
                elif (int(lastremotepe[-2:]) % 2) == 0 and (int(bdrn[-2:]) % 2) == 1:
                    oddremotepe = bdrn.split('_')[1] + bdrn[-2:]
                    evenremotepe = lastremotepe.split('_')[1] + lastremotepe[-2:]
            lastremotepe = bdrn
    localac = '%s-%s' %(oddlocalpe,evenlocalpe)
    remoteac = '%s-%s'%(oddremotepe,evenremotepe)
    #print(bdrnode)
    #print(localac)
    #print(remoteac)
    vid = l2relatedws['D%s' %i].value.split(':')[-1]
    vpnid = f"{int(l2relatedws['D%s' %i].value.split(':')[-1]):06d}"
    for p in range(3, l2mapsheet.max_row + 1):
        if l2relatedws['A%s' %i].value == l2mapsheet['A%s' % p].value:
            peerid = l2mapsheet['C%s' % p].value
    for v in range(3, miportws.max_row + 1):
        if miportws['A%s' %v].value == l2relatedws['G%s' % i].value: # if bdrnode match
            if miportws['C%s' %v].value == 'EPIPE-EVPN': # if service-type evpn
                if miportws['D%s' % v].value == l2relatedws['I%s' % i].value: # if service name match
                    miportsystem = miportws['B%s' % v].value
                    if l2relatedws['E%s' % i].value == 'primary':
                        preference = '200'
                    elif l2relatedws['E%s' % i].value == 'backup':
                        preference = '100'
                    if ':' not in miportws['O%s' % v].value:
                        sap = miportws['O%s' % v].value
                        port = miportws['O%s' % v].value
                        vlan = ''
                        vlanid = '0000'
                    else:
                        sap = miportws['O%s' % v].value
                        port = miportws['O%s' % v].value.split(':')[0]
                        vlan = miportws['O%s' % v].value.split(':')[-1]
                        vlanid = f"{int(miportws['O%s' % v].value.split(':')[-1]):04d}"
    for e in range(3, l2relatedws.max_row + 1):
        if l2relatedws['D%s' % i].value == l2relatedws['D%s' % e].value:
            if l2relatedws['G%s' % i].value.split('_')[1] != l2relatedws['G%s' % e].value.split('_')[1]:
                localethtag = l2relatedws['F%s' % i].value
                remoteethtag = l2relatedws['F%s' % e].value
    for er in range(3, noentropyws.max_row + 1):
        if l2relatedws['A%s' %i].value == noentropyws['A%s' % er].value and l2relatedws['C%s' %i].value == noentropyws['B%s' % er].value:
            entropy = 'yes'

    esi = '00'+peerid+vpnid+vlanid+'000000'
    ethersegment = localac+':'+sap

    if sap:
        l2evpnws['A%s' % row] = l2relatedws['G%s' % i].value #'nokia-node'
        l2evpnws['B%s' % row] = miportsystem #'system-ip'
        l2evpnws['C%s' % row] = l2relatedws['I%s' % i].value #'service-name'
        l2evpnws['D%s' % row] = l2relatedws['J%s' % i].value #'service-id'
        l2evpnws['E%s' % row] = ethersegment #'ethernet-segment'
        l2evpnws['F%s' % row] = esi #'esi'
        l2evpnws['G%s' % row] = preference #'preference'
        l2evpnws['H%s' % row] = port #'port/lag'
        l2evpnws['I%s' % row] = vlan  # 'vlan'
        l2evpnws['J%s' % row] = localac #'local-ac-name'
        l2evpnws['K%s' % row] = localethtag #'local-eth-tag'
        l2evpnws['L%s' % row] = remoteac #'remote-ac-name'
        l2evpnws['M%s' % row] = remoteethtag #'remote-eth-tag'
        l2evpnws['N%s' % row] = vid #'evi'
        l2evpnws['O%s' % row] = l2relatedws['K%s' %i].value #'customer-id'
        l2evpnws['P%s' % row] = entropy #
        createl2interasevpnworksheet(misum,l2relatedws,i,miportsystem,ethersegment,esi,preference,port,vlan,vlanid,localac,localethtag,vid,vpnid,entropy)
        if l2evpnws['C%s' % str(int(row) - 1)].value == l2evpnws['C%s' % row].value:
            if l2evpnws['O%s' % str(int(row) - 1)].value != l2evpnws['O%s' % row].value:
                lastrow = int(row) - 1
                l2evpnws['O%s' % lastrow].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                l2evpnws['O%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957', fill_type='solid')

        row += 1
        l2evpnws['A1'] = row

def createl2interasevpnworksheet(misum,l2relatedws,i,miportsystem,ethersegment,esi,preference,port,vlan,vlanid,localac,localethtag,vid,vpnid,entropy):
    l2evpninterws = misum['l2-evpn-inter-as']
    asbrethtag = '9'

    for asbrline in range(3, asbrmapws.max_row + 1):
        asbrdict = {}
        asbrlist = []
        if l2relatedws['G%s' % i].value == asbrmapws['A%s' % asbrline].value:
            asbrdict['asbr_node'] = asbrmapws['C%s' % asbrline].value
            asbrdict['asbr_preference'] = asbrmapws['D%s' % asbrline].value
            asbrdict['asbr_sap'] = asbrmapws['E%s' % asbrline].value
            asbr1 = asbrmapws['C%s' % asbrline].value.split('_')[1] + asbrmapws['C%s' % asbrline].value[-2:]
            asbrlist.append(asbrdict.copy())
            asbrdict['asbr_node'] = asbrmapws['F%s' % asbrline].value
            asbrdict['asbr_preference'] = asbrmapws['G%s' % asbrline].value
            asbrdict['asbr_sap'] = asbrmapws['H%s' % asbrline].value
            asbr2 = asbrmapws['F%s' % asbrline].value.split('_')[1] + asbrmapws['F%s' % asbrline].value[-2:]
            asbrlist.append(asbrdict.copy())
            asbrac = '%s-%s' % (asbr1, asbr2)
            asbrbridgevlan = 'xx%s' %i
            for ve in range(2, asbrepipevlansheet.max_row + 1):
                if asbrepipevlansheet['A%s' % ve].value != None:
                    if l2relatedws['C%s' % i].value == asbrepipevlansheet['A%s' % ve].value:
                        asbrbridgevlan = asbrepipevlansheet['B%s' % ve].value

            for vv in range(2, asbrvplsvlansheet.max_row + 1):
                if asbrvplsvlansheet['A%s' % vv].value != None:
                    if l2relatedws['C%s' % i].value == asbrvplsvlansheet['A%s' % vv].value:
                        asbrbridgevlan = asbrvplsvlansheet['B%s' % vv].value

            if 'xx' in str(asbrbridgevlan):
                asbrvlanethseg = f"{int(i):04d}"
            else:
                asbrvlanethseg = f"{int(asbrbridgevlan):04d}"
            asbresi = '00' + '99' + vpnid + str(asbrvlanethseg) + '000000'
            asbrethersegment = asbrac + ':' + asbrdict['asbr_sap'] + ':' + str(asbrbridgevlan)
            for asbr in asbrlist:
                interasrow = l2evpninterws['A1'].value
                l2evpninterws['A%s' % interasrow] = l2relatedws['G%s' % i].value #'nokia-node'
                l2evpninterws['B%s' % interasrow] = miportsystem #'system-ip'
                l2evpninterws['C%s' % interasrow] = l2relatedws['I%s' % i].value #'service-name'
                l2evpninterws['D%s' % interasrow] = l2relatedws['J%s' % i].value #'service-id'
                l2evpninterws['E%s' % interasrow] = ethersegment #'ethernet-segment'
                l2evpninterws['F%s' % interasrow] = esi #'esi'
                l2evpninterws['G%s' % interasrow] = preference #'preference'
                l2evpninterws['H%s' % interasrow] = port #'port/lag'
                l2evpninterws['I%s' % interasrow] = vlan #'vlan'
                l2evpninterws['J%s' % interasrow] = localac #'pe-ac-name'
                l2evpninterws['K%s' % interasrow] = localethtag #'pe-eth-tag'
                l2evpninterws['L%s' % interasrow] = asbrac #'asbr-ac-name'
                l2evpninterws['M%s' % interasrow] = asbrethtag #'asbr-eth-tag'
                l2evpninterws['N%s' % interasrow] = vid #'evi'
                l2evpninterws['O%s' % interasrow] = asbr['asbr_node'] #'asbr-node'
                l2evpninterws['P%s' % interasrow] = asbrethersegment #'asbr-ethernet-segment'
                l2evpninterws['Q%s' % interasrow] = asbresi #'asbr-esi'
                if asbr['asbr_preference'] == 'primary':
                    l2evpninterws['R%s' % interasrow] = '200' #'asbr-preference'
                elif asbr['asbr_preference'] == 'backup':
                    l2evpninterws['R%s' % interasrow] = '100' #'asbr-preference'
                l2evpninterws['S%s' % interasrow] = asbr['asbr_sap'] #'inter-as-lag'
                l2evpninterws['T%s' % interasrow] = vid #'md-index'
                l2evpninterws['U%s' % interasrow] = vid #'md-name'
                l2evpninterws['V%s' % interasrow] = vid #'ma-index'
                l2evpninterws['W%s' % interasrow] = l2relatedws['I%s' % i].value #'bridge-id'
                l2evpninterws['X%s' % interasrow] = asbrbridgevlan #'bridge-vlan'
                l2evpninterws['Y%s' % interasrow] = '20' #'local-mep-id'
                l2evpninterws['Z%s' % interasrow] = '10' #'remote-mep-id'
                l2evpninterws['AA%s' % interasrow] = l2relatedws['K%s' %i].value #'customer-id'
                l2evpninterws['AB%s' % interasrow] = entropy
                interasrow += 1
                l2evpninterws['A1'] = interasrow

def createl2interassdpworksheet(misum,sdpdict,l2relatedws,i,n, asbr,noentropyws):
    l2interasws = misum['l2inter-as']
    interasrow = l2interasws['A1'].value
    entropy = 'yes'
    asbrbridgevlan = 'xx%s'% i
    for sd in range(2, sdpmapws.max_row + 1):
        if l2relatedws['G%s' %i].value == sdpmapws['E%s' % sd].value:
            pesdp = sdpmapws['H%s' % sd].value
        if asbr['asbr_node'] == sdpmapws['E%s' % sd].value:
            asbrip = sdpmapws['F%s' % sd].value
            asbrsdp = sdpmapws['H%s' % sd].value
            asbrsdpdesc = '%s:%s' %(asbr['asbr_node'],sdpdict['%s_id' % n])

    for er in range(3, noentropyws.max_row + 1):
        if l2relatedws['A%s' %i].value == noentropyws['A%s' % er].value and l2relatedws['C%s' %i].value == noentropyws['B%s' % er].value:
            entropy = 'yes'

    for ve in range(2, asbrepipevlansheet.max_row + 1):
        if asbrepipevlansheet['A%s' % ve].value != None:
            if l2relatedws['C%s' % i].value == asbrepipevlansheet['A%s' % ve].value:
                asbrbridgevlan = asbrepipevlansheet['B%s' % ve].value

    for vv in range(2, asbrvplsvlansheet.max_row + 1):
        if asbrvplsvlansheet['A%s' % vv].value != None:
            if l2relatedws['C%s' % i].value == asbrvplsvlansheet['A%s' % vv].value:
                asbrbridgevlan = asbrvplsvlansheet['B%s' % vv].value

    vpnid = l2relatedws['D%s' %i].value.split(':')[-1]
    l2interasws['A%s' % interasrow] = l2relatedws['G%s' %i].value #'PE-node'
    l2interasws['B%s' % interasrow] = sdpdict['%s_type' %n] #'service-type'
    l2interasws['C%s' % interasrow] = sdpdict['%s_name' %n] #'service-name'
    l2interasws['D%s' % interasrow] = sdpdict['%s_id' % n] #'service-id'
    l2interasws['E%s' % interasrow] = pesdp #'PE-sdp-number'
    l2interasws['F%s' % interasrow] = '%s:%s' %(l2relatedws['G%s' %i].value,sdpdict['%s_id' % n]) #'PE-sdp-description'
    l2interasws['G%s' % interasrow] = asbr['asbr_node'] #'asbr-node'
    l2interasws['H%s' % interasrow] = asbrip #'asbr-ip'
    l2interasws['I%s' % interasrow] = asbrsdp #'asbr-sdp-number'
    l2interasws['J%s' % interasrow] = asbrsdpdesc #'asbr-sdp-description'
    l2interasws['K%s' % interasrow] = sdpdict['%s_id' % n] #'vc-id'
    l2interasws['L%s' % interasrow] = asbr['asbr_preference'] # 'sdp-precedence'
    l2interasws['M%s' % interasrow] = 'IPTN' # 'endpoint-name'
    l2interasws['N%s' % interasrow] = asbr['asbr_sap'] # 'inter-as-sap'
    l2interasws['O%s' % interasrow] = vpnid # 'md-index'
    l2interasws['P%s' % interasrow] = vpnid # 'md-name'
    l2interasws['Q%s' % interasrow] = vpnid # 'ma-index'
    l2interasws['R%s' % interasrow] = sdpdict['%s_name' %n] # 'bridge-id'
    l2interasws['S%s' % interasrow] = asbrbridgevlan # 'bridge-vlan'
    l2interasws['T%s' % interasrow] = '20' # 'local-mep-id'
    l2interasws['U%s' % interasrow] = '10' # 'remote-mep-id'
    l2interasws['V%s' % interasrow] = l2relatedws['K%s' %i].value # 'customer-id'
    l2interasws['W%s' % interasrow] = entropy
    interasrow += 1
    l2interasws['A1'] = interasrow



def createl2servicesdpworksheet(misum):
    l2relatedws = misum['l2service-related-pe']
    noentropyws = openpyxl.load_workbook(inputpath + wbmap)['no-entropy']
    sdpws = misum['l2service-sdp']
    intersdp = misum['l2inter-as']
    row = sdpws['A1'].value
    sdpdict = {}
    sdpdict['service'] = []
    lastcustid = ''
    lastiptnservice = ''
    lastbdrn = []
    l2interaslist = []


    for i in range(3,l2relatedws.max_row + 1):
        n = l2relatedws['I%s' %i].value
        sdpdict['%s_name' %n] = n
        sdpdict['%s_id' %n] = l2relatedws['J%s' %i].value
        sdpdict['%s_type' %n] = l2relatedws['H%s' %i].value
        sdpdict.setdefault('%s_node' % n,[]).append(l2relatedws['G%s' %i].value)
        sdpdict['%s_entropy' % n] = 'yes'
        if n not in sdpdict['service']:
            sdpdict.setdefault('service', []).append(n)
        for er in range(3, noentropyws.max_row + 1):
            if l2relatedws['A%s' %i].value == noentropyws['A%s' % er].value and l2relatedws['C%s' %i].value == noentropyws['B%s' % er].value:
                sdpdict['%s_entropy' % n] = 'yes'

    for i in range(3, l2relatedws.max_row + 1):
        n = l2relatedws['I%s' % i].value # n = service name
        #print(sdpdict['%s_name' % n])
        #print(sdpdict['%s_id' % n])
        #print(sdpdict['%s_node' % n])
        for bdrn in sdpdict['%s_node' % n]:
            if bdrn != l2relatedws['G%s' %i].value:
                if sdpdict['%s_type' %n] != 'EPIPE-EVPN':
                    sdpws['A%s' % row] = l2relatedws['G%s' %i].value #'nokia-node'
                    sdpws['B%s' % row] = sdpdict['%s_type' %n] #'service-type'
                    sdpws['C%s' % row] = sdpdict['%s_name' %n] #'service-name'
                    sdpws['D%s' % row] = sdpdict['%s_id' % n] #'service-id'
                    sdpws['E%s' % row] = bdrn #'farend-node'
                    for sdp in range(2, sdpmapws.max_row+1):
                        #print(sdpmapws['E%s' % sdp].value)
                        if bdrn == sdpmapws['E%s'%sdp].value:
                            sdpws['F%s' % row] = sdpmapws['F%s'%sdp].value #'farend-ip'
                            sdpws['G%s' % row] = sdpmapws['H%s'%sdp].value #'sdp-number'
                            sdpws['H%s' % row] = '%s:%s' %(bdrn,sdpdict['%s_id' % n]) #'sdp-description'
                    sdpws['I%s' % row] = sdpdict['%s_id' % n] #'vc-id'
                    sdpws['J%s' % row] = l2relatedws['K%s' %i].value #'customer-id'
                    sdpws['K%s' % row] = sdpdict['%s_entropy' % n] #entropy
                    if lastiptnservice == sdpdict['%s_name' % n]:
                        if lastcustid != l2relatedws['K%s' %i].value:
                            lastrow = int(row) - 1
                            sdpws['J%s' % lastrow].fill = PatternFill(start_color='FFfff957',end_color='FFfff957', fill_type='solid')
                            sdpws['J%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                    lastiptnservice = sdpdict['%s_name' % n]
                    lastcustid = l2relatedws['K%s' %i].value
                    row += 1
                    sdpws['A1'] = row
                else:
                    if l2relatedws['G%s' %i].value+sdpdict['%s_name' %n] not in lastbdrn:
                        createbdrevpnworksheet(misum,i,n,sdpdict['%s_node' % n],noentropyws)
                    lastbdrn.append(l2relatedws['G%s' %i].value+sdpdict['%s_name' %n])
            if sdpdict['%s_type' %n] == 'VPLS':
                l2interas = '%s%s%s%s' %(l2relatedws['G%s' %i].value,sdpdict['%s_type' %n],sdpdict['%s_name' %n],sdpdict['%s_id' % n])
                if l2interas not in l2interaslist:
                    for asbrline in range(3, asbrmapws.max_row + 1):
                        asbrdict = {}
                        asbrlist = []
                        if l2relatedws['G%s' % i].value == asbrmapws['A%s' % asbrline].value:
                            asbrdict['asbr_node'] = asbrmapws['C%s' % asbrline].value
                            asbrdict['asbr_preference'] = asbrmapws['D%s' % asbrline].value
                            asbrdict['asbr_sap'] = asbrmapws['E%s' % asbrline].value
                            asbrlist.append(asbrdict.copy())
                            asbrdict['asbr_node'] = asbrmapws['F%s' % asbrline].value
                            asbrdict['asbr_preference'] = asbrmapws['G%s' % asbrline].value
                            asbrdict['asbr_sap'] = asbrmapws['H%s' % asbrline].value
                            asbrlist.append(asbrdict.copy())
                            for asbr in asbrlist:
                                createl2interassdpworksheet(misum, sdpdict, l2relatedws, i, n,asbr,noentropyws)
                                l2interaslist.append(l2interas)
            elif sdpdict['%s_type' %n] == 'EPIPE':
                for pm in range(3, misum['port-lag'].max_row + 1):
                    if l2relatedws['G%s' %i].value == misum['port-lag']['A%s' % pm].value and sdpdict['%s_id' % n] == misum['port-lag']['E%s' % pm].value:
                        l2interas = '%s%s%s%s' %(l2relatedws['G%s' %i].value,sdpdict['%s_type' %n],sdpdict['%s_name' %n],sdpdict['%s_id' % n])
                        if l2interas not in l2interaslist:
                            for asbrline in range(3,asbrmapws.max_row + 1):
                                asbrdict = {}
                                asbrlist = []
                                if l2relatedws['G%s' %i].value == asbrmapws['A%s' %asbrline].value:
                                    asbrdict['asbr_node'] = asbrmapws['I%s' % asbrline].value
                                    asbrdict['asbr_preference'] = asbrmapws['J%s' % asbrline].value
                                    asbrdict['asbr_sap'] = asbrmapws['K%s' % asbrline].value
                                    asbrlist.append(asbrdict.copy())
                                    for asbr in asbrlist:
                                        createl2interassdpworksheet(misum, sdpdict, l2relatedws, i, n, asbr,noentropyws)
                                        l2interaslist.append(l2interas)

def createpolicynamingworksheet(policynamews,bdrnode,bdrservice,iptnpolname,bdrpolname, iptnterm,bdrtentry,bdrprefixlistname,routefilter):
    global policyprefixduplist, policyprefixdupdict
    row = policynamews['A1'].value
    policynamews['E%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957', fill_type='solid')
    policynamews['H%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957', fill_type='solid')
    policynamews['K%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957', fill_type='solid')
    policynamews['A%s' % row] = bdrnode #'nokia-node'
    policynamews['B%s' % row] = 'VPRN' #'service-type'
    policynamews['C%s' % row] = bdrservice #'service-name'
    policynamews['D%s' % row] = iptnpolname #'iptn-policy-name'
    policynamews['E%s' % row] = bdrpolname #'bdr-policy-name'
    policynamews['F%s' % row] = '=len(E%s)'  % row #'length(max 64)'
    policynamews['G%s' % row] = iptnterm #'iptn-term-name'
    if '_local-pref_new' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_local-pref_new','_Lp') #'bdr-entry-name'
    elif '_local-pref' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_local-pref', '_Lp')  # 'bdr-entry-name'
    elif '_local_pref' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_local_pref', '_Lp')  # 'bdr-entry-name'
    elif '_localpref' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_localpref', '_Lp')  # 'bdr-entry-name'
    elif '_local-prep' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_local-prep', '_Lp')  # 'bdr-entry-name'
    elif '_export_route' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_export_route', '_Exp')  # 'bdr-entry-name'
    elif '_export' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_export', '_Exp')  # 'bdr-entry-name'
    elif '-export' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('-export', '_Exp')  # 'bdr-entry-name'
    elif '_import_route' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_import_route', '_Imp')  # 'bdr-entry-name'
    elif '_import_prefix' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_import_prefix', '_Imp')  # 'bdr-entry-name'
    elif '_import' in bdrtentry:
        policynamews['H%s' % row] = bdrtentry.replace('_import', '_Imp')  # 'bdr-entry-name'
    else:
        policynamews['H%s' % row] = bdrtentry

    policynamews['I%s' % row] = '=len(H%s)'  % row #'length(max 255)'
    if routefilter == 'no':
        policynamews['J%s' % row] = bdrprefixlistname #'iptn-prefix-list-name'
    elif routefilter == 'yes':
        policynamews['J%s' % row] = bdrprefixlistname  # 'iptn-prefix-list-name'

    if '_local-pref_new' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_local-pref_new','_Lp') #'bdr-entry-name'
    elif '_local-pref' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_local-pref', '_Lp')  # 'bdr-entry-name'
    elif '_local_pref' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_local_pref', '_Lp')  # 'bdr-entry-name'
    elif '_localpref' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_localpref', '_Lp')  # 'bdr-entry-name'
    elif '_local-prep' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_local-prep', '_Lp')  # 'bdr-entry-name'
    elif '_export_route' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_export_route', '_Exp')  # 'bdr-entry-name'
    elif '_export_prefix' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_export_prefix', '_Exp')  # 'bdr-entry-name'
    elif '_export' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_export', '_Exp')  # 'bdr-entry-name'
    elif '-export' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('-export', '_Exp')  # 'bdr-entry-name'
    elif '_import_route' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_import_route', '_Imp')  # 'bdr-entry-name'
    elif '_import_prefix' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_import_prefix', '_Imp')  # 'bdr-entry-name'
    elif '_import' in bdrprefixlistname:
        policynamews['K%s' % row] = bdrprefixlistname.replace('_import', '_Imp')  # 'bdr-entry-name'
    else:
        policynamews['K%s' % row] = bdrprefixlistname

    if bdrpolname not in policyprefixduplist:
        policyprefixduplist.append(bdrpolname)
    if bdrnode not in policyprefixdupdict.setdefault('%s_node' % bdrpolname, []):
        policyprefixdupdict.setdefault('%s_node' % bdrpolname, []).append(bdrnode)
    if policynamews['K%s' % row].value not in policyprefixdupdict.setdefault('%s_%s' % (bdrnode,bdrpolname), []):
        policyprefixdupdict.setdefault('%s_%s' % (bdrnode,bdrpolname), []).append(policynamews['K%s' % row].value)
    if len(policynamews['K%s' % row].value) >= 33:
        policynamews['K%s' % row].fill = PatternFill(start_color='FFff957f', end_color='FFff957f', fill_type='solid')
        policynamews['L%s' % row].fill = PatternFill(start_color='FFff957f', end_color='FFff957f', fill_type='solid')
    policynamews['L%s' % row] = '=len(K%s)' % row #'length(max 32)'
    row += 1
    policynamews['A1'] = row

def checkpolicyprefixdup():
    global policyprefixduplist, policyprefixdupdict
    bdrtextractwb = openpyxl.load_workbook(filename=inputpath + 'bdrt-extraction.xlsx')
    bdrtpolicyws = bdrtextractwb['route-policy']
    livebdrtdict = {}
    for row in range(3, bdrtpolicyws.max_row + 1):
        #livebdrtdict[f"{bdrtpolicyws[f'A{row}'].value}_policy"]
        livenode = bdrtpolicyws[f'A{row}'].value
        livepolicy = bdrtpolicyws[f'B{row}'].value
        liveprefixlist = bdrtpolicyws[f'D{row}'].value
        if livepolicy not in livebdrtdict.setdefault('%s_policy' % livenode, []):
            livebdrtdict.setdefault('%s_policy' % livenode, []).append(livepolicy)
        if liveprefixlist not in livebdrtdict.setdefault('%s_%s_prefix' % (livenode,livepolicy), []):
            livebdrtdict.setdefault('%s_%s_prefix' % (livenode,livepolicy), []).append(liveprefixlist)
    misum = openpyxl.load_workbook(filename=migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)
    policynamews = misum['policy-naming']
    for row in range(3, policynamews.max_row + 1):
        bdrnode = policynamews['A%s' % row].value
        bdrpolicy = policynamews['E%s' % row].value
        bdrprefixlist = policynamews['K%s' % row].value
        if bdrprefixlist != None:
            for bdrpolname_c in policyprefixduplist:
                if bdrpolicy != bdrpolname_c:
                    for node_c in policyprefixdupdict['%s_node' % bdrpolname_c]:
                        if bdrnode == node_c:
                            if bdrprefixlist in policyprefixdupdict['%s_%s' % (node_c,bdrpolname_c)]:
                                policynamews['H%s' % row].fill = PatternFill(start_color='FF9370db', end_color='FF9370db', fill_type='solid')
                                policynamews['K%s' % row].fill = PatternFill(start_color='FF9370db', end_color='FF9370db', fill_type='solid')
            if '%s_policy' % bdrnode in livebdrtdict:
                for bdrpolname_l in livebdrtdict['%s_policy' % bdrnode]:
                    if bdrpolicy != bdrpolname_l:
                        if bdrprefixlist in livebdrtdict['%s_%s_prefix' % (bdrnode, bdrpolname_l)]:
                            #print(f'prefix-list {bdrprefixlist} duplicated with production configuration prefix-list of policy:{bdrpolname_l}')
                            policynamews['M%s' % row] = f'prefix-list "{bdrprefixlist}" duplicated with production configuration prefix-list of policy : {bdrpolname_l}'

    misum.save(migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)

def getlocalprefpolicydictcfg(file,bdrtid):
    iptnpolicylines = []
    bdrtpolicydict = {}
    lines = file.readlines()
    file.close()
    for a in range(len(lines)):
        lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
        lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
        lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
        iptnpolicylines.append(lines[a])

    for policyline in iptnpolicylines:
        if 'set policy-options policy-statement' in policyline:
            iptnpolname = policyline.split(' ')[3]
            if 'set policy-options policy-statement %s term' %iptnpolname in policyline:
                fromstatement = ''
                thenstatement = ''
                iptnterm = policyline.split(' ')[5]
                bdrtentry = iptnterm + bdrtid
                if bdrtentry not in bdrtpolicydict.setdefault('%s_entry' % iptnpolname, []):
                    bdrtpolicydict.setdefault('%s_entry' % iptnpolname, []).append(bdrtentry)
                    bdrtpolicydict['%s_%s_term' %(iptnpolname,bdrtentry)] = iptnterm
                if 'set policy-options policy-statement %s term %s from' % (iptnpolname, iptnterm) in policyline:
                    fromstatement = policyline.split('set policy-options policy-statement %s term %s from ' % (iptnpolname, iptnterm))[-1]
                    if 'set policy-options policy-statement %s term %s from route-filter' % (iptnpolname, iptnterm) in policyline:
                        bdrtpolicydict.setdefault('%s_%s_route-filter' % (iptnpolname, bdrtentry), []).append(policyline.split('set policy-options policy-statement %s term %s from route-filter ' % (iptnpolname, iptnterm))[-1])
                    if 'route-filter' in fromstatement:
                        fromstatement = 'route-filter'
                    if 'community' in fromstatement:
                        bdrtpolicydict.setdefault('%s_community' % (iptnpolname), []).append(policyline.split('set policy-options policy-statement %s term %s from community ' % (iptnpolname, iptnterm))[-1])
                    if 'prefix-list' in fromstatement:
                        bdrtpolicydict.setdefault('%s_%s_prefix-list' % (iptnpolname, bdrtentry), []).append(policyline.split('set policy-options policy-statement %s term %s from prefix-list ' % (iptnpolname, iptnterm))[-1]+bdrtid)
                if fromstatement not in bdrtpolicydict.setdefault('%s_%s_from' % (iptnpolname, bdrtentry), []):
                    bdrtpolicydict.setdefault('%s_%s_from' % (iptnpolname, bdrtentry), []).append(fromstatement)
                if 'set policy-options policy-statement %s term %s then' % (iptnpolname, iptnterm) in policyline:
                    thenstatement = policyline.split('set policy-options policy-statement %s term %s then ' % (iptnpolname, iptnterm))[-1]
                    if thenstatement not in bdrtpolicydict.setdefault('%s_%s_then' % (iptnpolname, bdrtentry), []):
                        bdrtpolicydict.setdefault('%s_%s_then' % (iptnpolname, bdrtentry), []).append(thenstatement)
                    if 'community set' in thenstatement:
                        bdrtpolicydict.setdefault('%s_community' % (iptnpolname), []).append(policyline.split('set policy-options policy-statement %s term %s then community set ' % (iptnpolname, iptnterm))[-1])
                        bdrtpolicydict.setdefault('%s_%s_community' % (iptnpolname,bdrtentry), []).append(policyline.split('set policy-options policy-statement %s term %s then community add ' % (iptnpolname, iptnterm))[-1])
                    elif 'community add' in thenstatement:
                        bdrtpolicydict.setdefault('%s_community' % (iptnpolname), []).append(policyline.split('set policy-options policy-statement %s term %s then community add ' % (iptnpolname, iptnterm))[-1])
                        bdrtpolicydict.setdefault('%s_%s_community' % (iptnpolname,bdrtentry), []).append(policyline.split('set policy-options policy-statement %s term %s then community add ' % (iptnpolname, iptnterm))[-1])
                    if 'local-preference' in thenstatement:
                        bdrtpolicydict.setdefault('%s_%s_then_local-pref' % (iptnpolname, bdrtentry), []).append(thenstatement.split()[-1])
        if 'deactivate policy-options policy-statement' in policyline:
            iptnterm = policyline.split(' ')[5]
            if 'deactivate policy-options policy-statement %s term' %iptnpolname in policyline:
                iptnterm = policyline.split(' ')[5]
                bdrtentry = iptnterm + bdrtid
                if bdrtentry not in bdrtpolicydict.setdefault('%s_%s_deactivate' % (iptnpolname, bdrtentry), []):
                    #print('deactivate %s' % policyline)
                    bdrtpolicydict.setdefault('%s_%s_deactivate' % (iptnpolname, bdrtentry), []).append(bdrtentry)

    return bdrtpolicydict, iptnpolname

def createvrfexportpolicy(file_node_a, bdrnode,bdrservice, bdrpolname, prefixlist, community,policynamews,iptnnode):

    bdrpolicylist = []
    bdrprefixcfglist = []
    fromstatement = ''
    bdrcommunitycheck = []
    lastentry = ''
    peerlist = []
    bdrtid = {}
    bdrtpolicydict = {}
    currentbdrtid = ''
    iptnpolnamenodea = ''
    iptnpolnamenodeb = ''

    for n in range(3, l2mapsheet.max_row + 1):
        if iptnnode == l2mapsheet['A%s' % n].value:
            peerid = l2mapsheet['C%s' % n].value
            currentbdrtid = l2mapsheet['D%s' % n].value[-2:]

    for m in range(3, l2mapsheet.max_row + 1):
        if peerid == l2mapsheet['C%s' % m].value:
            peerlist.append(l2mapsheet['A%s' % m].value)
            bdrtid['%s_id' % l2mapsheet['A%s' % m].value] = l2mapsheet['D%s' % m].value[-2:]

    #print(peerlist)
    #print(bdrtid)
    for node in peerlist:
        if node == iptnnode:
            bdrtpolicydictnodea, iptnpolnamenodea = getlocalprefpolicydictcfg(file_node_a,bdrtid['%s_id' % node])
        else:
            startrow, endrow = findextractnodeindex(exindex, node, 'F', 'G', 'H')
            for exv in range(startrow,endrow +1):
                exvprnglobalnodeb = createvprnglobaldict(exvprn, node, exv)
                if node == exvprnglobalnodeb['iptnnode'] and community == exvprnglobalnodeb['iptnrtexpcommu']:
                    expnodeb = open(extractpath + exvprnglobalnodeb['iptnrtexppol'], 'r')
                    bdrtpolicydictnodeb, iptnpolnamenodeb = getlocalprefpolicydictcfg(expnodeb,bdrtid['%s_id' % node])

    #print(bdrtpolicydictnodea)
    #print(bdrtpolicydictnodeb)
    iptnpolname = iptnpolnamenodea
    if iptnpolnamenodea:
        for entry in bdrtpolicydictnodea['%s_entry' % iptnpolnamenodea]:
            if '%s_%s_then_local-pref' %(iptnpolname,entry) in bdrtpolicydictnodea:
                #print(iptnnode, iptnpolname, entry, bdrtpolicydictnodea['%s_%s_then_local-pref' %(iptnpolname,entry)][0],bdrtpolicydictnodea['%s_%s_term' %(iptnpolname,entry)])
                bdrtpolicydict.setdefault('%s_entry' % iptnpolname, []).append(entry)
                bdrtpolicydict['%s_%s_term' %(iptnpolname,entry)] = bdrtpolicydictnodea['%s_%s_term' % (iptnpolname, entry)]
                if '%s_%s_route-filter' % (iptnpolname, entry) in bdrtpolicydictnodea:
                    bdrtpolicydict.setdefault('%s_%s_route-filter' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_route-filter' % (iptnpolname, entry)])
                if '%s_%s_prefix-list' % (iptnpolname, entry) in bdrtpolicydictnodea:
                    bdrtpolicydict.setdefault('%s_%s_prefix-list' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_prefix-list' % (iptnpolname, entry)])
                if '%s_community' % (iptnpolname) in bdrtpolicydictnodea:
                    bdrtpolicydict.setdefault('%s_community' % (iptnpolname), []).extend(bdrtpolicydictnodea['%s_community' % (iptnpolname)])
                if '%s_%s_from' % (iptnpolname, entry) in bdrtpolicydictnodea:
                    bdrtpolicydict.setdefault('%s_%s_from' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_from' % (iptnpolname, entry)])
                if '%s_%s_then' % (iptnpolname, entry) in bdrtpolicydictnodea:
                    bdrtpolicydict.setdefault('%s_%s_then' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_then' % (iptnpolname, entry)])
                if '%s_%s_deactivate' % (iptnpolname, entry) in bdrtpolicydictnodea:
                    bdrtpolicydict.setdefault('%s_%s_deactivate' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_deactivate' % (iptnpolname, entry)])
            if '%s_%s_community' % (iptnpolname,entry) in bdrtpolicydictnodea:
                if len(bdrtpolicydictnodea['%s_%s_community' % (iptnpolname,entry)]) >= 2:
                    bdrtpolicydict.setdefault('%s_entry' % iptnpolname, []).append(entry)
                    bdrtpolicydict['%s_%s_term' %(iptnpolname,entry)] = bdrtpolicydictnodea['%s_%s_term' % (iptnpolname, entry)]
                    if '%s_%s_route-filter' % (iptnpolname, entry) in bdrtpolicydictnodea:
                        bdrtpolicydict.setdefault('%s_%s_route-filter' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_route-filter' % (iptnpolname, entry)])
                    if '%s_%s_prefix-list' % (iptnpolname, entry) in bdrtpolicydictnodea:
                        bdrtpolicydict.setdefault('%s_%s_prefix-list' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_prefix-list' % (iptnpolname, entry)])
                    if '%s_community' % (iptnpolname) in bdrtpolicydictnodea:
                        bdrtpolicydict.setdefault('%s_community' % (iptnpolname), []).extend(bdrtpolicydictnodea['%s_community' % (iptnpolname)])
                    if '%s_%s_from' % (iptnpolname, entry) in bdrtpolicydictnodea:
                        bdrtpolicydict.setdefault('%s_%s_from' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_from' % (iptnpolname, entry)])
                    if '%s_%s_then' % (iptnpolname, entry) in bdrtpolicydictnodea:
                        bdrtpolicydict.setdefault('%s_%s_then' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_then' % (iptnpolname, entry)])
                    if '%s_%s_deactivate' % (iptnpolname, entry) in bdrtpolicydictnodea:
                        bdrtpolicydict.setdefault('%s_%s_deactivate' % (iptnpolname, entry), []).extend(bdrtpolicydictnodea['%s_%s_deactivate' % (iptnpolname, entry)])
        if iptnpolnamenodeb:
            for entry in bdrtpolicydictnodeb['%s_entry' % iptnpolnamenodeb]:
                if '%s_%s_then_local-pref' %(iptnpolname,entry) in bdrtpolicydictnodeb:
                    #print(iptnnode, iptnpolname, entry, bdrtpolicydictnodeb['%s_%s_then_local-pref' %(iptnpolname,entry)][0],bdrtpolicydictnodeb['%s_%s_term' %(iptnpolname,entry)])
                    bdrtpolicydict.setdefault('%s_entry' % iptnpolname, []).append(entry)
                    bdrtpolicydict['%s_%s_term' %(iptnpolname,entry)] = bdrtpolicydictnodeb['%s_%s_term' % (iptnpolname, entry)]
                    if '%s_%s_route-filter' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                        bdrtpolicydict.setdefault('%s_%s_route-filter' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_route-filter' % (iptnpolname, entry)])
                    if '%s_%s_prefix-list' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                        bdrtpolicydict.setdefault('%s_%s_prefix-list' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_prefix-list' % (iptnpolname, entry)])
                    if '%s_community' % (iptnpolname) in bdrtpolicydictnodeb:
                        bdrtpolicydict.setdefault('%s_community' % (iptnpolname), []).extend(bdrtpolicydictnodeb['%s_community' % (iptnpolname)])
                    if '%s_%s_from' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                        bdrtpolicydict.setdefault('%s_%s_from' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_from' % (iptnpolname, entry)])
                    if '%s_%s_then' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                        bdrtpolicydict.setdefault('%s_%s_then' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_then' % (iptnpolname, entry)])
                    if '%s_%s_deactivate' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                        bdrtpolicydict.setdefault('%s_%s_deactivate' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_deactivate' % (iptnpolname, entry)])
                if '%s_%s_community' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                    if len(bdrtpolicydictnodeb['%s_%s_community' % (iptnpolname, entry)]) >= 2:
                        #print(iptnnode, iptnpolname, entry, bdrtpolicydictnodeb['%s_%s_then_local-pref' %(iptnpolname,entry)][0],bdrtpolicydictnodeb['%s_%s_term' %(iptnpolname,entry)])
                        bdrtpolicydict.setdefault('%s_entry' % iptnpolname, []).append(entry)
                        bdrtpolicydict['%s_%s_term' %(iptnpolname,entry)] = bdrtpolicydictnodeb['%s_%s_term' % (iptnpolname, entry)]
                        if '%s_%s_route-filter' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                            bdrtpolicydict.setdefault('%s_%s_route-filter' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_route-filter' % (iptnpolname, entry)])
                        if '%s_%s_prefix-list' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                            bdrtpolicydict.setdefault('%s_%s_prefix-list' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_prefix-list' % (iptnpolname, entry)])
                        if '%s_community' % (iptnpolname) in bdrtpolicydictnodeb:
                            bdrtpolicydict.setdefault('%s_community' % (iptnpolname), []).extend(bdrtpolicydictnodeb['%s_community' % (iptnpolname)])
                        if '%s_%s_from' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                            bdrtpolicydict.setdefault('%s_%s_from' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_from' % (iptnpolname, entry)])
                        if '%s_%s_then' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                            bdrtpolicydict.setdefault('%s_%s_then' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_then' % (iptnpolname, entry)])
                        if '%s_%s_deactivate' % (iptnpolname, entry) in bdrtpolicydictnodeb:
                            bdrtpolicydict.setdefault('%s_%s_deactivate' % (iptnpolname, entry), []).extend(bdrtpolicydictnodeb['%s_%s_deactivate' % (iptnpolname, entry)])

        if '%s_community' % (iptnpolname) in bdrtpolicydictnodea:
            if '# Community list configuration' not in bdrpolicylist:
                bdrpolicylist.append('')
                bdrpolicylist.append('# Community list configuration')
                bdrpolicylist.append('')
            for commu in bdrtpolicydictnodea['%s_community' % (iptnpolname)]:
                if commu not in bdrcommunitycheck:
                    bdrcommunitycheck.append(commu)
                    for c in range(2,communitysheet.max_row + 1):
                        if commu == communitysheet['C%s' %c].value:
                            bdrcommunity = communitysheet['D%s' %c].value.replace('65000','65051')
                            bdrpolicylist.append('/configure policy-options community "L3_%s" member "%s"' % (commu, bdrcommunity))
            bdrtcommuname = 'L3_'+Counter(bdrtpolicydictnodea['%s_community' % (iptnpolname)]).most_common(1)[0][0]
        if '%s_entry' % iptnpolname in bdrtpolicydict:
            bdrtpolicydict['%s_entry' % iptnpolname].sort()
            #print(bdrtpolicydict['%s_entry' % iptnpolname])
            for entry in bdrtpolicydict['%s_entry' % iptnpolname]:
                if '%s_%s_prefix-list' % (iptnpolname, entry) in bdrtpolicydict:
                    if '# Prefix-list configuration' not in bdrprefixcfglist:
                        bdrprefixcfglist.append('')
                        bdrprefixcfglist.append('# Prefix-list configuration')
                        bdrprefixcfglist.append('')
                    for prefixlistname in bdrtpolicydict['%s_%s_prefix-list' % (iptnpolname, entry)]:
                        #print(prefixlistname)
                        #print(prefixlist)
                        prefixline = []
                        if prefixlistname in prefixlist:
                            bdrprefixcfglist.append('')
                            #print('match prefix-list')
                            p = open(extractpath + prefixlist, 'r')
                            lines = p.readlines()
                            p.close()
                            for a in range(len(lines)):
                                lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
                                lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
                                lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
                                pf = lines[a].split('set policy-options prefix-list %s ' % prefixlistname)[-1]
                                prefixline.append(pf)
                            for prefix in prefixline:
                                if '/configure policy-options prefix-list "%s" prefix %s type exact' %(prefixlistname,prefix) not in bdrprefixcfglist:
                                    bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type exact' %(prefixlistname,prefix))
                        #print(prefixline)
                if '%s_%s_route-filter' % (iptnpolname, entry) in bdrtpolicydict:
                    if '# Prefix-list (from IPTN route-filter) configuration' not in bdrprefixcfglist:
                        bdrprefixcfglist.append('')
                        bdrprefixcfglist.append('# Prefix-list (from IPTN route-filter) configuration')
                        bdrprefixcfglist.append('# Please verify prefix-list name (limit 32 characters)')
                        bdrprefixcfglist.append('')
                    for prefix in bdrtpolicydict['%s_%s_route-filter' % (iptnpolname, entry)]:
                        #print(prefix)
                        if entry != lastentry:
                            bdrprefixcfglist.append('')
                        if 'upto' in prefix:
                            p = prefix.split(' ')[0]
                            l = prefix.split(' ')[-1].lstrip('/')
                            if '/configure policy-options prefix-list "%s" prefix %s type through through-length %s' %(entry,p,l) not in bdrprefixcfglist:
                                bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type through through-length %s' %(entry,p,l))
                                #print('/configure policy-options prefix-list "%s" prefix %s type through through-length %s' %(entry,p,l))
                        elif 'prefix-length-range' in prefix:
                            p = prefixb.split(' ')[0]
                            l = prefixb.split(' ')[-1].lstrip('/')
                            s = l.split('-/')[0]
                            e = l.split('-/')[-1]
                            if '/configure policy-options prefix-list "%s" prefix %s type range start-length %s end-length %s' % (entry, p, s, e) not in bdrprefixcfglist:
                                #print('/configure policy-options prefix-list "%s" prefix %s type range start-length %s end-length %s' % (entry, p, s, e))
                                bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type range start-length %s end-length %s' % (entry, p, s, e))
                        elif 'orlonger' in prefix:
                            p = prefix.split(' ')[0]
                            if '/configure policy-options prefix-list "%s" prefix %s type longer' %(entry,p) not in bdrprefixcfglist:
                                bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type longer' %(entry,p))
                                #print('/configure policy-options prefix-list "%s" prefix %s type longer' %(entry,p))
                        elif 'exact' in prefix:
                            p = prefix.split(' ')[0]
                            if '/configure policy-options prefix-list "%s" prefix %s type exact' %(entry,p) not in bdrprefixcfglist:
                                bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type exact' %(entry,p))
                                #print('/configure policy-options prefix-list "%s" prefix %s type exact' %(entry,p))
                        lastentry = entry


        bdrprefixlistname = ''
        bdrpolicylist.append('')
        bdrpolicylist.append('# Policy-statement configuration')
        bdrpolicylist.append('')
        bdrpolicylist.append('/configure policy-options policy-statement "%s" entry-type named' % bdrpolname)
        if '%s_entry' % iptnpolname in bdrtpolicydict:
            for entry in bdrtpolicydict['%s_entry' % iptnpolname]:
                if '%s_%s_prefix-list' % (iptnpolname, entry) in bdrtpolicydict:
                    #print(bdrtpolicydict['%s_%s_prefix-list' % (iptnpolname, entry)])
                    bdrprefixlistname = bdrtpolicydict['%s_%s_prefix-list' % (iptnpolname, entry)][0]
                    createpolicynamingworksheet(policynamews,bdrnode,bdrservice,iptnpolname,bdrpolname, entry,entry,bdrprefixlistname, 'no')
                elif '%s_%s_route-filter' % (iptnpolname, entry) in bdrtpolicydict:
                    #print(bdrtpolicydict['%s_%s_route-filter' % (iptnpolname, entry)])
                    bdrprefixlistname = entry
                    createpolicynamingworksheet(policynamews,bdrnode, bdrservice, iptnpolname,bdrpolname, entry,entry, bdrprefixlistname, 'yes')
                else:
                    if 'Deny' not in entry:
                        bdrprefixlistname = ''
                        createpolicynamingworksheet(policynamews,bdrnode, bdrservice, iptnpolname,bdrpolname, entry,entry, bdrprefixlistname, 'no')
                bdrpolicylist.append('')
                for fromstatement in bdrtpolicydict['%s_%s_from' % (iptnpolname, entry)]:
                    #print(fromstatement)
                    if 'prefix-list' in fromstatement:
                        if '/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,entry,bdrprefixlistname) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,entry,bdrprefixlistname))
                    elif 'route-filter' in fromstatement:
                        if '/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,entry,bdrprefixlistname) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,entry,bdrprefixlistname))
                    elif 'next-hop' in fromstatement:
                        nh = fromstatement.split(' ')[-1]
                        if '/configure policy-options policy-statement "%s" named-entry "%s" from neighbor ip-address %s' %(bdrpolname,entry,nh) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from neighbor ip-address %s' %(bdrpolname,entry,nh))
                    elif 'protocol bgp' in fromstatement:
                        if '/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [bgp-vpn]' % (bdrpolname,entry) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [bgp-vpn]' % (bdrpolname,entry))
                    elif 'protocol static' in fromstatement:
                        if '/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [static]' %(bdrpolname,entry) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [static]' %(bdrpolname,entry))
                    elif 'community' in fromstatement:
                        co = fromstatement.split(' ')[-1]
                        if '/configure policy-options policy-statement "%s" named-entry "%s" from community name "L3_%s"' %(bdrpolname,entry,co) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from community name "L3_%s"' %(bdrpolname,entry,co))
                    elif 'neighbor' in fromstatement:
                        ne = fromstatement.split(' ')[-1]
                        if '/configure policy-options policy-statement "%s" named-entry "%s" from neighbor ip-address %s' %(bdrpolname,entry,ne) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from neighbor ip-address %s' %(bdrpolname,entry,ne))
                #print(bdrtpolicydict)
                for thenstatement in bdrtpolicydict['%s_%s_then' % (iptnpolname, entry)]:
                    #print(thenstatement)
                    if 'local-preference' in thenstatement:
                        lop = thenstatement.split(' ')[-1]
                        if currentbdrtid == entry[-2:]:
                            if '/configure policy-options policy-statement "%s" named-entry "%s" action local-preference %s' %(bdrpolname,entry,lop) not in bdrpolicylist:
                                bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action local-preference %s' %(bdrpolname,entry,lop))
                        else:
                            if '/configure policy-options policy-statement "%s" named-entry "%s" action local-preference 100' %(bdrpolname,entry) not in bdrpolicylist:
                                bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action local-preference 100' %(bdrpolname,entry))
                    elif 'community set' in thenstatement:
                        co = 'L3_'+thenstatement.split()[-1]
                        if '/configure policy-options policy-statement "%s" named-entry "%s" action community add ["%s"]' %(bdrpolname,entry,co) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action community add ["%s"]' %(bdrpolname,entry,co))
                    elif 'community add' in thenstatement:
                        co = 'L3_'+thenstatement.split()[-1]
                        if '/configure policy-options policy-statement "%s" named-entry "%s" action community add ["%s"]' %(bdrpolname,entry,co) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action community add ["%s"]' %(bdrpolname,entry,co))
                    elif 'as-path-prepend' in thenstatement:
                        repeat = thenstatement.count('65000')
                        if '/configure policy-options policy-statement "%s" named-entry "%s" action as-path-prepend as-path 65051' %(bdrpolname,entry) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action as-path-prepend as-path 65051' %(bdrpolname,entry))
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action as-path-prepend repeat %s' %(bdrpolname,entry,repeat))
                    elif 'accept' in thenstatement:
                        if '/configure policy-options policy-statement "%s" named-entry "%s" action action-type accept' %(bdrpolname,entry) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action action-type accept' %(bdrpolname,entry))
                            #bdrpolicylist.append('')
                    elif 'next term' in thenstatement:
                        if '/configure policy-options policy-statement "%s" named-entry "%s" action community add ["%s"]' %(bdrpolname,entry,bdrtcommuname) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action community add ["%s"]' %(bdrpolname,entry,bdrtcommuname))
                        if '/configure policy-options policy-statement "%s" named-entry "%s" action action-type accept' %(bdrpolname,entry) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action action-type accept' %(bdrpolname,entry))
                            #bdrpolicylist.append('')
                    elif 'reject' in thenstatement:
                        if '/configure policy-options policy-statement "%s" default-action action-type reject' % bdrpolname not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" default-action action-type reject' % bdrpolname)
                            #bdrpolicylist.append('')
                if '%s_%s_deactivate' % (iptnpolname, entry) in bdrtpolicydict:
                    for deactivate in bdrtpolicydict['%s_%s_deactivate' % (iptnpolname, entry)]:
                        if '/configure policy-options policy-statement "%s" delete named-entry "%s"' %(bdrpolname,entry) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" delete named-entry "%s"' %(bdrpolname,entry))
        bdrpolicylist.append('')
        bdrpolicylist.append('/configure policy-options policy-statement "%s" default-action action-type accept' % bdrpolname)
        bdrpolicylist.append('/configure policy-options policy-statement "%s" default-action community add ["%s"]' % (bdrpolname,bdrtcommuname))



    bdrpolicyfilename = '%s_%s_bdrpolicy_%s.txt' % (bdrnode, bdrservice,bdrpolname)
    bdrpolicyfilepath = '.\cfg\\'+ bdrpolicyfilename
    f = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + bdrpolicyfilename, "a")
    for policyline in bdrpolicylist:
        f.write(policyline + '\n')
    f.close()
    if bdrprefixcfglist:
        bdrprefixfilename = '%s_%s_bdrprefix-list_%s.txt' % (bdrnode, bdrservice,bdrpolname)
        bdrprefixfilepath = '.\cfg\\'+ bdrprefixfilename
        g = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + bdrprefixfilename, "a")
        for prefixlistline in bdrprefixcfglist:
            g.write(prefixlistline + '\n')
        g.close()
    else:
        bdrprefixfilepath = ''

    return bdrpolicyfilepath,bdrprefixfilepath

def createbdrpolicy(file, bdrnode,bdrservice, bdrpolname, prefixlist, community,policynamews,iptnstatic,poltype):
    iptnpolicylines = []
    bdrpolicylist = []
    bdrprefixcfglist = []
    iptnpolicydict = {}
    fromstatement = ''
    bdrcommunitycheck = []
    lastiptnterm = ''
    nextvrflist = []
    impvrfdict = {}
    lines = file.readlines()
    file.close()
    for a in range(len(lines)):
        lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
        lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
        lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
        iptnpolicylines.append(lines[a])

    if iptnstatic != None:
        st = open(extractpath + iptnstatic, 'r')
        lines = st.readlines()
        st.close()
        for a in range(len(lines)):
            lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
            lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
            lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
            if 'next-table' in lines[a]:
                vpn = lines[a].split()[-1].split('.')[0]
                prefix = lines[a].split()[-3]
                if vpn not in nextvrflist:
                    nextvrflist.append(vpn)
                    for u in range(2,communitysheet.max_row + 1):
                        if vpn == communitysheet['A%s' %u].value:
                            impvrfdict['%s_communame' % vpn] = communitysheet['C%s' %u].value
                            impvrfdict['%s_commuval' % vpn] = communitysheet['D%s' % u].value
                if prefix not in impvrfdict.setdefault('%s_imp_prefix' % vpn, []):
                    impvrfdict.setdefault('%s_imp_prefix' % vpn, []).append(prefix)


    for policyline in iptnpolicylines:
        if 'set policy-options policy-statement' in policyline:
            iptnpolname = policyline.split(' ')[3]
            if 'set policy-options policy-statement %s term' %iptnpolname in policyline:
                fromstatement = ''
                thenstatement = ''
                iptnterm = policyline.split(' ')[5]
                if iptnterm not in iptnpolicydict.setdefault('%s_term' % iptnpolname, []):
                    iptnpolicydict.setdefault('%s_term' % iptnpolname, []).append(iptnterm)
                if 'set policy-options policy-statement %s term %s from' % (iptnpolname, iptnterm) in policyline:
                    fromstatement = policyline.split('set policy-options policy-statement %s term %s from ' % (iptnpolname, iptnterm))[-1]
                    if 'set policy-options policy-statement %s term %s from route-filter' % (iptnpolname, iptnterm) in policyline:
                        iptnpolicydict.setdefault('%s_%s_route-filter' % (iptnpolname, iptnterm), []).append(policyline.split('set policy-options policy-statement %s term %s from route-filter ' % (iptnpolname, iptnterm))[-1])
                    if 'route-filter' in fromstatement:
                        fromstatement = 'route-filter'
                    if 'community' in fromstatement:
                        iptnpolicydict.setdefault('%s_%s_community' % (iptnpolname, iptnterm), []).append(policyline.split('set policy-options policy-statement %s term %s from community ' % (iptnpolname, iptnterm))[-1])
                    if 'prefix-list' in fromstatement:
                        iptnpolicydict.setdefault('%s_%s_prefix-list' % (iptnpolname, iptnterm), []).append(policyline.split('set policy-options policy-statement %s term %s from prefix-list ' % (iptnpolname, iptnterm))[-1])
                if fromstatement not in iptnpolicydict.setdefault('%s_%s_from' % (iptnpolname, iptnterm), []):
                    iptnpolicydict.setdefault('%s_%s_from' % (iptnpolname, iptnterm), []).append(fromstatement)
                if 'set policy-options policy-statement %s term %s then' % (iptnpolname, iptnterm) in policyline:
                    thenstatement = policyline.split('set policy-options policy-statement %s term %s then ' % (iptnpolname, iptnterm))[-1]
                    if thenstatement not in iptnpolicydict.setdefault('%s_%s_then' % (iptnpolname, iptnterm), []):
                        iptnpolicydict.setdefault('%s_%s_then' % (iptnpolname, iptnterm), []).append(thenstatement)
                    if 'community set' in thenstatement:
                        iptnpolicydict.setdefault('%s_%s_community' % (iptnpolname, iptnterm), []).append(policyline.split('set policy-options policy-statement %s term %s then community set ' % (iptnpolname, iptnterm))[-1])
                    elif 'community add' in thenstatement:
                        iptnpolicydict.setdefault('%s_%s_community' % (iptnpolname, iptnterm), []).append(policyline.split('set policy-options policy-statement %s term %s then community add ' % (iptnpolname, iptnterm))[-1])
        if 'deactivate policy-options policy-statement' in policyline:
            iptnterm = policyline.split(' ')[5]
            if 'deactivate policy-options policy-statement %s term' %iptnpolname in policyline:
                iptnterm = policyline.split(' ')[5]
                if iptnterm not in iptnpolicydict.setdefault('%s_%s_deactivate' % (iptnpolname, iptnterm), []):
                    #print('deactivate %s' % policyline)
                    iptnpolicydict.setdefault('%s_%s_deactivate' % (iptnpolname, iptnterm), []).append(iptnterm)

    for termnum in range(len(iptnpolicydict['%s_term' % iptnpolname])):
        iptnterm = iptnpolicydict['%s_term' % iptnpolname][termnum]
        if '%s_%s_prefix-list' % (iptnpolname, iptnterm) in iptnpolicydict:
            if '# Prefix-list configuration' not in bdrprefixcfglist:
                bdrprefixcfglist.append('')
                bdrprefixcfglist.append('# Prefix-list configuration')
                bdrprefixcfglist.append('')
            for prefixlistname in iptnpolicydict['%s_%s_prefix-list' % (iptnpolname, iptnterm)]:
                #print(prefixlistname)
                #print(prefixlist)
                prefixline = []
                if prefixlistname in prefixlist:
                    bdrprefixcfglist.append('')
                    if len(prefixlistname) >= 33:
                        bdrprefixcfglist.append('# The name length of this prefix-list is %s characters' % str(len(prefixlistname)))
                    #print('match prefix-list')
                    p = open(extractpath + prefixlist, 'r')
                    lines = p.readlines()
                    p.close()
                    for a in range(len(lines)):
                        lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
                        lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
                        lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
                        pf = lines[a].split('set policy-options prefix-list %s ' % prefixlistname)[-1]
                        prefixline.append(pf)
                    for prefix in prefixline:
                        if '/configure policy-options prefix-list "%s" prefix %s type exact' %(prefixlistname,prefix) not in bdrprefixcfglist:
                            bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type exact' %(prefixlistname,prefix))

                #print(prefixline)
        if '%s_%s_route-filter' % (iptnpolname, iptnterm) in iptnpolicydict:
            if '# Prefix-list (from IPTN route-filter) configuration' not in bdrprefixcfglist:
                bdrprefixcfglist.append('')
                bdrprefixcfglist.append('# Prefix-list (from IPTN route-filter) configuration')
                bdrprefixcfglist.append('# Please verify prefix-list name (limit 32 characters)')
                bdrprefixcfglist.append('')
            for prefix in iptnpolicydict['%s_%s_route-filter' % (iptnpolname, iptnterm)]:
                #print(prefix)
                if iptnterm != lastiptnterm:
                    bdrprefixcfglist.append('')
                if 'upto' in prefix:
                    p = prefix.split(' ')[0]
                    l = prefix.split(' ')[-1].lstrip('/')
                    if '/configure policy-options prefix-list "%s" prefix %s type through through-length %s' %(entry,p,l) not in bdrprefixcfglist:
                        bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type through through-length %s' %(entry,p,l))
                        #print('/configure policy-options prefix-list "%s" prefix %s type through through-length %s' %(entry,p,l))
                elif 'prefix-length-range' in prefix:
                    p = prefixb.split(' ')[0]
                    l = prefixb.split(' ')[-1].lstrip('/')
                    s = l.split('-/')[0]
                    e = l.split('-/')[-1]
                    if '/configure policy-options prefix-list "%s" prefix %s type range start-length %s end-length %s' % (entry, p, s, e) not in bdrprefixcfglist:
                        #print('/configure policy-options prefix-list "%s" prefix %s type range start-length %s end-length %s' % (entry, p, s, e))
                        bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type range start-length %s end-length %s' % (entry, p, s, e))
                elif 'orlonger' in prefix:
                    p = prefix.split(' ')[0]
                    if '/configure policy-options prefix-list "%s" prefix %s type longer' %(iptnterm,p) not in bdrprefixcfglist:
                        bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type longer' %(iptnterm,p))
                        #print('/configure policy-options prefix-list "%s" prefix %s type longer' %(iptnterm,p))
                elif 'exact' in prefix:
                    p = prefix.split(' ')[0]
                    if '/configure policy-options prefix-list "%s" prefix %s type exact' %(iptnterm,p) not in bdrprefixcfglist:
                        bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type exact' %(iptnterm,p))
                        #print('/configure policy-options prefix-list "%s" prefix %s type exact' %(iptnterm,p))
                lastiptnterm = iptnterm
        if nextvrflist:
            for vr in nextvrflist:
                leakprefixname = bdrservice.lstrip('L3_') + '_leak_' + vr + '_Imp'
                for lp in impvrfdict['%s_imp_prefix' % vr]:
                    if '/configure policy-options prefix-list "%s" prefix %s type exact' %(leakprefixname,lp) not in bdrprefixcfglist:
                        bdrprefixcfglist.append('/configure policy-options prefix-list "%s" prefix %s type exact' %(leakprefixname,lp))
        if '%s_%s_community' % (iptnpolname, iptnterm) in iptnpolicydict:
            if '# Community list configuration' not in bdrpolicylist:
                bdrpolicylist.append('')
                bdrpolicylist.append('# Community list configuration')
                bdrpolicylist.append('')
            for commu in iptnpolicydict['%s_%s_community' % (iptnpolname, iptnterm)]:
                if commu not in bdrcommunitycheck:
                    bdrcommunitycheck.append(commu)
                    if 'target:65000' in community:
                        for c in range(2,communitysheet.max_row + 1):
                            if commu == communitysheet['C%s' %c].value:
                                bdrcommunity = communitysheet['D%s' %c].value.replace('65000','65051')
                                bdrpolicylist.append('/configure policy-options community "L3_%s" member "%s"' % (commu, bdrcommunity))
                    else:
                        ebgpcommu = community.split(':')[0] + ':' + commu[-4:]
                        bdrpolicylist.append('/configure policy-options community "L3_%s" member "%s"' % (commu, ebgpcommu))
        if nextvrflist:
            for vr in nextvrflist:
                leakcommuname = impvrfdict['%s_communame' % vr]
                leakcommuval = impvrfdict['%s_commuval' % vr].replace('65000','65051')
                if '/configure policy-options community "L3_%s" member "%s"' % (leakcommuname, leakcommuval) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options community "L3_%s" member "%s"' % (leakcommuname, leakcommuval))

    bdrprefixlistname = ''
    bdrpolicylist.append('')
    bdrpolicylist.append('# Policy-statement configuration')
    bdrpolicylist.append('')
    bdrpolicylist.append('/configure policy-options policy-statement "%s" entry-type named' % bdrpolname)
    for termnum in range(len(iptnpolicydict['%s_term' % iptnpolname])):
        iptnterm = iptnpolicydict['%s_term' % iptnpolname][termnum]
        bdrterm = iptnterm
        if poltype == 'vrfimp':
            bdrterm = 'L3_' + iptnterm
        #print(iptnpolicydict['%s_term' % iptnpolname][termnum])
        if '%s_%s_prefix-list' % (iptnpolname, iptnterm) in iptnpolicydict:
            #print(iptnpolicydict['%s_%s_prefix-list' % (iptnpolname, iptnterm)])
            bdrprefixlistname = iptnpolicydict['%s_%s_prefix-list' % (iptnpolname, iptnterm)][0]
            createpolicynamingworksheet(policynamews,bdrnode,bdrservice,iptnpolname,bdrpolname, bdrterm,bdrterm,bdrprefixlistname, 'no')
        elif '%s_%s_route-filter' % (iptnpolname, iptnterm) in iptnpolicydict:
            #print(iptnpolicydict['%s_%s_route-filter' % (iptnpolname, iptnterm)])
            bdrprefixlistname = iptnterm
            createpolicynamingworksheet(policynamews,bdrnode, bdrservice, iptnpolname,bdrpolname, bdrterm,bdrterm, bdrprefixlistname, 'yes')
        else:
            if 'Deny' not in iptnterm:
                bdrprefixlistname = ''
                createpolicynamingworksheet(policynamews,bdrnode, bdrservice, iptnpolname,bdrpolname, bdrterm,bdrterm, bdrprefixlistname, 'no')
        bdrpolicylist.append('')
        for fromstatement in iptnpolicydict['%s_%s_from' % (iptnpolname, iptnterm)]:
            #print(fromstatement)
            if 'prefix-list' in fromstatement:
                if '/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,bdrterm,bdrprefixlistname) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,bdrterm,bdrprefixlistname))
            elif 'route-filter' in fromstatement:
                if '/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,bdrterm,bdrprefixlistname) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,bdrterm,bdrprefixlistname))
            elif 'next-hop' in fromstatement:
                nh = fromstatement.split(' ')[-1]
                if '/configure policy-options policy-statement "%s" named-entry "%s" from neighbor ip-address %s' %(bdrpolname,bdrterm,nh) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from neighbor ip-address %s' %(bdrpolname,bdrterm,nh))
            elif 'protocol bgp' in fromstatement:
                if '/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [bgp-vpn]' % (bdrpolname,bdrterm) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [bgp-vpn]' % (bdrpolname,bdrterm))
            elif 'protocol static' in fromstatement:
                if '/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [static]' %(bdrpolname,bdrterm) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [static]' %(bdrpolname,bdrterm))
            elif 'community' in fromstatement:
                co = fromstatement.split(' ')[-1]
                if '/configure policy-options policy-statement "%s" named-entry "%s" from community name "L3_%s"' %(bdrpolname,bdrterm,co) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from community name "L3_%s"' %(bdrpolname,bdrterm,co))
            elif 'neighbor' in fromstatement:
                ne = fromstatement.split(' ')[-1]
                if '/configure policy-options policy-statement "%s" named-entry "%s" from neighbor ip-address %s' %(bdrpolname,bdrterm,ne) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from neighbor ip-address %s' %(bdrpolname,bdrterm,ne))
        #print(iptnpolicydict)
        for thenstatement in iptnpolicydict['%s_%s_then' % (iptnpolname, iptnterm)]:
            #print(thenstatement)
            if 'local-preference' in thenstatement:
                lop = thenstatement.split(' ')[-1]
                if '/configure policy-options policy-statement "%s" named-entry "%s" action local-preference %s' %(bdrpolname,bdrterm,lop) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action local-preference %s' %(bdrpolname,bdrterm,lop))
            elif 'community set' in thenstatement:
                comset = thenstatement.split(' ')[-1]
                if '/configure policy-options policy-statement "%s" named-entry "%s" action community add ["L3_%s"]' %(bdrpolname,bdrterm,comset) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action community add ["L3_%s"]' %(bdrpolname,bdrterm,comset))
            elif 'community add' in thenstatement:
                comadd = thenstatement.split(' ')[-1]
                if '/configure policy-options policy-statement "%s" named-entry "%s" action community add ["L3_%s"]' %(bdrpolname,bdrterm,comadd) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action community add ["L3_%s"]' %(bdrpolname,bdrterm,comadd))
            elif 'as-path-prepend' in thenstatement:
                repeat = thenstatement.count('65000')
                if '/configure policy-options policy-statement "%s" named-entry "%s" action as-path-prepend as-path 65051' %(bdrpolname,bdrterm) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action as-path-prepend as-path 65051' %(bdrpolname,bdrterm))
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action as-path-prepend repeat %s' %(bdrpolname,bdrterm,repeat))
            elif 'accept' in thenstatement:
                if '/configure policy-options policy-statement "%s" named-entry "%s" action action-type accept' %(bdrpolname,bdrterm) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action action-type accept' %(bdrpolname,bdrterm))
                    bdrpolicylist.append('')
            elif 'next term' in thenstatement:
                if '/configure policy-options policy-statement "%s" named-entry "%s" action action-type next-entry' %(bdrpolname,bdrterm) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action action-type next-entry' %(bdrpolname,bdrterm))
                    bdrpolicylist.append('')
            elif 'reject' in thenstatement:
                if nextvrflist:
                    for vr in nextvrflist:
                        leakentry = bdrservice[3:] + '_leak_' + vr + '_Imp'
                        if '/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,leakentry,leakentry) not in bdrpolicylist:
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from prefix-list ["%s"]' %(bdrpolname,leakentry,leakentry))
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from community name "L3_%s"' %(bdrpolname,leakentry,impvrfdict['%s_communame' % vr]))
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" from protocol name [bgp-vpn]' %(bdrpolname,leakentry))
                            bdrpolicylist.append('/configure policy-options policy-statement "%s" named-entry "%s" action action-type accept' %(bdrpolname,leakentry))
                            bdrpolicylist.append('')
                            createpolicynamingworksheet(policynamews,bdrnode, bdrservice, iptnpolname,bdrpolname, leakentry,leakentry, leakentry, 'yes')
                if '/configure policy-options policy-statement "%s" default-action action-type reject' % bdrpolname not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" default-action action-type reject' % bdrpolname)
                    bdrpolicylist.append('')
        if '%s_%s_deactivate' % (iptnpolname, iptnterm) in iptnpolicydict:
            for deactivate in iptnpolicydict['%s_%s_deactivate' % (iptnpolname, iptnterm)]:
                if '/configure policy-options policy-statement "%s" delete named-entry "%s"' %(bdrpolname,bdrterm) not in bdrpolicylist:
                    bdrpolicylist.append('/configure policy-options policy-statement "%s" delete named-entry "%s"' %(bdrpolname,bdrterm))



    #print(bdrprefixcfglist)
    #print(bdrpolicylist)

    bdrpolicyfilename = '%s_%s_bdrpolicy_%s.txt' % (bdrnode, bdrservice,bdrpolname)
    bdrpolicyfilepath = '.\cfg\\'+ bdrpolicyfilename
    f = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + bdrpolicyfilename, "a")
    for policyline in bdrpolicylist:
        f.write(policyline + '\n')
    f.close()
    if bdrprefixcfglist:
        bdrprefixfilename = '%s_%s_bdrprefix-list_%s.txt' % (bdrnode, bdrservice,bdrpolname)
        bdrprefixfilepath = '.\cfg\\'+ bdrprefixfilename
        g = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + bdrprefixfilename, "a")
        for prefixlistline in bdrprefixcfglist:
            g.write(prefixlistline + '\n')
        g.close()
    else:
        bdrprefixfilepath = ''

    return bdrpolicyfilepath,bdrprefixfilepath


def createbdrstaticconfig(file,namingdict,bdrvprnservice,family):
    staticlines = []
    staticcfglist = []
    bfdintervallist = []
    bfdmultilist = []
    staticentry = []
    iptnstatic = {}
    staticdict = {}
    bdrbfdinterval = ''
    bdrbfdmulti = ''
    lines = file.readlines()
    file.close()
    for a in range(len(lines)):
        lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
        lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
        lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
        staticlines.append(lines[a])

    for iptnstaticline in staticlines:
        if 'inet6.0' in iptnstaticline:
            if iptnstaticline.split(' ')[8] not in staticentry:
                staticentry.append(iptnstaticline.split(' ')[8])
                staticdict[iptnstaticline.split(' ')[8]] = ''
            if 'bfd-liveness-detection' in iptnstaticline:
                staticdict[iptnstaticline.split(' ')[8]] = 'enable'
                if 'bfd-liveness-detection minimum-interval' in iptnstaticline:
                    bfdintervallist.append(iptnstaticline.split(' ')[-1])
                elif 'bfd-liveness-detection multiplier' in iptnstaticline:
                    bfdmultilist.append(iptnstaticline.split(' ')[-1])
        else:
            if iptnstaticline.split(' ')[6] not in staticentry:
                staticentry.append(iptnstaticline.split(' ')[6])
                staticdict[iptnstaticline.split(' ')[6]] = ''
            if 'bfd-liveness-detection' in iptnstaticline:
                staticdict[iptnstaticline.split(' ')[6]] = 'enable'
                if 'bfd-liveness-detection minimum-interval' in iptnstaticline:
                    bfdintervallist.append(iptnstaticline.split(' ')[-1])
                elif 'bfd-liveness-detection multiplier' in iptnstaticline:
                    bfdmultilist.append(iptnstaticline.split(' ')[-1])

    if bfdintervallist:
        bdrbfdinterval = Counter(bfdintervallist).most_common(1)[0][0]
    if bfdmultilist:
        bdrbfdmulti = Counter(bfdmultilist).most_common(1)[0][0]
    #print(bdrbfdinterval)
    #print(bdrbfdmulti)
    #print(staticentry)
    for destination in staticentry:
        nexthop = ''
        nexttable = ''
        bfd = ''
        #print('dest = %s' % destination)
        for iptnstaticline in staticlines:
            if destination in iptnstaticline:
                if 'next-table' in iptnstaticline:
                    nexttable = iptnstaticline.split(' ')[-1]
                    #print('# this vprn has a static route for %s to another vprn (%s) please do it by manually' % (destination, nexttable))
                    staticcfglist.append('# this vprn has a static route for %s leak from another vprn (%s) please recheck ' % (destination, nexttable))
                elif staticdict[destination] == 'enable':
                    if 'next-hop' in iptnstaticline and 'qualified-next-hop' not in iptnstaticline:
                        nexthop = iptnstaticline.split(' ')[-1]
                        if '/configure service vprn "%s" interface "%s" %s bfd admin-state enable' %(bdrvprnservice,namingdict['bdrl3int'],family) not in staticcfglist:
                            staticcfglist.append('')
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd admin-state enable' %(bdrvprnservice,namingdict['bdrl3int'],family))
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd transmit-interval %s'%(bdrvprnservice,namingdict['bdrl3int'],family,bdrbfdinterval))
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd receive %s'%(bdrvprnservice,namingdict['bdrl3int'],family,bdrbfdinterval))
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd multiplier %s'%(bdrvprnservice,namingdict['bdrl3int'],family,bdrbfdmulti))
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd type cpm-np'%(bdrvprnservice,namingdict['bdrl3int'],family))
                        staticcfglist.append('')
                        staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" admin-state enable'%(bdrvprnservice,destination,nexthop))
                        staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" description "%s"'%(bdrvprnservice,destination,nexthop,namingdict['bdrl3int']))
                        staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" bfd-liveness true'%(bdrvprnservice,destination,nexthop))
                    if 'qualified-next-hop' in iptnstaticline:
                        nexthop = iptnstaticline.split(' ')[-3]
                        preference = iptnstaticline.split(' ')[-1]
                        if '/configure service vprn "%s" interface "%s" %s bfd admin-state enable' %(bdrvprnservice,namingdict['bdrl3int'],family) not in staticcfglist:
                            staticcfglist.append('')
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd admin-state enable' %(bdrvprnservice,namingdict['bdrl3int'],family))
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd transmit-interval %s'%(bdrvprnservice,namingdict['bdrl3int'],family,bdrbfdinterval))
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd receive %s'%(bdrvprnservice,namingdict['bdrl3int'],family,bdrbfdinterval))
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd multiplier %s'%(bdrvprnservice,namingdict['bdrl3int'],family,bdrbfdmulti))
                            staticcfglist.append('/configure service vprn "%s" interface "%s" %s bfd type cpm-np'%(bdrvprnservice,namingdict['bdrl3int'],family))
                        staticcfglist.append('')
                        staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" admin-state enable'%(bdrvprnservice,destination,nexthop))
                        staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" description "%s"'%(bdrvprnservice,destination,nexthop,namingdict['bdrl3int']))
                        staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" bfd-liveness true'%(bdrvprnservice,destination,nexthop))
                        staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" preference %s' %(bdrvprnservice,destination,nexthop,preference))

                elif 'next-hop' in iptnstaticline and 'qualified-next-hop' not in iptnstaticline:
                    nexthop = iptnstaticline.split(' ')[-1]
                    staticcfglist.append('')
                    staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" admin-state enable'%(bdrvprnservice,destination,nexthop))
                    staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" description "%s"'%(bdrvprnservice,destination,nexthop,namingdict['bdrl3int']))
                    #print(nexthop)
                elif 'qualified-next-hop' in iptnstaticline:
                    nexthop = iptnstaticline.split(' ')[-3]
                    preference = iptnstaticline.split(' ')[-1]
                    staticcfglist.append('')
                    staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" admin-state enable'%(bdrvprnservice,destination,nexthop))
                    staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" description "%s"'%(bdrvprnservice,destination,nexthop,namingdict['bdrl3int']))
                    staticcfglist.append('/configure service vprn "%s" static-routes route %s route-type unicast next-hop "%s" preference %s' %(bdrvprnservice,destination,nexthop,preference))


    relatedstaticcheck = []
    iptnstaticrelatedpelist = []
    relatestaticlastpe = ''
    iptnstaticrelatedpelist.append('#' * 120)
    if os.path.exists(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'related-static-pe.txt'):
        rs = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'related-static-pe.txt', "r")
        for lrs in rs.readlines():
            lrs = lrs.rstrip('\n')
            relatedstaticcheck.append(lrs)

    for i in range(3, exvprnint.max_row + 1):
        if bdrvprnservice.lstrip('L3_') == exvprnint['C%s' % i].value:
            if family == exvprnint['M%s' % i].value:
                if exvprnint['R%s' % i].value != None:
                    f = open(extractpath + exvprnint['R%s' % i].value)
                    lines = f.readlines()
                    for line in lines:
                        line = line.rstrip('\n')
                        if 'next-hop' in line:
                            if family == 'ipv4' and 'inet6' not in line:
                                iptndest = line.split()[6]
                                for bdrtdest in staticentry:
                                    if iptndest == bdrtdest:
                                        iptnintunit = exvprnint['E%s' % i].value + '.' + exvprnint['F%s' % i].value
                                        if 'STATIC to %s %s %s %s' %(bdrtdest,exvprnint['A%s' % i].value,exvprnint['C%s' % i].value,iptnintunit) not in relatedstaticcheck:
                                            if exvprnint['A%s' % i].value != relatestaticlastpe:
                                                iptnstaticrelatedpelist.append('')
                                            iptnstaticrelatedpelist.append('STATIC to %s %s %s %s' %(bdrtdest,exvprnint['A%s' % i].value,exvprnint['C%s' % i].value,iptnintunit))
                                            relatestaticlastpe = exvprnint['A%s' % i].value
                            elif family == 'ipv6' and 'inet6' in line:
                                iptndest = line.split()[8]
                                for bdrtdest in staticentry:
                                    if iptndest == bdrtdest:
                                        iptnintunit = exvprnint['E%s' % i].value + '.' + exvprnint['F%s' % i].value
                                        if 'STATIC to %s %s %s %s' %(bdrtdest,exvprnint['A%s' % i].value,exvprnint['C%s' % i].value,iptnintunit) not in relatedstaticcheck:
                                            if exvprnint['A%s' % i].value != relatestaticlastpe:
                                                iptnstaticrelatedpelist.append('')
                                            iptnstaticrelatedpelist.append('STATIC to %s %s %s %s' %(bdrtdest,exvprnint['A%s' % i].value,exvprnint['C%s' % i].value,iptnintunit))
                                            relatestaticlastpe = exvprnint['A%s' % i].value

    sf = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'related-static-pe.txt', "a")
    for ls in iptnstaticrelatedpelist:
        sf.write(ls + '\n')
    sf.close()
    staticfilename = '%s_%s_%s_static-route.txt' % (namingdict['bdrnode'], bdrvprnservice,namingdict['bdrl3int'])
    staticfilepath = '.\cfg\\'+ staticfilename
    f = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'cfg'+'\\'+ staticfilename, "a")
    for staticline in staticcfglist:
        f.write(staticline + '\n')
    f.close()


    return staticfilepath

def createfilternamingworksheet(filternamews,namingdict,filterpol, entry,prefixlist,portlist):
    row = filternamews['A1'].value
    filternamews['E%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957', fill_type='solid')
    filternamews['I%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957', fill_type='solid')
    filternamews['L%s' % row].fill = PatternFill(start_color='FFfff957', end_color='FFfff957', fill_type='solid')
    filternamews['A%s' % row ] = namingdict['bdrnode'] #'nokia-node'
    filternamews['B%s' % row ] = '' #'service-type'
    filternamews['C%s' % row ] = namingdict['bdrservice'] #'service-name'
    filternamews['D%s' % row ] = filterpol #'iptn-filter-name'
    filternamews['E%s' % row ] = filterpol #'bdr-filter-name'
    filternamews['F%s' % row ] = '=len(E%s)' % row #'length(max 64)'
    filternamews['G%s' % row ] = entry
    filternamews['H%s' % row] = prefixlist  # 'existing-prefix-name'
    if len(prefixlist) >= 32:
        filternamews['I%s' % row].fill = PatternFill(start_color='FFff957f', end_color='FFff957f', fill_type='solid')
    filternamews['I%s' % row ] = prefixlist #'bdr-prefix-name'
    filternamews['J%s' % row ] = '=len(I%s)' % row#'length(max 32)'
    filternamews['K%s' % row] = portlist  # 'existing-port-name'
    if len(portlist) >= 32:
        filternamews['L%s' % row].fill = PatternFill(start_color='FFff957f', end_color='FFff957f', fill_type='solid')
    filternamews['L%s' % row ] = portlist#'bdr-port-name'
    filternamews['M%s' % row ] = '=len(L%s)' % row#'length(max 32)'
    row += 1
    filternamews['A1'] = row

def createbdrfilter(file,namingdict,filternamews,direction):
    filterlines = []
    bdrlines = []
    filterlist = []
    policy = {}
    embed = []
    embedname = ''
    lowerentrynum = 1
    serviceentrynum = 21
    filterfilename = ''
    filtername = ''
    mirror_entry = []
    mirrorallentry = False
    lines = file.readlines()
    file.close()
    for a in range(len(lines)):
        lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
        lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
        lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
        filterlines.append(lines[a])
    for f in filterlines:
        if re.search('set firewall filter',f): # for ipv4 filter
            filtername = f.split(' ')[3]
            if filtername not in filterlist:
                filterlist.append(filtername)
            policy[filtername] = []
        elif re.search('set firewall family inet6 filter',f): # for ipv6 filter
            filtername = f.split(' ')[5]
            if filtername not in filterlist:
                filterlist.append(filtername)
            policy[filtername] = []
    for p in filterlist:
        for q in range(len(filterlines)):
            if re.search('set firewall filter %s' %p, filterlines[q]):  # for ipv4 filter
                filterterm = filterlines[q].split(' ')[5]
                policy.setdefault('%s_%s_termline' % (p,filterterm), []).append(q)
                if filterterm not in policy[p]:
                    policy.setdefault(p, []).append(filterterm)
            elif re.search('set firewall family inet6 filter %s' %p, filterlines[q]):  # for ipv6 filter
                filterterm = filterlines[q].split(' ')[7]
                policy.setdefault('%s_%s_termline' % (p,filterterm), []).append(q)
                if filterterm not in policy[p]:
                    policy.setdefault(p, []).append(filterterm)
    for r in filterlist:
        for s in policy[r]:
            srcaddr = 0
            dstaddr = 0
            portterm = 0
            #policy['%s_%s_src4_list' % (r, s)] = ''
            #policy['%s_%s_dst4_list' % (r, s)] = ''
            #policy['%s_%s_src6_list' % (r, s)] = ''
            #policy['%s_%s_dst6_list' % (r, s)] = ''
            #policy['%s_%s_protocol_list' % (r, s)] = ''
            #policy['%s_%s_port_list' % (r, s)] = ''
            for n in policy['%s_%s_termline' %(r,s)]:
                if re.search('set firewall filter %s term %s' %(r,s), filterlines[n]):  # for ipv4 filter
                    if 'ipv4' not in policy.setdefault('%s_%s_family' % (r, s), []):
                        policy.setdefault('%s_%s_family' % (r, s), []).append('ipv4')
                    if re.search('from source-address',filterlines[n]):
                        #srcaddr += 1
                        policy.setdefault('%s_%s_src_addr' % (r, s), []).append(filterlines[n].split(' ')[-1]) # get src ipv4 addr of each term
                        #if srcaddr >= 2:
                        #    policy['%s_%s_src4_list'%(r,s)] = '%s_%s_src' %(r,s)
                    elif re.search('from destination-address',filterlines[n]):
                        #dstaddr += 1
                        policy.setdefault('%s_%s_dst_addr' % (r, s), []).append(filterlines[n].split(' ')[-1]) # get dst ipv4 addr of each term
                        #if dstaddr >= 2:
                        #    policy['%s_%s_dst4_list'%(r,s)] = '%s_%s_dst' %(r,s)
                    elif re.search('from protocol',filterlines[n]):
                        policy.setdefault('%s_%s_protocol' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get protocol list
                        #if filterlines[n].split(' ')[-1] == 'tcp':
                        #    if 'tcp' in policy['%s_%s_protocol' % (r, s)] and 'udp' in policy['%s_%s_protocol' % (r, s)]:
                        #        policy['%s_%s_protocol_list'%(r,s)] = 'tcp-udp'
                        #elif filterlines[n].split(' ')[-1] == 'udp':
                        #    if 'tcp' in policy['%s_%s_protocol' % (r, s)] and 'udp' in policy['%s_%s_protocol' % (r, s)]:
                        #        policy['%s_%s_protocol_list'%(r,s)] = 'tcp-udp'
                    elif re.search('from port',filterlines[n]):
                        #portterm += 1
                        if filterlines[n].split(' ')[-1] not in policy.setdefault('%s_%s_port' % (r, s), []):
                            policy.setdefault('%s_%s_port' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get port list
                        #if portterm >= 2:
                        #    policy['%s_%s_port_list'%(r,s)] = '%s_%s_port_list'%(r,s)
                    elif re.search('from source-port', filterlines[n]):
                        if filterlines[n].split(' ')[-1] not in policy.setdefault('%s_%s_port' % (r, s), []):
                            policy.setdefault('%s_%s_port' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get port list
                    elif re.search('from destination-port', filterlines[n]):
                        if filterlines[n].split(' ')[-1] not in policy.setdefault('%s_%s_port' % (r, s), []):
                            policy.setdefault('%s_%s_port' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get port list
                    elif re.search('from is-fragment', filterlines[n]):
                        policy.setdefault('%s_%s_frag' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get fragment list
                    elif re.search('then', filterlines[n]):
                        if 'next term' not in filterlines[n].split('then ')[-1]:
                            policy.setdefault('%s_%s_then' % (r, s), []).append(filterlines[n].split('then ')[-1])  # get then list
                elif re.search('set firewall family inet6 filter %s term %s' %(r,s), filterlines[n]):  # for ipv6 filter
                    if 'ipv6' not in policy.setdefault('%s_%s_family' % (r, s), []):
                        policy.setdefault('%s_%s_family' % (r, s), []).append('ipv6')
                    if re.search('from source-address', filterlines[n]):
                        #srcaddr += 1
                        policy.setdefault('%s_%s_src_addr' % (r, s), []).append(filterlines[n].split(' ')[-1]) # get src ipv6 addr of each term
                        #policy['%s_%s_src_list' % (r, s)] = ''
                        #if srcaddr >= 2:
                        #    policy['%s_%s_src6_list' % (r, s)] = '%s_%s_src' % (r, s)
                    elif re.search('from destination-address', filterlines[n]):
                        #dstaddr += 1
                        policy.setdefault('%s_%s_dst_addr' % (r, s), []).append(filterlines[n].split(' ')[-1]) # get dst ipv6 addr of each term
                        #policy['%s_%s_dst_list' % (r, s)] = ''
                        #if dstaddr >= 2:
                        #    policy['%s_%s_dst6_list' % (r, s)] = '%s_%s_dst' % (r, s)
                    elif re.search('from next-header',filterlines[n]):
                        policy.setdefault('%s_%s_protocol' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get protocol list
                        #if filterlines[n].split(' ')[-1] == 'tcp':
                        #    #policy['%s_%s_protocol_list' % (r, s)] = ''
                        #    if 'tcp' in policy['%s_%s_protocol' % (r, s)] and 'udp' in policy['%s_%s_protocol' % (r, s)]:
                        #        policy['%s_%s_protocol_list'%(r,s)] = 'tcp-udp'
                        #elif filterlines[n].split(' ')[-1] == 'udp':
                        #    #policy['%s_%s_protocol_list' % (r, s)] = ''
                        #    if 'tcp' in policy['%s_%s_protocol' % (r, s)] and 'udp' in policy['%s_%s_protocol' % (r, s)]:
                        #        policy['%s_%s_protocol_list'%(r,s)] = 'tcp-udp'
                    elif re.search('from port',filterlines[n]):
                        #portterm += 1
                        if filterlines[n].split(' ')[-1] not in policy.setdefault('%s_%s_port' % (r, s), []):
                            policy.setdefault('%s_%s_port' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get port list
                        #policy['%s_%s_port_list' % (r, s)] = ''
                        #if portterm >= 2:
                        #    policy['%s_%s_port_list'%(r,s)] = '%s_%s_port_list'%(r,s)
                    elif re.search('from source-port', filterlines[n]):
                        if filterlines[n].split(' ')[-1] not in policy.setdefault('%s_%s_port' % (r, s), []):
                            policy.setdefault('%s_%s_port' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get port list
                    elif re.search('from destination-port', filterlines[n]):
                        if filterlines[n].split(' ')[-1] not in policy.setdefault('%s_%s_port' % (r, s), []):
                            policy.setdefault('%s_%s_port' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get port list
                    elif re.search('from is-fragment', filterlines[n]):
                        policy.setdefault('%s_%s_frag' % (r, s), []).append(filterlines[n].split(' ')[-1])  # get fragment list
                    elif re.search('then', filterlines[n]):
                        if 'next term' not in filterlines[n].split('then ')[-1]:
                            policy.setdefault('%s_%s_then' % (r, s), []).append(filterlines[n].split('then ')[-1])  # get then list

    #print('filterlist %s' % filterlist)
    #print('#'*80)
    #print(policy)
    #print('#'*80)
    #print(namingdict)
    #print('#'*80)

    bdrlines.append('#################################################################################################')
    bdrlines.append('')
    bdrlines.append('#### Filter policy for %s:%s:%s:%s ####' %(namingdict['bdrnode'],namingdict['bdrservice'],namingdict['bdrport'], namingdict['bdrvlan']))
    bdrlines.append('')
    if len(filterlist) >= 2:
            filtername = 'ACL_' + namingdict['bdrl3desc'] + '_ingress'
            filterfilename = '%s_%s.%s_acl_%s.txt' % (namingdict['bdrnode'], namingdict['bdrport'].replace('/','-'), namingdict['bdrvlan'], filtername)
    else:
            filtername = 'ACL_' + filterlist[0]
            filterfilename = '%s_%s.%s_acl_%s.txt' % (namingdict['bdrnode'], namingdict['bdrport'].replace('/','-'), namingdict['bdrvlan'], filtername)

    for iptnfilterpolicy in filterlist:
        for iptnterm in policy[iptnfilterpolicy]:
            iptnpol_family = []
            iptnpol_src_addr = []
            iptnpol_dst_addr = []
            iptnpol_protocol = []
            iptnpol_port = []
            iptnpol_frag = []
            iptnpol_then = []
            if not iptnterm.isnumeric():
                entry = lowerentrynum
                lowerentrynum += 1
            elif re.match('0',iptnterm):
                if '000' == iptnterm:
                    entry = 0
                else:
                    entry = lowerentrynum
                    lowerentrynum += 1
            elif iptnterm == '99998' or iptnterm == '99999':
                entry = int(iptnterm)
            else:
                entry = serviceentrynum
            iptnpol_family = policy['%s_%s_family' % (iptnfilterpolicy, iptnterm)][0]
            iptnpol_then = policy['%s_%s_then' % (iptnfilterpolicy, iptnterm)]

            if '%s_%s_src_addr' % (iptnfilterpolicy, iptnterm) in policy:
                iptnpol_src_addr = policy['%s_%s_src_addr' % (iptnfilterpolicy, iptnterm)]
            if '%s_%s_dst_addr' % (iptnfilterpolicy, iptnterm) in policy:
                iptnpol_dst_addr = policy['%s_%s_dst_addr' % (iptnfilterpolicy, iptnterm)]
            if '%s_%s_protocol' % (iptnfilterpolicy, iptnterm) in policy:
                iptnpol_protocol = policy['%s_%s_protocol' % (iptnfilterpolicy, iptnterm)]
            if '%s_%s_port' % (iptnfilterpolicy, iptnterm) in policy:
                iptnpol_port = policy['%s_%s_port' % (iptnfilterpolicy, iptnterm)]
            if '%s_%s_frag' % (iptnfilterpolicy, iptnterm) in policy:
                iptnpol_frag = policy['%s_%s_frag' % (iptnfilterpolicy, iptnterm)]

            if not iptnpol_src_addr:
                iptnpol_src_addr.append('n/a')
            if not iptnpol_dst_addr:
                iptnpol_dst_addr.append('n/a')
            if not iptnpol_port:
                iptnpol_port.append('n/a')

            if iptnpol_family == 'ipv4':
                for srcaddress in iptnpol_src_addr:
                    for dstaddress in iptnpol_dst_addr:
                        for iptnport in iptnpol_port:
                            protocol = 'n/a'
                            port = 'n/a'
                            frag = 'n/a'
                            action = 'n/a'
                            if iptnpol_protocol:
                                if 'tcp' in iptnpol_protocol and 'udp' in iptnpol_protocol:
                                    protocol = 'tcp-udp'
                                elif 'icmp' in iptnpol_protocol or 'icmp6' in iptnpol_protocol:
                                    protocol = 'icmp'
                                else:
                                    protocol = iptnpol_protocol[0]
                            if iptnport != 'n/a':
                                if '-' in iptnport:
                                    portrange = iptnport.split(' ')[-1].split('-')
                                    port = 'range start %s end %s' %(portrange[0],portrange[1])
                                else:
                                    port = 'eq %s' % iptnport.split(' ')[-1]
                            if iptnpol_frag:
                                frag = 'fragment true'
                            if iptnpol_then:
                                if 'accept' in iptnpol_then:
                                    action = 'accept'
                                elif 'discard' in iptnpol_then:
                                    action = 'drop'
                                if 'port-mirror' in iptnpol_then[0]:
                                    if '000' == iptnterm:
                                        mirrorallentry = True
                                    elif re.match('0',iptnterm):
                                        action = 'mirror-accept'
                            if not iptnterm.isnumeric():
                                #print('/configure filter ip-filter "%s" entry %s description "%s"' % (filtername, entry, iptnterm))
                                bdrlines.append('/configure filter ip-filter "%s" entry %s description "%s"' % (filtername, entry, iptnterm))
                            elif len(filterlist) >= 2:
                                if iptnterm != '99998' and iptnterm != '99999':
                                    #print('/configure filter ip-filter "%s" entry %s description "%s"' % (filtername, entry, iptnfilterpolicy))
                                    bdrlines.append('/configure filter ip-filter "%s" entry %s description "%s"' % (filtername, entry, iptnfilterpolicy))
                            elif action == 'mirror-accept':
                                bdrlines.append('/configure filter ip-filter "%s" entry %s description "MIRROR"' % (filtername, entry))
                            if srcaddress != 'n/a':
                                #print('/configure filter ip-filter "%s" entry %s match src-ip address %s' % (filtername, entry, srcaddress))
                                bdrlines.append('/configure filter ip-filter "%s" entry %s match src-ip address %s' % (filtername, entry, srcaddress))
                            if dstaddress != 'n/a':
                                #print('/configure filter ip-filter "%s" entry %s match dst-ip address %s' % (filtername, entry, dstaddress))
                                bdrlines.append('/configure filter ip-filter "%s" entry %s match dst-ip address %s' % (filtername, entry, dstaddress))
                            if protocol != 'n/a':
                                #print('/configure filter ip-filter "%s" entry %s match protocol %s' % (filtername, entry, protocol))
                                bdrlines.append('/configure filter ip-filter "%s" entry %s match protocol %s' % (filtername, entry, protocol))
                            if port != 'n/a':
                                if protocol == 'n/a':
                                    #print('/configure filter ip-filter "%s" entry %s match protocol tcp-udp' % (filtername, entry))
                                    bdrlines.append('/configure filter ip-filter "%s" entry %s match protocol tcp-udp' % (filtername, entry))
                                #print('/configure filter ip-filter "%s" entry %s match port %s' %(filtername, entry, port))
                                bdrlines.append('/configure filter ip-filter "%s" entry %s match port %s' %(filtername, entry, port))
                            if frag != 'n/a':
                                #print('/configure filter ip-filter "%s" entry %s match fragment true'%(filtername, entry))
                                bdrlines.append('/configure filter ip-filter "%s" entry %s match fragment true'%(filtername, entry))
                            if action != 'n/a':
                                #print('/configure filter ip-filter "%s" entry %s action %s'%(filtername, entry,action))
                                if action == 'mirror-accept':
                                    mirror_entry.append(entry)
                                    action = 'accept'
                                bdrlines.append('/configure filter ip-filter "%s" entry %s action %s'%(filtername, entry,action))
                                if action == 'accept' and mirrorallentry:
                                    mirror_entry.append(entry)
                            entry += 1
                            if iptnterm.isnumeric() and not re.match('0',iptnterm):
                                serviceentrynum = entry
            elif iptnpol_family == 'ipv6':
                for srcaddress in iptnpol_src_addr:
                    for dstaddress in iptnpol_dst_addr:
                        for iptnport in iptnpol_port:
                            protocol = 'n/a'
                            port = 'n/a'
                            frag = 'n/a'
                            action = 'n/a'
                            if iptnpol_protocol:
                                if 'tcp' in iptnpol_protocol and 'udp' in iptnpol_protocol:
                                    protocol = 'tcp-udp'
                                elif 'icmp' in iptnpol_protocol or 'icmp6' in iptnpol_protocol:
                                    protocol = 'ipv6-icmp'
                                else:
                                    protocol = iptnpol_protocol[0]
                            if iptnport != 'n/a':
                                if '-' in iptnport:
                                    portrange = iptnport.split(' ')[-1].split('-')
                                    port = 'range start %s end %s' %(portrange[0],portrange[1])
                                else:
                                    port = 'eq %s' % iptnport.split(' ')[-1]
                            if iptnpol_frag:
                                frag = 'fragment true'
                            if iptnpol_then:
                                if 'accept' in iptnpol_then:
                                    action = 'accept'
                                elif 'discard' in iptnpol_then:
                                    action = 'drop'
                                if 'port-mirror' in iptnpol_then[0]:
                                    if '000' == iptnterm:
                                        mirrorallentry = True
                                    elif re.match('0',iptnterm):
                                        action = 'mirror-accept'
                            if not iptnterm.isnumeric():
                                #print('/configure filter ipv6-filter "%s" entry %s description "%s"' % (filtername, entry, iptnterm))
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s description "%s"' % (filtername, entry, iptnterm))
                            elif len(filterlist) >= 2:
                                if iptnterm != '99998' and iptnterm != '99999':
                                    #print('/configure filter ipv6-filter "%s" entry %s description "%s"' % (filtername, entry, iptnfilterpolicy))
                                    bdrlines.append('/configure filter ipv6-filter "%s" entry %s description "%s"' % (filtername, entry, iptnfilterpolicy))
                            elif action == 'mirror-accept':
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s description "MIRROR"' % (filtername, entry))
                            if srcaddress != 'n/a':
                                #print('/configure filter ipv6-filter "%s" entry %s match src-ip address %s' % (filtername, entry, srcaddress))
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s match src-ip address %s' % (filtername, entry, srcaddress))
                            if dstaddress != 'n/a':
                                #print('/configure filter ipv6-filter "%s" entry %s match dst-ip address %s' % (filtername, entry, dstaddress))
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s match dst-ip address %s' % (filtername, entry, dstaddress))
                            if protocol != 'n/a':
                                #print('/configure filter ipv6-filter "%s" entry %s match next-header %s' % (filtername, entry, protocol))
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s match next-header %s' % (filtername, entry, protocol))
                            if port != 'n/a':
                                if protocol == 'n/a':
                                    #print('/configure filter ipv6-filter "%s" entry %s match next-header tcp-udp' % (filtername, entry))
                                    bdrlines.append('/configure filter ipv6-filter "%s" entry %s match next-header tcp-udp' % (filtername, entry))
                                #print('/configure filter ipv6-filter "%s" entry %s match port %s' %(filtername, entry, port))
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s match port %s' %(filtername, entry, port))
                            if frag != 'n/a':
                                #print('/configure filter ipv6-filter "%s" entry %s match fragment true'%(filtername, entry))
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s match fragment true'%(filtername, entry))
                            if action != 'n/a':
                                #print('/configure filter ipv6-filter "%s" entry %s action %s'%(filtername, entry,action))
                                if action == 'mirror-accept':
                                    mirror_entry.append(entry)
                                    action = 'accept'
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s action %s'%(filtername, entry,action))
                                if action == 'accept' and mirrorallentry:
                                    mirror_entry.append(entry)
                            entry += 1
                            if iptnterm.isnumeric() and not re.match('0',iptnterm):
                                serviceentrynum = entry


    if mirror_entry:
        bdrlines.append('')
        bdrlines.append('#################################################################################################')
        bdrlines.append('# BDRT Mirror Configuration')
        bdrlines.append('')
        bdrlines.append('/configure port x/x/x admin-state enable')
        bdrlines.append('/configure port x/x/x description "Mirror_Destination_Port"')
        bdrlines.append('/configure port x/x/x ethernet mode access')
        bdrlines.append('/configure port x/x/x ethernet encap-type dot1q')
        bdrlines.append('')
        bdrlines.append('/configure mirror mirror-dest "%s"' % filtername)
        bdrlines.append('/configure mirror mirror-dest "%s" admin-state enable' % filtername)
        bdrlines.append('/configure mirror mirror-dest "%s" sap x/x/x:x' % filtername)
        bdrlines.append('/configure mirror mirror-source "%s" admin-state enable' % filtername)
        if not mirrorallentry:
            if iptnpol_family == 'ipv4':
                for mirrorentry in mirror_entry:
                    bdrlines.append('/configure mirror mirror-source "%s" ip-filter "%s" entry %s' %(filtername,filtername,mirrorentry))
            elif iptnpol_family == 'ipv6':
                for mirrorentry in mirror_entry:
                    bdrlines.append('/configure mirror mirror-source "%s" ipv6-filter "%s" entry %s' %(filtername,filtername,mirrorentry))
        else:
            if namingdict['bdrvlan']:
                sapid = '%s:%s' %(namingdict['bdrport'],namingdict['bdrvlan'])
            else:
                sapid = '%s' % (namingdict['bdrport'])
            if direction == 'input':
                bdrlines.append('/configure mirror mirror-source "%s" sap %s ingress true' %(filtername,sapid))
            elif direction == 'output':
                bdrlines.append('/configure mirror mirror-source "%s" sap %s egress true' % (filtername, sapid))

    """                       
    bdrlines.append('')
    bdrlines.append('#################################################################################################')
    bdrlines.append('')
    bdrlines.append('#### Filter policy for %s:%s:%s:%s ####' %(namingdict['bdrnode'],namingdict['bdrservice'],namingdict['bdrport'], namingdict['bdrvlan']))
    bdrlines.append('')
    if len(filterlist) >= 2:
        filterfilename = '%s_%s.%s_acl-embed-filter.txt' % (namingdict['bdrnode'], namingdict['bdrport'].replace('/','-'), namingdict['bdrvlan'])
        filtername = '%s.%s_embed-filter' % (namingdict['bdrport'], namingdict['bdrvlan'])
        bdrlines.append('#### Please verify embed-filter policy naming and offset of each child policy ####')
    else:
        filterfilename = '%s_%s.%s_acl_%s.txt' % (namingdict['bdrnode'], namingdict['bdrport'].replace('/','-'), namingdict['bdrvlan'], filterlist[0])
        filtername = filterlist[0]
    for r in filterlist:
        for s in policy[r]:
            if policy['%s_%s_src4_list' % (r, s)]:  # this policy having ip-prefix-list for source address
                bdrlines.append('#### Please verify ip-prefix-list naming (limit 32 character) ####')
                bdrlines.append('')
                for src in policy['%s_%s_src_addr' % (r, s)]:
                    #print ('/configure filter match-list ip-prefix-list "%s" prefix %s' %(policy['%s_%s_src_list'%(r,s)],src))
                    bdrlines.append('/configure filter match-list ip-prefix-list "%s" prefix %s' %(policy['%s_%s_src4_list'%(r,s)],src))
            if policy['%s_%s_dst4_list' % (r, s)]:  # this policy having ip-prefix-list for destination address
                bdrlines.append('#### Please verify ip-prefix-list naming (limit 32 character) ####')
                bdrlines.append('')
                for dst in policy['%s_%s_dst_addr' % (r, s)]:
                    bdrlines.append('/configure filter match-list ip-prefix-list "%s" prefix %s' % (policy['%s_%s_dst4_list' % (r, s)], dst))
            if policy['%s_%s_src6_list' % (r, s)]:  # this policy having ip-prefix-list for source address
                bdrlines.append('#### Please verify ipv6-prefix-list naming (limit 32 character) ####')
                bdrlines.append('')
                for src in policy['%s_%s_src_addr' % (r, s)]:
                    #print ('/configure filter match-list ipv6-prefix-list "%s" prefix %s' %(policy['%s_%s_src6_list'%(r,s)],src))
                    bdrlines.append('/configure filter match-list ipv6-prefix-list "%s" prefix %s' %(policy['%s_%s_src6_list'%(r,s)],src))
            if policy['%s_%s_dst6_list' % (r, s)]:  # this policy having ip-prefix-list for destination address
                bdrlines.append('#### Please verify ipv6-prefix-list naming (limit 32 character) ####')
                bdrlines.append('')
                for dst in policy['%s_%s_dst_addr' % (r, s)]:
                    bdrlines.append('/configure filter match-list ipv6-prefix-list "%s" prefix %s' % (policy['%s_%s_dst6_list' % (r, s)], dst))
            if policy['%s_%s_port_list' % (r, s)]:  # this policy having port-list
                bdrlines.append('#### Please verify port-list naming (limit 32 character) ####')
                bdrlines.append('')
                for por in policy['%s_%s_port' % (r, s)]:
                    if '-' in por:
                        bdrlines.append('/configure filter match-list port-list %s range start %s end %s' % (policy['%s_%s_port_list' % (r, s)], por.split('-')[0], por.split('-')[1]))
                        #print('/configure filter match-list port-list %s range start %s end %s' % (policy['%s_%s_port_list' % (r, s)], por.split('-')[0], por.split('-')[1]))
                    else:
                        bdrlines.append('/configure filter match-list port-list %s port %s' % (policy['%s_%s_port_list' % (r, s)], por))
                        #print('/configure filter match-list port-list %s port %s' % (policy['%s_%s_port_list' % (r, s)], por))

    if len(filterlist) >= 2:
        bdrlines.append('')
        bdrlines.append('#### This policy is embed filter policy:Please verify embed policy name ####')
        for r in filterlist:
            embed.append(r)
            embedname =  'Embed_%s:%s'% (namingdict['bdrport'], namingdict['bdrvlan'])
    for r in filterlist:
        bdrlines.append('')
        bdrlines .append('#### Filter policy:%s configuration ####' %r)
        bdrlines.append('')
        for s in policy[r]:
            if not s.isnumeric():
                entry = dynamicentrynum
                dynamicentrynum += 1
            else:
                entry = s
            for n in policy['%s_%s_termline' %(r,s)]:
                if re.search('set firewall filter %s term %s' %(r,s), filterlines[n]):  # for ipv4 filter
                    if r in embed:
                        emcfg = '/configure filter ip-filter "%s" scope embedded' %r
                        if emcfg not in bdrlines:
                            bdrlines.append(emcfg)
                            bdrlines.append('')
                    if not s.isnumeric():
                        desc = '/configure filter ip-filter "%s" entry %s description "%s"' % (r, entry, s)
                        if desc not in bdrlines:
                            bdrlines.append(desc)
                    if re.search('from source-address',filterlines[n]):
                        if policy['%s_%s_src4_list' % (r, s)]:
                            #print('/configure filter ip-filter %s entry %s match src-ip ip-prefix-list "%s"' % (r,s,policy['%s_%s_src_list' % (r, s)]))
                            fcfg = '/configure filter ip-filter "%s" entry %s match src-ip ip-prefix-list "%s"' % (r,entry,policy['%s_%s_src4_list' % (r, s)])
                            if fcfg not in bdrlines:
                                bdrlines.append(fcfg)
                                createfilternamingworksheet(filternamews,namingdict,r, entry,policy['%s_%s_src4_list' % (r, s)],'')
                        else:
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match src-ip address %s' % (r,entry,policy['%s_%s_src_addr' % (r, s)][0]))
                    elif re.search('from destination-address', filterlines[n]):
                        if policy['%s_%s_dst4_list' % (r, s)]:
                            fcfg = '/configure filter ip-filter "%s" entry %s match dst-ip ip-prefix-list "%s"' % (r, entry, policy['%s_%s_dst4_list' % (r, s)])
                            if fcfg not in bdrlines:
                                bdrlines.append(fcfg)
                                createfilternamingworksheet(filternamews,namingdict,r, entry,policy['%s_%s_dst4_list' % (r, s)],'')
                        else:
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match dst-ip address %s' % (r, entry, policy['%s_%s_dst_addr' % (r, s)][0]))
                    elif re.search('from protocol', filterlines[n]):
                        if policy['%s_%s_protocol_list'%(r,s)]:
                            fcfg = '/configure filter ip-filter "%s" entry %s match protocol tcp-udp' % (r,entry)
                            if fcfg not in bdrlines:
                                bdrlines.append(fcfg)
                        else:
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match protocol %s' % (r, entry, filterlines[n].split(' ')[-1] ))
                    elif re.search('from port', filterlines[n]):
                        if policy['%s_%s_port_list'%(r,s)]:
                            fcfg = '/configure filter ip-filter "%s" entry %s match port port-list "%s"' %(r, entry, policy['%s_%s_port_list'%(r,s)])
                            if fcfg not in bdrlines:
                                bdrlines.append(fcfg)
                                createfilternamingworksheet(filternamews,namingdict,r, entry,'',policy['%s_%s_port_list'%(r,s)])
                        elif '-' in filterlines[n].split(' ')[-1]:
                            #print(filterlines[n].split(' ')[-1])
                            portrange = filterlines[n].split(' ')[-1].split('-')
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match port range start %s end %s' %(r, entry, portrange[0], portrange[1]))
                        else:
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match port eq %s' %(r, entry, filterlines[n].split(' ')[-1]))
                    elif re.search('from source-port', filterlines[n]):
                        if '-' in filterlines[n].split(' ')[-1]:
                            portrange = filterlines[n].split(' ')[-1].split('-')
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match src-port range start %s end %s' % (r, entry, portrange[0], portrange[1]))
                        else:
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match src-port eq %s' % (r, entry, filterlines[n].split(' ')[-1]))
                    elif re.search('from destination-port', filterlines[n]):
                        if '-' in filterlines[n].split(' ')[-1]:
                            portrange = filterlines[n].split(' ')[-1].split('-')
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match dst-port range start %s end %s' % (r, entry, portrange[0], portrange[1]))
                        else:
                            bdrlines.append('/configure filter ip-filter "%s" entry %s match dst-port eq %s' % (r, entry, filterlines[n].split(' ')[-1]))
                    elif re.search('from is-fragment', filterlines[n]):
                        bdrlines.append('/configure filter ip-filter "%s" entry %s match fragment true'%(r,entry))
                    elif re.search('then accept', filterlines[n]):
                        bdrlines.append('/configure filter ip-filter "%s" entry %s action accept'%(r,entry))
                    elif re.search('then discard', filterlines[n]):
                        bdrlines.append('/configure filter ip-filter "%s" entry %s action drop' % (r, entry))
                elif re.search('set firewall family inet6 filter %s term %s' % (r, s),filterlines[n]):  # for ipv6 filter
                    if r in embed:
                        emcfg = '/configure filter ipv6-filter "%s" scope embedded' %r
                        if emcfg not in bdrlines:
                            bdrlines.append(emcfg)
                            bdrlines.append('')
                    if not s.isnumeric():
                        desc = '/configure filter ipv6-filter "%s" entry %s description "%s"' % (r, entry, s)
                        if desc not in bdrlines:
                            bdrlines.append(desc)
                    if re.search('from source-address', filterlines[n]):
                        if policy['%s_%s_src6_list' % (r, s)]:
                            #print('/configure filter ipv6-filter %s entry %s match src-ip ipv6-prefix-list "%s"' % (r,s,policy['%s_%s_src_list' % (r, s)]))
                            fcfg = '/configure filter ipv6-filter "%s" entry %s match src-ip ipv6-prefix-list "%s"' % (r,entry,policy['%s_%s_src6_list' % (r, s)])
                            if fcfg not in bdrlines:
                                bdrlines.append(fcfg)
                                createfilternamingworksheet(filternamews,namingdict,r, entry,policy['%s_%s_src6_list' % (r, s)],'')
                        else:
                            bdrlines.append('/configure filter ipv6-filter "%s" entry %s match src-ip address %s' % (r, entry, policy['%s_%s_src_addr' % (r, s)][0]))
                    elif re.search('from destination-address', filterlines[n]):
                        if policy['%s_%s_dst6_list' % (r, s)]:
                            fcfg = '/configure filter ipv6-filter "%s" entry %s match dst-ip ipv6-prefix-list "%s"' % (r, entry, policy['%s_%s_dst6_list' % (r, s)])
                            if fcfg not in bdrlines:
                                bdrlines.append(fcfg)
                                createfilternamingworksheet(filternamews,namingdict,r, entry,policy['%s_%s_dst6_list' % (r, s)],'')
                        else:
                            bdrlines.append('/configure filter ipv6-filter "%s" entry %s match dst-ip address %s' % (r, entry, policy['%s_%s_dst_addr' % (r, s)][0]))
                    elif re.search('from next-header', filterlines[n]):
                        if policy['%s_%s_protocol_list'%(r,s)]:
                            fcfg = '/configure filter ipv6-filter "%s" entry %s match next-header tcp-udp' % (r,entry)
                            if fcfg not in bdrlines:
                                bdrlines.append(fcfg)
                        else:
                            if re.search('icmp',filterlines[n].split(' ')[-1]):
                                nh = '/configure filter ipv6-filter "%s" entry %s match next-header ipv6-icmp' % (r, entry)
                                if nh not in bdrlines:
                                    bdrlines.append(nh)
                            else:
                                bdrlines.append('/configure filter ipv6-filter "%s" entry %s match next-header %s' % (r, entry, filterlines[n].split(' ')[-1] ))
                    elif re.search('from port', filterlines[n]):
                        if policy['%s_%s_port_list'%(r,s)]:
                            fcfg = '/configure filter ipv6-filter "%s" entry %s match port port-list "%s"' %(r, entry, policy['%s_%s_port_list'%(r,s)])
                            if fcfg not in bdrlines:
                                bdrlines.append(fcfg)
                                createfilternamingworksheet(filternamews,namingdict,r, entry,'',policy['%s_%s_port_list'%(r,s)])
                        elif '-' in filterlines[n].split(' ')[-1]:
                            #print(filterlines[n].split(' ')[-1])
                            portrange = filterlines[n].split(' ')[-1].split('-')
                            bdrlines.append('/configure filter ipv6-filter "%s" entry %s match port range start %s end %s' %(r, entry, portrange[0], portrange[1]))
                        else:
                            bdrlines.append('/configure filter ipv6-filter "%s" entry %s match port eq %s' %(r, entry, filterlines[n].split(' ')[-1]))
                    elif re.search('from source-port', filterlines[n]):
                        if '-' in filterlines[n].split(' ')[-1]:
                            portrange = filterlines[n].split(' ')[-1].split('-')
                            bdrlines.append('/configure filter ipv6-filter "%s" entry %s match src-port range start %s end %s' % (r, entry, portrange[0], portrange[1]))
                        else:
                            bdrlines.append('/configure filter ipv6-filter "%s" entry %s match src-port eq %s' % (r, entry, filterlines[n].split(' ')[-1]))
                    elif re.search('from destination-port', filterlines[n]):
                        if '-' in filterlines[n].split(' ')[-1]:
                            portrange = filterlines[n].split(' ')[-1].split('-')
                            bdrlines.append('/configure filter ipv6-filter "%s" entry %s match dst-port range start %s end %s' % (r, entry, portrange[0], portrange[1]))
                        else:
                            bdrlines.append('/configure filter ipv6-filter "%s" entry %s match dst-port eq %s' % (r, entry, filterlines[n].split(' ')[-1]))
                    elif re.search('from is-fragment', filterlines[n]):
                        bdrlines.append('/configure filter ipv6-filter "%s" entry %s match fragment true'%(r,entry))
                    elif re.search('then accept', filterlines[n]):
                        bdrlines.append('/configure filter ipv6-filter "%s" entry %s action accept'%(r,entry))
                    elif re.search('then discard', filterlines[n]):
                        bdrlines.append('/configure filter ipv6-filter "%s" entry %s action drop' % (r, entry))


    if embedname:
        offset = 10000
        bdrlines.append('')
        bdrlines.append('#### embed policy configuration ####')
        bdrlines.append('')
        for r in filterlist:
            bdrlines.append('/configure filter ip-filter "%s" embed filter "%s" offset %s' %(embedname,r,offset))
            offset += 20000
    """
    #print(bdrlines)
    filterfilepath = '.\cfg\\'+ filterfilename
    f = open(migratepath +'\\'+ t.strftime("%Y%m%d-%H%M") +'\\'+'cfg'+'\\'+ filterfilename, "a")
    for bdrl in bdrlines:
        f.write(bdrl + '\n')
    f.close()

    return filterfilepath, filtername


def createbdrportlagworksheet(portws,r,namingdict,exportdict,bdrservicetype, bdrserviceid, iptnservicerd,filternamews):
    sapqos = ''
    filterinfilepath = ''
    filterinname = ''
    filteroutfilepath = ''
    filteroutname = ''
    if exportdict['iptninputfilter'] != None:
        i = open(extractpath + exportdict['iptninputfilter'], 'r')
        filterinfilepath, filterinname = createbdrfilter(i,namingdict,filternamews,'input')
    if exportdict['iptnoutputfilter'] != None:
        o = open(extractpath + exportdict['iptnoutputfilter'], 'r')
        filteroutfilepath, filteroutname = createbdrfilter(o,namingdict,filternamews,'output')
    if exportdict['iptnfc'] == 'mobile':
        sapqos = 'mobile'
    elif exportdict['iptnfc'] == 'corporate':
        sapqos = 'corporate'
    elif exportdict['iptnfc'] == 'internal':
        sapqos = 'internal'
    elif exportdict['iptnfc'] == 'network-control':
        sapqos = 'network-control'
    elif exportdict['iptnfc'] == 'best-effort':
        sapqos = 'best-effort'
    elif exportdict['iptnfc'] == 'exp_classifier_ipcbb':
        sapqos = 'ipcbb'


    portws['A%s' % r] = namingdict['bdrnode'] #'nokia-node'
    portws['B%s' % r] = namingdict['bdrsystem'] #'system-ip'
    if not bdrservicetype and namingdict['bdrlag']:
        portws['C%s' % r] = 'lag-member'  # 'service-type'
    else:
        portws['C%s' % r] = bdrservicetype #'service-type'
    portws['D%s' % r] = namingdict['bdrservice'] #'service-name'
    portws['E%s' % r] = bdrserviceid #'service-id'
    portws['F%s' % r] = namingdict['bdrport'] #'port/lag'
    portws['G%s' % r] = namingdict['bdrportencap']  # 'port-encapsulation'
    portws['H%s' % r] = exportdict['iptnportadminstate'] #'admin-state'
    portws['I%s' % r] = namingdict['bdrportdesc'] #'physical-description'
    portws['J%s' % r] = exportdict['iptnportspeed'] #'speed'
    portws['K%s' % r] = exportdict['iptnportautonego'] #'auto-negotiation'
    portws['L%s' % r] = namingdict['bdrlag'] #'LAG'
    portws['M%s' % r] = exportdict['iptnaedesc'] #'LAG-protocol'
    if exportdict['iptnaeminlink'] != None:
        portws['N%s' % r] = int(exportdict['iptnaeminlink'])-1 #'LAG-port-threshold'
    if (namingdict['bdrvlan'] != None) and (namingdict['bdrportencap'] != None):
        if namingdict['bdrportencap'] == 'null':
            portws['O%s' % r] = namingdict['bdrport'] #'SAP'
        elif namingdict['bdrportencap'] == 'dot1q':
            portws['O%s' % r] = '%s:%s' % (namingdict['bdrport'], namingdict['bdrvlan'])  # 'SAP'
    portws['P%s' % r] = exportdict['iptnunitstate'] #'SAP-state'
    portws['Q%s' % r] = namingdict['bdrsapdesc'] #'SAP-description'
    portws['R%s' % r] = filterinname #'input-filter-policy-name'
    portws['S%s' % r].hyperlink = filterinfilepath #'input-filter-configuration'
    portws['T%s' % r] = filteroutname #'output-filter-policy-name'
    portws['U%s' % r].hyperlink = filteroutfilepath #'output-filter-configuration'
    portws['V%s' % r] = sapqos #'sap-ingress qos'
    portws['W%s' % r] = sapqos #'sap-egress qos'
    portws['X%s' % r] = namingdict['iptnnode'] #'iptn-node'
    portws['Y%s' % r] = exportdict['iptnservicetype'] #'iptn-service-type'
    portws['Z%s' % r] = namingdict['iptnservice'] #'iptn-service'
    if exportdict['iptnunit']:
        portws['AA%s' % r] = '%s.%s'%(exportdict['iptnport'],exportdict['iptnunit']) #'iptn-interface'
    else:
        portws['AA%s' % r] = exportdict['iptnport']  # 'iptn-interface'

    if not bdrservicetype and namingdict['bdrlag']:
        portws['D%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['E%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['M%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['N%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['O%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['P%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['Q%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['R%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['S%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['T%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['U%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['V%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        portws['W%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')

    r+=1
    portws['A1'] = r


def createsummaryworksheet(sumws,r,namingdict,exportdict,bdrservicetype, bdrserviceid, iptnservicerd):
    if namingdict['bdrvlan'] == None:
        sumws['C%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        sumws['D%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        sumws['E%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        sumws['I%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        sumws['J%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        sumws['K%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        sumws['O%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        sumws['P%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')
        sumws['Q%s' % r].fill = PatternFill(start_color='FF565656', end_color='FF565656', fill_type='solid')

    sumws['A1'] = r+1
    sumws['A%s' % r] = namingdict['bdrnode'] #'nokia-node'
    sumws['B%s' % r] = namingdict['bdrsystem'] #'system-ip'
    sumws['C%s' % r] = namingdict['bdrservice'] #'service-name'
    if namingdict['bdrvlan'] == '41' and namingdict['bdrservice'] == None:
        sumws['C%s' % r] = 'Please manual migrate this SAP(switch mgmt)'
    sumws['D%s' % r] = bdrserviceid #'service-id'
    sumws['E%s' % r] = bdrservicetype #'service-type'
    sumws['F%s' % r] = namingdict['bdrport'] #'port'
    sumws['G%s' % r] = exportdict['iptnportadminstate'] #'port-state'
    sumws['H%s' % r] = namingdict['bdrlag'] #'LAG'
    if (namingdict['bdrvlan'] != None) and (namingdict['bdrportencap'] != None): # port member of lag have not any vlan
        if namingdict['bdrportencap'] == 'null':
            sumws['I%s' % r] = namingdict['bdrport'] #'SAP'
        elif namingdict['bdrportencap'] == 'dot1q':
            sumws['I%s' % r] = '%s:%s' % (namingdict['bdrport'], namingdict['bdrvlan'])  # 'SAP'
    sumws['J%s' % r] = exportdict['iptnunitstate'] #'SAP-state'
    sumws['K%s' % r] = namingdict['bdrl3int'] #'vprn-interface-name'
    sumws['L%s' % r] = namingdict['iptnnode'] #'iptn-node'
    sumws['M%s' % r] = namingdict['iptnloopback'] #'iptn-loopback'
    sumws['N%s' % r] = namingdict['iptnintunit'] #'iptn-int'
    sumws['O%s' % r] = exportdict['iptnservicetype'] #'iptn-service-type'
    sumws['P%s' % r] = exportdict['iptnservice'] #'iptn-service-name'
    sumws['Q%s' % r] = iptnservicerd #'iptn-service-rd'

def maintask(misum,summaryrow, namingdict,naming,node,namingrow):
    namingdict = createnamingdict(naming, node, namingrow)
    exvprnintdict = {}
    exl2vpndict = {}
    exvplsdict = {}
    exbddict = {}
    exvswdict = {}
    noderow = {}
    l2map = {}
    portrow = misum['port-lag']['A1'].value
    noderow['exportstartrow'], noderow['exportendrow'] = findextractnodeindex(exindex, namingdict['iptnnode'], 'B', 'C', 'D')
    noderow['exvprnstartrow'], noderow['exvprnendrow'] = findextractnodeindex(exindex, namingdict['iptnnode'], 'F', 'G', 'H')
    noderow['exvprnintstartrow'], noderow['exvprnintendrow'] = findextractnodeindex(exindex, namingdict['iptnnode'], 'J', 'K', 'L')
    noderow['exl2startrow'], noderow['exl2endrow'] = findextractnodeindex(exindex, namingdict['iptnnode'], 'N', 'O', 'P')
    noderow['exvplsstartrow'], noderow['exvplsendrow'] = findextractnodeindex(exindex, namingdict['iptnnode'], 'R', 'S', 'T')
    noderow['exbdstartrow'], noderow['exbdendrow'] = findextractnodeindex(exindex, namingdict['iptnnode'], 'V', 'W', 'X')
    noderow['exvswstartrow'], noderow['exvswendrow'] = findextractnodeindex(exindex, namingdict['iptnnode'], 'Z', 'AA', 'AB')
    if '.' in namingdict['iptnintunit']:
        iptnp = namingdict['iptnintunit'].split('.')[0]
        iptnu = namingdict['iptnintunit'].split('.')[1]
    else:
        iptnp = namingdict['iptnintunit']
        iptnu = ''

    for p in range(noderow['exportstartrow'],noderow['exportendrow']+1):
        exportdict = createextractionportdict(export,node,p)
        #print(exvprnintdict)
        #print(exportdict)
        if iptnp == exportdict['iptnport']:
            if exportdict['iptnunit'] == None:
                createsummaryworksheet(misum['summary'], summaryrow, namingdict, exportdict,'','','')
                createbdrportlagworksheet(misum['port-lag'], portrow, namingdict, exportdict, '','','',misum['filter-naming'])
        if iptnp == exportdict['iptnport']:
            if iptnu == exportdict['iptnunit']:
                if exportdict['iptnservicetype'] == 'VPRN':
                    exvprnintdict,bdrservicetype, bdrserviceid = checkvprnint(namingdict, exportdict, iptnp, iptnu, noderow)
                    iptnservicerd = exvprnintdict['iptnrd']
                    #print(exvprnintdict['iptnservice'])
                elif exportdict['iptnservicetype'] == 'EPIPE':
                    exl2vpndict, bdrservicetype, bdrserviceid  = checkl2vpn(namingdict, exportdict, iptnp, iptnu, noderow)
                    iptnservicerd = exl2vpndict['iptnrd']
                    #print(exl2vpndict['iptnrt'])
                    #print(exl2vpndict)
                    createl2servicerelatedworksheet(misum['l2service-related-pe'], exl2vpn, namingdict, exl2vpndict['iptnrt'],exportdict['iptnservicetype'],bdrservicetype,bdrserviceid,'U','V','W')
                    # print('service-type : %s service-name : %s' % (exportdict['iptnservicetype'], exportdict['iptnservice']))
                elif exportdict['iptnservicetype'] == 'VPLS':
                    exvplsdict, bdrservicetype, bdrserviceid = checkvpls(namingdict, exportdict, iptnp, iptnu, noderow)
                    iptnservicerd = exvplsdict['iptnrd']
                    #print(exvplsdict['iptnrt'])
                    #print(exvplsdict)
                    createl2servicerelatedworksheet(misum['l2service-related-pe'], exvpls, namingdict,exvplsdict['iptnrt'], exportdict['iptnservicetype'],bdrservicetype, bdrserviceid,'T', 'U', 'V')
                elif exportdict['iptnservicetype'] == 'Bridge-Domain':
                    exbddict, l2map, bdrservicetype, bdrserviceid = checkbd(namingdict, exportdict, iptnp, iptnu, noderow)
                    iptnservicerd = 'N/A'
                    createbdservicerelateworksheet(misum['l2service-related-pe'],exbridge,namingdict,exbddict['iptnservice'],exportdict['iptnservicetype'],bdrservicetype,bdrserviceid)
                    #if exbddict['iptnint'] == 'ae0':
                    #print(exbddict['iptnservice'])
                    #print(exbddict)
                    # print('service-type : %s service-name : %s' % (exportdict['iptnservicetype'], exportdict['iptnservice']))
                elif exportdict['iptnservicetype'] == 'Virtual-Switch':
                    exvswdict, l2map, bdrservicetype, bdrserviceid = checkvsw(namingdict, exportdict, iptnp, iptnu, noderow)
                    iptnservicerd = 'N/A'
                    createvswservicerelatedworksheet(misum['l2service-related-pe'], exvsw, namingdict,exvswdict['iptnrt'],exvswdict['iptnbd'], exportdict['iptnservicetype'], bdrservicetype,bdrserviceid, 'P', 'T', 'U')
                    #print(exvswdict['iptnrt'])
                    #print(exvswdict)
                    # print('service-type : %s service-name : %s' % (exportdict['iptnservicetype'], exportdict['iptnservice']))
                else: # for unservice port
                    bdrserviceid = ''
                    bdrservicetype = ''
                    iptnservicerd = ''

                #print(iptnp)
                createsummaryworksheet(misum['summary'], summaryrow, namingdict, exportdict, bdrservicetype, bdrserviceid, iptnservicerd)
                createbdrportlagworksheet(misum['port-lag'], portrow, namingdict, exportdict, bdrservicetype, bdrserviceid, iptnservicerd,misum['filter-naming'])
                filterin = misum['port-lag']['R%s'%portrow].value
                filterout = misum['port-lag']['T%s'%portrow].value
                sapqos = misum['port-lag']['V%s' % portrow].value
                for customerrow in range(2, customerws.max_row + 1):
                    if bdrserviceid == customerws['B%s' % customerrow].value:
                        sapqos = customerws['C%s' % customerrow].value
                serviceidepipesapcheck = []

                if exportdict['iptnservicetype'] == 'VPRN':
                    ef = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidvprnint.txt',"a")
                    ef.write(str(bdrserviceid) + '\n')
                    createvprnintinfo(misum['vprn-interface'], sapqos, namingdict, filterin, filterout, bdrservicetype,bdrserviceid, exvprnintdict,misum['policy-naming'])

                elif exportdict['iptnservicetype'] == 'EPIPE':
                    ef = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidepipesap.txt',"a")
                    ef.write(str(bdrserviceid) + '\n')
                    createepipesapinfo(misum['epipe-sap'], sapqos, namingdict, filterin, filterout, bdrservicetype,bdrserviceid, exl2vpndict)

                elif exportdict['iptnservicetype'] == 'VPLS':
                    if bdrservicetype == 'rVPLS':
                        ef = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidrvplssap.txt',"a")
                        ef.write(str(bdrserviceid) + '\n')
                        creatervplssapinfo(misum['rvpls-sap'], sapqos, namingdict, filterin, filterout, bdrservicetype,bdrserviceid,exportdict['iptnservicetype'])
                    elif bdrservicetype == 'VPLS':
                        ef = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidvplssap.txt',"a")
                        ef.write(str(bdrserviceid) + '\n')
                        createvplssapinfo(misum['vpls-sap'], sapqos, namingdict, filterin, filterout, bdrservicetype,bdrserviceid)

                elif exportdict['iptnservicetype'] == 'Bridge-Domain':
                    ef = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidrvplssap.txt',"a")
                    ef.write(str(bdrserviceid) + '\n')
                    creatervplssapinfo(misum['rvpls-sap'], sapqos, namingdict, filterin, filterout, bdrservicetype,bdrserviceid,exportdict['iptnservicetype'])

                elif exportdict['iptnservicetype'] == 'Virtual-Switch':
                    ef = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidrvplssap.txt',"a")
                    ef.write(str(bdrserviceid) + '\n')
                    creatervplssapinfo(misum['rvpls-sap'], sapqos, namingdict, filterin, filterout, bdrservicetype,bdrserviceid,exportdict['iptnservicetype'])



    #print('iptnnode : %s start %s end %s' %(namingdict['iptnnode'],exportstartrow, exportendrow))

def bdrepipesap():
    misum = openpyxl.load_workbook(filename=migratepath+t.strftime("%Y%m%d-%H%M")+'\\'+misumname)
    serviceidepipesapcheck = []
    if os.path.exists(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidepipesap.txt'):
        ff = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidepipesap.txt','r').readlines()
        fg = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceepipesapinfo.txt', 'r').readlines()
        # print('g : %s' % g)
        for a in range(len(ff)):
            ff[a] = ff[a].rstrip('\n')  # remove newline('\n') from end of line
            if ff[a] not in serviceidepipesapcheck:
                serviceidepipesapcheck.append(ff[a])
        #print(serviceidepipesapcheck)
        for serviceid in serviceidepipesapcheck:
            #print(serviceid)
            for epipesapline in range(len(fg)):
                fg[epipesapline] = fg[epipesapline].rstrip('\n')
                epipesap = fg[epipesapline].split('$%')
                if epipesap[3] == serviceid:
                    createepipesapworksheet(misum['epipe-sap'],epipesap)
        misum.save(migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)

def bdrvplssap():
    misum = openpyxl.load_workbook(filename=migratepath+t.strftime("%Y%m%d-%H%M")+'\\'+misumname)
    serviceidvplssapcheck = []
    if os.path.exists(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidvplssap.txt'):
        ff = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidvplssap.txt','r').readlines()
        fg = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'servicevplssapinfo.txt', 'r').readlines()
        # print('g : %s' % g)
        for a in range(len(ff)):
            ff[a] = ff[a].rstrip('\n')  # remove newline('\n') from end of line
            if ff[a] not in serviceidvplssapcheck:
                serviceidvplssapcheck.append(ff[a])
        #print(serviceidvplssapcheck)
        for serviceid in serviceidvplssapcheck:
            #print(serviceid)
            for vplssapline in range(len(fg)):
                fg[vplssapline] = fg[vplssapline].rstrip('\n')
                vplssap = fg[vplssapline].split('$%')
                if vplssap[3] == serviceid:
                    createvplssapworksheet(misum['vpls-sap'],vplssap)
        misum.save(migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)

def bdrrvplssap():
    misum = openpyxl.load_workbook(filename=migratepath+t.strftime("%Y%m%d-%H%M")+'\\'+misumname)
    serviceidrvplssapcheck = []
    nm = open(migratepath + inputpath + "latestnaming.txt", 'r')
    latestnaming = nm.readlines()[0]
    naws = openpyxl.load_workbook(filename=migratepath + inputpath + latestnaming)
    naming = naws['naming']
    exportdict = {}
    if os.path.exists(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidrvplssap.txt'):
        ff = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidrvplssap.txt','r').readlines()
        fg = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'servicervplssapinfo.txt', 'r').readlines()
        # print('g : %s' % g)
        for a in range(len(ff)):
            ff[a] = ff[a].rstrip('\n')  # remove newline('\n') from end of line
            if ff[a] not in serviceidrvplssapcheck:
                serviceidrvplssapcheck.append(ff[a])
        #print(serviceidrvplssapcheck)
        for serviceid in serviceidrvplssapcheck:
            #print(serviceid)
            for rvplssapline in range(len(fg)):
                noderow = {}
                iptnvprnint = {}
                fg[rvplssapline] = fg[rvplssapline].rstrip('\n')
                rvplssap = fg[rvplssapline].split('$%')
                if rvplssap[3] == serviceid:
                    creatervplssapworksheet(misum['rvpls-sap'], rvplssap)
                    iptnsapcheck = rvplssap[9] + rvplssap[10] + rvplssap[11]
                    for nl in range(3, naming.max_row + 1):
                        if naming['O%s' % nl].value != None:
                            namingcheck = naming['A%s' % nl].value + naming['O%s' % nl].value + naming['C%s' % nl].value
                            if namingcheck == iptnsapcheck:
                                namingdict = createnamingdict(naming, rvplssap[0], nl)
                    if 'Virtual-Switch' in rvplssap[12]:
                        noderow['exvprnintstartrow'], noderow['exvprnintendrow'] = findextractnodeindex(exindex, rvplssap[9], 'J', 'K', 'L')
                        noderow['exvswstartrow'], noderow['exvswendrow'] = findextractnodeindex(exindex, rvplssap[9], 'Z','AA', 'AB')
                        noderow['exportstartrow'], noderow['exportendrow'] = findextractnodeindex(exindex, rvplssap[9], 'B', 'C', 'D')
                        for vl in range(noderow['exvswstartrow'], noderow['exvswendrow']+1):
                            exvswcheck = exvsw['A%s' %vl].value+exvsw['C%s' %vl].value+exvsw['K%s' %vl].value+'.'+exvsw['L%s' %vl].value
                            if iptnsapcheck == exvswcheck:
                                exvswirb = exvsw['Q%s' %vl].value
                                exvswirbvrf = exvsw['R%s' %vl].value
                                if exvswirbvrf != None:
                                    exvswcheck = rvplssap[9]+exvswirb+exvswirbvrf
                                for vrfl in range(noderow['exvprnintstartrow'], noderow['exvprnintendrow']+1):
                                    vprnintcheck = exvprnint['A%s'%vrfl].value+exvprnint['E%s'%vrfl].value+'.'+exvprnint['F%s'%vrfl].value+exvprnint['C%s'%vrfl].value
                                    if exvswcheck == vprnintcheck:
                                        iptnvprnint = createvprnintdict(exvprnint, rvplssap[9], vrfl)
                                for exp in range(noderow['exportstartrow'], noderow['exportendrow']+1):
                                    if 'irb' == export['C%s'%exp].value:
                                        if exvswirb.split('.')[-1] == export['M%s'%exp].value:
                                            exportdict['iptninputfiltername'] = export['Q%s'%exp].value
                                            exportdict['iptninputfilter'] = export['R%s'%exp].value
                                            exportdict['iptnoutputfiltername'] = export['S%s'%exp].value
                                            exportdict['iptnoutputfilter'] = export['T%s'%exp].value
                                    elif exvswirb.split('.')[0] == export['C%s'%exp].value:
                                        if exvswirb.split('.')[-1] == export['M%s'%exp].value:
                                            exportdict['iptninputfiltername'] = export['Q%s'%exp].value
                                            exportdict['iptninputfilter'] = export['R%s'%exp].value
                                            exportdict['iptnoutputfiltername'] = export['S%s'%exp].value
                                            exportdict['iptnoutputfilter'] = export['T%s'%exp].value
                        if iptnvprnint:
                            creatervplsinterfaceworksheet(misum['rvpls-interface'],rvplssap,namingdict,iptnvprnint,exportdict,rvplssap[8],misum['filter-naming'])
                    elif 'Bridge-Domain' in rvplssap[12]:
                        noderow['exvprnintstartrow'], noderow['exvprnintendrow'] = findextractnodeindex(exindex, rvplssap[9], 'J', 'K', 'L')
                        noderow['exbdstartrow'], noderow['exbdendrow'] = findextractnodeindex(exindex, rvplssap[9], 'V', 'W', 'X')
                        noderow['exportstartrow'], noderow['exportendrow'] = findextractnodeindex(exindex, rvplssap[9], 'B', 'C', 'D')
                        for vl in range(noderow['exbdstartrow'], noderow['exbdendrow']+1):
                            exbridgecheck = exbridge['A%s' %vl].value+exbridge['C%s' %vl].value+exbridge['H%s' %vl].value+'.'+exbridge['I%s' %vl].value
                            print(iptnsapcheck)
                            if iptnsapcheck == exbridgecheck:
                                exbridgeirb = exbridge['F%s' %vl].value
                                exbridgeirbvrf = exbridge['G%s' %vl].value
                                if exbridgeirbvrf != None:
                                    exbridgecheck = rvplssap[9]+exbridgeirb+exbridgeirbvrf
                                for vrfl in range(noderow['exvprnintstartrow'], noderow['exvprnintendrow']+1):
                                    vprnintcheck = exvprnint['A%s'%vrfl].value+exvprnint['E%s'%vrfl].value+'.'+exvprnint['F%s'%vrfl].value+exvprnint['C%s'%vrfl].value
                                    if exbridgecheck == vprnintcheck:
                                        iptnvprnint = createvprnintdict(exvprnint, rvplssap[9], vrfl)
                                for exp in range(noderow['exportstartrow'], noderow['exportendrow']+1):
                                    if 'irb' == export['C%s'%exp].value:
                                        if exbridgeirb != None:
                                            if exbridgeirb.split('.')[-1] == export['M%s'%exp].value:
                                                exportdict['iptninputfiltername'] = export['Q%s'%exp].value
                                                exportdict['iptninputfilter'] = export['R%s'%exp].value
                                                exportdict['iptnoutputfiltername'] = export['S%s'%exp].value
                                                exportdict['iptnoutputfilter'] = export['T%s'%exp].value
                        if iptnvprnint:
                            creatervplsinterfaceworksheet(misum['rvpls-interface'],rvplssap,namingdict,iptnvprnint,exportdict,rvplssap[8],misum['filter-naming'])
                    elif 'VPLS' in rvplssap[12]:
                        noderow['exvprnintstartrow'], noderow['exvprnintendrow'] = findextractnodeindex(exindex, rvplssap[9], 'J', 'K', 'L')
                        noderow['exvplsstartrow'], noderow['exvplsendrow'] = findextractnodeindex(exindex, rvplssap[9], 'R', 'S', 'T')
                        noderow['exportstartrow'], noderow['exportendrow'] = findextractnodeindex(exindex, rvplssap[9], 'B', 'C', 'D')
                        for vl in range(noderow['exvplsstartrow'], noderow['exvplsendrow']+1):
                            exvplscheck = exvpls['A%s' %vl].value+exvpls['C%s' %vl].value+exvpls['I%s' %vl].value+'.'+exvpls['J%s' %vl].value
                            if iptnsapcheck == exvplscheck:
                                exvplsirb = exvpls['P%s' %vl].value
                                exvplsirbvrf = exvpls['O%s' %vl].value
                                if exvplsirbvrf != None:
                                    exvplscheck = rvplssap[9]+exvplsirb+exvplsirbvrf
                                for vrfl in range(noderow['exvprnintstartrow'], noderow['exvprnintendrow']+1):
                                    vprnintcheck = exvprnint['A%s'%vrfl].value+exvprnint['E%s'%vrfl].value+'.'+exvprnint['F%s'%vrfl].value+exvprnint['C%s'%vrfl].value
                                    if exvplscheck == vprnintcheck:
                                        iptnvprnint = createvprnintdict(exvprnint, rvplssap[9], vrfl)
                                for exp in range(noderow['exportstartrow'], noderow['exportendrow']+1):
                                    if exvplsirb.split('.')[0] == export['C%s'%exp].value:
                                        if exvplsirb.split('.')[-1] == export['M%s'%exp].value:
                                            exportdict['iptninputfiltername'] = export['Q%s'%exp].value
                                            exportdict['iptninputfilter'] = export['R%s'%exp].value
                                            exportdict['iptnoutputfiltername'] = export['S%s'%exp].value
                                            exportdict['iptnoutputfilter'] = export['T%s'%exp].value
                        if iptnvprnint:
                            creatervplsinterfaceworksheet(misum['rvpls-interface'],rvplssap,namingdict,iptnvprnint,exportdict,rvplssap[8],misum['filter-naming'])


        misum.save(migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)

def bdrvprnint():
    misum = openpyxl.load_workbook(filename=migratepath+t.strftime("%Y%m%d-%H%M")+'\\'+misumname)
    serviceidvprnintcheck = []
    if os.path.exists(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidvprnint.txt'):
        ff = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'serviceidvprnint.txt','r').readlines()
        fg = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'servicevprnintinfo.txt', 'r').readlines()
        # print('g : %s' % g)
        for a in range(len(ff)):
            ff[a] = ff[a].rstrip('\n')  # remove newline('\n') from end of line
            if ff[a] not in serviceidvprnintcheck:
                serviceidvprnintcheck.append(ff[a])
        #print(serviceidvprnintcheck)
        for serviceid in serviceidvprnintcheck:
            #print(serviceid)
            for vprnintline in range(len(fg)):
                fg[vprnintline] = fg[vprnintline].rstrip('\n')
                vprnint = fg[vprnintline].split('$%')
                if vprnint[3] == serviceid:
                    createvprnintworksheet(misum['vprn-interface'],vprnint)
        misum.save(migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)

def bdrvprnglobal():
    misum = openpyxl.load_workbook(filename=migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)
    vprnglobalcusdict = {}
    vprnglobal = []
    serviceidcheck = []
    if os.path.exists(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'allvprn.txt'):
        f = open(migratepath + '\\' + t.strftime("%Y%m%d-%H%M") + '\\' + 'cfg' + '\\' + 'allvprn.txt', "r").readlines()
        for v in range(len(f)):
            f[v] = f[v].rstrip('\n')
            vprnglobalid = f[v].split('$%')[3]
            vprnglobalcus = f[v].split('$%')[-1]
            vprnglobalcusdict.setdefault('%s' % vprnglobalid,[]).append(vprnglobalcus)
        #print(vprnglobalcusdict)
        for v in range(len(f)):
            f[v] = f[v].rstrip('\n')
            vg = f[v].split('$%')
            vg[-1] = Counter(vprnglobalcusdict['%s' % vg[3]]).most_common(1)[0][0]
            if vg not in vprnglobal:
                #print(vg)
                vprnglobal.append(vg)
            if vg[3] not in serviceidcheck:
                serviceidcheck.append(vg[3])
        #print(vprnglobal)
        #print(serviceidcheck)
        for serviceid in serviceidcheck:
            #print(serviceid)
            for vprn in vprnglobal:
                noderow = {}
                if vprn[3] == serviceid:
                    #print(vprn)
                    noderow['exvprnstartrow'], noderow['exvprnendrow'] = findextractnodeindex(exindex, vprn[4], 'F', 'G', 'H')
                    for exv in range(noderow['exvprnstartrow'],noderow['exvprnendrow'] +1):
                        exvprnglobal = createvprnglobaldict(exvprn, vprn[4], exv)
                        if vprn[4] == exvprnglobal['iptnnode'] and vprn[5] == exvprnglobal['iptnrd']:
                            #print(exvprnglobal)
                            createbdrvprnglobalworksheet(misum['global-vprn'],vprn,exvprnglobal,vprn[6],misum['policy-naming'])

        misum.save(migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)

def nokiasummary():
    misum = createnokiasummaryfile()
    dummyport = []
    nodelist = []
    nodeline = {}
    # open naming xls file
    nm = open(migratepath + inputpath + "latestnaming.txt", 'r')
    latestnaming = nm.readlines()[0]
    naws = openpyxl.load_workbook(filename=migratepath + inputpath + latestnaming)
    naming = naws['naming']
    namingrow = []
    for i in range(3, naming.max_row+1):
        namingrow.append(i)

    nodelist, nodeline = findnamingobject(nodelist,nodeline, naming, namingrow,'E')
    #print(nodelist, nodeline)
    summaryrow = misum['summary']['A1'].value
    for n in nodelist:
        namingdict = {}
        servicelist = []
        serviceline = {}
        servicelist, serviceline = findnamingobject(servicelist, serviceline, naming, nodeline[n], 'O')
        #print(servicelist, serviceline)
        if None in serviceline: # for physical port member of LAG and port that not in any service binging such as mgmt for j// sw (unit 40 or 41)
            for nonserviceportline in serviceline[None]:
                maintask(misum,summaryrow, namingdict,naming,n,nonserviceportline)
                summaryrow += 1
        for s in servicelist: # for port in service (SAP)
            if s != None: # if service not None (none mean port member of lag or port that not in any service
                for r in serviceline[s]:
                    maintask(misum,summaryrow, namingdict, naming, n, r)
                    summaryrow += 1

    #print(misum['l2service-related-pe']['I34'].value)
    createl2servicesdpworksheet(misum)
    misum['summary']['B1'] = t.strftime("%Y%m%d-%H%M")
    misum.save(migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)

    if os.path.exists(migratepath + inputpath + 'latestmigrationsummary.txt'):  # create text file for specific latest workbook
        os.remove(migratepath + inputpath + 'latestmigrationsummary.txt')
    f = open(migratepath + inputpath + "latestmigrationsummary.txt", "w")
    f.write(migratepath + t.strftime("%Y%m%d-%H%M") + '\\' + misumname)
    f.close()

def main():
    print('Start-time : %s' %t.strftime("%Y%m%d-%H%M"))
    if not os.path.exists(migratepath):
        os.mkdir(migratepath)
    if not os.path.exists(migratepath+inputpath):
        os.mkdir(migratepath+inputpath)
    if not os.path.exists(migratepath+t.strftime("%Y%m%d-%H%M")+'\\'):
        os.mkdir(migratepath+t.strftime("%Y%m%d-%H%M"))
    if not os.path.exists(migratepath+t.strftime("%Y%m%d-%H%M")+'\\'+'cfg'+'\\'):
        os.mkdir(migratepath+t.strftime("%Y%m%d-%H%M")+'\\'+'cfg'+'\\')
    nokiasummary()
    bdrepipesap()
    bdrvplssap()
    bdrrvplssap()
    bdrvprnint()
    bdrvprnglobal()
    checkpolicyprefixdup()
    print('Finished-time : %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
if __name__ == "__main__":
    main()