from threading import Thread
from netaddr import IPNetwork, IPAddress
from collections import Counter
import os
import openpyxl
import datetime
import time
import re

t = datetime.datetime.now()
wbname = 'service-extraction-%s.xlsx' % t.strftime("%Y%m%d-%H%M")
wbmap = 'nodemapping.xlsx'
cfgpath = '.\iptn-cfg\\'
extractpath = '.\extraction\\'
inputpath = '.\input\\'
attrpath = 'attr-cfg\\'
wsindex = 'index'
wsport = 'port-lag'
wsvprn = 'vprn'
wsvprnint = 'vprn-int'
wsl2vpn = 'l2vpn'
wsvpls = 'vpls'
wsbridge = 'bridge-domain'
wsvsw = 'virtual-switch'

def create_excel():
    # create workbook
    row = 3 # initial row for write data
    excel = openpyxl.Workbook()
    excel.remove(excel['Sheet'])
    #############################
    # create index worksheet
    #############################
    index = excel.create_sheet(wsindex)
    # print row
    for i in range(row,200,1):
        index['A%s' %i] = i
    # port row index
    index['B1'] = row
    index['C1'] = wsport
    index['B2'] = 'node'
    index['C2'] = 'start-row'
    index['D2'] = 'end-row'
    index['E2'] = '###'
    # vprn global row index
    index['F1'] = row
    index['G1'] = wsvprn
    index['F2'] = 'node'
    index['G2'] = 'start-row'
    index['H2'] = 'end-row'
    index['I2'] = '###'
    # vprn-interface row index
    index['J1'] = row
    index['K1'] = wsvprnint
    index['J2'] = 'node'
    index['K2'] = 'start-row'
    index['L2'] = 'end-row'
    index['M2'] = '###'
    # l2vpn row index
    index['N1'] = row
    index['O1'] = wsl2vpn
    index['N2'] = 'node'
    index['O2'] = 'start-row'
    index['P2'] = 'end-row'
    index['Q2'] = '###'
    # vpls row index
    index['R1'] = row
    index['S1'] = wsvpls
    index['R2'] = 'node'
    index['S2'] = 'start-row'
    index['T2'] = 'end-row'
    index['U2'] = '###'
    # bridge row index
    index['V1'] = row
    index['W1'] = wsbridge
    index['V2'] = 'node'
    index['W2'] = 'start-row'
    index['X2'] = 'end-row'
    index['Y2'] = '###'
    # virtual switch row index
    index['Z1'] = row
    index['AA1'] = wsvsw
    index['Z2'] = 'node'
    index['AA2'] = 'start-row'
    index['AB2'] = 'end-row'
    index['AC2'] = '###'
    index.auto_filter.ref = 'A2:AC2'
    index.freeze_panes = index['A3']
    index.column_dimensions['A'].width = 10.0
    index.column_dimensions['B'].width = 30.0
    index.column_dimensions['C'].width = 12.0
    index.column_dimensions['D'].width = 12.0
    index.column_dimensions['E'].width = 6.0
    index.column_dimensions['F'].width = 30.0
    index.column_dimensions['G'].width = 12.0
    index.column_dimensions['H'].width = 12.0
    index.column_dimensions['I'].width = 6.0
    index.column_dimensions['J'].width = 30.0
    index.column_dimensions['K'].width = 12.0
    index.column_dimensions['L'].width = 12.0
    index.column_dimensions['M'].width = 6.0
    index.column_dimensions['N'].width = 30.0
    index.column_dimensions['O'].width = 12.0
    index.column_dimensions['P'].width = 12.0
    index.column_dimensions['Q'].width = 6.0
    index.column_dimensions['R'].width = 30.0
    index.column_dimensions['S'].width = 12.0
    index.column_dimensions['T'].width = 12.0
    index.column_dimensions['U'].width = 6.0
    index.column_dimensions['V'].width = 30.0
    index.column_dimensions['W'].width = 12.0
    index.column_dimensions['X'].width = 12.0
    index.column_dimensions['Y'].width = 6.0
    index.column_dimensions['Z'].width = 30.0
    index.column_dimensions['AA'].width = 12.0
    index.column_dimensions['AB'].width = 12.0
    index.column_dimensions['AC'].width = 6.0

    #############################
    # create port-lag worksheet
    #############################
    port = excel.create_sheet(wsport)
    port['A1'] = row
    port['A2'] = 'node'
    port['B2'] = 'system-ip'
    port['C2'] = 'physical-port'
    port['D2'] = 'admin-state'
    port['E2'] = 'port-encapsulation'
    port['F2'] = 'physical-description'
    port['G2'] = 'LAG'
    port['H2'] = 'LAG-protocol'
    port['I2'] = 'LAG-minimum-link'
    port['J2'] = 'speed'
    port['K2'] = 'mtu'
    port['L2'] = 'auto-negotiation'
    port['M2'] = 'unit'
    port['N2'] = 'unit-description'
    port['O2'] = 'unit-state'
    port['P2'] = 'vlan'
    port['Q2'] = 'input-filter-policy-name'
    port['R2'] = 'input-filter-configuration'
    port['S2'] = 'output-filter-policy-name'
    port['T2'] = 'output-filter-configuration'
    port['U2'] = 'forwarding-class'
    port['V2'] = 'service'
    port['W2'] = 'service-type'
    port['X2'] = 'nokia-node'
    port['Y2'] = 'nokia-system-ip'
    port['Z2'] = 'nokia-sap'
    port.auto_filter.ref = 'A2:Z2'
    port.freeze_panes = port['A3']
    port.column_dimensions['A'].width = 28.0
    port.column_dimensions['B'].width = 14.0
    port.column_dimensions['C'].width = 15.0
    port.column_dimensions['D'].width = 14.0
    port.column_dimensions['E'].width = 20.0
    port.column_dimensions['F'].width = 70.0
    port.column_dimensions['G'].width = 8.0
    port.column_dimensions['H'].width = 15.0
    port.column_dimensions['I'].width = 20.0
    port.column_dimensions['J'].width = 8.0
    port.column_dimensions['K'].width = 7.0
    port.column_dimensions['L'].width = 19.0
    port.column_dimensions['M'].width = 7.0
    port.column_dimensions['N'].width = 70.0
    port.column_dimensions['O'].width = 12.0
    port.column_dimensions['P'].width = 7.0
    port.column_dimensions['Q'].width = 50.0
    port.column_dimensions['R'].width = 70.0
    port.column_dimensions['S'].width = 28.0
    port.column_dimensions['T'].width = 70.0
    port.column_dimensions['U'].width = 19.0
    port.column_dimensions['V'].width = 65.0
    port.column_dimensions['W'].width = 14.0
    port.column_dimensions['X'].width = 20.0
    port.column_dimensions['Y'].width = 17.0
    port.column_dimensions['Z'].width = 12.0
    #############################
    # create vprn-global worksheet
    #############################
    vprn = excel.create_sheet(wsvprn)
    vprn['A1'] = row
    vprn['A2'] = 'node'
    vprn['B2'] = 'system-ip'
    vprn['C2'] = 'vpn-name'
    vprn['D2'] = 'admin-state'
    vprn['E2'] = 'rd'
    vprn['F2'] = 'rt-import-policy-name'
    vprn['G2'] = 'rt-import-policy'
    vprn['H2'] = 'rt-import-prefix'
    vprn['I2'] = 'rt-import-commu'
    vprn['J2'] = 'rt-export-policy-name'
    vprn['K2'] = 'rt-export-policy'
    vprn['L2'] = 'rt-export-prefix'
    vprn['M2'] = 'rt-export-commu'
    vprn['N2'] = 'static-route'
    vprn['O2'] = 'aggregate-route'
    vprn['P2'] = 'bgp'
    vprn['Q2'] = 'bgp-imp-policy'
    vprn['R2'] = 'bgp-imp-policy-prefix'
    vprn['S2'] = 'bgp-imp-policy-commu'
    vprn['T2'] = 'bgp-exp-policy'
    vprn['U2'] = 'bgp-exp-policy-prefix'
    vprn['V2'] = 'bgp-exp-policy-commu'
    vprn.auto_filter.ref = 'A2:V2'
    vprn.freeze_panes = vprn['A3']
    vprn.column_dimensions['A'].width = 28.0
    vprn.column_dimensions['B'].width = 14.0
    vprn.column_dimensions['C'].width = 19.0
    vprn.column_dimensions['D'].width = 14.0
    vprn.column_dimensions['E'].width = 11.0
    vprn.column_dimensions['F'].width = 27.0
    vprn.column_dimensions['G'].width = 70.0
    vprn.column_dimensions['H'].width = 17.0
    vprn.column_dimensions['I'].width = 19.0
    vprn.column_dimensions['J'].width = 26.0
    vprn.column_dimensions['K'].width = 70.0
    vprn.column_dimensions['L'].width = 17.0
    vprn.column_dimensions['M'].width = 18.0
    vprn.column_dimensions['N'].width = 70.0
    vprn.column_dimensions['O'].width = 70.0
    vprn.column_dimensions['P'].width = 70.0
    vprn.column_dimensions['Q'].width = 70.0
    vprn.column_dimensions['R'].width = 70.0
    vprn.column_dimensions['S'].width = 24.0
    vprn.column_dimensions['T'].width = 70.0
    vprn.column_dimensions['U'].width = 70.0
    vprn.column_dimensions['V'].width = 24.0
    #############################
    # create vprn-interface worksheet
    #############################
    vprnint = excel.create_sheet(wsvprnint)
    vprnint['A1'] = row
    vprnint['A2'] = 'node'
    vprnint['B2'] = 'system-ip'
    vprnint['C2'] = 'vpn-name'
    vprnint['D2'] = 'rd'
    vprnint['E2'] = 'layer 3 interface'
    vprnint['F2'] = 'unit'
    vprnint['G2'] = 'unit-state'
    vprnint['H2'] = 'vlan'
    vprnint['I2'] = 'interface description'
    vprnint['J2'] = 'forwarding-class'
    vprnint['K2'] = 'ip/ipv6 address'
    vprnint['L2'] = 'secondary ip/ipv6'
    vprnint['M2'] = 'family'
    vprnint['N2'] = 'vrrp id'
    vprnint['O2'] = 'vrrp vip'
    vprnint['P2'] = 'vrrp priority'
    vprnint['Q2'] = 'vrrp interval'
    vprnint['R2'] = 'static-route'
    vprnint['S2'] = 'bgp'
    vprnint['T2'] = 'group'
    vprnint['U2'] = 'peer-as'
    vprnint['V2'] = 'neighbor'
    vprnint['W2'] = 'neighbor-description'
    vprnint['X2'] = 'bgp-imp-policy-name'
    vprnint['Y2'] = 'bgp-imp-policy'
    vprnint['Z2'] = 'bgp-imp-policy-prefix'
    vprnint['AA2'] = 'bgp-imp-policy-commu'
    vprnint['AB2'] = 'bgp-exp-policy-name'
    vprnint['AC2'] = 'bgp-exp-policy'
    vprnint['AD2'] = 'bgp-exp-policy-prefix'
    vprnint['AE2'] = 'bgp-exp-policy-commu'
    vprnint['AF2'] = 'bgp-bfd-interval'
    vprnint['AG2'] = 'bgp-bfd-multiply'
    vprnint['AH2'] = 'nokia-node'
    vprnint['AI2'] = 'nokia-system-ip'
    vprnint['AJ2'] = 'nokia-sap'
    vprnint.auto_filter.ref = 'A2:AJ2'
    vprnint.freeze_panes = vprnint['A3']
    vprnint.column_dimensions['A'].width = 28.0
    vprnint.column_dimensions['B'].width = 14.0
    vprnint.column_dimensions['C'].width = 19.0
    vprnint.column_dimensions['D'].width = 11.0
    vprnint.column_dimensions['E'].width = 17.0
    vprnint.column_dimensions['F'].width = 6.0
    vprnint.column_dimensions['G'].width = 12.0
    vprnint.column_dimensions['H'].width = 7.0
    vprnint.column_dimensions['I'].width = 70.0
    vprnint.column_dimensions['J'].width = 18.0
    vprnint.column_dimensions['K'].width = 24.0
    vprnint.column_dimensions['L'].width = 19.0
    vprnint.column_dimensions['M'].width = 9.0
    vprnint.column_dimensions['N'].width = 9.0
    vprnint.column_dimensions['O'].width = 21.0
    vprnint.column_dimensions['P'].width = 14.0
    vprnint.column_dimensions['Q'].width = 14.0
    vprnint.column_dimensions['R'].width = 70.0
    vprnint.column_dimensions['S'].width = 70.0
    vprnint.column_dimensions['T'].width = 36.0
    vprnint.column_dimensions['U'].width = 11.0
    vprnint.column_dimensions['V'].width = 21.0
    vprnint.column_dimensions['W'].width = 56.0
    vprnint.column_dimensions['X'].width = 48.0
    vprnint.column_dimensions['Y'].width = 70.0
    vprnint.column_dimensions['Z'].width = 70.0
    vprnint.column_dimensions['AA'].width = 24.0
    vprnint.column_dimensions['AB'].width = 48.0
    vprnint.column_dimensions['AC'].width = 70.0
    vprnint.column_dimensions['AD'].width = 70.0
    vprnint.column_dimensions['AE'].width = 24.0
    vprnint.column_dimensions['AF'].width = 18.0
    vprnint.column_dimensions['AG'].width = 18.0
    vprnint.column_dimensions['AH'].width = 29.0
    vprnint.column_dimensions['AI'].width = 17.0
    vprnint.column_dimensions['AJ'].width = 12.0
    #############################
    # create l2vpn worksheet
    #############################
    l2vpn = excel.create_sheet(wsl2vpn)
    l2vpn['A1'] = row
    l2vpn['A2'] = 'node'
    l2vpn['B2'] = 'system-ip'
    l2vpn['C2'] = 'l2vpn-name'
    l2vpn['D2'] = 'rd'
    l2vpn['E2'] = 'vrf-target'
    l2vpn['F2'] = 'site'
    l2vpn['G2'] = 'site-id'
    l2vpn['H2'] = 'remote-site-id'
    l2vpn['I2'] = 'site-preference'
    l2vpn['J2'] = 'interface'
    l2vpn['K2'] = 'unit'
    l2vpn['L2'] = 'unit-state'
    l2vpn['M2'] = 'unit-description'
    l2vpn['N2'] = 'vlan'
    l2vpn['O2'] = 'input-vlan-map'
    l2vpn['P2'] = 'input-map-vlan-id'
    l2vpn['Q2'] = 'output-vlan-map'
    l2vpn['R2'] = 'output-map-vlan-id'
    l2vpn['S2'] = 'policer-input'
    l2vpn['T2'] = 'policer-output'
    l2vpn['U2'] = 'forwarding-class'
    l2vpn['V2'] = 'nokia-node'
    l2vpn['W2'] = 'nokia-system-ip'
    l2vpn['X2'] = 'nokia-sap'
    l2vpn.auto_filter.ref = 'A2:X2'
    l2vpn.freeze_panes = l2vpn['A3']
    l2vpn.column_dimensions['A'].width = 28.0
    l2vpn.column_dimensions['B'].width = 14.0
    l2vpn.column_dimensions['C'].width = 70.0
    l2vpn.column_dimensions['D'].width = 11.0
    l2vpn.column_dimensions['E'].width = 17.0
    l2vpn.column_dimensions['F'].width = 20.0
    l2vpn.column_dimensions['G'].width = 9.0
    l2vpn.column_dimensions['H'].width = 16.0
    l2vpn.column_dimensions['I'].width = 17.0
    l2vpn.column_dimensions['J'].width = 11.0
    l2vpn.column_dimensions['K'].width = 7.0
    l2vpn.column_dimensions['L'].width = 12.0
    l2vpn.column_dimensions['M'].width = 70.0
    l2vpn.column_dimensions['N'].width = 7.0
    l2vpn.column_dimensions['O'].width = 17.0
    l2vpn.column_dimensions['P'].width = 19.0
    l2vpn.column_dimensions['Q'].width = 18.0
    l2vpn.column_dimensions['R'].width = 20.0
    l2vpn.column_dimensions['S'].width = 20.0
    l2vpn.column_dimensions['T'].width = 20.0
    l2vpn.column_dimensions['U'].width = 19.0
    l2vpn.column_dimensions['V'].width = 20.0
    l2vpn.column_dimensions['W'].width = 17.0
    l2vpn.column_dimensions['X'].width = 12.0
    #############################
    # create vpls worksheet
    #############################
    vpls = excel.create_sheet(wsvpls)
    vpls['A1'] = row
    vpls['A2'] = 'node'
    vpls['B2'] = 'system-ip'
    vpls['C2'] = 'vpls-name'
    vpls['D2'] = 'rd'
    vpls['E2'] = 'vrf-target'
    vpls['F2'] = 'site'
    vpls['G2'] = 'site-id'
    vpls['H2'] = 'site-preference'
    vpls['I2'] = 'interface'
    vpls['J2'] = 'unit'
    vpls['K2'] = 'unit-state'
    vpls['L2'] = 'unit-description'
    vpls['M2'] = 'vlan'
    vpls['N2'] = 'physical-description'
    vpls['O2'] = 'L3-vrf'
    vpls['P2'] = 'routing-interface'
    vpls['Q2'] = 'L3-int-description'
    vpls['R2'] = 'policer-input'
    vpls['S2'] = 'policer-output'
    vpls['T2'] = 'forwarding-class'
    vpls['U2'] = 'nokia-node'
    vpls['V2'] = 'nokia-system-ip'
    vpls['W2'] = 'nokia-sap'
    vpls.auto_filter.ref = 'A2:W2'
    vpls.freeze_panes = vpls['A3']
    vpls.column_dimensions['A'].width = 28.0
    vpls.column_dimensions['B'].width = 14.0
    vpls.column_dimensions['C'].width = 38.0
    vpls.column_dimensions['D'].width = 10.0
    vpls.column_dimensions['E'].width = 16.0
    vpls.column_dimensions['F'].width = 15.0
    vpls.column_dimensions['G'].width = 9.0
    vpls.column_dimensions['H'].width = 17.0
    vpls.column_dimensions['I'].width = 11.0
    vpls.column_dimensions['J'].width = 7.0
    vpls.column_dimensions['K'].width = 12.0
    vpls.column_dimensions['L'].width = 41.0
    vpls.column_dimensions['M'].width = 7.0
    vpls.column_dimensions['N'].width = 60.0
    vpls.column_dimensions['O'].width = 30.0
    vpls.column_dimensions['P'].width = 18.0
    vpls.column_dimensions['Q'].width = 60.0
    vpls.column_dimensions['R'].width = 15.0
    vpls.column_dimensions['S'].width = 16.0
    vpls.column_dimensions['T'].width = 18.0
    vpls.column_dimensions['U'].width = 20.0
    vpls.column_dimensions['V'].width = 17.0
    vpls.column_dimensions['W'].width = 12.0
    #############################
    # create bridge-domain worksheet
    #############################
    bd = excel.create_sheet(wsbridge)
    bd['A1'] = row
    bd['A2'] = 'node'
    bd['B2'] = 'system-ip'
    bd['C2'] = 'bridge-domain'
    bd['D2'] = 'bd-vlan'
    bd['E2'] = 'filter'
    bd['F2'] = 'routing-interface'
    bd['G2'] = 'irb-vprn'
    bd['H2'] = 'interface'
    bd['I2'] = 'unit'
    bd['J2'] = 'unit-state'
    bd['K2'] = 'unit-description'
    bd['L2'] = 'int-vlan'
    bd['M2'] = 'forwarding-class'
    bd['N2'] = 'nokia-node'
    bd['O2'] = 'nokia-system-ip'
    bd['P2'] = 'nokia-sap'
    bd.auto_filter.ref = 'A2:P2'
    bd.freeze_panes = bd['A3']
    bd.column_dimensions['A'].width = 28.0
    bd.column_dimensions['B'].width = 14.0
    bd.column_dimensions['C'].width = 59.0
    bd.column_dimensions['D'].width = 10.0
    bd.column_dimensions['E'].width = 13.0
    bd.column_dimensions['F'].width = 18.0
    bd.column_dimensions['G'].width = 14.0
    bd.column_dimensions['H'].width = 11.0
    bd.column_dimensions['I'].width = 7.0
    bd.column_dimensions['J'].width = 12.0
    bd.column_dimensions['K'].width = 59.0
    bd.column_dimensions['L'].width = 10.0
    bd.column_dimensions['M'].width = 18.0
    bd.column_dimensions['N'].width = 20.0
    bd.column_dimensions['O'].width = 17.0
    bd.column_dimensions['P'].width = 12.0
    #############################
    # create virtual-sw worksheet
    #############################
    vsw = excel.create_sheet(wsvsw)
    vsw['A1'] = row
    vsw['A2'] = 'node'
    vsw['B2'] = 'system-ip'
    vsw['C2'] = 'virtual-switch-name'
    vsw['D2'] = 'rd'
    vsw['E2'] = 'vrf-target'
    vsw['F2'] = 'site'
    vsw['G2'] = 'site-id'
    vsw['H2'] = 'interface'
    vsw['I2'] = 'bridge-domain'
    vsw['J2'] = 'bridge-vlan'
    vsw['K2'] = 'interface'
    vsw['L2'] = 'unit'
    vsw['M2'] = 'unit-state'
    vsw['N2'] = 'unit-mode'
    vsw['O2'] = 'unit-vlan-id'
    vsw['P2'] = 'forwarding-class'
    vsw['Q2'] = 'routing-interface'
    vsw['R2'] = 'irb-vprn'
    vsw['S2'] = 'filter'
    vsw['T2'] = 'nokia-node'
    vsw['U2'] = 'nokia-system-ip'
    vsw['V2'] = 'nokia-sap'
    vsw.auto_filter.ref = 'A2:V2'
    vsw.freeze_panes = vsw['A3']
    vsw.column_dimensions['A'].width = 28.0
    vsw.column_dimensions['B'].width = 14.0
    vsw.column_dimensions['C'].width = 30.0
    vsw.column_dimensions['D'].width = 10.0
    vsw.column_dimensions['E'].width = 16.0
    vsw.column_dimensions['F'].width = 6.0
    vsw.column_dimensions['G'].width = 9.0
    vsw.column_dimensions['H'].width = 11.0
    vsw.column_dimensions['I'].width = 41.0
    vsw.column_dimensions['J'].width = 12.0
    vsw.column_dimensions['K'].width = 11.0
    vsw.column_dimensions['L'].width = 7.0
    vsw.column_dimensions['M'].width = 12.0
    vsw.column_dimensions['N'].width = 18.0
    vsw.column_dimensions['O'].width = 7.0
    vsw.column_dimensions['P'].width = 18.0
    vsw.column_dimensions['Q'].width = 18.0
    vsw.column_dimensions['R'].width = 11.0
    vsw.column_dimensions['S'].width = 18.0
    vsw.column_dimensions['T'].width = 20.0
    vsw.column_dimensions['U'].width = 17.0
    vsw.column_dimensions['V'].width = 12.0
    #############################
    excel.save(extractpath+wbname)
    if os.path.exists(extractpath+'latest.txt'):  # create text file for specific latest workbook
        os.remove(extractpath+'latest.txt')
    f = open(extractpath+"latest.txt", "w")
    f.write(wbname)
    f.close()

def checkline(cfglines):
    node = ''
    system = ''
    portname = []
    portlines = []
    vprnname = []
    vprnline = []
    l2vpnname = []
    l2vpnline = []
    vplsname = []
    vplsline = []
    bridgename = []
    bridgeline = []
    vswname = []
    vswline = []
    prefixname = []
    prefixline = []
    communame = []
    commuline = []
    policyname = []
    policyline = []
    aclname = []
    aclline = []
    qosline = []
    serviceline = []
    #############################
    # Get hostname and system IP
    #############################
    for b in range(len(cfglines)):
        if re.match('set groups re0 system host-name',cfglines[b]):
            node = cfglines[b].rstrip('_re0').split(' ')[-1]
            print(node)
        elif re.match('set interfaces lo0 unit 0 family inet address',cfglines[b]) and '127.0.0.1' not in cfglines[b]:
            system = cfglines[b].split(' ')[-1].split('/')[0]
            print(system)
        #############################
        # Get Port Index
        #############################
        if re.match('set interfaces',cfglines[b]) or re.match('deactivate interfaces',cfglines[b]):
            portlines.append(b)
            p = cfglines[b].split(' ')[2]
            if p not in portname:
                portname.append(p)
        #############################
        # Get Policy Index
        #############################
        elif re.match('set policy-options',cfglines[b]) or re.match('deactivate policy-options',cfglines[b]):
            if re.match('set policy-options policy-statement',cfglines[b]) or re.match('deactivate policy-options policy-statement',cfglines[b]):
                policyline.append(b)
                policy = cfglines[b].split(' ')[3]
                if policy not in policyname:
                    policyname.append(policy)
            elif re.match('set policy-options prefix-list',cfglines[b]) or re.match('deactivate policy-options prefix-list',cfglines[b]):
                prefixline.append(b)
                prefix = cfglines[b].split(' ')[3]
                if prefix not in prefixname:
                    prefixname.append(prefix)
            elif re.match('set policy-options community',cfglines[b]) or re.match('deactivate policy-options community',cfglines[b]):
                commuline.append(b)
                commu = cfglines[b].split(' ')[3]
                if commu not in communame:
                    communame.append(commu)
        #############################
        # Get ACL Index
        #############################
        elif re.match('set firewall filter',cfglines[b]) or re.match('deactivate firewall filter',cfglines[b]):
            aclline.append(b)
            acl = cfglines[b].split(' ')[3]
            if acl not in aclname:
                aclname.append(acl)
        elif re.match('set firewall family inet6 filter',cfglines[b]) or re.match('deactivate firewall family inet6 filter',cfglines[b]):
            aclline.append(b)
            acl = cfglines[b].split(' ')[5]
            if acl not in aclname:
                aclname.append(acl)
        #############################
        # Get Service Type and Name
        #############################
        elif re.match('set routing-instances',cfglines[b]) or re.match('deactivate routing-instances',cfglines[b]):
            serviceline.append(b)
            if re.search('instance-type vrf',cfglines[b]):
                vprnname.append(cfglines[b].split(' ')[2])
            elif re.search('instance-type l2vpn',cfglines[b]):
                l2vpnname.append(cfglines[b].split(' ')[2])
            elif re.search('instance-type vpls',cfglines[b]):
                vplsname.append(cfglines[b].split(' ')[2])
            elif re.search('instance-type virtual-switch',cfglines[b]):
                vswname.append(cfglines[b].split(' ')[2])
        #############################
        # Get Bridge Domain Information
        #############################
        elif re.match('set bridge-domains',cfglines[b]) or re.match('deactivate bridge-domains',cfglines[b]):
            bridgeline.append(b)
            if re.search('domain-type bridge',cfglines[b]):
                bridgename.append(cfglines[b].split(' ')[2])
        #############################
        # Get QOS Line Index
        #############################
        elif re.match('set class-of-service',cfglines[b]) or re.match('deactivate class-of-service',cfglines[b]):
            qosline.append(b)

    #############################
    # Get Each Service Type Line
    #############################
    for c in serviceline:
        for d in vprnname:
            if d == cfglines[c].split(' ')[2]:
                vprnline.append(c)
        for e in l2vpnname:
            if e == cfglines[c].split(' ')[2]:
                l2vpnline.append(c)
        for f in vplsname:
            if f == cfglines[c].split(' ')[2]:
                vplsline.append(c)
        for g in vswname:
            if g == cfglines[c].split(' ')[2]:
                vswline.append(c)


    #print('port = %s' % portname)
    #print('vprn = %s' % vprnname)
    #print('vprnline = %s' %vprnline)
    #print('l2vpn = %s' % l2vpnname)
    #print('vpls = %s' % vplsname)
    #print('vsw = %s' % vswname)
    #print('bridge = %s' % bridgename)
    #print('prefix = %s' % prefixname)
    #print('commu = %s' % communame)
    #print('policy = %s' % policyname)
    #print('acl = %s' % aclname)
    f = open(extractpath + attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node +'_lines.txt', "a")
    f.write('portname:' + str(portname) + '\n')
    f.write('portlines:' + str(portlines) + '\n')
    f.write('vprnname:' + str(vprnname) + '\n')
    f.write('vprnline:' + str(vprnline) + '\n')
    f.write('l2vpnname:' + str(l2vpnname) + '\n')
    f.write('l2vpnline:' + str(l2vpnline) + '\n')
    f.write('vplsname:' + str(vplsname) + '\n')
    f.write('vplsline:' + str(vplsline) + '\n')
    f.write('bridgename:' + str(bridgename) + '\n')
    f.write('bridgeline:' + str(bridgeline) + '\n')
    f.write('vswname:' + str(vswname) + '\n')
    f.write('vswline:' + str(vswline) + '\n')
    f.write('prefixname:' + str(prefixname) + '\n')
    f.write('prefixline:' + str(prefixline) + '\n')
    f.write('communame:' + str(communame) + '\n')
    f.write('commuline:' + str(commuline) + '\n')
    f.write('policyname:' + str(policyname) + '\n')
    f.write('policyline:' + str(policyline) + '\n')
    f.write('aclname:' + str(aclname) + '\n')
    f.write('aclline:' + str(aclline) + '\n')
    f.write('qosline:' + str(qosline) + '\n')
    f.close()
    return node, system, portname, portlines, vprnname, vprnline, l2vpnname, \
           l2vpnline, vplsname, vplsline, bridgename, bridgeline, vswname, vswline, \
           prefixname, prefixline, communame, commuline, policyname, policyline, aclname, aclline, qosline


def writeport(port, row, node, system, physicalport, adminstate, portencap, physicaldesc, LAG, LAGprotocol, LAGminlink, speed, mtu, autonego, subint):
    if len(subint['vlan'].split(', ')) >= 2:
        for w in subint['vlan'].split(', '):
            subint['vlan'] = w
            row = writeport(port, row, node, system, physicalport, adminstate, portencap, physicaldesc, LAG, LAGprotocol,LAGminlink, speed, mtu, autonego, subint)
    else:
        port['A%s' % row] = node
        port['B%s' % row] = system
        port['C%s' % row] = physicalport
        port['D%s' % row] = adminstate
        port['E%s' % row] = portencap
        port['F%s' % row] = physicaldesc
        port['G%s' % row] = LAG
        port['H%s' % row] = LAGprotocol
        port['I%s' % row] = LAGminlink
        port['J%s' % row] = speed
        port['K%s' % row] = mtu
        port['L%s' % row] = autonego
        if portencap == 'trunk':
            port['M%s' % row] = subint['vlan']
        else:
            port['M%s' % row] = subint['unit']
        port['N%s' % row] = subint['unitdesc']
        port['O%s' % row] = subint['unitstate']
        port['P%s' % row] = subint['vlan']
        if re.search('apply-acl', subint['input-filter-policy-name']):
            port['Q%s' % row].hyperlink = subint['input-filter-policy-name']
        else:
            port['Q%s' % row] = subint['input-filter-policy-name']
        port['R%s' % row].hyperlink = subint['input-filter-configuration']
        port['S%s' % row] = subint['output-filter-policy-name']
        port['T%s' % row].hyperlink = subint['output-filter-configuration']
        port['U%s' % row] = subint['forwardingclass']
        port['V%s' % row] = subint['service']
        port['W%s' % row] = subint['servicetype']
        port['X%s' % row] = subint['nokianode']
        port['Y%s' % row] = subint['nokiasystemip']
        port['Z%s' % row] = subint['nokiasap']
        row += 1

    return row


def port(excel, cfglines, node, system,portname, portlines, vprnline, l2vpnline, vplsline, bridgeline, vswline, aclline, qosline):
    port = excel[wsport]
    row = port['A1'].value
    start_row = row
    lastport = ''
    filterinputlist = []
    portname.append('lastport')
    mapsheet = openpyxl.load_workbook(inputpath + wbmap)['node-map']
    vswvlan = []
    for x in range(3, (mapsheet.max_row + 1)):
        if re.match(node, mapsheet['A%s' % x].value):
            nokianode = mapsheet['C%s' % x].value
            nokiasystemip = mapsheet['D%s' % x].value
            # print('### nokia node : %s' % subint['nokianode'])
            # print('### nokia system ip : %s' % subint['nokiasystemip']

    #p = 'ae17'
    #if p:
    for p in portname:
        if p != lastport and lastport:
            row = writeport(port, row, node, system, physicalport, adminstate, portencap, physicaldesc, LAG, LAGprotocol, LAGminlink, speed, mtu, autonego, subint)
            #row += 1

        physicalport    = p
        adminstate      = ''
        portencap       = ''
        physicaldesc    = ''
        LAG             = ''
        LAGprotocol     = ''
        LAGminlink      = ''
        speed           = ''
        mtu             = ''
        autonego        = ''
        lastsubint      = ''
        subint          = {
            'unit' : '' ,
            'unitdesc' : '' ,
            'unitstate' : '' ,
            'vlan' : '' ,
            'input-filter-policy-name' : '' ,
            'input-filter-configuration' : '' ,
            'output-filter-policy-name' : '' ,
            'output-filter-configuration' : '' ,
            'forwardingclass' : '' ,
            'service' : '' ,
            'servicetype' : '' ,
            'nokianode' : nokianode ,
            'nokiasystemip' : nokiasystemip ,
            'nokiasap' : ''
        }
        lastport = physicalport
        vswvlan = []

        for l in portlines:
            if p == cfglines[l].split(' ')[2]:

                if 'ae' in p:
                    LAG = p

                if (re.search('deactivate interfaces %s' % p, cfglines[l]) and 'unit' not in cfglines[l]):
                    adminstate = 'deactivate'
                    #print('### deactivate : %s' %cfglines[l])
                    #print(adminstate)
                elif re.search('interfaces %s disable' % p, cfglines[l]):
                    adminstate = 'disable'
                    #print('### disable : %s' % cfglines[l])
                    #print(adminstate)
                elif re.search('interfaces %s vlan-tagging' % p, cfglines[l]):
                    portencap = 'vlan-tagging'
                    #print('### portencap : %s' % cfglines[l])
                    #print(portencap)
                elif re.search('interfaces %s unit 0 family bridge interface-mode trunk' % p, cfglines[l]):
                    portencap = 'trunk'
                    # print('### portencap : %s' % cfglines[l])
                    # print(portencap)
                elif re.search('interfaces %s unit 0 family bridge interface-mode access' % p, cfglines[l]):
                    portencap = 'access'
                    # print('### portencap : %s' % cfglines[l])
                    # print(portencap)
                elif re.search('interfaces %s encapsulation' % p, cfglines[l]):
                    portencap = cfglines[l].split('interfaces %s encapsulation ' % p)[1]
                    #print('### portencap : %s' % cfglines[l])
                    #print(portencap)
                elif re.search('interfaces %s description' % p, cfglines[l]):
                    physicaldesc = cfglines[l].split('interfaces %s description ' % p)[1]
                    #print('### desc : %s' % cfglines[l])
                    #print(physicaldesc)
                elif re.search('interfaces %s gigether-options 802.3ad' % p, cfglines[l]):
                    LAG = cfglines[l].split('interfaces %s gigether-options 802.3ad ' % p)[1]
                    #print('### LAG : %s' % cfglines[l])
                    #print(LAG)
                elif re.search('interfaces %s aggregated-ether-options lacp active' % p, cfglines[l]):
                    LAGprotocol = 'lacp'
                    #print('### LAGprotocol : %s' % cfglines[l])
                    #print(LAGprotocol)
                elif re.search('interfaces %s aggregated-ether-options minimum-links' % p, cfglines[l]):
                    LAGminlink = cfglines[l].split('interfaces %s aggregated-ether-options minimum-links ' % p)[1]
                    #print('### LAGminlink : %s' % cfglines[l])
                    #print(LAGminlink)
                elif re.search('interfaces %s speed' % p, cfglines[l]):
                    speed = cfglines[l].split('interfaces %s speed ' % p)[1]
                    #print('### speed : %s' % cfglines[l])
                    #print(speed)
                elif re.search('interfaces %s mtu' % p, cfglines[l]):
                    mtu = cfglines[l].split('interfaces %s mtu ' % p)[1]
                    #print('### mtu : %s' % cfglines[l])
                    #print(mtu)
                elif re.search('auto-negotiation', cfglines[l]):
                    autonego = cfglines[l].split(' ')[-1]
                    #print('### autonego : %s' % cfglines[l])
                    #print(autonego)
                elif re.search('interfaces %s unit' % p, cfglines[l]):
                    if cfglines[l].split(' ')[4] != lastsubint:
                        if lastsubint:
                            row = writeport(port, row, node, system, physicalport, adminstate, portencap, physicaldesc, LAG, LAGprotocol, LAGminlink, speed, mtu, autonego, subint)
                            #row += 1

                        subint['unit'] = cfglines[l].split(' ')[4]
                        #print('### subint-unit : %s' % cfglines[l])
                        #print(subint['unit'])
                        subint['unitdesc'] = ''
                        subint['unitstate'] = ''
                        subint['vlan'] = ''
                        subint['input-filter-policy-name'] = ''
                        subint['input-filter-configuration'] = ''
                        subint['output-filter-policy-name'] = ''
                        subint['output-filter-configuration'] = ''
                        subint['forwardingclass'] = ''
                        subint['service'] = ''
                        subint['servicetype'] = ''
                        subint['nokianode'] = nokianode
                        subint['nokiasystemip'] = nokiasystemip
                        subint['nokiasap'] = ''



                        for vprn in vprnline:
                            if re.search('routing-instances', cfglines[vprn]) and re.search('interface %s.%s' % (p, subint['unit']),cfglines[vprn]):
                                if '%s.%s' % (p, subint['unit']) == cfglines[vprn].split(' ')[-1]:
                                    subint['service'] = cfglines[vprn].split(' ')[2]
                                    subint['servicetype'] = 'VPRN'
                                    #print('### subint-service : %s' % cfglines[vprn])
                                    #print('%s.%s' % (p, subint['unit']))
                                    #print(subint['service'])
                        for l2vpn in  l2vpnline:
                            if re.search('routing-instances', cfglines[l2vpn]) and re.search('interface %s.%s' % (p, subint['unit']),cfglines[l2vpn]) and 'protocols l2vpn site' not in cfglines[l2vpn]:
                                if '%s.%s' % (p, subint['unit']) == cfglines[l2vpn].split(' ')[-1]:
                                    subint['service'] = cfglines[l2vpn].split(' ')[2]
                                    subint['servicetype'] = 'EPIPE'
                                    #print('### subint-service : %s' % cfglines[l2vpn])
                                    #print(subint['service'])
                        for vpls in vplsline:
                            if re.search('routing-instances', cfglines[vpls]) and re.search('interface %s.%s' % (p, subint['unit']),cfglines[vpls]) and 'protocols vpls site' not in cfglines[vpls]:
                                if '%s.%s' % (p, subint['unit']) == cfglines[vpls].split(' ')[-1]:
                                    subint['service'] = cfglines[vpls].split(' ')[2]
                                    subint['servicetype'] = 'VPLS'
                                    #print('### subint-service : %s' % cfglines[vpls])
                                    #print(subint['service'])
                        for bridge in bridgeline:
                            if re.search('bridge-domains', cfglines[bridge]) and re.search('interface %s.%s' % (p, subint['unit']),cfglines[bridge]):
                                if '%s.%s' % (p, subint['unit']) == cfglines[bridge].split(' ')[-1]:
                                    subint['service'] = cfglines[bridge].split(' ')[2]
                                    subint['servicetype'] = 'Bridge-Domain'
                                    #print('### subint-service : %s' % cfglines[bridge])
                                    #print(subint['service'])
                        for vsw in vswline:
                            if re.search('routing-instances', cfglines[vsw]) and re.search('interface %s.%s' % (p, subint['unit']),cfglines[vsw]):
                                if '%s.%s' % (p, subint['unit']) == cfglines[vsw].split(' ')[-1]:
                                    subint['service'] = cfglines[vsw].split(' ')[2]
                                    subint['servicetype'] = 'Virtual-Switch'
                                    #print('### subint-service : %s' % cfglines[vsw])
                                    #print(subint['service'])
                        for qos in qosline:
                            if re.search('class-of-service interfaces %s unit %s forwarding-class' % (p, subint['unit']),cfglines[qos]):
                                subint['forwardingclass'] = cfglines[qos].split('class-of-service interfaces %s unit %s forwarding-class ' % (p, subint['unit']))[-1]
                                #print('### subint-forwardingclass : %s' % cfglines[qos])
                                #print(subint['forwardingclass'])
                            elif re.search('class-of-service interfaces %s unit %s classifiers exp' % (p, subint['unit']),cfglines[qos]):
                                subint['forwardingclass'] = cfglines[qos].split('class-of-service interfaces %s unit %s classifiers exp ' % (p, subint['unit']))[-1]
                                #print('### subint-forwardingclass : %s' % cfglines[qos])
                                #print(subint['forwardingclass'])
                        lastsubint = subint['unit']

                    if re.search('interfaces %s unit %s description' % (p,subint['unit']), cfglines[l]):
                        subint['unitdesc'] = cfglines[l].split('interfaces %s unit %s description ' % (p,subint['unit']))[-1]
                        #print('### subint-unitdesc : %s' % cfglines[l])
                        #print(subint['unitdesc'])
                    elif re.search('deactivate interfaces %s unit %s' % (p,subint['unit']), cfglines[l]):
                        subint['unitstate'] = 'deactivate'
                        #print('### subint-unitstate : %s' % cfglines[l])
                        #print(subint['unitstate'])
                    elif re.search('interfaces %s unit %s disable' % (p,subint['unit']), cfglines[l]):
                        subint['unitstate'] = 'disable'
                        #print('### subint-unitstate : %s' % cfglines[l])
                        #print(subint['unitstate'])
                    elif re.search('interfaces %s unit %s vlan-id' % (p,subint['unit']), cfglines[l]):
                        subint['vlan'] = cfglines[l].split('interfaces %s unit %s vlan-id ' % (p,subint['unit']))[-1]
                        #print('### subint-vlan : %s' % cfglines[l])
                        #print(subint['vlan'])
                    elif re.search('interfaces %s unit %s family bridge vlan-id-list' % (p, subint['unit']), cfglines[l]):
                        vswvlan.append(cfglines[l].split('interfaces %s unit %s family bridge vlan-id-list ' % (p, subint['unit']))[-1])
                        subint['vlan'] = ', '.join(vswvlan)
                        # print('### subint-vlan : %s' % cfglines[l])
                        # print(subint['vlan'])
                    elif re.search('interfaces %s unit %s family inet filter input' % (p,subint['unit']), cfglines[l]) or re.search('interfaces %s unit %s family inet6 filter input' % (p,subint['unit']), cfglines[l]):
                        if not re.search('filter input-list',cfglines[l]):
                            subint['input-filter-policy-name'] = cfglines[l].split('filter input ')[-1]
                            po = '%s-%s' %(p.replace('/', '-'),subint['unit'])
                            #print('### subint-input-filter-policy-name : %s' % cfglines[l])
                            #print(subint['input-filter-policy-name'])
                            subint['input-filter-configuration'] = attrpath+t.strftime("%Y%m%d-%H%M")+'\\'+node +'_acl_'+ po + '_' +subint['input-filter-policy-name']+'.txt'
                            for acl in aclline:
                                if 'interface-specific' not in cfglines[acl]:
                                    if subint['input-filter-policy-name'] == cfglines[acl].split(' ')[3]:
                                        if not os.path.exists(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")):
                                            os.mkdir(extractpath+attrpath+t.strftime("%Y%m%d-%H%M"))
                                        f = open(extractpath+subint['input-filter-configuration'], "a")
                                        f.write(cfglines[acl]+'\n')
                                        f.close()
                                    elif subint['input-filter-policy-name'] == cfglines[acl].split(' ')[5]:
                                        if not os.path.exists(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")):
                                            os.mkdir(extractpath+attrpath+t.strftime("%Y%m%d-%H%M"))
                                        f = open(extractpath+subint['input-filter-configuration'], "a")
                                        f.write(cfglines[acl]+'\n')
                                        f.close()
                        else:
                            inputlist = cfglines[l].split('filter input-list ')[-1]
                            po = '%s-%s' %(p.replace('/', '-'),subint['unit'])
                            subint['input-filter-policy-name'] = attrpath+t.strftime("%Y%m%d-%H%M")+'\\'+node +'_apply-acl_'+ 'interfaces_%s_unit_%s_input-list' % (po,subint['unit']) +'.txt'
                            f = open(extractpath + subint['input-filter-policy-name'], "a")
                            f.write(cfglines[l] + '\n')
                            f.close()
                            subint['input-filter-configuration'] = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_acl_' + 'interfaces_%s_unit_%s_input-list' % (po,subint['unit']) + '.txt'
                            for acl in aclline:
                                if 'interface-specific' not in cfglines[acl]:
                                    if inputlist == cfglines[acl].split(' ')[3]:
                                        if not os.path.exists(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")):
                                            os.mkdir(extractpath+attrpath+t.strftime("%Y%m%d-%H%M"))
                                        f = open(extractpath + subint['input-filter-configuration'], "a")
                                        f.write(cfglines[acl] + '\n')
                                        f.close()
                                    elif inputlist == cfglines[acl].split(' ')[5]:
                                        if not os.path.exists(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")):
                                            os.mkdir(extractpath+attrpath+t.strftime("%Y%m%d-%H%M"))
                                        f = open(extractpath + subint['input-filter-configuration'], "a")
                                        f.write(cfglines[acl] + '\n')
                                        f.close()
                                    #print('### subint-input-filter-policy-name : %s' % cfglines[l])
                    elif re.search('interfaces %s unit %s family inet filter output' % (p,subint['unit']), cfglines[l]) or re.search('interfaces %s unit %s family inet6 filter output' % (p,subint['unit']), cfglines[l]):
                        subint['output-filter-policy-name'] = cfglines[l].split('filter output ')[-1]
                        po = '%s-%s' %(p.replace('/', '-'),subint['unit'])
                        #print('### subint-output-filter-policy-name : %s' % cfglines[l])
                        #print(subint['output-filter-policy-name'])
                        subint['output-filter-configuration'] = attrpath+t.strftime("%Y%m%d-%H%M")+'\\'+node +'_acl_'+ po + '_' + subint['output-filter-policy-name']+'.txt'
                        for acl in aclline:
                            if 'interface-specific' not in cfglines[acl]:
                                if subint['output-filter-policy-name'] == cfglines[acl].split(' ')[3]:
                                    if not os.path.exists(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")):
                                        os.mkdir(extractpath+attrpath+t.strftime("%Y%m%d-%H%M"))
                                    f = open(extractpath+subint['output-filter-configuration'], "a")
                                    f.write(cfglines[acl]+'\n')
                                    f.close()
                                elif subint['output-filter-policy-name'] == cfglines[acl].split(' ')[5]:
                                    if not os.path.exists(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")):
                                        os.mkdir(extractpath+attrpath+t.strftime("%Y%m%d-%H%M"))
                                    f = open(extractpath+subint['output-filter-configuration'], "a")
                                    f.write(cfglines[acl]+'\n')
                                    f.close()


    port['A1'] = row
    print('port:finish %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
    if row > start_row:
        index_row = excel['index']['B1'].value
        excel['index']['B%s' % index_row] = node        # update hostname index
        excel['index']['C%s' % index_row] = start_row   # update start row index
        excel['index']['D%s' % index_row] = row - 1     # update end row index
        excel['index']['E%s' % index_row] = '###'
        excel['index']['B1'] = index_row + 1            # update row of index


def vprn(excel, cfglines, node, system, portlines, vprnname, vprnline, prefixline, commuline, policyline):
    vprn = excel[wsvprn]
    row = vprn['A1'].value
    start_row = row
    lastvpn = ''
    vprnname.append('lastvpn')

    for vrf in vprnname:
        if vrf != lastvpn and lastvpn:
            vprn['A%s' % row] = node
            vprn['B%s' % row] = system
            vprn['C%s' % row] = vpnname
            vprn['D%s' % row] = adminstate
            vprn['E%s' % row] = rd
            vprn['F%s' % row] = rtimportpolicyname
            vprn['G%s' % row].hyperlink = rtimportpolicy
            vprn['H%s' % row].hyperlink = rtimportprefix
            vprn['I%s' % row] = rtimportcommu
            vprn['J%s' % row] = rtexportpolicyname
            vprn['K%s' % row].hyperlink = rtexportpolicy
            vprn['L%s' % row].hyperlink = rtexportprefix
            vprn['M%s' % row] = rtexportcommu
            vprn['N%s' % row].hyperlink = staticroute
            vprn['O%s' % row].hyperlink = aggregateroute
            vprn['P%s' % row].hyperlink = bgp
            vprn['Q%s' % row].hyperlink = impbgppolicy
            vprn['R%s' % row].hyperlink = impbgppolicyprefix
            vprn['S%s' % row] = impbgppolicycommu
            vprn['T%s' % row].hyperlink = expbgppolicy
            vprn['U%s' % row].hyperlink = expbgppolicyprefix
            vprn['V%s' % row] = expbgppolicycommu
            row += 1

        vpnname  = vrf
        adminstate  = ''
        rd  = ''
        rtimportpolicyname  = ''
        rtimportpolicy  = ''
        rtimportprefix  = ''
        rtimportcommu  = ''
        rtexportpolicyname  = ''
        rtexportpolicy  = ''
        rtexportprefix  = ''
        rtexportcommu  = ''
        staticroute  = ''
        aggregateroute = ''
        bgp  = ''
        impbgppolicy  = ''
        impbgppolicyprefix  = ''
        impbgppolicycommu  = ''
        expbgppolicy  = ''
        expbgppolicyprefix  = ''
        expbgppolicycommu  = ''
        lastvpn = vpnname

        for l in vprnline:
            if re.search('deactivate routing-instances %s' % vrf, cfglines[l]):
                adminstate = 'deactivate'
                #print('### deactivate : %s' %cfglines[l])
                #print(adminstate)
            elif re.search('routing-instances %s disable' % vrf, cfglines[l]):
                adminstate = 'disable'
                #print('### disable : %s' % cfglines[l])
                #print(adminstate)
            elif re.search('routing-instances %s route-distinguisher' % vrf, cfglines[l]):
                rd = cfglines[l].split('routing-instances %s route-distinguisher ' % vrf)[1]
                #print('### rd : %s' % cfglines[l])
                #print(rd)
            elif re.search('routing-instances %s vrf-import' % vrf, cfglines[l]):
                rtimportpolicyname = cfglines[l].split('routing-instances %s vrf-import ' % vrf)[1]
                #print('### rtimportpolicyname : %s' % cfglines[l])
                #print(rtimportpolicyname)
                rtimportpolicy = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_policy_' + rtimportpolicyname + '.txt'
                for p in policyline:
                    if rtimportpolicyname == cfglines[p].split(' ')[3]:
                        if not os.path.exists(extractpath + attrpath + t.strftime("%Y%m%d-%H%M")):
                            os.mkdir(extractpath + attrpath + t.strftime("%Y%m%d-%H%M"))
                        f = open(extractpath + rtimportpolicy, "a")
                        f.write(cfglines[p] + '\n')
                        f.close()
                        if re.search('prefix-list',cfglines[p]):
                            prefix = cfglines[p].split('prefix-list ')[-1]
                            rtimportprefix = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_prefix-list_' + prefix + '.txt'
                            for q in prefixline:
                                if prefix == cfglines[q].split(' ')[3]:
                                    f = open(extractpath + rtimportprefix, "a")
                                    f.write(cfglines[q] + '\n')
                                    f.close()
                        elif re.search('community', cfglines[p]):
                            commu = cfglines[p].split(' ')[-1]
                            for r in commuline:
                                if commu == cfglines[r].split(' ')[3]:
                                    rtimportcommu = cfglines[r].split(' ')[5]
                                    #print(rtimportcommu)
            elif re.search('routing-instances %s vrf-export' % vrf, cfglines[l]):
                rtexportpolicyname = cfglines[l].split('routing-instances %s vrf-export ' % vrf)[1]
                #print('### rtexportpolicyname : %s' % cfglines[l])
                #print(rtexportpolicyname)
                rtexportpolicy = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_policy_' + rtexportpolicyname + '.txt'
                rtexpcommulist = []
                for p in policyline:
                    if rtexportpolicyname == cfglines[p].split(' ')[3]:
                        if not os.path.exists(extractpath + attrpath + t.strftime("%Y%m%d-%H%M")):
                            os.mkdir(extractpath + attrpath + t.strftime("%Y%m%d-%H%M"))
                        f = open(extractpath + rtexportpolicy, "a")
                        f.write(cfglines[p] + '\n')
                        f.close()
                        if re.search('prefix-list',cfglines[p]):
                            prefix = cfglines[p].split('prefix-list ')[-1]
                            rtexportprefix = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_prefix-list_' + prefix + '.txt'
                            for q in prefixline:
                                if prefix == cfglines[q].split(' ')[3]:
                                    f = open(extractpath + rtexportprefix, "a")
                                    f.write(cfglines[q] + '\n')
                                    #print(cfglines[q])
                                    f.close()
                        elif re.search('community', cfglines[p]):
                            commu = cfglines[p].split(' ')[-1]
                            for r in commuline:
                                if commu == cfglines[r].split(' ')[3]:
                                    #rtexportcommu = cfglines[r].split(' ')[5]
                                    rtexpcommulist.append(cfglines[r].split(' ')[5])
                                    #print(rtexportcommu)
                if rtexpcommulist:
                    rtexportcommu = Counter(rtexpcommulist).most_common(1)[0][0]
            elif re.search('routing-instances %s routing-options static route' % vrf, cfglines[l]) or re.search('routing-instances %s routing-options rib %s.inet6.0 static route' % (vrf,vrf), cfglines[l]):
                staticroute = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_static-route_' + vrf + '.txt'
                f = open(extractpath + staticroute, "a")
                f.write(cfglines[l] + '\n')
                f.close()
            elif re.search('routing-instances %s routing-options aggregate route' % vrf, cfglines[l]) or re.search('routing-instances %s routing-options rib %s.inet6.0 aggregate route' % (vrf,vrf), cfglines[l]):
                aggregateroute = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_aggregate-route_' + vrf + '.txt'
                f = open(extractpath + aggregateroute, "a")
                f.write(cfglines[l] + '\n')
                f.close()
            elif re.search('routing-instances %s protocols bgp' % vrf, cfglines[l]):
                bgp = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + 'global-vprn' + '_bgp-to-ce_' + vrf + '.txt'
                f = open(extractpath + bgp, "a")
                f.write(cfglines[l] + '\n')
                f.close()
                if re.match('import',cfglines[l].split(' ')[-2]):
                    imppolicy = cfglines[l].split(' ')[-1]
                    impbgppolicy = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + 'global-vprn' + '_bgp-imp_' + vrf + '.txt'
                    for i in policyline:
                        if imppolicy == cfglines[i].split(' ')[3]:
                            f = open(extractpath + impbgppolicy, "a")
                            f.write(cfglines[i] + '\n')
                            f.close()
                            if re.search('prefix-list', cfglines[i]):
                                prefix = cfglines[i].split('prefix-list ')[-1]
                                impbgppolicyprefix = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + 'global-vprn' + '_bgp-imp-prefix-list_' + prefix + '.txt'
                                for q in prefixline:
                                    if prefix == cfglines[q].split(' ')[3]:
                                        f = open(extractpath + impbgppolicyprefix, "a")
                                        f.write(cfglines[q] + '\n')
                                        f.close()
                                        #print(impbgppolicyprefix)
                            elif re.search('community', cfglines[i]):
                                commu = cfglines[i].split(' ')[-1]
                                for r in commuline:
                                    if commu == cfglines[r].split(' ')[3]:
                                        impbgppolicycommu = cfglines[r].split(' ')[5]
                                        #print(impbgppolicycommu)
                elif re.match('export',cfglines[l].split(' ')[-2]):
                    exppolicy = cfglines[l].split(' ')[-1]
                    expbgppolicy = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + 'global-vprn' + '_bgp-exp_' + vrf + '.txt'
                    for i in policyline:
                        if exppolicy == cfglines[i].split(' ')[3]:
                            f = open(extractpath + expbgppolicy, "a")
                            f.write(cfglines[i] + '\n')
                            f.close()
                            if re.search('prefix-list', cfglines[i]):
                                prefix = cfglines[i].split('prefix-list ')[-1]
                                expbgppolicyprefix = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + 'global-vprn' + '_bgp-exp-prefix-list_' + prefix + '.txt'
                                for q in prefixline:
                                    if prefix == cfglines[q].split(' ')[3]:
                                        f = open(extractpath + expbgppolicyprefix, "a")
                                        f.write(cfglines[q] + '\n')
                                        f.close()
                                        #print(impbgppolicyprefix)
                            elif re.search('community', cfglines[i]):
                                commu = cfglines[i].split(' ')[-1]
                                for r in commuline:
                                    if commu == cfglines[r].split(' ')[3]:
                                        expbgppolicycommu = cfglines[r].split(' ')[5]
                                        #print(expbgppolicycommu)

    vprn['A1'] = row
    print('vprn:finish %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
    if row > start_row:
        index_row = excel['index']['F1'].value
        excel['index']['F%s' % index_row] = node       # update hostname index
        excel['index']['G%s' % index_row] = start_row  # update start row index
        excel['index']['H%s' % index_row] = row - 1    # update end row index
        excel['index']['I%s' % index_row] = '###'
        excel['index']['F1'] = index_row + 1           # update row of index


def vprnint(excel, cfglines, node, system, portlines, vprnname, vprnline, prefixline, commuline, policyline, qosline):
    vprnint = excel[wsvprnint]
    row = vprnint['A1'].value
    start_row = row
    lastaddr = ''
    address = []
    imppolilist = []
    exppolilist = []
    bgpce = {}
    mapsheet = openpyxl.load_workbook(inputpath + wbmap)['node-map']

    for vrf in vprnname:
        for v in vprnline:
            if re.search('routing-instances %s interface' % vrf,cfglines[v]):
                int = cfglines[v].split(' ')[-1].split('.')[0]
                uni = cfglines[v].split(' ')[-1].split('.')[1]
                #print('int %s unit %s' % (int,unit))
                vrrp = ''
                for u in portlines:
                    primary = 'none'
                    if re.search('interfaces %s unit %s family' % (int,uni),cfglines[u]):
                        if re.search('address',cfglines[u]):
                            if re.search('primary',cfglines[u]):
                                primary = 'primary'
                            if 'inet' == cfglines[u].split(' ')[6]:
                                fam = 'ipv4'
                            elif 'inet6' == cfglines[u].split(' ')[6]:
                                fam = 'ipv6'
                            if re.search('vrrp', cfglines[u]):
                                if cfglines[u].split(' ')[10] != vrrp:
                                    vrrp = cfglines[u].split(' ')[10]
                                    address.append(vrf+'$&'+int+'$&'+uni+'$&'+cfglines[u].split(' ')[8]+'$&'+fam+'$&'+primary)
                            else:
                                address.append(vrf+'$&'+int+'$&'+uni+'$&'+cfglines[u].split(' ')[8]+'$&'+fam+'$&'+primary)
                                #print(cfglines[u])

    if not os.path.exists(extractpath + attrpath + t.strftime("%Y%m%d-%H%M")):
        os.mkdir(extractpath + attrpath + t.strftime("%Y%m%d-%H%M"))

    for l in vprnline:
        if re.search('routing-instances', cfglines[l]):
            vpn = cfglines[l].split(' ')[2]
            if re.search('routing-instances %s protocols bgp group' % vpn, cfglines[l]):
                group = cfglines[l].split(' ')[6]
                if re.search('routing-instances %s protocols bgp group %s' % (vpn, group), cfglines[l]):
                    bgp = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-' + vpn + '_group-' + group + '.txt'
                    f = open(extractpath + bgp, "a")
                    f.write(cfglines[l] + '\n')
                    f.close()
                    if re.search('routing-instances %s protocols bgp group %s peer-as' % (vpn, group), cfglines[l]):
                        peeras = cfglines[l].split('routing-instances %s protocols bgp group %s peer-as ' % (vpn, group))[-1]
                        bgpce[vpn+group] = []
                        bgpce[vpn+group].append(vpn+'$&'+group+'$&'+bgp+'$&'+peeras)


    address.append('vrf$&int$&unit$&addr$&family$&pri')
    #print(address)
    #print(bgpce)
    for a in range(len(address)):
        if address[a] != lastaddr and lastaddr:
            vprnint['A%s' % row] = node
            vprnint['B%s' % row] = system
            vprnint['C%s' % row] = vpn
            vprnint['D%s' % row] = rd
            vprnint['E%s' % row] = layer3int
            vprnint['F%s' % row] = unit
            vprnint['G%s' % row] = unitstate
            vprnint['H%s' % row] = vlan
            vprnint['I%s' % row] = interfacedesc
            vprnint['J%s' % row] = forwardingclass
            vprnint['K%s' % row] = addr
            vprnint['L%s' % row] = secondary
            vprnint['M%s' % row] = family
            vprnint['N%s' % row] = vrrpid
            vprnint['O%s' % row] = vrrpvip
            vprnint['P%s' % row] = vrrppriority
            vprnint['Q%s' % row] = vrrpinterval
            vprnint['R%s' % row].hyperlink = staticroute
            vprnint['S%s' % row].hyperlink = bgp
            vprnint['T%s' % row] = group
            vprnint['U%s' % row] = peeras
            vprnint['V%s' % row] = neighbor
            vprnint['W%s' % row] = neighbordesc
            if neighbor:
                if IPAddress(neighbor) in IPNetwork(addr):
                    vprnint['X%s' % row] = bgpimppolicyname
                    vprnint['Y%s' % row].hyperlink = bgpimpolicy
                    #print(bgpimpolicyprefix)
                    vprnint['Z%s' % row].hyperlink = bgpimpolicyprefix
                    vprnint['AA%s' % row] = bgpimpolicycommu
                    vprnint['AB%s' % row] = bgpexppolicyname
                    vprnint['AC%s' % row].hyperlink = bgpexpolicy
                    vprnint['AD%s' % row].hyperlink = bgpexpolicyprefix
                    vprnint['AE%s' % row] = bgpexpolicycommu
            vprnint['AF%s' % row] = bgpbfdinterval
            vprnint['AG%s' % row] = bgpbfdmultiply
            vprnint['AH%s' % row] = nokianode
            vprnint['AI%s' % row] = nokiasystemip
            vprnint['AJ%s' % row] = nokiasap
            row += 1
        lastrow = row - 1
        vpn = address[a].split('$&')[0]
        rd = ''
        layer3int = address[a].split('$&')[1]
        unit = address[a].split('$&')[2]
        unitstate = ''
        vlan = ''
        interfacedesc = ''
        forwardingclass = ''
        addr = address[a].split('$&')[3]
        secondary = ''
        if address[a].split('$&')[1] == address[a-1].split('$&')[1]:
            if address[a].split('$&')[2] == address[a-1].split('$&')[2]:
                if address[a].split('$&')[4] == address[a-1].split('$&')[4]:
                    if address[a].split('$&')[3] != address[a-1].split('$&')[3]:
                        if address[a].split('$&')[5] == 'primary':
                            secondary = 'yes primary'
                        else:
                            secondary = 'yes'
                            if address[a].split('$&')[2] == address[a - 1].split('$&')[2]:
                                vprnint['L%s' % lastrow] = 'yes'
        family = address[a].split('$&')[4]
        vrrpid = ''
        vrrpvip = ''
        vrrppriority = ''
        vrrpinterval = ''
        staticroute = ''
        bgp = ''
        group = ''
        peeras = ''
        neighbor = ''
        neighbordesc = ''
        bgpimppolicyname = ''
        bgpimpolicy = ''
        bgpimpolicyprefix = ''
        bgpimpolicycommu = ''
        bgpexppolicyname = ''
        bgpexpolicy = ''
        bgpexpolicyprefix = ''
        bgpexpolicycommu = ''
        bgpbfdinterval = ''
        bgpbfdmultiply = ''
        nokianode = ''
        nokiasystemip = ''
        nokiasap = ''
        lastaddr = address[a]
        nh = '1.1.1.1'
        nhc = '1.1.1.1'


        for p in portlines:
            if re.search('interfaces %s unit %s' % (layer3int, unit), cfglines[p]):
                if re.search('deactivate interfaces %s unit %s' % (layer3int, unit), cfglines[p]):
                    unitstate = 'deactivate'
                    #print('### deactivate : %s' %cfglines[p])
                    #print(unitstate)
                elif re.search('set interfaces %s unit %s disable' % (layer3int, unit), cfglines[p]):
                    unitstate = 'disable'
                    #print('### disable : %s' % cfglines[p])
                    #print(unitstate)
                elif re.search('interfaces %s unit %s vlan-id' % (layer3int, unit), cfglines[p]):
                    vlan = cfglines[p].split('interfaces %s unit %s vlan-id ' % (layer3int, unit))[-1]
                    #print('### vlan : %s' % cfglines[p])
                    #print(vlan)
                elif re.search('interfaces %s unit %s description' % (layer3int, unit), cfglines[p]):
                    interfacedesc = cfglines[p].split('interfaces %s unit %s description ' % (layer3int, unit))[-1]
                    #print('### interfacedesc : %s' % cfglines[p])
                    #print(interfacedesc)
                elif re.search('interfaces %s unit %s family inet address %s vrrp-group' % (layer3int, unit, addr), cfglines[p]):
                    vrrpid = cfglines[p].split(' ')[10]
                    #print('### vrrpid : %s' % cfglines[p])
                    #print(vrrpid)
                    if re.search('interfaces %s unit %s family inet address %s vrrp-group %s virtual-address' % (layer3int, unit, addr, vrrpid),cfglines[p]):
                        vrrpvip = cfglines[p].split(' ')[-1]
                        #print('### vrrpvip : %s' % cfglines[p])
                        #print(vrrpvip)
                    elif re.search('interfaces %s unit %s family inet address %s vrrp-group %s priority' % (layer3int, unit, addr, vrrpid),cfglines[p]):
                        vrrppriority = cfglines[p].split(' ')[-1]
                        #print('### vrrppriority : %s' % cfglines[p])
                        #print(vrrppriority)
                    elif re.search('interfaces %s unit %s family inet address %s vrrp-group %s fast-interval' % (layer3int, unit, addr, vrrpid),cfglines[p]):
                        vrrpinterval = cfglines[p].split(' ')[-1]
                        #print('### vrrpinterval : %s' % cfglines[p])
                        #print(vrrpinterval)
                elif re.search('interfaces %s unit %s family inet6 address %s vrrp-inet6-group' % (layer3int, unit, addr), cfglines[p]):
                    vrrpid = cfglines[p].split(' ')[10]
                    #print('### vrrpid : %s' % cfglines[p])
                    #print(vrrpid)
                    if re.search('interfaces %s unit %s family inet6 address %s vrrp-inet6-group %s virtual-inet6-address' % (layer3int, unit, addr, vrrpid),cfglines[p]):
                        vrrpvip = cfglines[p].split(' ')[-1]
                        #print('### vrrpvip : %s' % cfglines[p])
                        #print(vrrpvip)
                    elif re.search('interfaces %s unit %s family inet6 address %s vrrp-inet6-group %s priority' % (layer3int, unit, addr, vrrpid),cfglines[p]):
                        vrrppriority = cfglines[p].split(' ')[-1]
                        #print('### vrrppriority : %s' % cfglines[p])
                        #print(vrrppriority)
                    elif re.search('interfaces %s unit %s family inet6 address %s vrrp-inet6-group %s fast-interval' % (layer3int, unit, addr, vrrpid),cfglines[p]):
                        vrrpinterval = cfglines[p].split(' ')[-1]
                        #print('### vrrpinterval : %s' % cfglines[p])
                        #print(vrrpinterval)

        for qos in qosline:
            if re.search('class-of-service interfaces %s unit %s forwarding-class' % (layer3int, unit),cfglines[qos]):
                forwardingclass = cfglines[qos].split('class-of-service interfaces %s unit %s forwarding-class ' % (layer3int, unit))[-1]
                #print('### forwardingclass : %s' % cfglines[qos])
                #print(forwardingclass)
            elif re.search('class-of-service interfaces %s unit %s classifiers exp' % (layer3int, unit),cfglines[qos]):
                forwardingclass = cfglines[qos].split('class-of-service interfaces %s unit %s classifiers exp ' % (layer3int, unit))[-1]
                #print('### forwardingclass : %s' % cfglines[qos])
                #print(forwardingclass)

        for l in vprnline:
            if re.search('routing-instances %s route-distinguisher' % vpn, cfglines[l]):
                rd = cfglines[l].split('routing-instances %s route-distinguisher ' % vpn)[-1]
                #print('### rd : %s' % cfglines[l])
                #print(rd)
            elif re.search('routing-instances %s routing-options static route' % vpn, cfglines[l]):
                if re.search('next-hop', cfglines[l]):
                    nh = cfglines[l].split('next-hop ')[-1].split()[0]
                    if IPAddress(nh) in IPNetwork(addr):
                        nhc = nh
                        dest = cfglines[l].split(' ')[6]
                    #print('ip route %s next-hop %s' %(dest,nh))
                if IPAddress(nhc) in IPNetwork(addr):
                    staticroute = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_static-route_' + vpn + '_' + layer3int.replace('/','-') + '.' + unit + '.txt'
                    f = open(extractpath + staticroute, "a")
                    if 'next-hop' in cfglines[l]:
                        if nhc == cfglines[l].split('next-hop ')[-1].split()[0]:
                            f.write(cfglines[l] + '\n')
                    else:
                        if dest in cfglines[l]:
                            f.write(cfglines[l] + '\n')
                    f.close()
            elif re.search('routing-instances %s routing-options rib ' % vpn, cfglines[l]) and 'inet6.0 static route' in cfglines[l]:
                if re.search('next-hop', cfglines[l]):
                    nh = cfglines[l].split('next-hop ')[-1].split()[0]
                    if IPAddress(nh) in IPNetwork(addr):
                        nhc = nh
                        dest = cfglines[l].split(' ')[8]
                if IPAddress(nhc) in IPNetwork(addr):
                    staticroute = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_static-route_' + vpn + '_' + layer3int.replace('/','-') + '.' + unit + '.txt'
                    f = open(extractpath + staticroute, "a")
                    if 'next-hop' in cfglines[l]:
                        if nhc == cfglines[l].split('next-hop ')[-1].split()[0]:
                            f.write(cfglines[l] + '\n')
                    else:
                        if dest in cfglines[l]:
                            f.write(cfglines[l] + '\n')
                    f.close()
            elif re.search('routing-instances %s protocols bgp group' % vpn, cfglines[l]):
                gr = cfglines[l].split(' ')[6]
                if re.search('bgp group %s import' % gr,cfglines[l]):
                    bgpimppolicyname = cfglines[l].split(' ')[-1]
                    bgpimpolicy = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-imp-' + vpn + '_' + bgpimppolicyname + '.txt'
                    for i in policyline:
                        if bgpimppolicyname == cfglines[i].split(' ')[3]:
                                f = open(extractpath + bgpimpolicy, "a")
                                if cfglines[i] not in imppolilist:
                                    imppolilist.append(cfglines[i])
                                    f.write(cfglines[i] + '\n')
                                f.close()
                                if re.search('prefix-list', cfglines[i]):
                                    #print(cfglines[l])
                                    #print(cfglines[i])
                                    prefix = cfglines[i].split('prefix-list ')[-1]
                                    bgpimpolicyprefix = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-imp-prefix-list_' + prefix + '.txt'
                                    if not os.path.exists(extractpath + bgpimpolicyprefix):
                                        for q in prefixline:
                                            if prefix == cfglines[q].split(' ')[3]:
                                                f = open(extractpath + bgpimpolicyprefix, "a")
                                                f.write(cfglines[q] + '\n')
                                                f.close()
                                                # print(bgpimpolicyprefix)
                                elif re.search('community', cfglines[i]):
                                    commu = cfglines[i].split(' ')[-1]
                                    for r in commuline:
                                        if commu == cfglines[r].split(' ')[3]:
                                            bgpimpolicycommu = cfglines[r].split(' ')[5]
                                            # print(bgpimpolicycommu)
                elif re.search('bgp group %s export' % gr, cfglines[l]):
                    bgpexppolicyname = cfglines[l].split(' ')[-1]
                    # print(bgpexppolicyname)
                    bgpexpolicy = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-exp-' + vpn + '_' + bgpexppolicyname + '.txt'
                    for i in policyline:
                        if bgpexppolicyname == cfglines[i].split(' ')[3]:
                            f = open(extractpath + bgpexpolicy, "a")
                            if cfglines[i] not in exppolilist:
                                exppolilist.append(cfglines[i])
                                f.write(cfglines[i] + '\n')
                            f.close()
                            if re.search('prefix-list', cfglines[i]):
                                prefix = cfglines[i].split('prefix-list ')[-1]
                                bgpexpolicyprefix = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-exp-prefix-list_' + prefix + '.txt'
                                if not os.path.exists(extractpath + bgpexpolicyprefix):
                                    for q in prefixline:
                                        if prefix == cfglines[q].split(' ')[3]:
                                            f = open(extractpath + bgpexpolicyprefix, "a")
                                            f.write(cfglines[q] + '\n')
                                            f.close()
                                            # print(bgpexpolicyprefix)
                            elif re.search('community', cfglines[i]):
                                commu = cfglines[i].split(' ')[-1]
                                for r in commuline:
                                    if commu == cfglines[r].split(' ')[3]:
                                        bgpexpolicycommu = cfglines[r].split(' ')[5]
                                        # print(bgpexpolicycommu)
                if re.search('routing-instances %s protocols bgp group %s neighbor'% (vpn,gr), cfglines[l]):
                    nei = cfglines[l].split(' ')[8]
                    if IPAddress(nei) in IPNetwork(addr):
                        #print(bgpce[vpn+gr][0])
                        neighbor = nei
                        bgp = bgpce[vpn+gr][0].split('$&')[2]
                        group = bgpce[vpn+gr][0].split('$&')[1]
                        peeras = bgpce[vpn+gr][0].split('$&')[3]
                        #print('int addr : %s , neighbor : %s' %(addr,neighbor))
                        #print('bgp path : %s' % bgp)
                        #print('group : %s' % group)
                        #print('peeras : %s' % peeras)
                        #print('neighbor : %s' % neighbor)
                        if re.search('routing-instances %s protocols bgp group %s neighbor %s description' % (vpn, group, neighbor), cfglines[l]):
                            neighbordesc = cfglines[l].split('routing-instances %s protocols bgp group %s neighbor %s description ' % (vpn, group, neighbor))[-1]
                            #print(neighbordesc)
                        elif re.match('import', cfglines[l].split(' ')[-2]):
                            bgpimppolicyname = cfglines[l].split(' ')[-1]
                            bgpimpolicy = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-imp-' + vpn + '_' + bgpimppolicyname + '.txt'
                            for i in policyline:
                                if bgpimppolicyname == cfglines[i].split(' ')[3]:
                                    f = open(extractpath + bgpimpolicy, "a")
                                    if cfglines[i] not in imppolilist:
                                        imppolilist.append(cfglines[i])
                                        f.write(cfglines[i] + '\n')
                                    f.close()
                                    if re.search('prefix-list', cfglines[i]):
                                        prefix = cfglines[i].split('prefix-list ')[-1]
                                        bgpimpolicyprefix = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-imp-prefix-list_' + prefix + '.txt'
                                        if not os.path.exists(extractpath + bgpimpolicyprefix):
                                            for q in prefixline:
                                                if prefix == cfglines[q].split(' ')[3]:
                                                    #print(cfglines[q])
                                                    f = open(extractpath + bgpimpolicyprefix, "a")
                                                    f.write(cfglines[q] + '\n')
                                                    f.close()
                                                    #print(bgpimpolicyprefix)
                                    elif re.search('community', cfglines[i]):
                                        commu = cfglines[i].split(' ')[-1]
                                        for r in commuline:
                                            if commu == cfglines[r].split(' ')[3]:
                                                bgpimpolicycommu = cfglines[r].split(' ')[5]
                                                #print(bgpimpolicycommu)
                        elif re.match('export', cfglines[l].split(' ')[-2]):
                            bgpexppolicyname = cfglines[l].split(' ')[-1]
                            #print(bgpexppolicyname)
                            bgpexpolicy = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-exp-' + vpn + '_' + bgpexppolicyname + '.txt'
                            for i in policyline:
                                if bgpexppolicyname == cfglines[i].split(' ')[3]:
                                    f = open(extractpath + bgpexpolicy, "a")
                                    if cfglines[i] not in exppolilist:
                                        exppolilist.append(cfglines[i])
                                        f.write(cfglines[i] + '\n')
                                    f.close()
                                    if re.search('prefix-list', cfglines[i]):
                                        prefix = cfglines[i].split('prefix-list ')[-1]
                                        bgpexpolicyprefix = attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + node + '_bgp-exp-prefix-list_' + prefix + '.txt'
                                        if not os.path.exists(extractpath + bgpexpolicyprefix):
                                            for q in prefixline:
                                                if prefix == cfglines[q].split(' ')[3]:
                                                    f = open(extractpath + bgpexpolicyprefix, "a")
                                                    f.write(cfglines[q] + '\n')
                                                    f.close()
                                                    #print(bgpexpolicyprefix)
                                    elif re.search('community', cfglines[i]):
                                        commu = cfglines[i].split(' ')[-1]
                                        for r in commuline:
                                            if commu == cfglines[r].split(' ')[3]:
                                                bgpexpolicycommu = cfglines[r].split(' ')[5]
                                                #print(bgpexpolicycommu)
                        elif re.search('routing-instances %s protocols bgp group %s neighbor %s bfd-liveness-detection minimum-interval' % (vpn, group, neighbor), cfglines[l]):
                            bgpbfdinterval = cfglines[l].split(' ')[-1]
                            #print(bgpbfdinterval)
                        elif re.search('routing-instances %s protocols bgp group %s neighbor %s bfd-liveness-detection multiplier' % (vpn, group, neighbor), cfglines[l]):
                            bgpbfdmultiply = cfglines[l].split(' ')[-1]
                            #print(bgpbfdmultiply)


        for x in range(3, (mapsheet.max_row + 1)):
            if re.match(node, mapsheet['A%s' % x].value):
                nokianode = mapsheet['C%s' % x].value
                nokiasystemip = mapsheet['D%s' % x].value
                #print('### nokia node : %s' % nokianode)
                #print('### nokia system ip : %s' % nokiasystemip)


    vprnint['A1'] = row
    print('vprnint:finish %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
    if row > start_row:
        index_row = excel['index']['J1'].value
        excel['index']['J%s' % index_row] = node       # update hostname index
        excel['index']['K%s' % index_row] = start_row  # update start row index
        excel['index']['L%s' % index_row] = row - 1    # update end row index
        excel['index']['M%s' % index_row] = '###'
        excel['index']['J1'] = index_row + 1           # update row of index


def l2vpn(excel,cfglines,node, system, portlines, l2vpnname, l2vpnline, qosline):
    l2vpn = excel[wsl2vpn]
    row = l2vpn['A1'].value
    start_row = row
    unitlist = []
    lastunit = ''
    mapsheet = openpyxl.load_workbook(inputpath + wbmap)['node-map']

    #print(l2vpnname)
    for l2 in l2vpnname:
        for l in l2vpnline:
            if re.search('routing-instances %s interface' % l2, cfglines[l]):
                int = cfglines[l].split(' ')[-1].split('.')[0]
                uni = cfglines[l].split(' ')[-1].split('.')[1]
                #print('int %s unit %s' %(int,uni))
                sta = ''
                des = ''
                vl = ''
                inmap = ''
                inmapvlan = ''
                outmap = ''
                outmapvlan = ''
                polin = ''
                polou = ''
                fc = ''
                for p in portlines:
                    if re.search('interfaces %s unit %s description' %(int,uni), cfglines[p]):
                        des = cfglines[p].split('interfaces %s unit %s description '%(int,uni))[-1]
                        #print(cfglines[p])
                        #print(des)
                    elif re.search('interfaces %s unit %s disable' %(int,uni), cfglines[p]):
                        sta = 'disable'
                        #print(cfglines[p])
                        #print(sta)
                    elif re.search('deactivate interfaces %s unit %s' %(int,uni), cfglines[p]):
                        sta = 'deactivate'
                        #print(cfglines[p])
                        #print(sta)
                    elif re.search('interfaces %s unit %s vlan-id' %(int,uni), cfglines[p]):
                        vl = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(vl)
                    elif re.search('interfaces %s unit %s input-vlan-map' %(int,uni), cfglines[p]) and 'vlan-id' not in cfglines[p]:
                        inmap = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(inmap)
                    elif re.search('interfaces %s unit %s input-vlan-map' %(int,uni), cfglines[p]) and 'vlan-id' in cfglines[p]:
                        inmapvlan = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(inmapvlan)
                    elif re.search('interfaces %s unit %s output-vlan-map' %(int,uni), cfglines[p]) and 'vlan-id' not in cfglines[p]:
                        outmap = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(outmap)
                    elif re.search('interfaces %s unit %s output-vlan-map' %(int,uni), cfglines[p]) and 'vlan-id' in cfglines[p]:
                        outmapvlan = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(outmapvlan)
                    elif re.search('interfaces %s unit %s family ccc policer input' %(int,uni), cfglines[p]):
                        polin = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(polin)
                    elif re.search('interfaces %s unit %s family ccc policer output' %(int,uni), cfglines[p]):
                        polou = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(polou)
                for qos in qosline:
                    if re.search('class-of-service interfaces %s unit %s forwarding-class' % (int, uni),cfglines[qos]):
                        fc = cfglines[qos].split('class-of-service interfaces %s unit %s forwarding-class ' % (int, uni))[-1]
                        #print('### forwardingclass : %s' % cfglines[qos])
                        #print(fc)
                    elif re.search('class-of-service interfaces %s unit %s classifiers exp' % (int, uni),cfglines[qos]):
                        fc = cfglines[qos].split('class-of-service interfaces %s unit %s classifiers exp ' % (int, uni))[-1]
                        #print('### forwardingclass : %s' % cfglines[qos])
                        #print(fc)
                unitlist.append(l2+'$&'+int+'$&'+uni+'$&'+sta+'$&'+des+'$&'+vl+'$&'+inmap+'$&'+inmapvlan+'$&'+outmap+'$&'+outmapvlan+'$&'+polin+'$&'+polou+'$&'+fc)

    unitlist.append('l2'+'$&'+'int'+'$&'+'uni'+'$&'+'sta'+'$&'+'des'+'$&'+'vl'+'$&'+'inmap'+'$&'+'inmapvlan'+'$&'+'outmap'+'$&'+'outmapvlan'+'$&'+'polin'+'$&'+'polou'+'$&'+'fc')
    #print(unitlist)

    for u in range(len(unitlist)):
        if unitlist[u] != lastunit and lastunit:
            l2vpn['A%s' % row] = node
            l2vpn['B%s' % row] = system
            l2vpn['C%s' % row] = l2vpnname
            l2vpn['D%s' % row] = rd
            l2vpn['E%s' % row] = vrftarget
            l2vpn['F%s' % row] = site
            l2vpn['G%s' % row] = siteid
            l2vpn['H%s' % row] = remotesiteid
            l2vpn['I%s' % row] = sitepreference
            l2vpn['J%s' % row] = interface
            l2vpn['K%s' % row] = unit
            l2vpn['L%s' % row] = unitstate
            l2vpn['M%s' % row] = unitdesc
            l2vpn['N%s' % row] = vlan
            l2vpn['O%s' % row] = inputvlanmap
            l2vpn['P%s' % row] = inputmapvlanid
            l2vpn['Q%s' % row] = outputvlanmap
            l2vpn['R%s' % row] = outputmapvlanid
            l2vpn['S%s' % row] = policerin
            l2vpn['T%s' % row] = policerout
            l2vpn['U%s' % row] = forwardingclass
            l2vpn['V%s' % row] = nokianode
            l2vpn['W%s' % row] = nokiasystemip
            l2vpn['X%s' % row] = nokiasap
            row += 1

        l2vpnname = unitlist[u].split('$&')[0]
        rd = ''
        vrftarget = ''
        site = ''
        siteid = ''
        remotesiteid = ''
        sitepreference = ''
        interface = unitlist[u].split('$&')[1]
        unit = unitlist[u].split('$&')[2]
        unitstate = unitlist[u].split('$&')[3]
        unitdesc = unitlist[u].split('$&')[4]
        vlan = unitlist[u].split('$&')[5]
        inputvlanmap = unitlist[u].split('$&')[6]
        inputmapvlanid = unitlist[u].split('$&')[7]
        outputvlanmap = unitlist[u].split('$&')[8]
        outputmapvlanid = unitlist[u].split('$&')[9]
        policerin = unitlist[u].split('$&')[10]
        policerout = unitlist[u].split('$&')[11]
        forwardingclass = unitlist[u].split('$&')[12]
        nokianode = ''
        nokiasystemip = ''
        nokiasap = ''
        lastunit = unitlist[u]

        for l in l2vpnline:
            if re.search('routing-instances %s' %l2vpnname, cfglines[l]):
                if re.search('routing-instances %s route-distinguisher' % l2vpnname, cfglines[l]):
                    rd = cfglines[l].split('routing-instances %s route-distinguisher ' % l2vpnname)[-1]
                    #print(cfglines[l])
                    #print(rd)
                elif re.search('routing-instances %s vrf-export' % l2vpnname, cfglines[l]):
                    vrftarget = 'target:%s' % rd
                elif re.search('routing-instances %s vrf-target' % l2vpnname, cfglines[l]):
                    vrftarget = cfglines[l].split('routing-instances %s vrf-target ' % l2vpnname)[-1]
                    #print(cfglines[l])
                    #print(vrftarget)
                elif re.search('routing-instances %s protocols l2vpn site' % l2vpnname, cfglines[l]):
                    site = cfglines[l].split(' ')[6]
                    #print(cfglines[l])
                    #print(site)
                if re.search('routing-instances %s protocols l2vpn site %s site-identifier' % (l2vpnname,site), cfglines[l]):
                    siteid = cfglines[l].split('routing-instances %s protocols l2vpn site %s site-identifier ' % (l2vpnname,site))[-1]
                    #print(cfglines[l])
                    #print(siteid)
                if re.search('routing-instances %s protocols l2vpn site %s interface %s.%s remote-site-id' % (l2vpnname,site,interface,unit), cfglines[l]):
                    remotesiteid = cfglines[l].split('routing-instances %s protocols l2vpn site %s interface %s.%s remote-site-id ' % (l2vpnname,site,interface,unit))[-1]
                    #print(cfglines[l])
                    #print(remotesiteid)
                if re.search('routing-instances %s protocols l2vpn site %s site-preference' % (l2vpnname, site), cfglines[l]):
                    sitepreference = cfglines[l].split('routing-instances %s protocols l2vpn site %s site-preference ' % (l2vpnname, site))[-1]
                    #print(cfglines[l])
                    #print(sitepreference)
        for x in range(3, (mapsheet.max_row + 1)):
            if re.match(node, mapsheet['A%s' % x].value):
                nokianode = mapsheet['C%s' % x].value
                nokiasystemip = mapsheet['D%s' % x].value
                #print('### nokia node : %s' % nokianode)
                #print('### nokia system ip : %s' % nokiasystemip)


    l2vpn['A1'] = row
    print('l2vpn:finish %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
    if row > start_row:
        index_row = excel['index']['N1'].value
        excel['index']['N%s' % index_row] = node       # update hostname index
        excel['index']['O%s' % index_row] = start_row  # update start row index
        excel['index']['P%s' % index_row] = row - 1    # update end row index
        excel['index']['Q%s' % index_row] = '###'
        excel['index']['N1'] = index_row + 1           # update row of index


def phylooprvpls(excel):
    if os.path.exists(extractpath + attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + "phylooprvpls.txt"):
        fv = open(extractpath + attrpath + t.strftime("%Y%m%d-%H%M") + '\\' + "phylooprvpls.txt", "r")
        phylooplines = fv.readlines()
        fv.close()
        for phyloop in phylooplines:
            node = phyloop.split('$&')[0]
            servicetype = phyloop.split('$&')[1]
            if 'VPLS' == servicetype:
                servicename = phyloop.split('$&')[2].rstrip('\n')
                l3vrf = phyloop.split('$&')[3].rstrip('\n')
                l3int = phyloop.split('$&')[4].rstrip('\n')
                l3intdesc = phyloop.split('$&')[5].rstrip('\n')
                for i in range(3,excel[wsvpls].max_row + 1):
                    if node == excel[wsvpls]['A%s' %i].value and servicename == excel[wsvpls]['C%s' %i].value:
                        excel[wsvpls]['O%s' % i] = l3vrf
                        excel[wsvpls]['P%s' % i] = l3int
                        excel[wsvpls]['Q%s' % i] = l3intdesc
            elif 'VSW' == servicetype:
                servicename = phyloop.split('$&')[2].rstrip('\n')
                vlan = phyloop.split('$&')[3].rstrip('\n')
                l3vrf = phyloop.split('$&')[4].rstrip('\n')
                l3int = phyloop.split('$&')[5].rstrip('\n')
                for i in range(3,excel[wsvsw].max_row + 1):
                    if node == excel[wsvsw]['A%s' %i].value and servicename == excel[wsvsw]['C%s' %i].value and vlan == excel[wsvsw]['J%s' %i].value:
                        excel[wsvsw]['Q%s' % i] = l3int
                        excel[wsvsw]['R%s' % i] = l3vrf


def vpls(excel, cfglines, node, system, portlines, vplsname, vplsline, qosline,vprnline):
    vpls = excel[wsvpls]
    row = vpls['A1'].value
    start_row = row
    unitlist = []
    lastunit = ''
    mapsheet = openpyxl.load_workbook(inputpath + wbmap)['node-map']

    for v in vplsname:
        for l in vplsline:
            if re.search('routing-instances %s interface' % v, cfglines[l]):
                int = cfglines[l].split(' ')[-1].split('.')[0]
                uni = cfglines[l].split(' ')[-1].split('.')[1]
                #print('int %s unit %s' %(int,uni))
                sta = ''
                des = ''
                vl = ''
                phydes = ''
                l3vr = ''
                l3in = ''
                l3intdes = ''
                polin = ''
                polou = ''
                fc = ''
                for p in portlines:
                    if re.search('interfaces %s unit %s description' %(int,uni), cfglines[p]):
                        des = cfglines[p].split('interfaces %s unit %s description '%(int,uni))[-1]
                        #print(cfglines[p])
                        #print(des)
                    elif re.search('interfaces %s unit %s disable' %(int,uni), cfglines[p]):
                        sta = 'disable'
                        #print(cfglines[p])
                        #print(sta)
                    elif re.search('deactivate interfaces %s unit %s' %(int,uni), cfglines[p]):
                        sta = 'deactivate'
                        #print(cfglines[p])
                        #print(sta)
                    elif re.search('interfaces %s unit %s vlan-id' %(int,uni), cfglines[p]):
                        vl = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(vl)
                    elif re.search('interfaces %s description' %int, cfglines[p]) and re.search('Physical loop', cfglines[p], re.IGNORECASE):
                        phydes = cfglines[p].split('interfaces %s description ' %int)[-1]
                        #print(cfglines[p])
                        #print(phydes)
                        destloopport = phydes.split(' ')[3]
                        for pli in portlines:
                            if re.search('interfaces %s unit %s description' % (destloopport, uni), cfglines[pli]):
                                l3intdes = cfglines[pli].split('interfaces %s unit %s description ' % (destloopport, uni))[-1]
                        for vprnl in vprnline:
                            if '%s.%s' %(destloopport,uni) in cfglines[vprnl]:
                                #print(cfglines[vprnl])
                                l3vr = cfglines[vprnl].split(' ')[2]
                                l3in = cfglines[vprnl].split(' ')[-1]
                                #print(l3vr)
                                #print(l3in)
                        fv = open(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")+'\\'+"phylooprvpls.txt", "a")
                        fv.write(node+'$&'+'VPLS'+'$&'+v+'$&'+l3vr+'$&'+l3in+'$&'+l3intdes+'\n')
                        fv.close()
                    elif re.search('interfaces %s unit %s family vpls policer input' %(int,uni), cfglines[p]):
                        polin = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(polin)
                    elif re.search('interfaces %s unit %s family vpls policer output' %(int,uni), cfglines[p]):
                        polou = cfglines[p].split(' ')[-1]
                        #print(cfglines[p])
                        #print(polou)
                for qos in qosline:
                    if re.search('class-of-service interfaces %s unit %s forwarding-class' % (int, uni),cfglines[qos]):
                        fc = cfglines[qos].split('class-of-service interfaces %s unit %s forwarding-class ' % (int, uni))[-1]
                        #print('### forwardingclass : %s' % cfglines[qos])
                        #print(fc)
                    elif re.search('class-of-service interfaces %s unit %s classifiers exp' % (int, uni),cfglines[qos]):
                        fc = cfglines[qos].split('class-of-service interfaces %s unit %s classifiers exp ' % (int, uni))[-1]
                        #print('### forwardingclass : %s' % cfglines[qos])
                        #print(fc)

                unitlist.append(v + '$&' + int + '$&' + uni + '$&' + sta + '$&' + des + '$&' + vl + '$&' + phydes + '$&' + l3vr + '$&' + l3in + '$&' + l3intdes + '$&' + polin + '$&' + polou + '$&' + fc)

    unitlist.append('v' + '$&' + 'int' + '$&' + 'uni' + '$&' + 'sta' + '$&' + 'des' + '$&' + 'vl' + '$&' + 'phydes' + '$&' + 'l3vr' + '$&' + 'l3in' + '$&' + 'l3intdes' + '$&' + 'polin' + '$&' + 'polou' + '$&' + 'fc')
    #print(unitlist)
    for u in range(len(unitlist)):
        if unitlist[u] != lastunit and lastunit:
            vpls['A%s' % row] = node
            vpls['B%s' % row] = system
            vpls['C%s' % row] = vplsname
            vpls['D%s' % row] = rd
            vpls['E%s' % row] = vrftarget
            vpls['F%s' % row] = site
            vpls['G%s' % row] = siteid
            vpls['H%s' % row] = sitepreference
            vpls['I%s' % row] = interface
            vpls['J%s' % row] = unit
            vpls['K%s' % row] = unitstate
            vpls['L%s' % row] = unitdescription
            vpls['M%s' % row] = vlan
            vpls['N%s' % row] = phydesc
            vpls['O%s' % row] = l3vrf
            vpls['P%s' % row] = l3int
            vpls['Q%s' % row] = l3intdesc
            vpls['R%s' % row] = policerinput
            vpls['S%s' % row] = policeroutput
            vpls['T%s' % row] = forwardingclass
            vpls['U%s' % row] = nokianode
            vpls['V%s' % row] = nokiasystemip
            vpls['W%s' % row] = nokiasap
            row += 1

        vplsname = unitlist[u].split('$&')[0]
        rd = ''
        vrftarget = ''
        site = ''
        siteid = ''
        sitepreference = ''
        interface = unitlist[u].split('$&')[1]
        unit = unitlist[u].split('$&')[2]
        unitstate = unitlist[u].split('$&')[3]
        unitdescription = unitlist[u].split('$&')[4]
        vlan = unitlist[u].split('$&')[5]
        phydesc = unitlist[u].split('$&')[6]
        l3vrf = unitlist[u].split('$&')[7]
        l3int = unitlist[u].split('$&')[8]
        l3intdesc = unitlist[u].split('$&')[9]
        policerinput = unitlist[u].split('$&')[10]
        policeroutput = unitlist[u].split('$&')[11]
        forwardingclass = unitlist[u].split('$&')[12]
        nokianode = ''
        nokiasystemip = ''
        nokiasap = ''
        lastunit = unitlist[u]
        for l in vplsline:
            if re.search('routing-instances %s' % vplsname, cfglines[l]):
                if re.search('routing-instances %s route-distinguisher' % vplsname, cfglines[l]):
                    rd = cfglines[l].split('routing-instances %s route-distinguisher ' % vplsname)[-1]
                    # print(cfglines[l])
                    # print(rd)
                elif re.search('routing-instances %s vrf-target' % vplsname, cfglines[l]):
                    vrftarget = cfglines[l].split('routing-instances %s vrf-target ' % vplsname)[-1]
                    # print(cfglines[l])
                    # print(vrftarget)
                elif re.search('routing-instances %s protocols vpls site' % vplsname, cfglines[l]):
                    site = cfglines[l].split(' ')[6]
                    # print(cfglines[l])
                    # print(site)
                if re.search('routing-instances %s protocols vpls site %s site-identifier' % (vplsname, site),cfglines[l]):
                    siteid = cfglines[l].split('routing-instances %s protocols vpls site %s site-identifier ' % (vplsname, site))[-1]
                    # print(cfglines[l])
                    # print(siteid)
                if re.search('routing-instances %s protocols vpls site %s site-preference' % (vplsname, site),cfglines[l]):
                    sitepreference = cfglines[l].split('routing-instances %s protocols vpls site %s site-preference ' % (vplsname, site))[-1]
                    # print(cfglines[l])
                    # print(sitepreference)
        for x in range(3, (mapsheet.max_row + 1)):
            if re.match(node, mapsheet['A%s' % x].value):
                nokianode = mapsheet['C%s' % x].value
                nokiasystemip = mapsheet['D%s' % x].value
                # print('### nokia node : %s' % nokianode)
                # print('### nokia system ip : %s' % nokiasystemip)




    vpls['A1'] = row
    print('vpls:finish %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
    if row > start_row:
        index_row = excel['index']['R1'].value
        excel['index']['R%s' % index_row] = node       # update hostname index
        excel['index']['S%s' % index_row] = start_row  # update start row index
        excel['index']['T%s' % index_row] = row - 1    # update end row index
        excel['index']['U%s' % index_row] = '###'
        excel['index']['R1'] = index_row + 1           # update row of index


def bridge(excel,cfglines,node, system, portlines, vprnline, bridgename, bridgeline, qosline):
    bd = excel[wsbridge]
    row = bd['A1'].value
    start_row = row
    unitlist = []
    lastunit = ''
    mapsheet = openpyxl.load_workbook(inputpath + wbmap)['node-map']

    for br in bridgename:
        for u in bridgeline:
            if re.search('bridge-domains %s interface' % br,cfglines[u]):
                #print(cfglines[u])
                int = cfglines[u].split(' ')[-1].split('.')[0]
                uni = cfglines[u].split(' ')[-1].split('.')[1]
                #print('int %s unit %s' %(int,uni))
                sta = ''
                des = ''
                vl = ''
                fc = ''
                for p in portlines:
                    if re.search('interfaces %s unit %s description' % (int, uni), cfglines[p]):
                        des = cfglines[p].split('interfaces %s unit %s description ' % (int, uni))[-1]
                        # print(cfglines[p])
                        # print(des)
                    elif re.search('interfaces %s unit %s disable' % (int, uni), cfglines[p]):
                        sta = 'disable'
                        # print(cfglines[p])
                        # print(sta)
                    elif re.search('deactivate interfaces %s unit %s' % (int, uni), cfglines[p]):
                        sta = 'deactivate'
                        # print(cfglines[p])
                        # print(sta)
                    elif re.search('interfaces %s unit %s vlan-id' % (int, uni), cfglines[p]):
                        vl = cfglines[p].split(' ')[-1]
                        # print(cfglines[p])
                        # print(vl)
                for qos in qosline:
                    if re.search('class-of-service interfaces %s unit %s forwarding-class' % (int, uni), cfglines[qos]):
                        fc = cfglines[qos].split('class-of-service interfaces %s unit %s forwarding-class ' % (int, uni))[-1]
                        # print('### forwardingclass : %s' % cfglines[qos])
                        # print(fc)
                    elif re.search('class-of-service interfaces %s unit %s classifiers exp' % (int, uni),cfglines[qos]):
                        fc = cfglines[qos].split('class-of-service interfaces %s unit %s classifiers exp ' % (int, uni))[-1]
                        # print('### forwardingclass : %s' % cfglines[qos])
                        # print(fc)

                unitlist.append(br + '$&' + int + '$&' + uni + '$&' + sta + '$&' + des + '$&' + vl + '$&' + fc)

    unitlist.append('br' + '$&' + 'int' + '$&' + 'uni' + '$&' + 'sta' + '$&' + 'des' + '$&' + 'vl' + '$&' + 'fc')
    #print(unitlist)

    for u in range(len(unitlist)):
        if unitlist[u] != lastunit and lastunit:
            bd['A%s' % row] = node
            bd['B%s' % row] = system
            bd['C%s' % row] = bridgedomain
            bd['D%s' % row] = vlan
            bd['E%s' % row] = filter
            bd['F%s' % row] = routinginterface
            bd['G%s' % row] = irbvprn
            bd['H%s' % row] = interface
            bd['I%s' % row] = unit
            bd['J%s' % row] = unitstate
            bd['K%s' % row] = unitdescription
            bd['L%s' % row] = intvlan
            bd['M%s' % row] = forwardingclass
            bd['N%s' % row] = nokianode
            bd['O%s' % row] = nokiasystemip
            bd['P%s' % row] = nokiasap
            row += 1

        bridgedomain = unitlist[u].split('$&')[0]
        vlan = ''
        filter = ''
        routinginterface = ''
        irbvprn = ''
        interface = unitlist[u].split('$&')[1]
        unit = unitlist[u].split('$&')[2]
        unitstate = unitlist[u].split('$&')[3]
        unitdescription = unitlist[u].split('$&')[4]
        intvlan = unitlist[u].split('$&')[5]
        forwardingclass = unitlist[u].split('$&')[6]
        nokianode = ''
        nokiasystemip = ''
        nokiasap = ''
        lastunit = unitlist[u]

        for l in bridgeline:
            if re.search('bridge-domains %s vlan-id' % bridgedomain,cfglines[l]):
                vlan = cfglines[l].split(' ')[-1]
                #print(cfglines[l])
                #print(vlan)
            elif re.search('bridge-domains %s forwarding-options filter input' % bridgedomain,cfglines[l]):
                filter = cfglines[l].split(' ')[-1]
                #print(cfglines[l])
                #print(filter)
            elif re.search('bridge-domains %s routing-interface' % bridgedomain,cfglines[l]):
                routinginterface = cfglines[l].split(' ')[-1]
                #print(cfglines[l])
                #print(routinginterface)
                for v in vprnline:
                    if re.search('interface %s' % routinginterface,cfglines[v]):
                        irbvprn = cfglines[v].split(' ')[2]
                        #print(cfglines[v])
                        #print(irbvprn)
        for x in range(3, (mapsheet.max_row + 1)):
            if re.match(node, mapsheet['A%s' % x].value):
                nokianode = mapsheet['C%s' % x].value
                nokiasystemip = mapsheet['D%s' % x].value
                # print('### nokia node : %s' % nokianode)
                # print('### nokia system ip : %s' % nokiasystemip)


    bd['A1'] = row
    print('bridge:finish %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
    if row > start_row:
        index_row = excel['index']['V1'].value
        excel['index']['V%s' % index_row] = node       # update hostname index
        excel['index']['W%s' % index_row] = start_row  # update start row index
        excel['index']['X%s' % index_row] = row - 1    # update end row index
        excel['index']['Y%s' % index_row] = '###'
        excel['index']['V1'] = index_row + 1           # update row of index


def vsw(excel, cfglines, node, system, portlines, vprnline, vswname, vswline, qosline):
    vsw = excel[wsvsw]
    row = vsw['A1'].value
    start_row = row
    unitlist = []
    vswlist = []
    lastunit = ''
    mapsheet = openpyxl.load_workbook(inputpath + wbmap)['node-map']
    vswbdline = []

    for vs in vswname:
        for i in vswline:
            if re.search('routing-instances %s interface' % vs, cfglines[i]):
                int = cfglines[i].split(' ')[-1]
                por = cfglines[i].split(' ')[-1].split('.')[0]
                uni = cfglines[i].split(' ')[-1].split('.')[1]
                #print('int %s : port %s unit %s' %(int,por,uni))
                sta = ''
                mod = ''
                vl = ''
                fc = ''
                l3vr = ''
                l3in = ''
                destloopport = ''
                for qos in qosline:
                    if re.search('class-of-service interfaces %s unit %s forwarding-class' % (por, uni), cfglines[qos]):
                        fc = cfglines[qos].split('class-of-service interfaces %s unit %s forwarding-class ' % (por, uni))[-1]
                        # print('### forwardingclass : %s' % cfglines[qos])
                        # print(fc)
                    elif re.search('class-of-service interfaces %s unit %s classifiers exp' % (por, uni),cfglines[qos]):
                        fc = cfglines[qos].split('class-of-service interfaces %s unit %s classifiers exp ' % (por, uni))[-1]
                        # print('### forwardingclass : %s' % cfglines[qos])
                        # print(fc)
                for p in portlines:
                    if re.search('interfaces %s unit %s family bridge interface-mode' % (por, uni), cfglines[p]):
                        mod = cfglines[p].split('interfaces %s unit %s family bridge interface-mode ' % (por, uni))[-1]
                        # print(cfglines[p])
                        #print(mod)
                    elif re.search('interfaces %s unit %s disable' % (por, uni), cfglines[p]):
                        sta = 'disable'
                        # print(cfglines[p])
                        # print(sta)
                    elif re.search('deactivate interfaces %s unit %s' % (por, uni), cfglines[p]):
                        sta = 'deactivate'
                        # print(cfglines[p])
                        # print(sta)
                    elif re.search('interfaces %s description' %por, cfglines[p]) and re.search('Physical loop', cfglines[p], re.IGNORECASE):
                        phydes = cfglines[p].split('interfaces %s description ' %por)[-1]
                        #print(cfglines[p])
                        #print(phydes)
                        destloopport = phydes.split(' ')[3]
                    elif re.search('interfaces %s unit %s family bridge vlan-id-list' % (por, uni), cfglines[p]):
                        vl = cfglines[p].split(' ')[-1]
                        # print(cfglines[p])
                        # print(vl)
                        if destloopport:
                            #print(destloopport)
                            for vprnl in vprnline:
                                if '%s.%s' %(destloopport,vl) in cfglines[vprnl]:
                                    #print(cfglines[vprnl])
                                    l3vr = cfglines[vprnl].split(' ')[2]
                                    l3in = cfglines[vprnl].split(' ')[-1]
                                    #print(l3vr)
                                    #print(l3in)
                            fv = open(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")+'\\'+"phylooprvpls.txt", "a")
                            fv.write(node+'$&'+'VSW'+'$&'+vs+'$&'+vl+'$&'+l3vr+'$&'+l3in+'\n')
                            fv.close()
                        unitlist.append(vs + '$&' + int + '$&' + por + '$&' + uni + '$&' + sta + '$&' + mod + '$&' + vl + '$&' + fc + '$&' + l3vr  + '$&' + l3in)
                    elif re.search('interfaces %s unit %s family bridge vlan-id' % (por, uni), cfglines[p]):
                        vl = cfglines[p].split(' ')[-1]
                        # print(cfglines[p])
                        # print(vl)
                        if destloopport:
                            #print(destloopport)
                            for vprnl in vprnline:
                                if '%s.%s' %(destloopport,vl) in cfglines[vprnl]:
                                    #print(cfglines[vprnl])
                                    l3vr = cfglines[vprnl].split(' ')[2]
                                    l3in = cfglines[vprnl].split(' ')[-1]
                                    #print(l3vr)
                                    #print(l3in)
                            fv = open(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")+'\\'+"phylooprvpls.txt", "a")
                            fv.write(node+'$&'+'VSW'+'$&'+vs+'$&'+vl+'$&'+l3vr+'$&'+l3in+'\n')
                            fv.close()
                        unitlist.append(vs + '$&' + int + '$&' + por + '$&' + uni + '$&' + sta + '$&' + mod + '$&' + vl + '$&' + fc + '$&' + l3vr  + '$&' + l3in )

        for b in vswline:
            if re.search('routing-instances %s route-distinguisher' % vs, cfglines[b]):
                ro = cfglines[b].split(' ')[-1]
                #print(cfglines[l])
                #print(rd)
            elif re.search('routing-instances %s vrf-target' % vs, cfglines[b]):
                vrft = cfglines[b].split(' ')[-1]
                #print(cfglines[l])
                #print(vrftarget)
            elif re.search('routing-instances %s protocols vpls site' % vs, cfglines[b]):
                s = cfglines[b].split(' ')[6]
                #print(cfglines[l])
                #print(site)
                if re.search('routing-instances %s protocols vpls site %s site-identifier' % (vs,s), cfglines[b]):
                    sid = cfglines[b].split(' ')[-1]
                    #print(cfglines[l])
                    #print(siteid)
            if re.search('routing-instances %s bridge-domains' %vs,cfglines[b]):
                vswbdline.append(b)
        vswlist.append(vs + '$&' + ro + '$&' + vrft + '$&' + s + '$&' + sid)


    unitlist.append('vs' + '$&' + 'int' + '$&' + 'por' + '$&' + 'uni' + '$&' + 'sta' + '$&' + 'mod' + '$&' + 'vl' + '$&' + 'fc' + '$&' + 'l3vr'  + '$&' + 'l3in')
    #print(unitlist)
    for u in range(len(unitlist)):
        if unitlist[u] != lastunit and lastunit:
            vsw['A%s' % row] = node
            vsw['B%s' % row] = system
            vsw['C%s' % row] = virtualswname
            vsw['D%s' % row] = rd
            vsw['E%s' % row] = vrftarget
            vsw['F%s' % row] = site
            vsw['G%s' % row] = siteid
            vsw['H%s' % row] = interfaceunit
            vsw['I%s' % row] = bridgedomain
            vsw['J%s' % row] = bridgevlan
            vsw['K%s' % row] = interface
            if unitmode == 'trunk':
                vsw['L%s' % row] = vlan
            else:
                vsw['L%s' % row] = unit
            vsw['M%s' % row] = unitstate
            vsw['N%s' % row] = unitmode
            vsw['O%s' % row] = vlan
            vsw['P%s' % row] = forwardingclass
            vsw['Q%s' % row] = routinginterface
            vsw['R%s' % row] = irbvprn
            vsw['S%s' % row] = filter
            vsw['T%s' % row] = nokianode
            vsw['U%s' % row] = nokiasystemip
            vsw['V%s' % row] = nokiasap
            row += 1


        virtualswname = unitlist[u].split('$&')[0]
        rd = ''
        vrftarget = ''
        site = ''
        siteid = ''
        interfaceunit = unitlist[u].split('$&')[1]
        bridgedomain = ''
        bridgevlan = ''
        interface = unitlist[u].split('$&')[2]
        unit = unitlist[u].split('$&')[3]
        unitstate = unitlist[u].split('$&')[4]
        unitmode = unitlist[u].split('$&')[5]
        vlan = unitlist[u].split('$&')[6]
        forwardingclass = unitlist[u].split('$&')[7]
        routinginterface = unitlist[u].split('$&')[9]
        irbvprn = unitlist[u].split('$&')[8]
        filter = ''
        nokianode = ''
        nokiasystemip = ''
        nokiasap = ''
        lastunit = unitlist[u]

        for l in vswbdline:
            for v in vswlist:
                if v.split('$&')[0] == virtualswname:
                    rd = v.split('$&')[1]
                    vrftarget = v.split('$&')[2]
                    site = v.split('$&')[3]
                    siteid = v.split('$&')[4]
            if re.search('vlan-id', cfglines[l]):
                if cfglines[l].split(' ')[-1] == vlan:
                    bridgevlan = cfglines[l].split(' ')[-1]
                    #print(cfglines[l])
                    #print(bridgevlan
                    if re.search('routing-instances %s bridge-domains' % virtualswname, cfglines[l]):
                        bridgedomain = cfglines[l].split(' ')[4]
                        #print(cfglines[l])
                        #print(bridgedomain)
                        for b in vswbdline:
                            if re.search('routing-instances %s bridge-domains %s routing-interface' % (virtualswname,bridgedomain), cfglines[b]):
                                routinginterface = cfglines[b].split(' ')[-1]
                                #print(cfglines[b])
                                #print(routinginterface)
                            elif re.search('routing-instances %s bridge-domains %s forwarding-options filter input' % (virtualswname,bridgedomain), cfglines[b]):
                                filter = cfglines[b].split(' ')[-1]
                                #print(cfglines[b])
                                #print(filter)
                                for v in vprnline:
                                    if routinginterface:
                                        if re.search('interface %s' % routinginterface, cfglines[v]):
                                            irbvprn = cfglines[v].split(' ')[2]
                                            #print(cfglines[v])
                                            #print(irbvprn)
        for x in range(3, (mapsheet.max_row + 1)):
            if re.match(node, mapsheet['A%s' % x].value):
                nokianode = mapsheet['C%s' % x].value
                nokiasystemip = mapsheet['D%s' % x].value
                # print('### nokia node : %s' % nokianode)
                # print('### nokia system ip : %s' % nokiasystemip)


    vsw['A1'] = row
    print('vsw:finish %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
    if row > start_row:
        index_row = excel['index']['Z1'].value
        excel['index']['Z%s' % index_row] = node        # update hostname index
        excel['index']['AA%s' % index_row] = start_row  # update start row index
        excel['index']['AB%s' % index_row] = row - 1    # update end row index
        excel['index']['AC%s' % index_row] = '###'
        excel['index']['Z1'] = index_row + 1            # update row of index


def main():
    print(t.strftime("%Y%m%d-%H%M"))
    if not os.path.exists(inputpath):
        os.mkdir(inputpath)
    if not os.path.exists(cfgpath):
        os.mkdir(cfgpath)
    if not os.path.exists(extractpath):
        os.mkdir(extractpath)
    if not os.path.exists(extractpath+attrpath):
        os.mkdir(extractpath+attrpath)
    if not os.path.exists(extractpath+attrpath+t.strftime("%Y%m%d-%H%M")+'\\'):
        os.mkdir(extractpath+attrpath+t.strftime("%Y%m%d-%H%M"))
    create_excel()
    for filename in os.listdir(os.getcwd()+cfgpath):
        cfgtype = filename.split('.')[-1]
        cfglines = []
        node = ''
        system = ''
        portname = []
        portlines = []
        vprnname = []
        vprnline = []
        l2vpnname = []
        l2vpnline = []
        vplsname = []
        vplsline = []
        bridgename = []
        bridgeline = []
        vswname = []
        vswline = []
        prefixname = []
        prefixline = []
        communame = []
        commuline = []
        policyname = []
        policyline = []
        aclname = []
        aclline = []
        qosline = []
        # Put 'txt' or 'cfg' files in iptn-cfg directory
        if re.match('txt',cfgtype) or re.match('cfg',cfgtype):
            starttime = time.time()
            #try:
            #    f = open(cfgpath+filename, 'r',encoding="utf-8")
            #except:
            f = open(cfgpath + filename, 'r', encoding="cp1252")
            lines = f.readlines()
            f.close()
            for a in range(len(lines)):
                lines[a] = re.sub(r"^\s+", "", lines[a])        # remove space from beginning
                lines[a] = lines[a].rstrip('\n')                # remove newline('\n') from end of line
                lines[a] = re.sub(r"\s+$", "", lines[a])        # remove space from ending
                cfglines.append(lines[a])

            node, system, portname, portlines, vprnname, vprnline, l2vpnname, \
            l2vpnline, vplsname, vplsline, bridgename, bridgeline, vswname, vswline, \
            prefixname, prefixline, communame, commuline, policyname, policyline, \
            aclname, aclline, qosline = checkline(cfglines)

            excel = openpyxl.load_workbook(extractpath+wbname)
            #excel = openpyxl.load_workbook(extractpath + 'service-extraction-20210202-1009.xlsx')
            print('port-lag:start %s'%datetime.datetime.now().strftime("%Y%m%d-%H%M"))
            t1 = Thread(target=port, args=(excel, cfglines, node, system, portname, portlines, vprnline, l2vpnline, vplsline, bridgeline, vswline, aclline, qosline))
            print('vprn:start %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
            t2 = Thread(target=vprn, args=(excel, cfglines, node, system, portlines, vprnname, vprnline, prefixline, commuline, policyline))
            print('vprnint:start %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
            t3 = Thread(target=vprnint, args=(excel, cfglines, node, system, portlines, vprnname, vprnline, prefixline, commuline, policyline, qosline))
            print('l2vpn:start %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
            t4 = Thread(target=l2vpn, args=(excel, cfglines, node, system, portlines, l2vpnname, l2vpnline, qosline))
            print('vpls:start %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
            t5 = Thread(target=vpls, args=(excel, cfglines, node, system, portlines, vplsname, vplsline, qosline,vprnline))
            print('bridge:start %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
            t6 = Thread(target=bridge, args=(excel,cfglines, node, system, portlines, vprnline, bridgename, bridgeline, qosline))
            print('vsw:start %s' % datetime.datetime.now().strftime("%Y%m%d-%H%M"))
            t7 = Thread(target=vsw, args=(excel, cfglines, node, system, portlines, vprnline, vswname, vswline, qosline))
            t1.start()
            t2.start()
            t3.start()
            t4.start()
            t5.start()
            t6.start()
            t7.start()
            t1.join()
            t2.join()
            t3.join()
            t4.join()
            t5.join()
            t6.join()
            t7.join()
            phylooprvpls(excel)
            excel.save(extractpath+wbname)
    print(datetime.datetime.now().strftime("%Y%m%d-%H%M"))

if __name__ == "__main__":
    main()