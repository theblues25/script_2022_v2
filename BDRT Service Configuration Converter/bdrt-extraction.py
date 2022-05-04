from threading import Thread
from netaddr import IPNetwork, IPAddress
from collections import Counter
import os
import openpyxl
import datetime
import time
import re

today = datetime.date.today()
currentdate = today.strftime("%Y%m%d")
yesterday = today - datetime.timedelta(days=1)
configpath = os.path.join(os.getcwd(),'backup_cfg',currentdate,'XRS')

def create_excel():
    # create workbook
    row = 3 # initial row for write data
    excel = openpyxl.Workbook()
    excel.remove(excel['Sheet'])
    #############################
    # create policyws-lag worksheet
    #############################
    policyws = excel.create_sheet('route-policy')
    policyws['A1'] = row
    policyws['A2'] = 'bdrt-node'
    policyws['B2'] = 'policy'
    policyws['C2'] = 'entry'
    policyws['D2'] = 'prefix-list'
    policyws.auto_filter.ref = 'A2:D2'
    policyws.freeze_panes = policyws['A3']
    policyws.column_dimensions['A'].width = 18.0
    policyws.column_dimensions['B'].width = 53.0
    policyws.column_dimensions['C'].width = 42.0
    policyws.column_dimensions['D'].width = 37.0

    exportcommumissingws = excel.create_sheet('export_commu_missing')
    exportcommumissingws['A1'] = row
    exportcommumissingws['A2'] = 'bdrt-node'
    exportcommumissingws['B2'] = 'policy'
    exportcommumissingws['C2'] = 'entry'
    exportcommumissingws['D2'] = 'all vrf-export policy that missing community configuration'
    exportcommumissingws.auto_filter.ref = 'A2:D2'
    exportcommumissingws.freeze_panes = exportcommumissingws['A3']
    exportcommumissingws.column_dimensions['A'].width = 18.0
    exportcommumissingws.column_dimensions['B'].width = 53.0
    exportcommumissingws.column_dimensions['C'].width = 42.0
    exportcommumissingws.column_dimensions['D'].width = 37.0
    return excel


def getallpehost():
    allhostdict = {}
    allhost = []
    nodeoobwb = openpyxl.load_workbook('node-oob-ip.xlsx')
    nodeoobws = nodeoobwb['active_pe_node']
    for i in range(3, nodeoobws.max_row + 1):
        if nodeoobws['A%s' % i].value != None:
            allhost.append(nodeoobws['A%s' % i].value)
    return allhost

if __name__ == "__main__":
    excel = create_excel()
    policyws = excel['route-policy']
    exportcommumissingws = excel['export_commu_missing']
    allhost = getallpehost()
    configfilepathdict = {}
    for nodename in allhost:
        #print(nodename)
        if f'full-context_{nodename}.txt' in os.listdir(configpath):
            configfile = os.path.join(os.getcwd(), 'backup_cfg', currentdate,'XRS', f'full-context_{nodename}.txt')
            #print(configfile)
            configfilepathdict[f'{nodename}'] = configfile
        else:
            currentday = today
            notfoundconfig = True
            while notfoundconfig:
                lastday = currentday - datetime.timedelta(days=1)
                lastdayconfigpath = os.path.join(os.getcwd(),'backup_cfg',lastday.strftime("%Y%m%d"),'XRS')
                if f'full-context_{nodename}.txt' in os.listdir(lastdayconfigpath):
                    configfile = os.path.join(os.getcwd(), 'backup_cfg', lastday.strftime("%Y%m%d"),'XRS', f'full-context_{nodename}.txt')
                    #print(configfile)
                    configfilepathdict[f'{nodename}'] = configfile
                    notfoundconfig = False
                else:
                    currentday = lastday

    for nodename in allhost:
        checkentrydict = {}
        pollist = []
        print(nodename)
        f = open(configfilepathdict[f'{nodename}'], 'r')
        alllines = f.readlines()
        f.close()
        for line in alllines:
            line = re.sub(r"^\s+", "", line)  # remove space from beginning
            line = line.rstrip('\n')  # remove newline('\n') from end of line
            line = re.sub(r"\s+$", "", line)  # remove space from ending
            if '/configure policy-options policy-statement' in line and 'prefix-list' in line:
                bdrtpolicy = line.split()[3].replace('"','')
                bdrtentry = line.split()[5].replace('"','')
                bdrtprefixlist = line.split()[-1].replace('["','').replace('"]','')
                row = policyws['A1'].value
                policyws[f'A{row}'] = nodename
                policyws[f'B{row}'] = bdrtpolicy
                policyws[f'C{row}'] = bdrtentry
                policyws[f'D{row}'] = bdrtprefixlist
                row += 1
                policyws['A1'] = row
            if '/configure policy-options policy-statement "vrf' in line and '_export" ' in line and '_ebgp' not in line:
                if ' named-entry ' in line:
                    polname = line.split()[3]
                    entryname = line.split()[5]
                    if entryname not in checkentrydict.setdefault('%s' % (polname), []):
                        checkentrydict.setdefault('%s' % (polname), []).append(entryname)
                    if polname not in pollist:
                        pollist.append(polname)
        for line in alllines:
            line = re.sub(r"^\s+", "", line)  # remove space from beginning
            line = line.rstrip('\n')  # remove newline('\n') from end of line
            line = re.sub(r"\s+$", "", line)  # remove space from ending
            if '/configure policy-options policy-statement "vrf' in line and '_export" ' in line and '_ebgp' not in line:
                if ' named-entry ' in line:
                    polname = line.split()[3]
                    entryname = line.split()[5]
                    #print(checkentrydict['%s' % (polname)])
                    if polname in line and entryname in line and 'action community' in line:
                        commuentry = line.split()[5]
                        checkentrydict.setdefault('%s_commu' % (polname), []).append(commuentry)
        for pol in pollist:
            for entry in checkentrydict['%s' % (pol)]:
                if entry not in checkentrydict['%s_commu' % (pol)]:
                    row = exportcommumissingws['A1'].value
                    exportcommumissingws[f'A{row}'] = nodename
                    exportcommumissingws[f'B{row}'] = pol
                    exportcommumissingws[f'C{row}'] = entry
                    exportcommumissingws[f'D{row}'] = 'vrf-export missing community'
                    row += 1
                    exportcommumissingws['A1'] = row


    filepath = os.path.join(os.getcwd(), 'backup_cfg', currentdate, 'bdrt-extraction.xlsx')
    excel.save(filepath)





