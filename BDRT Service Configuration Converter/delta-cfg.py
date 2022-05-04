import openpyxl
from openpyxl.utils import get_column_letter
from netaddr import IPNetwork, IPAddress
import os



def main():
    configdict = {}
    allnode = []
    headerlist = []
    nodeexecdict = {}
    ipaddr = ''

    extendfilename = input("Extended filename:")
    if not os.path.exists('./delta-cfg/'):
        os.mkdir('./delta-cfg/')
    for filename in os.listdir('./delta-cfg/'):
        if 'CE_filter_source_ip' not in filename:
            if '.xlsx' in filename or '.xlsm' in filename:
                if not '~$' in filename:
                    mopwb = openpyxl.load_workbook(filename='./delta-cfg/'+filename)
                    allsheet = mopwb.sheetnames
                    for worksheet in allsheet:
                        if worksheet != 'iptncfg' and '#' not in worksheet:
                            columnnumber = 2
                            maxcolumn = get_column_letter(mopwb[worksheet].max_column)
                            for allcolumn in mopwb[worksheet]['B2:%s2' % maxcolumn]:
                                for headercell in allcolumn:
                                    if headercell.value != None and headercell.value != 'ASBR':
                                        node = headercell.value
                                        if node not in allnode:
                                            allnode.append(node)
                                        headercheck = node+filename
                                        if headercheck not in headerlist:
                                            configdict.setdefault('%s_cfg' % node, []).append('')
                                            configdict.setdefault('%s_cfg' % node, []).append('#' * 100)
                                            configdict.setdefault('%s_cfg' % node, []).append('### Configuration Service:%s' % filename)
                                            configdict.setdefault('%s_cfg' % node, []).append('### Node:%s' % node)
                                            configdict.setdefault('%s_cfg' % node, []).append('#' * 100)
                                            configdict.setdefault('%s_cfg' % node, []).append('')
                                            headerlist.append(headercheck)
                                        columnletter = get_column_letter(columnnumber)
                                        for row in range(3,mopwb[worksheet].max_row + 1):
                                            if mopwb[worksheet]['%s%s'% (columnletter,row)].value == None:
                                                config = ''
                                            else:
                                                config = mopwb[worksheet]['%s%s' % (columnletter, row)].value
                                            configdict.setdefault('%s_cfg' % node, []).append(config)
                                            if worksheet == 'bdr-configuration' or worksheet == 'l2inter-as-configuration':
                                                if 'disable' in config:
                                                    if ' sap ' in config:
                                                        sapservice = config.split()[3]
                                                        if config not in nodeexecdict.setdefault('%s_disable_sap' % node, []):
                                                            nodeexecdict.setdefault('%s_disable_sap' % node, []).append(config)
                                                            nodeexecdict.setdefault('%s_enable_sap' % node, []).append(config.replace('admin-state disable','admin-state enable'))
                                                            nodeexecdict.setdefault('%s_show_sap' % node, []).append('/show service id %s base' % sapservice)
                                                            if ' vpls ' in config:
                                                                nodeexecdict.setdefault('%s_show_sap' % node, []).append('/show service id %s fdb detail' % sapservice)
                                                    if ' interface ' in config:
                                                        intvprn = config.split()[3]
                                                        interface = config.split()[5]
                                                        nodeexecdict.setdefault('%s_disable_intf' % node, []).append(config)
                                                        nodeexecdict.setdefault('%s_enable_intf' % node, []).append(config.replace('admin-state disable','admin-state enable'))
                                                        nodeexecdict.setdefault('%s_show_intf' % node, []).append('/show router service-name %s interface %s' % (intvprn,interface))
                                                        nodeexecdict.setdefault('%s_show_arp' % node, []).append('/show router service-name %s arp %s' % (intvprn,interface))
                                                        if '/show router service-name %s route-table' % intvprn not in nodeexecdict.setdefault('%s_show_route' % node, []):
                                                            nodeexecdict.setdefault('%s_show_route' % node, []).append('/show router service-name %s route-table' % intvprn)
                                                        if '/show router service-name %s static-route' % intvprn not in nodeexecdict.setdefault('%s_show_static' % node, []):
                                                            nodeexecdict.setdefault('%s_show_static' % node, []).append('/show router service-name %s static-route' % intvprn)
                                                    if ' bgp neighbor ' in config:
                                                        bgpvprn = config.split()[3]
                                                        bgpneighbor = config.split()[6]
                                                        nodeexecdict.setdefault('%s_disable_bgp_neighbor' % node, []).append(config)
                                                        nodeexecdict.setdefault('%s_enable_bgp_neighbor' % node, []).append(config.replace('admin-state disable','admin-state enable'))
                                                        nodeexecdict.setdefault('%s_show_bgp_neighbor' % node, []).append('/show router service-name %s bgp neighbor %s received-routes brief | match *' %(bgpvprn,bgpneighbor))
                                                        nodeexecdict.setdefault('%s_show_bgp_neighbor' % node, []).append('/show router service-name %s bgp neighbor %s advertised-routes brief' %(bgpvprn,bgpneighbor))
                                                if ' vrrp ' in config and ' admin-state ' in config:
                                                    vrrpvprn = config.split()[3]
                                                    vrrpinterface = config.split()[5]
                                                    nodeexecdict.setdefault('%s_show_vrrp' % node, []).append('/show router service-name %s vrrp instance | match %s' %(vrrpvprn,vrrpinterface))
                                                if 'bfd' in config and ' admin-state ' in config:
                                                    bfdvprn = config.split()[3]
                                                    bfdinterface = config.split()[5]
                                                    nodeexecdict.setdefault('%s_show_bfd' % node, []).append('/show router service-name %s bfd session | match %s' %(bfdvprn,bfdinterface))
                                                if '/configure port ' in config and ' admin-state ' in config:
                                                    port = config.split()[2]
                                                    if '/show port %s' % port not in nodeexecdict.setdefault('%s_show_port' % node, []):
                                                        nodeexecdict.setdefault('%s_show_port' % node, []).append('/show port %s' % port)
                                                    if config not in nodeexecdict.setdefault('%s_enable_port' % node, []):
                                                        nodeexecdict.setdefault('%s_enable_port' % node, []).append(config)
                                                    if config.replace('admin-state enable','admin-state disable') not in nodeexecdict.setdefault('%s_disable_port' % node, []):
                                                        nodeexecdict.setdefault('%s_disable_port' % node, []).append(config.replace('admin-state enable','admin-state disable'))
                                                if '/configure lag ' in config and ' admin-state ' in config:
                                                    slag = config.split()[2]
                                                    if '/show lag %s' % slag not in nodeexecdict.setdefault('%s_show_lag' % node, []):
                                                        nodeexecdict.setdefault('%s_show_lag' % node, []).append('/show lag %s' % slag)
                                                    if config not in nodeexecdict.setdefault('%s_enable_lag' % node, []):
                                                        nodeexecdict.setdefault('%s_enable_lag' % node, []).append(config)
                                                    if config.replace('admin-state enable','admin-state disable') not in nodeexecdict.setdefault('%s_disable_lag' % node, []):
                                                        nodeexecdict.setdefault('%s_disable_lag' % node, []).append(config.replace('admin-state enable','admin-state disable'))
                                                if ' hold-time ' in config:
                                                    nodeexecdict.setdefault('%s_hold-time' % node, []).append(config.lstrip('#'))
                                                if '"rVPLS' in config and 'service-id' in config:
                                                    rvplsservice = config.split()[3]
                                                    rvplsvlan = int(config.split()[-1][-4:])
                                                    nodeexecdict.setdefault('%s_enable_rvpls-bridge' % node, []).append('/configure service vpls %s sap lag-700:%s admin-state enable' %(rvplsservice,rvplsvlan))
                                                    nodeexecdict.setdefault('%s_disable_rvpls-bridge' % node, []).append('/configure service vpls %s sap lag-700:%s admin-state disable' %(rvplsservice,rvplsvlan))
                                                    nodeexecdict.setdefault('%s_delete_rvpls-bridge' % node, []).append('/configure service vpls %s delete sap lag-700:%s' %(rvplsservice,rvplsvlan))
                                                if 'ipv4' in config:
                                                    if 'primary address' in config:
                                                        ipint = config.split()[-1]
                                                    if 'prefix-length' in config:
                                                        intf = config.split()[5]
                                                        vrf = config.split()[3]
                                                        lengthint = config.split()[-1]
                                                        ipaddr = ipint + '/' + lengthint
                                                        nodeexecdict.setdefault('%s_prefix_adv_rr' % node, []).append(ipaddr)
                                                        if lengthint == '30':
                                                            for ip in IPNetwork(ipaddr):
                                                                if ip != IPNetwork(ipaddr).network and ip != IPNetwork(ipaddr).broadcast and ip != IPAddress(ipint):
                                                                    iptn_ping_cmd = 'ping routing-instance %s inet %s count 2' % (vrf.replace('"','').lstrip('L3_'),ip)
                                                                    ping_cmd = '/ping %s router-instance %s count 2 # intf:%s' % (ip,vrf,intf)
                                                                    if ping_cmd not in configdict.setdefault('%s_ping' % node, []):
                                                                        configdict.setdefault('%s_ping' % node, []).append(ping_cmd)
                                                                    if iptn_ping_cmd not in configdict.setdefault('%s_iptn_ping' % node, []):
                                                                        configdict.setdefault('%s_iptn_ping' % node, []).append(iptn_ping_cmd)
                                                        elif lengthint == '31':
                                                            for ip in IPNetwork(ipaddr):
                                                                if ip != IPAddress(ipint):
                                                                    iptn_ping_cmd = 'ping routing-instance %s inet %s count 2' % (vrf.replace('"','').lstrip('L3_'),ip)
                                                                    ping_cmd = '/ping %s router-instance %s count 2 # intf:%s' % (ip,vrf,intf)
                                                                    if ping_cmd not in configdict.setdefault('%s_ping' % node, []):
                                                                        configdict.setdefault('%s_ping' % node, []).append(ping_cmd)
                                                                    if iptn_ping_cmd not in configdict.setdefault('%s_iptn_ping' % node, []):
                                                                        configdict.setdefault('%s_iptn_ping' % node, []).append(iptn_ping_cmd)
                                                        else:
                                                            ip = IPNetwork(ipaddr)[4]
                                                            iptn_ping_cmd = 'ping routing-instance %s inet %s count 2' % (vrf.replace('"','').lstrip('L3_'),ip)
                                                            ping_cmd = '/ping %s router-instance %s count 2 # intf:%s' % (ip, vrf, intf)
                                                            if ping_cmd not in configdict.setdefault('%s_ping' % node, []):
                                                                configdict.setdefault('%s_ping' % node, []).append(ping_cmd)
                                                            if iptn_ping_cmd not in configdict.setdefault('%s_iptn_ping' % node, []):
                                                                configdict.setdefault('%s_iptn_ping' % node, []).append(iptn_ping_cmd)
                                                if 'ipv6' in config:
                                                    if 'prefix-length' in config:
                                                        intf = config.split()[5]
                                                        vrf = config.split()[3]
                                                        ipint = config.split()[8]
                                                        lengthint = config.split()[-1]
                                                        ipaddr = ipint + '/' + lengthint
                                                        nodeexecdict.setdefault('%s_prefix_adv_rr' % node, []).append(ipaddr)
                                                        ip = IPNetwork(ipaddr)[4]
                                                        iptn_ping_cmd = 'ping routing-instance %s inet6 %s count 2' % (vrf.replace('"','').lstrip('L3_'),ip)
                                                        ping_cmd = '/ping %s router-instance %s count 2 # intf:%s' % (ip, vrf, intf)
                                                        if ping_cmd not in configdict.setdefault('%s_ping' % node, []):
                                                            configdict.setdefault('%s_ping' % node, []).append(ping_cmd)
                                                        if iptn_ping_cmd not in configdict.setdefault('%s_iptn_ping' % node, []):
                                                            configdict.setdefault('%s_iptn_ping' % node, []).append(iptn_ping_cmd)
                                                if 'next-hop' in config and 'description' in config:
                                                    intf = config.split()[-1]
                                                    vrf = config.split()[3]
                                                    ipaddr = config.split()[6]
                                                    nodeexecdict.setdefault('%s_prefix_adv_rr' % node, []).append(ipaddr)
                                                    ip = config.split()[10].replace('"','')
                                                    if ':' in ip:
                                                        iptn_ping_cmd = 'ping routing-instance %s inet6 %s count 2' % (vrf.replace('"','').lstrip('L3_'),ip)
                                                    else:
                                                        iptn_ping_cmd = 'ping routing-instance %s inet %s count 2' % (vrf.replace('"','').lstrip('L3_'),ip)
                                                    ping_cmd = '/ping %s router-instance %s count 2 # intf:%s' % (ip, vrf, intf)
                                                    if ping_cmd not in configdict.setdefault('%s_ping' % node, []):
                                                        configdict.setdefault('%s_ping' % node, []).append(ping_cmd)
                                                    if iptn_ping_cmd not in configdict.setdefault('%s_iptn_ping' % node, []):
                                                        configdict.setdefault('%s_iptn_ping' % node, []).append(iptn_ping_cmd)
                                                if 'ingress filter' in config or 'ingress routed-override-filter' in config:
                                                    vrf = config.split()[3]
                                                    if vrf not in configdict.setdefault('%s_vrf' % node, []):
                                                        configdict.setdefault('%s_vrf' % node, []).append(vrf)
                                                    aclpol = config.split()[-1]
                                                    if aclpol not in configdict.setdefault('%s_%s_vrf_acl' % (node,vrf), []):
                                                        configdict.setdefault('%s_%s_vrf_acl' % (node,vrf), []).append(aclpol)
                                                        int_acl = config.split()[5]
                                                        configdict['int_acl_%s' % aclpol] = int_acl
                                            if worksheet == 'bdr-filter':
                                                if 'Filter policy for' in config:
                                                    sap = config.split(':', 2)[2].split()[0]
                                                    configdict.setdefault('%s_sap' % node, []).append(sap)
                                                if 'configure filter' in config:
                                                    filterpol = config.split()[3]
                                                    if filterpol not in configdict.setdefault('%s_%s_filter_policy' % (node,sap), []):
                                                        configdict.setdefault('%s_%s_filter_policy' % (node,sap), []).append(filterpol)
                                                    configdict.setdefault('%s_%s_%s_filter_policy_cfg' % (node,sap,filterpol), []).append(config)
                                                    if 'icmp' in config:
                                                        icmpentry = config.split()[5]
                                                        if icmpentry not in configdict.setdefault('%s_%s_%s_filter_icmp_entry' % (node,sap,filterpol), []):
                                                            configdict.setdefault('%s_%s_%s_filter_icmp_entry' % (node,sap,filterpol), []).append(icmpentry)
                                                if 'ip-filter' in config and 'src-ip address' in config:
                                                    host = config.split()[-1]
                                                    filterp = config.split()[3]
                                                    if host not in configdict.setdefault('%s_%s_source-ip' % (node,filterp), []):
                                                        configdict.setdefault('%s_%s_source-ip' % (node,filterp), []).append(host)
                                                    if filterp not in configdict.setdefault('%s_filter' % node, []):
                                                        configdict.setdefault('%s_filter' % node, []).append(filterp)
                                                if 'ipv6-filter' in config and 'src-ip address' in config:
                                                    host = config.split()[-1]
                                                    filterp = config.split()[3]
                                                    if host not in configdict.setdefault('%s_%s_source-ip' % (node,filterp), []):
                                                        configdict.setdefault('%s_%s_source-ip' % (node,filterp), []).append(host)
                                                    if filterp not in configdict.setdefault('%s_filter' % node, []):
                                                        configdict.setdefault('%s_filter' % node, []).append(filterp)
                                    columnnumber += 1

            elif '.txt' in filename:
                os.remove('./delta-cfg/'+filename)
        elif 'CE_filter_source_ip' in filename:
            os.remove('./delta-cfg/'+filename)
    aclwb = openpyxl.Workbook()
    aclwb.remove(aclwb['Sheet'])
    aclws = aclwb.create_sheet('CE_filter_src_ip')
    aclws['A1'] = 3
    aclws['A2'] = 'vrf'
    aclws['B2'] = 'interface'
    aclws['C2'] = 'ingress_filter_policy'
    aclws['D2'] = 'source_ip'
    aclws.auto_filter.ref = 'A2:D2'
    aclws.freeze_panes = aclws['A3']
    aclws.column_dimensions['A'].width = 25.0
    aclws.column_dimensions['B'].width = 40.0
    aclws.column_dimensions['C'].width = 50.0
    aclws.column_dimensions['D'].width = 25.0
    sourcedup = []
    for node in allnode:
        if '%s_vrf' % node in configdict:
            for vrf in configdict['%s_vrf' % node]:
                for filterp in configdict['%s_filter' % node]:
                    if filterp in configdict['%s_%s_vrf_acl' % (node,vrf)]:
                        for srcip in configdict['%s_%s_source-ip' % (node,filterp)]:
                            srccheck = vrf + srcip + filterp
                            if srccheck not in sourcedup:
                                sourcedup.append(srccheck)
                                aclrow = aclws['A1'].value
                                aclws['A%s' % aclrow] = vrf.replace('"','')
                                #print(vrf.replace('"',''))
                                aclws['B%s' % aclrow] = configdict['int_acl_%s' % filterp].replace('"','')
                                #print(configdict['int_acl_%s' % filterp].replace('"',''))
                                aclws['C%s' % aclrow] = filterp.replace('"','')
                                #print(filterp.replace('"',''))
                                aclws['D%s' % aclrow] = srcip
                                #print(srcip)
                                #configdict.setdefault('%s_iptn_ping' % node, []).append('ping routing-instance %s inet6 %s count 2' % (vrf.replace('"','').lstrip('L3_'),host))
                                aclrow += 1
                                aclws['A1'] = aclrow

    aclwb.save('./delta-cfg/CE_filter_source_ip_%s.xlsx' % extendfilename)


    for node in allnode:
        filterdup = []
        icmpentrycfg = []
        f = open('./delta-cfg/' + node + '_delta-configuration_%s.txt' % extendfilename, "a", encoding="utf-8")
        for cfg in configdict['%s_cfg' % node]:
            f.write(cfg+'\n')
        f.close()
        ex = open('./delta-cfg/' + node + '_execute-cmd_%s.txt' % extendfilename, "a", encoding="utf-8")
        if '%s_ping' % node in configdict:
            ex.write('# IPTN ping connected and next-hop' + '\n')
            ex.write('\n')
            for iptn_ping in configdict['%s_iptn_ping' % node]:
                ex.write(iptn_ping + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
            ex.write('# BDRT ping connected and next-hop' + '\n')
            ex.write('\n')
            for ping in configdict['%s_ping' % node]:
                ex.write(ping + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_disable_port' % node in nodeexecdict:
            ex.write('# disable port of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_disable_port' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_enable_port' % node in nodeexecdict:
            ex.write('# enable port of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_enable_port' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_port' % node in nodeexecdict:
            ex.write('# show port of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_port' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_disable_lag' % node in nodeexecdict:
            ex.write('# disable lag of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_disable_lag' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_enable_lag' % node in nodeexecdict:
            ex.write('# enable lag of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_enable_lag' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_lag' % node in nodeexecdict:
            ex.write('# show lag of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_lag' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_disable_sap' % node in nodeexecdict:
            ex.write('# disable sap of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_disable_sap' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_enable_sap' % node in nodeexecdict:
            ex.write('# enable sap of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_enable_sap' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_sap' % node in nodeexecdict:
            ex.write('# show L2 service of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_sap' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')

        if '%s_enable_rvpls-bridge' % node in nodeexecdict:
            ex.write('# enable bridge link sap of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_enable_rvpls-bridge' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')

        if '%s_disable_rvpls-bridge' % node in nodeexecdict:
            ex.write('# disable bridge link sap of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_disable_rvpls-bridge' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')

        if '%s_delete_rvpls-bridge' % node in nodeexecdict:
            ex.write('# delete bridge link sap of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_delete_rvpls-bridge' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')

        if '%s_disable_intf' % node in nodeexecdict:
            ex.write('# disable interface of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_disable_intf' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_enable_intf' % node in nodeexecdict:
            ex.write('# enable interface of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_enable_intf' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_intf' % node in nodeexecdict:
            ex.write('# show interface of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_intf' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_arp' % node in nodeexecdict:
            ex.write('# show arp table of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_arp' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_vrrp' % node in nodeexecdict:
            ex.write('# show vrrp interface of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_vrrp' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_bfd' % node in nodeexecdict:
            ex.write('# show bfd interface of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_bfd' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_disable_bgp_neighbor' % node in nodeexecdict:
            ex.write('# disable bgp neighbor of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_disable_bgp_neighbor' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_enable_bgp_neighbor' % node in nodeexecdict:
            ex.write('# enable bgp neighbor of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_enable_bgp_neighbor' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_bgp_neighbor' % node in nodeexecdict:
            ex.write('# show bgp neighbor of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_bgp_neighbor' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')

        if '%s_prefix_adv_rr' % node in nodeexecdict:
            ex.write('# show bgp route advertise to RR of %s' % node + '\n')
            ex.write('\n')
            for advrrprefix in nodeexecdict['%s_prefix_adv_rr' % node]:
                if ':' in advrrprefix:
                    advrrcmd = f"/show router bgp routes {advrrprefix} vpn-ipv6 hunt brief | match 'Community|Advertised'"
                else:
                    advrrcmd = f"/show router bgp routes {advrrprefix} vpn-ipv4 hunt brief | match 'Community|Advertised'"
                ex.write(advrrcmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')

        if '%s_show_static' % node in nodeexecdict:
            ex.write('# show static-route of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_static' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_show_route' % node in nodeexecdict:
            ex.write('# show route-table of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_show_route' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        if '%s_hold-time' % node in nodeexecdict:
            ex.write('# enable hold-time of %s' % node + '\n')
            ex.write('\n')
            for exec_cmd in nodeexecdict['%s_hold-time' % node]:
                ex.write(exec_cmd + '\n')
            ex.write('\n')
            ex.write('#' * 80 + '\n')
            ex.write('\n')
        ex.close()
        if '%s_sap' % node in configdict:
            for sap in configdict['%s_sap' % node]:
                for filter_policy in configdict['%s_%s_filter_policy' % (node,sap)]:
                    for particular_sap in configdict['%s_sap' % node]:
                        for particular_filter_policy in configdict['%s_%s_filter_policy' % (node, particular_sap)]:
                            if filter_policy == particular_filter_policy:
                                if configdict['%s_%s_%s_filter_policy_cfg' % (node,sap,filter_policy)] != configdict['%s_%s_%s_filter_policy_cfg' % (node, particular_sap, particular_filter_policy)]:
                                    print('!!!!! Found duplicate policy name but detail of entry not match !!!!!')
                                    print('Node:%s SAP:%s Filter Policy:%s and Node:%s SAP:%s Filter Policy:%s' %(node,sap,filter_policy,node, particular_sap, particular_filter_policy))
                                    print('!!!!! Please re-check filter policy !!!!!')
                                    print('#########################################')
                                    filterdup.append('!!!!! Found duplicate policy name but detail of entry not match !!!!!')
                                    filterdup.append('Node:%s SAP:%s Filter Policy:%s and Node:%s SAP:%s Filter Policy:%s' %(node,sap,filter_policy,node, particular_sap, particular_filter_policy))
                                    filterdup.append('!!!!! Please re-check filter policy !!!!!')
                                    filterdup.append('#########################################')
            if filterdup:
                fi = open('./delta-cfg/' + node + '_filter_duplicate_%s.txt' % extendfilename, "a", encoding="utf-8")
                for filter_alert in filterdup:
                    fi.write(filter_alert+'\n')
                fi.close()


            for sap in configdict['%s_sap' % node]:
                for filter_policy in configdict['%s_%s_filter_policy' % (node,sap)]:
                    if '%s_%s_%s_filter_icmp_entry' % (node,sap,filter_policy) in configdict:
                        for icmp_entry in configdict['%s_%s_%s_filter_icmp_entry' % (node,sap,filter_policy)]:
                            icmpdropcfg = '/configure filter ip-filter %s entry %s action drop' % (filter_policy,icmp_entry)
                            if icmpdropcfg not in icmpentrycfg:
                                icmpentrycfg.append(icmpdropcfg)

            #if icmpentrycfg:
            #    icmpentrycfg.append('')
            #    icmpentrycfg.append('#' * 80)
            #    icmpentrycfg.append('')
            #    for sap in configdict['%s_sap' % node]:
            #        for filter_policy in configdict['%s_%s_filter_policy' % (node,sap)]:
            #            if '%s_%s_%s_filter_icmp_entry' % (node,sap,filter_policy) in configdict:
            #                for icmp_entry in configdict['%s_%s_%s_filter_icmp_entry' % (node,sap,filter_policy)]:
            #                    icmpacceptcfg = '/configure filter ip-filter %s entry %s action accept' % (filter_policy,icmp_entry)
            #                    if icmpacceptcfg not in icmpentrycfg:
            #                        icmpentrycfg.append(icmpacceptcfg)
            #
            #    ic = open('./delta-cfg/' + node + '_ping_%s.txt' % extendfilename, "a", encoding="utf-8")
            #    for icmpencfg in icmpentrycfg:
            #        ic.write(icmpencfg+'\n')
            #    ic.close()


if __name__ == "__main__":
    main()