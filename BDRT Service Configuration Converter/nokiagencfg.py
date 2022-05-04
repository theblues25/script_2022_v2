import sys
import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from netaddr import IPNetwork, IPAddress

migratepath = '.\migration\\'
inputpath = '.\input\\'
iptncfgpath = '.\iptn-cfg\\'
extractpath = '.\extraction\\'
attrpath = 'attr-cfg\\'
l = open(migratepath + inputpath + "latestmigrationsummary.txt", 'r')
latestmisum = l.readlines()[0]
misumtime = latestmisum.split('\\')[2]
misum = openpyxl.load_workbook(filename=latestmisum)
time = misum['summary']['B1'].value
columnlist = ['D','G','J','M','P','S','V','Y','AB','AE','AH','AK','AN','AQ','AT','AW','AZ']

def createbdrconfigfile():
    row = 3
    bdrcfgwb = openpyxl.Workbook()
    bdrcfgwb.remove(bdrcfgwb['Sheet'])
    iptncfgws = bdrcfgwb.create_sheet('iptncfg')
    iptncfgws['A1'] = row
    iptncfgws.auto_filter.ref = 'A2:Q2'
    iptncfgws.freeze_panes = iptncfgws['A3']

    bdrconfigws = bdrcfgwb.create_sheet('bdr-configuration')
    bdrconfigws['A1'] = 'A'
    bdrconfigws['B1'] = row
    bdrconfigws['D1'] = 'A'
    bdrconfigws.auto_filter.ref = 'A2:Q2'
    bdrconfigws.freeze_panes = bdrconfigws['A3']
    bdrconfigws.column_dimensions['A'].width = 2.0
    bdrconfigws.column_dimensions['B'].width = 2.0

    bdrinterasws = bdrcfgwb.create_sheet('l2inter-as-configuration')
    bdrinterasws['A1'] = 'A' # for port
    bdrinterasws['B1'] = row # for pe row
    bdrinterasws['D1'] = 'A' # for pe column
    bdrinterasws['J1'] = row # for asbr row
    bdrinterasws.auto_filter.ref = 'A2:Q2'
    bdrinterasws.freeze_panes = bdrinterasws['A3']
    bdrinterasws.column_dimensions['A'].width = 2.0
    bdrinterasws.column_dimensions['B'].width = 2.0

    bdrpolicyws = bdrcfgwb.create_sheet('bdr-policy')
    bdrpolicyws['A1'] = 'A'
    bdrpolicyws['B1'] = row
    bdrpolicyws['D1'] = 'A'
    bdrpolicyws.auto_filter.ref = 'A2:Q2'
    bdrpolicyws.freeze_panes = bdrpolicyws['A3']
    bdrpolicyws.column_dimensions['A'].width = 2.0
    bdrpolicyws.column_dimensions['B'].width = 2.0

    bdrfilterws = bdrcfgwb.create_sheet('bdr-filter')
    bdrfilterws['A1'] = 'A'
    bdrfilterws['B1'] = row
    bdrfilterws['D1'] = 'A'
    bdrfilterws.auto_filter.ref = 'A2:Q2'
    bdrfilterws.freeze_panes = bdrfilterws['A3']
    bdrfilterws.column_dimensions['A'].width = 2.0
    bdrfilterws.column_dimensions['B'].width = 2.0

    return bdrcfgwb

def createlistservice():
    listserviceid = []
    listiptnnode = []
    miiptndict = {}
    for i in range(3,misum['port-lag'].max_row + 1):
        serviceid = misum['port-lag']['E%s' %i].value
        if serviceid != None:
            if serviceid not in listserviceid:
                listserviceid.append(serviceid)

    for id in listserviceid:
        miiptndict['%s_node' % id] = []
        for i in range(3,misum['port-lag'].max_row + 1):
            if id == misum['port-lag']['E%s' %i].value:
                iptnnode = misum['port-lag']['X%s' %i].value
                servicetype = misum['port-lag']['Y%s' % i].value
                servicename = misum['port-lag']['Z%s' % i].value
                iptnint = misum['port-lag']['AA%s' % i].value
                miiptndict['%s_type' % id] = servicetype
                if servicename not in miiptndict.setdefault('%s_service' % id, []):
                    miiptndict.setdefault('%s_service' % id, []).append(servicename)
                if iptnnode not in miiptndict.setdefault('%s_node' % id, []):
                    miiptndict.setdefault('%s_node' % id, []).append(iptnnode)
                miiptndict.setdefault('%s_%s_int' % (id,iptnnode), []).append(iptnint)

    return listserviceid, miiptndict

def createportdict(port,row):
    portdict = {}
    portdict['bdrnode'] = port['A%s' % row].value             #'nokia-node'
    portdict['bdrsystem'] = port['B%s' % row].value           #'system-ip'
    portdict['bdrservicetype'] = port['C%s' % row].value      #'service-type'
    portdict['bdrservicename'] = port['D%s' % row].value      #'service-name'
    portdict['bdrserviceid'] = port['E%s' % row].value        #'service-id'
    portdict['bdrport'] = port['F%s' % row].value             #'port/lag'
    portdict['bdrportencap'] = port['G%s' % row].value        #'port-encapsulation'
    portdict['bdrportstate'] = port['H%s' % row].value        #'admin-state'
    portdict['bdrportdesc'] = port['I%s' % row].value         #'physical-description'
    portdict['bdrportspeed'] = port['J%s' % row].value        #'speed'
    portdict['bdrportautonego'] = port['K%s' % row].value     #'auto-negotiation'
    portdict['bdrlag'] = port['L%s' % row].value              #'LAG'
    portdict['bgrlagprotocol'] = port['M%s' % row].value      #'LAG-protocol'
    portdict['bdrlagthreshold'] = port['N%s' % row].value     #'LAG-port-threshold'
    portdict['bdrsap'] = port['O%s' % row].value              #'SAP'
    portdict['bdrsapstate'] = port['P%s' % row].value         #'SAP-state'
    portdict['bdrsapdesc'] = port['Q%s' % row].value          #'SAP-description'
    portdict['bdrinfiltername'] = port['R%s' % row].value     #'input-filter-policy-name'
    portdict['bdrinfilterpolicy'] = port['S%s' % row].value   #'input-filter-configuration'
    portdict['bdroutfiltername'] = port['T%s' % row].value    #'output-filter-policy-name'
    portdict['bdroutfilterpolicy'] = port['U%s' % row].value  #'output-filter-configuration'
    portdict['bdrsapinqos'] = port['V%s' % row].value         #'sap-ingress qos'
    portdict['bdrsapoutqos'] = port['W%s' % row].value        #'sap-egress qos'
    portdict['iptnnode'] = port['X%s' % row].value            #'iptn-node'
    portdict['iptnservicetype'] = port['Y%s' % row].value     #'iptn-service-type'
    portdict['iptnservice'] = port['Z%s' % row].value         #'iptn-service'
    portdict['iptnint'] = port['AA%s' % row].value            #'iptn-interface'
    return portdict

def getiptnconfigline(extime,iptnnode):
    f = open(extractpath + attrpath + extime + '\\' + iptnnode +'_lines.txt', "r")
    nodeline = f.readlines()
    f.close()
    portname = nodeline[0].split('portname:')[-1].rstrip('\n').strip("[]").strip("'").split("', '")
    portline = nodeline[1].split('portlines:')[-1].rstrip('\n').strip("[]").split(', ')
    vprnname = nodeline[2].split('vprnname:')[-1].rstrip('\n').strip("[]").strip("'").split("', '")
    vprnline = nodeline[3].split('vprnline:')[-1].rstrip('\n').strip("[]").split(', ')
    l2vpnname = nodeline[4].split('l2vpnname:')[-1].rstrip('\n').strip("[]").strip("'").split("', '")
    l2vpnline = nodeline[5].split('l2vpnline:')[-1].rstrip('\n').strip("[]").split(', ')
    vplsname = nodeline[6].split('vplsname:')[-1].rstrip('\n').strip("[]").strip("'").split("', '")
    vplsline = nodeline[7].split('vplsline:')[-1].rstrip('\n').strip("[]").split(', ')
    bridgename = nodeline[8].split('bridgename:')[-1].rstrip('\n').strip("[]").strip("'").split("', '")
    bridgeline = nodeline[9].split('bridgeline:')[-1].rstrip('\n').strip("[]").split(', ')
    vswname = nodeline[10].split('vswname:')[-1].rstrip('\n').strip("[]").strip("'").split("', '")
    vswline = nodeline[11].split('vswline:')[-1].rstrip('\n').strip("[]").split(', ')
    prefixline = nodeline[13].split('prefixline:')[-1].rstrip('\n').strip("[]").split(', ')
    commuline = nodeline[15].split('commuline:')[-1].rstrip('\n').strip("[]").split(', ')
    policyline = nodeline[17].split('policyline:')[-1].rstrip('\n').strip("[]").split(', ')
    aclline = nodeline[19].split('aclline:')[-1].rstrip('\n').strip("[]").split(', ')
    qosline = nodeline[20].split('qosline:')[-1].rstrip('\n').strip("[]").split(', ')

    return portname,portline,vprnname,vprnline,l2vpnname,l2vpnline,vplsname,vplsline,bridgename,bridgeline,vswname,vswline,prefixline,commuline,policyline,aclline, qosline

def createiptncfgworksheet(serviceid,extime,miiptndict,bdrcfgwb):
    iptncfg = bdrcfgwb['iptncfg']
    iptncfg.column_dimensions['A'].width = 14.0
    column = 'B'
    for iptnnode in miiptndict['%s_node'%serviceid]:
        excelrow = iptncfg['A1'].value
        filterlist = []
        iptncfglines = []
        policylist = []
        iplist = []
        prefixlist = []
        commulist = []
        irblist = []
        vprnlist = []
        vlanlist = []
        configcheck = []
        iptncfg['%s2'%column] = iptnnode
        iptncfg.column_dimensions['%s' % column].width = 126.0
        iptncfg.column_dimensions['%s' % chr(ord(column) + 2)].width = 14.0
        #print('bdr-service-id %s iptnservice %s' %(serviceid, miiptndict['%s_service' %serviceid]))
        if iptnnode != None:
            portname, portline, vprnname, vprnline, l2vpnname, l2vpnline, vplsname, vplsline, bridgename, bridgeline, vswname, vswline, prefixline, commuline, policyline, aclline, qosline = getiptnconfigline(extime,iptnnode)
            f = open(iptncfgpath + iptnnode + ".txt", 'r',encoding="cp1252")
            lines = f.readlines()
            f.close()
            for a in range(len(lines)):
                lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
                lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
                lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
                iptncfglines.append(lines[a])
            iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'interface'
            excelrow += 1
            for iptnint in miiptndict['%s_%s_int' % (serviceid,iptnnode)]:
                #iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'interface'
                #excelrow += 1
                #iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'interface'
                #iptncfg['%s%s' % (column, excelrow)] = '#### iptn interface configuration ####'
                #excelrow += 1
                for pl in portline:
                    pl = int(pl)
                    if iptnint.split('.')[0] in iptncfglines[pl]:
                        if iptncfglines[pl] not in configcheck:
                            configcheck.append(iptncfglines[pl])
                            iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'interface'
                            iptncfg['%s%s' % (column, excelrow)] = iptncfglines[pl]
                            if re.search('family inet filter input',iptncfglines[pl]) or re.search('family inet6 filter input', iptncfglines[pl]):
                                filterlist.append(iptncfglines[pl].split(' ')[-1])
                            if re.search('family inet filter output',iptncfglines[pl]) or re.search('family inet6 filter output', iptncfglines[pl]):
                                filterlist.append(iptncfglines[pl].split(' ')[-1])
                            if re.search('family inet address',iptncfglines[pl]) or re.search('family inet6 address', iptncfglines[pl]):
                                iplist.append(iptncfglines[pl].split(' ')[8])
                            if re.search('vlan-id',iptncfglines[pl]):
                                vlanlist.append(iptncfglines[pl].split(' ')[-1])
                            excelrow += 1

                #iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'qos'
                #excelrow += 1
                #iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'qos'
                #iptncfg['%s%s' % (column, excelrow)] = '#### iptn interface qos configuration ####'
                #excelrow += 1
                for ql in qosline:
                    ql = int(ql)
                    if iptnint.split('.')[0] in iptncfglines[ql]:
                        if iptncfglines[ql] not in configcheck:
                            configcheck.append(iptncfglines[ql])
                            iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'qos'
                            iptncfg['%s%s' % (column, excelrow)] = iptncfglines[ql]
                            excelrow += 1


            #iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'service'
            #excelrow += 1
            #iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'service'
            #iptncfg['%s%s' % (column, excelrow)] = '#### iptn service configuration ####'
            #excelrow += 1
            if miiptndict['%s_type'%serviceid] == 'EPIPE':
                for el in l2vpnline:
                    el = int(el)
                    if iptncfglines[el].split()[2] in miiptndict['%s_service' % serviceid]:
                        if iptncfglines[el] not in configcheck:
                            configcheck.append(iptncfglines[el])
                            iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'l2vpn'
                            iptncfg['%s%s' % (column, excelrow)] = iptncfglines[el]
                            excelrow += 1
            elif miiptndict['%s_type'%serviceid] == 'VPLS':
                for vl in vplsline:
                    vl = int(vl)
                    if iptncfglines[vl].split()[2] in miiptndict['%s_service' % serviceid]:
                        if iptncfglines[vl] not in configcheck:
                            configcheck.append(iptncfglines[vl])
                            iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'vpls'
                            iptncfg['%s%s' % (column, excelrow)] = iptncfglines[vl]
                            excelrow += 1
            elif miiptndict['%s_type'%serviceid] == 'Virtual-Switch':
                for vswl in vswline:
                    vswl = int(vswl)
                    if iptncfglines[vswl].split()[2] in miiptndict['%s_service' % serviceid]:
                        if iptncfglines[vswl] not in configcheck:
                            configcheck.append(iptncfglines[vswl])
                            iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'vsw'
                            iptncfg['%s%s' % (column, excelrow)] = iptncfglines[vswl]
                            for vlan in vlanlist:
                                if 'irb.%s' %vlan in iptncfglines[vswl]:
                                    if iptncfglines[vswl].split(' ')[-1] not in irblist:
                                        irblist.append(iptncfglines[vswl].split(' ')[-1])
                            excelrow += 1
            elif miiptndict['%s_type'%serviceid] == 'Bridge-Domain':
                for brl in bridgeline:
                    brl = int(brl)
                    if iptncfglines[brl].split()[2] in miiptndict['%s_service' % serviceid]:
                        if iptncfglines[brl] not in configcheck:
                            iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'bridge-domain'
                            iptncfg['%s%s' % (column, excelrow)] = iptncfglines[brl]
                            if 'irb' in iptncfglines[brl]:
                                if iptncfglines[brl].split(' ')[-1] not in irblist:
                                    irblist.append(iptncfglines[brl].split(' ')[-1])
                                if iptncfglines[brl].split(' ')[-1].split('.')[-1] not in vlanlist:
                                    vlanlist.append(iptncfglines[brl].split(' ')[-1].split('.')[-1])
                            excelrow += 1
            elif miiptndict['%s_type'%serviceid] == 'VPRN':
                for vprnl in vprnline:
                    vprnl = int(vprnl)
                    if iptncfglines[vprnl].split()[2] in miiptndict['%s_service' % serviceid]:
                        if iptncfglines[vprnl] not in configcheck:
                            iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'vprn'
                            iptncfg['%s%s' % (column, excelrow)] = iptncfglines[vprnl]
                            if 'import' in iptncfglines[vprnl] or 'export' in iptncfglines[vprnl]:
                                if iptncfglines[vprnl].split(' ')[-1] not in policylist:
                                    policylist.append(iptncfglines[vprnl].split(' ')[-1])
                            if 'next-hop' in iptncfglines[vprnl]:
                                n = iptncfglines[vprnl].split(' ')[-1]
                                for ip in iplist:
                                    if IPAddress(n) in IPNetwork(ip):
                                        iptncfg['%s%s' % (column, excelrow)].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                            elif 'neighbor' in iptncfglines[vprnl]:
                                n = iptncfglines[vprnl].split(' ')[8]
                                for ip in iplist:
                                    if IPAddress(n) in IPNetwork(ip):
                                        iptncfg['%s%s' % (column, excelrow)].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                            excelrow += 1


            #if irblist:
            #    iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'irb interface'
            #    excelrow += 1
            #    iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'irb interface'
            #    iptncfg['%s%s' % (column, excelrow)] = '#### iptn interface irb configuration ####'
            #    excelrow += 1
            for irb in irblist:
                for pl in portline:
                    pl = int(pl)
                    for vlan in vlanlist:
                        if '%s unit %s'%(irb.split('.')[0],vlan) in iptncfglines[pl]:
                            if iptncfglines[pl] not in configcheck:
                                configcheck.append(iptncfglines[pl])
                                iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'irb interface'
                                iptncfg['%s%s' % (column, excelrow)] = iptncfglines[pl]
                                if re.search('family inet filter input',iptncfglines[pl]) or re.search('family inet6 filter input', iptncfglines[pl]):
                                    filterlist.append(iptncfglines[pl].split(' ')[-1])
                                if re.search('family inet filter output',iptncfglines[pl]) or re.search('family inet6 filter output', iptncfglines[pl]):
                                    filterlist.append(iptncfglines[pl].split(' ')[-1])
                                if re.search('family inet address',iptncfglines[pl]) or re.search('family inet6 address', iptncfglines[pl]):
                                    iplist.append(iptncfglines[pl].split(' ')[8])
                                excelrow += 1

                #iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'irb qos'
                #excelrow += 1
                #iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'irb qos'
                #iptncfg['%s%s' % (column, excelrow)] = '#### iptn interface irb qos configuration ####'
                #excelrow += 1
                for ql in qosline:
                    ql = int(ql)
                    if '%s unit %s'%(irb.split('.')[0],irb.split('.')[1]) in iptncfglines[ql]:
                        iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'qos'
                        iptncfg['%s%s' % (column, excelrow)] = iptncfglines[ql]
                        excelrow += 1
                for vpnl in vprnline:
                    vpnl = int(vpnl)
                    if irb in iptncfglines[vpnl]:
                        if iptncfglines[vpnl].split()[2] not in vprnlist:
                            vprnlist.append(iptncfglines[vpnl].split()[2])

            #if vprnlist:
            #    iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'irb vprn'
            #    excelrow += 1
            #    iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'irb vprn'
            #    iptncfg['%s%s' % (column, excelrow)] = '#### iptn irb vprn configuration ####'
            for vpn in vprnlist:
                for vprnl in vprnline:
                    vprnl = int(vprnl)
                    if vpn == iptncfglines[vprnl].split()[2]:
                        iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'vprn'
                        iptncfg['%s%s' % (column, excelrow)] = iptncfglines[vprnl]
                        if 'import' in iptncfglines[vprnl] or 'export' in iptncfglines[vprnl]:
                            if iptncfglines[vprnl].split(' ')[-1] not in policylist:
                                policylist.append(iptncfglines[vprnl].split(' ')[-1])
                        if 'next-hop' in iptncfglines[vprnl]:
                            n = iptncfglines[vprnl].split(' ')[-1]
                            for ip in iplist:
                                if IPAddress(n) in IPNetwork(ip):
                                    iptncfg['%s%s' % (column, excelrow)].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                        elif 'neighbor' in iptncfglines[vprnl]:
                            n = iptncfglines[vprnl].split(' ')[8]
                            for ip in iplist:
                                if IPAddress(n) in IPNetwork(ip):
                                    iptncfg['%s%s' % (column, excelrow)].fill = PatternFill(start_color='FFfff957', end_color='FFfff957',fill_type='solid')
                        excelrow += 1

            #if policylist:
            #    iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'policy'
            #    excelrow += 1
            #    iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'policy'
            #    iptncfg['%s%s' % (column, excelrow)] = '#### iptn policy configuration ####'
            #    excelrow += 1
            for policy in policylist:
                for pol in policyline:
                    pol = int(pol)
                    if policy == iptncfglines[pol].split(' ')[3]:
                        iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'policy'
                        iptncfg['%s%s' % (column, excelrow)] = iptncfglines[pol]
                        if 'prefix-list' in iptncfglines[pol]:
                            if iptncfglines[pol].split(' ')[-1] not in prefixlist:
                                prefixlist.append(iptncfglines[pol].split(' ')[-1])
                        elif 'community' in iptncfglines[pol]:
                            if iptncfglines[pol].split(' ')[-1] not in commulist:
                                commulist.append(iptncfglines[pol].split(' ')[-1])
                        excelrow += 1


            for prefix in prefixlist:
                for prl in prefixline:
                    prl = int(prl)
                    if prefix == iptncfglines[prl].split(' ')[3]:
                        iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'prefix-list'
                        iptncfg['%s%s' % (column, excelrow)] = iptncfglines[prl]
                        excelrow += 1

            for commu in commulist:
                for col in commuline:
                    col = int(col)
                    if commu == iptncfglines[col].split(' ')[3]:
                        iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'community'
                        iptncfg['%s%s' % (column, excelrow)] = iptncfglines[col]
                        excelrow += 1

            for filter in filterlist:
                for fl in aclline:
                    fl = int(fl)
                    if filter in iptncfglines[fl]:
                        iptncfg['%s%s' % (chr(ord(column) - 1), excelrow)] = 'filter'
                        iptncfg['%s%s' % (column, excelrow)] = iptncfglines[fl]
                        excelrow += 1

            column = chr(ord(column) + 3)
            #print(iptncfglines)

def portconfig(bdrconfiglist,portdict):

    bdrconfiglist.append('')
    bdrconfiglist.append('### BDR Port Configuration ###')
    bdrconfiglist.append('')
    if 'c' in portdict['bdrport']:
        bdrconfiglist.append('/configure port %s description "%s"' % (portdict['bdrport'][:-2], portdict['bdrportdesc']))
        bdrconfiglist.append('/configure port %s admin-state enable' % portdict['bdrport'][:-2])
        bdrconfiglist.append('/configure port %s connector breakout c1-100g' % portdict['bdrport'][:-2])
        bdrconfiglist.append('')
    bdrconfiglist.append('/configure port %s description "%s"' % (portdict['bdrport'], portdict['bdrportdesc']))
    if 'auto-negotiation' == portdict['bdrportautonego'] and 'ge' in portdict['iptnint']:
        bdrconfiglist.append('/configure port %s ethernet autonegotiate limited' % portdict['bdrport'])
    elif 'no-auto-negotiation' == portdict['bdrportautonego'] and 'ge' in portdict['iptnint']:
        bdrconfiglist.append('/configure port %s ethernet autonegotiate false' % portdict['bdrport'])
    bdrconfiglist.append('/configure port %s ethernet mode access' % portdict['bdrport'])
    bdrconfiglist.append('/configure port %s ethernet encap-type %s' % (portdict['bdrport'],portdict['bdrportencap']))
    bdrconfiglist.append('/configure port %s ethernet mtu 9212' % portdict['bdrport'])
    if '1g' == portdict['bdrportspeed'] and 'ge' in portdict['iptnint']:
        bdrconfiglist.append('/configure port %s ethernet speed 1000' % portdict['bdrport'])
    elif '100m' == portdict['bdrportspeed'] and 'ge' in portdict['iptnint']:
        bdrconfiglist.append('/configure port %s ethernet speed 100' % portdict['bdrport'])
    if '/c' in portdict['bdrport'] or 'esat' in portdict['bdrport']:
        bdrconfiglist.append('/configure port %s ethernet down-on-internal-error tx-laser on' % portdict['bdrport'])
    else:
        bdrconfiglist.append('/configure port %s ethernet down-on-internal-error tx-laser off' % portdict['bdrport'])
    bdrconfiglist.append('/configure port %s ethernet crc-monitor signal-degrade threshold 5' % portdict['bdrport'])
    bdrconfiglist.append('/configure port %s ethernet crc-monitor signal-degrade multiplier 5' % portdict['bdrport'])
    bdrconfiglist.append('/configure port %s ethernet symbol-monitor signal-degrade threshold 5' % portdict['bdrport'])
    bdrconfiglist.append('/configure port %s ethernet symbol-monitor signal-degrade multiplier 5' % portdict['bdrport'])
    bdrconfiglist.append('/configure port %s ethernet egress port-scheduler-policy policy-name "AIS_BDRT_Access_Scheduler"' % portdict['bdrport'])
    bdrconfiglist.append('/configure port %s admin-state enable' % portdict['bdrport'])

    return bdrconfiglist


def bdrportlagconfigworksheet(serviceid,miportdict,lastbdrnode,bdrcfgwb,portcheck,bdrfilterdict):
    bdrportws = bdrcfgwb['bdr-configuration']
    bdrinterasportws = bdrcfgwb['l2inter-as-configuration']
    bdrpolicyws = bdrcfgwb['bdr-policy']
    bdrfilterws = bdrcfgwb['bdr-filter']
    bdrconfiglist = []
    bdrfiltercfglist = []
    remarklist = []
    lagmemberlist = []
    #print('current node  :%s' %miportdict['bdrnode'])
    #print('last bdr node :%s' %lastbdrnode)
    if lastbdrnode != miportdict['bdrnode']:
        column = bdrportws['A1'].value
        column = chr(ord(column) + 3)
        bdrportws['A1'] = column
        bdrinterasportws['A1'] = column
        bdrportws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
        bdrportws.column_dimensions['%s' % column].width = 120.0
        bdrinterasportws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
        bdrinterasportws.column_dimensions['%s' % column].width = 120.0
        bdrpolicyws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
        bdrpolicyws.column_dimensions['%s' % column].width = 120.0
        bdrfilterws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
        bdrfilterws.column_dimensions['%s' % column].width = 120.0
    elif lastbdrnode:
        column = bdrportws['A1'].value

    bdrportws['%s2' % column] = miportdict['bdrnode']
    bdrpolicyws['%s2' % column] = miportdict['bdrnode']
    bdrfilterws['%s2' % column] = miportdict['bdrnode']
    bdrinterasportws['%s2' % column] = miportdict['bdrnode']
    #print('node %s lag %s' %(miportdict['bdrnode'],miportdict['bdrport']))
    if '%s%s' % (miportdict['bdrnode'], miportdict['bdrport']) not in portcheck:
        portcheck.append('%s%s' % (miportdict['bdrnode'], miportdict['bdrport']))
        if re.search('lag',miportdict['bdrport']):
            for i in range(3, misum['port-lag'].max_row + 1):
                if misum['port-lag']['L%s' % i].value != None and miportdict['bdrnode'] == misum['port-lag']['A%s' % i].value:
                    if not re.search('lag', misum['port-lag']['F%s' % i].value) and miportdict['bdrport'] == misum['port-lag']['L%s' % i].value and misum['port-lag']['C%s' % i].value == 'lag-member':
                        lagmemberdict = createportdict(misum['port-lag'],i)
                        lagmemberdict['bdrportencap'] = miportdict['bdrportencap']
                        lagmemberlist.append(lagmemberdict['bdrport'])
                        bdrconfiglist = portconfig(bdrconfiglist,lagmemberdict)
            lagnumber = miportdict['bdrport'].split('-')[-1]
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR LAG Configuration ###')
            if miportdict['bdrportdesc'] != None:
                bdrconfiglist.append('/configure lag %s description "%s"' % (lagnumber,miportdict['bdrportdesc'].replace('ae','LAG')))
            bdrconfiglist.append('/configure lag %s encap-type %s' % (lagnumber, miportdict['bdrportencap']))
            bdrconfiglist.append('/configure lag %s mode access' % lagnumber)
            if miportdict['bdrlagthreshold'] != None:
                if miportdict['bdrlagthreshold'] == 0 or miportdict['bdrlagthreshold'] == '0':
                    pass
                else:
                    bdrconfiglist.append('/configure lag %s port-threshold value %s' % (lagnumber,miportdict['bdrlagthreshold']))
                    bdrconfiglist.append('/configure lag %s port-threshold action down' % lagnumber)
            if miportdict['bgrlagprotocol'] == 'lacp':
                bdrconfiglist.append('/configure lag %s lacp mode active' % lagnumber)
                bdrconfiglist.append('/configure lag %s lacp administrative-key %s' % (lagnumber, lagnumber))
                bdrconfiglist.append('/configure lag %s lacp-mux-control independent' % lagnumber)
            bdrconfiglist.append('/configure lag %s access adapt-qos mode link' % lagnumber)
            for lagport in lagmemberlist:
                bdrconfiglist.append('/configure lag %s port %s' % (lagnumber, lagport))
            bdrconfiglist.append('/configure lag %s admin-state enable' % lagnumber)
        else:
            bdrconfiglist = portconfig(bdrconfiglist, miportdict)

    for i in bdrconfiglist:
        if 'LAG' in i or 'lag' in i:
            remarklist.append('lag')
        else:
            remarklist.append('port')

    if miportdict['bdrinfilterpolicy'] != None:
        bdrfiltercfglist.append('')
        miportdict['bdrinfilterpolicy'] = miportdict['bdrinfilterpolicy'].strip('.')
        f = open(migratepath + misumtime + miportdict['bdrinfilterpolicy'], "r")
        lines = f.readlines()
        f.close()
        for a in range(len(lines)):
            lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
            lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
            lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
            #print(lines[a])
            if 'prefix-list' in lines[a]:
                if '%s_prefix' % miportdict['bdrnode'] in bdrfilterdict:
                    for existingprefixlist in bdrfilterdict['%s_prefix' % miportdict['bdrnode']]:
                        if existingprefixlist != None:
                            #print(existingprefixlist)
                            bdrprefixlist = bdrfilterdict['%s_pr_%s' % (miportdict['bdrnode'], existingprefixlist)]
                            #print(bdrprefixlist)
                            lines[a] = lines[a].replace(existingprefixlist,bdrprefixlist)
            if 'port-list' in lines[a]:
                if '%s_port' % miportdict['bdrnode'] in bdrfilterdict:
                    for existingportlist in bdrfilterdict['%s_port' % miportdict['bdrnode']]:
                        if existingportlist != None:
                            #print(existingportlist)
                            bdrportlist = bdrfilterdict['%s_po_%s' % (miportdict['bdrnode'], existingportlist)]
                            #print(bdrportlist)
                            lines[a] = lines[a].replace(existingportlist, bdrportlist)

            bdrfiltercfglist.append(lines[a])

    if miportdict['bdroutfilterpolicy'] != None:
        bdrfiltercfglist.append('')
        miportdict['bdroutfilterpolicy'] = miportdict['bdroutfilterpolicy'].strip('.')
        f = open(migratepath + misumtime + miportdict['bdroutfilterpolicy'] , "r")
        lines = f.readlines()
        f.close()
        for a in range(len(lines)):
            lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
            lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
            lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
            #print(lines[a])
            if 'prefix-list' in lines[a]:
                if '%s_prefix' % miportdict['bdrnode'] in bdrfilterdict:
                    for existingprefixlist in bdrfilterdict['%s_prefix' % miportdict['bdrnode']]:
                        if existingprefixlist != None:
                            #print(existingprefixlist)
                            bdrprefixlist = bdrfilterdict['%s_pr_%s' % (miportdict['bdrnode'], existingprefixlist)]
                            #print(bdrprefixlist)
                            lines[a] = lines[a].replace(existingprefixlist,bdrprefixlist)
            if 'port-list' in lines[a]:
                if '%s_port' % miportdict['bdrnode'] in bdrfilterdict:
                    for existingportlist in bdrfilterdict['%s_port' % miportdict['bdrnode']]:
                        if existingportlist != None:
                            #print(existingportlist)
                            bdrportlist = bdrfilterdict['%s_po_%s' % (miportdict['bdrnode'], existingportlist)]
                            #print(bdrportlist)
                            lines[a] = lines[a].replace(existingportlist, bdrportlist)

            bdrfiltercfglist.append(lines[a])

    if bdrportws['%s1' % chr(ord(column) - 1)].value == None:
        portwsrow = bdrportws['B1'].value
    else:
        portwsrow = bdrportws['%s1' % chr(ord(column) - 1)].value

    if bdrfilterws['%s1'%(chr(ord(column) - 1))].value == None:
        bdrfilterrow = bdrfilterws['B1'].value
    else:
        bdrfilterrow = bdrfilterws['%s1' % (chr(ord(column) - 1))].value

    for line in range(len(bdrconfiglist)):
        bdrportws['%s%s' % (chr(ord(column) - 1), portwsrow)] = remarklist[line]
        bdrportws['%s%s' % (column,portwsrow)] = bdrconfiglist[line]
        bdrinterasportws['%s%s' % (chr(ord(column) - 1), portwsrow)] = remarklist[line]
        bdrinterasportws['%s%s' % (column,portwsrow)] = bdrconfiglist[line]
        portwsrow +=1

    for line in range(len(bdrfiltercfglist)):
        bdrfilterws['%s%s' % (column, bdrfilterrow)] = bdrfiltercfglist[line]
        bdrfilterrow += 1

    lastbdrnode = miportdict['bdrnode']
    bdrportws['%s1' % chr(ord(column) - 1)] = portwsrow
    bdrinterasportws['%s1' % chr(ord(column) - 1)] = portwsrow
    bdrfilterws['%s1' % (chr(ord(column) - 1))] = bdrfilterrow
    return portcheck, lastbdrnode

def createl2sdpdict(sdpws, row):
    sdpdict = {}
    sdpdict['bdrnode'] = sdpws['A%s' % row].value # 'nokia-node'
    sdpdict['bdrservicetype'] = sdpws['B%s' % row].value # 'service-type'
    sdpdict['bdrservice'] = sdpws['C%s' % row].value # 'service-name'
    sdpdict['bdrserviceid'] = sdpws['D%s' % row].value # 'service-id'
    sdpdict['bdrfarendnode'] = sdpws['E%s' % row].value # 'farend-node'
    sdpdict['bdrfarendip'] = sdpws['F%s' % row].value # 'farend-ip'
    sdpdict['bdrsdp'] = sdpws['G%s' % row].value # 'sdp-number'
    sdpdict['bdrsdpdesc'] = sdpws['H%s' % row].value # 'sdp-description'
    sdpdict['bdrvcid'] = sdpws['I%s' % row].value # 'vc-id'
    sdpdict['bdrcusid'] = sdpws['J%s' % row].value # 'customer-id'
    sdpdict['bdrentropy'] = sdpws['K%s' % row].value  # 'entropy'
    return sdpdict

def createl2sdpinterasdict(sdpinterasws, row):
    sdpinterasdict = {}
    sdpinterasdict['bdrnode'] = sdpinterasws['A%s' %row].value #'PE-node'
    sdpinterasdict['bdrservicetype'] = sdpinterasws['B%s' %row].value #'service-type'
    sdpinterasdict['bdrservice'] = sdpinterasws['C%s' %row].value #'service-name'
    sdpinterasdict['bdrserviceid'] = sdpinterasws['D%s' %row].value #'service-id'
    sdpinterasdict['bdrpesdp'] = sdpinterasws['E%s' %row].value #'PE-sdp-number'
    sdpinterasdict['bdrpesdpdesc'] = sdpinterasws['F%s' %row].value #'PE-sdp-description'
    sdpinterasdict['bdrasbrnode'] = sdpinterasws['G%s' %row].value #'asbr-node'
    sdpinterasdict['bdrasbrip'] = sdpinterasws['H%s' %row].value #'asbr-ip'
    sdpinterasdict['bdrasbrsdp'] = sdpinterasws['I%s' %row].value #'asbr-sdp-number'
    sdpinterasdict['bdrasbrsdpdesc'] = sdpinterasws['J%s' %row].value #'asbr-sdp-description'
    sdpinterasdict['bdrvcid'] = sdpinterasws['K%s' %row].value #'vc-id'
    sdpinterasdict['asbrsdpprecedence'] = sdpinterasws['L%s' %row].value #'sdp-precedence'
    sdpinterasdict['asbrendpoint'] = sdpinterasws['M%s' %row].value #'endpoint-name'
    sdpinterasdict['asbrinteraslag'] = sdpinterasws['N%s' %row].value #'inter-as-lag'
    sdpinterasdict['asbrmdindex'] = sdpinterasws['O%s' %row].value #'md-index'
    sdpinterasdict['asbrmdname'] = sdpinterasws['P%s' %row].value #'md-name'
    sdpinterasdict['asbrmaindex'] = sdpinterasws['Q%s' %row].value #'ma-index'
    sdpinterasdict['asbrbridgeid'] = sdpinterasws['R%s' %row].value #'bridge-id'
    sdpinterasdict['asbrbridgevlan'] = sdpinterasws['S%s' %row].value #'bridge-vlan'
    sdpinterasdict['asbrmepid'] = sdpinterasws['T%s' %row].value #'local-mep-id'
    sdpinterasdict['asbriptnmepid'] = sdpinterasws['U%s' %row].value #'remote-mep-id'
    sdpinterasdict['bdrcusid'] = sdpinterasws['V%s' %row].value #'customer-id'
    sdpinterasdict['bdrentropy'] = sdpinterasws['W%s' %row].value #'entropy'
    return sdpinterasdict

def createl2evpninterasdict(l2evpninterws, row):
    evpninterasdict = {}
    evpninterasdict['bdrnode'] = l2evpninterws['A%s' % row].value  #'nokia-node'
    evpninterasdict['bdrsystem'] = l2evpninterws['B%s' % row].value  #'system-ip'
    evpninterasdict['bdrservice'] = l2evpninterws['C%s' % row].value  #'service-name'
    evpninterasdict['bdrserviceid'] = l2evpninterws['D%s' % row].value  #'service-id'
    evpninterasdict['bdrpeethseg'] = l2evpninterws['E%s' % row].value  #'ethernet-segment'
    evpninterasdict['bdrpeesi'] = l2evpninterws['F%s' % row].value  #'esi'
    evpninterasdict['bdrpepreference'] = l2evpninterws['G%s' % row].value  #'preference'
    evpninterasdict['bdrpeport'] = l2evpninterws['H%s' % row].value  #'port/lag'
    evpninterasdict['bdrpevlan'] = l2evpninterws['I%s' % row].value  #'vlan'
    evpninterasdict['bdrpeacname'] = l2evpninterws['J%s' % row].value  #'pe-ac-name'
    evpninterasdict['bdrpeethtag'] = l2evpninterws['K%s' % row].value  #'pe-eth-tag'
    evpninterasdict['bdrasbracname'] = l2evpninterws['L%s' % row].value  #'asbr-ac-name'
    evpninterasdict['bdrasbrethtag'] = l2evpninterws['M%s' % row].value  #'asbr-eth-tag'
    evpninterasdict['bdrevi'] = l2evpninterws['N%s' % row].value  #'evi'
    evpninterasdict['asbrnode'] = l2evpninterws['O%s' % row].value  #'asbr-node'
    evpninterasdict['asbrethseg'] = l2evpninterws['P%s' % row].value  #'asbr-ethernet-segment'
    evpninterasdict['asbresi'] = l2evpninterws['Q%s' % row].value  #'asbr-esi'
    evpninterasdict['asbrpreference'] = l2evpninterws['R%s' % row].value  #'asbr-preference'
    evpninterasdict['asbrinteraslag'] = l2evpninterws['S%s' % row].value  #'inter-as-lag'
    evpninterasdict['asbrmdindex'] = l2evpninterws['T%s' % row].value  #'md-index'
    evpninterasdict['asbrmdname'] = l2evpninterws['U%s' % row].value  #'md-name'
    evpninterasdict['asbrmaindex'] = l2evpninterws['V%s' % row].value  #'ma-index'
    evpninterasdict['asbrbridgeid'] = l2evpninterws['W%s' % row].value  #'bridge-id'
    evpninterasdict['asbrbridgevlan'] = l2evpninterws['X%s' % row].value  #'bridge-vlan'
    evpninterasdict['asbrlocalmep'] = l2evpninterws['Y%s' % row].value  #'local-mep-id'
    evpninterasdict['asbriptnmep'] = l2evpninterws['Z%s' % row].value  #'remote-mep-id'
    evpninterasdict['bdrcustid'] = l2evpninterws['AA%s' % row].value #'customer-id'
    evpninterasdict['bdrentropy'] = l2evpninterws['AB%s' % row].value  # 'entropy'
    return evpninterasdict

def bdrl2sdpworksheet(serviceid,misdpdict,lastbdrnode, bdrcfgwb):
    bdrl2sdpws = bdrcfgwb['bdr-configuration']
    bdrconfiglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print('column %s' %bdrl2sdpws['A1'].value)
    if lastbdrnode != misdpdict['bdrnode']:
        for c in columnlist:
            if bdrl2sdpws['%s2'%c].value != None:
                nodecolumnlist.append(bdrl2sdpws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if misdpdict['bdrnode'] == nodecolumnlist[n]:
                #print(existcolumn[n])
                #print(misdpdict['bdrnode'])
                #print(misdpdict['bdrservice'])
                column = existcolumn[n]
                bdrl2sdpws['D1'] = existcolumn[-1]
                #print(column)
        if misdpdict['bdrnode'] not in nodecolumnlist:
            bdrl2sdpws['D1'] = existcolumn[-1]
            column = bdrl2sdpws['D1'].value
            column = chr(ord(column) + 3)
            bdrl2sdpws['D1'] = column
            bdrl2sdpws['%s2' % column] = misdpdict['bdrnode']
            bdrl2sdpws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrl2sdpws.column_dimensions['%s' % column].width = 120.0

        if 'EPIPE' == misdpdict['bdrservicetype']:
            bdrconfiglist.append('')
            #print(misdpdict['bdrservice'])
            #print(misdpdict['bdrservicetype'])
            bdrconfiglist.append('### BDR EPIPE service and sdp configuration to remote BDR node ###')
            bdrconfiglist.append('/configure service epipe "%s" admin-state enable' % misdpdict['bdrservice'])
            bdrconfiglist.append('/configure service epipe "%s" description "%s"' %(misdpdict['bdrservice'],misdpdict['bdrservice']))
            bdrconfiglist.append('/configure service epipe "%s" service-id %s' % (misdpdict['bdrservice'],misdpdict['bdrserviceid']))
            bdrconfiglist.append('/configure service epipe "%s" customer "%s"' % (misdpdict['bdrservice'], misdpdict['bdrcusid']))
            bdrconfiglist.append('/configure service epipe "%s" service-mtu 9194' % misdpdict['bdrservice'])

        elif 'VPLS' == misdpdict['bdrservicetype']:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR VPLS service and sdp configuration to remote BDR node ###')
            bdrconfiglist.append('/configure service vpls "%s" admin-state enable' % misdpdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" description "%s"' % (misdpdict['bdrservice'],misdpdict['bdrservice']))
            bdrconfiglist.append('/configure service vpls "%s" service-id %s' % (misdpdict['bdrservice'],misdpdict['bdrserviceid']))
            bdrconfiglist.append('/configure service vpls "%s" customer "%s"' % (misdpdict['bdrservice'], misdpdict['bdrcusid']))
            bdrconfiglist.append('/configure service vpls "%s" service-mtu 9194' % misdpdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" mac-flush tldp send-on-failure true' % misdpdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" fdb mac-move admin-state enable' % misdpdict['bdrservice'])

        elif 'rVPLS' == misdpdict['bdrservicetype']:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR rVPLS service and sdp configuration to remote BDR node ###')
            bdrconfiglist.append('/configure service vpls "%s" admin-state enable' % misdpdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" description "%s"' % (misdpdict['bdrservice'],misdpdict['bdrservice']))
            bdrconfiglist.append('/configure service vpls "%s" service-id %s' % (misdpdict['bdrservice'],misdpdict['bdrserviceid']))
            bdrconfiglist.append('/configure service vpls "%s" customer "%s"' % (misdpdict['bdrservice'], misdpdict['bdrcusid']))
            bdrconfiglist.append('/configure service vpls "%s" service-mtu 9194' % misdpdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" mac-flush tldp send-on-failure true' % misdpdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" fdb mac-move admin-state enable' % misdpdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" routed-vpls' % misdpdict['bdrservice'])

    elif lastbdrnode:
        for c in columnlist:
            if bdrl2sdpws['%s2' % c].value != None:
                nodecolumnlist.append(bdrl2sdpws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if misdpdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrl2sdpws['D1'] = existcolumn[-1]

    if 'EPIPE' == misdpdict['bdrservicetype']:
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s admin-state enable' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s description "%s"' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid'],misdpdict['bdrsdpdesc']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s entropy-label' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s control-word true' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        for r in bdrconfiglist:
            remarklist.append('epipe')

    elif 'VPLS' == misdpdict['bdrservicetype']:
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s admin-state enable' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s description "%s"' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid'],misdpdict['bdrsdpdesc']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s entropy-label' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s control-word true' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        for r in bdrconfiglist:
            remarklist.append('vpls')

    elif 'rVPLS' == misdpdict['bdrservicetype']:
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s admin-state enable' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s description "%s"' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid'],misdpdict['bdrsdpdesc']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s entropy-label' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s control-word true' % (misdpdict['bdrservice'],misdpdict['bdrsdp'], misdpdict['bdrvcid']))
        for r in bdrconfiglist:
            remarklist.append('rvpls')

    if bdrl2sdpws['%s1' % chr(ord(column) - 1)].value == None:
        l2sdpwsrow = bdrl2sdpws['B1'].value
    else:
        l2sdpwsrow = bdrl2sdpws['%s1' % chr(ord(column) - 1)].value

    for line in range(len(bdrconfiglist)):
        bdrl2sdpws['%s%s' % (chr(ord(column) - 1), l2sdpwsrow)] = remarklist[line]
        bdrl2sdpws['%s%s' % (column, l2sdpwsrow)] = bdrconfiglist[line]
        l2sdpwsrow += 1

    lastbdrnode = misdpdict['bdrnode']
    bdrl2sdpws['%s1' % chr(ord(column) - 1)] = l2sdpwsrow
    return lastbdrnode

def bdrl2sdpinterasworksheet(serviceid,mil2sdpinterasdict,lastbdrnode, bdrcfgwb):
    bdrl2sdpinterasws = bdrcfgwb['l2inter-as-configuration']
    bdrconfiglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    if lastbdrnode != mil2sdpinterasdict['bdrnode']:
        for c in columnlist:
            if bdrl2sdpinterasws['%s2'%c].value != None:
                nodecolumnlist.append(bdrl2sdpinterasws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mil2sdpinterasdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrl2sdpinterasws['D1'] = existcolumn[-1]
                #print(column)
        if mil2sdpinterasdict['bdrnode'] not in nodecolumnlist:
            bdrl2sdpinterasws['D1'] = existcolumn[-1]
            column = bdrl2sdpinterasws['D1'].value
            column = chr(ord(column) + 3)
            bdrl2sdpinterasws['D1'] = column
            bdrl2sdpinterasws['%s2' % column] = mil2sdpinterasdict['bdrnode']
            bdrl2sdpinterasws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrl2sdpinterasws.column_dimensions['%s' % column].width = 120.0

        if 'EPIPE' == mil2sdpinterasdict['bdrservicetype']:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR EPIPE service and sdp configuration to ASBR node ###')
            bdrconfiglist.append('/configure service epipe "%s" admin-state enable' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service epipe "%s" description "%s"' %(mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrservice']))
            bdrconfiglist.append('/configure service epipe "%s" service-id %s' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrserviceid']))
            bdrconfiglist.append('/configure service epipe "%s" customer "%s"' % (mil2sdpinterasdict['bdrservice'], mil2sdpinterasdict['bdrcusid']))
            bdrconfiglist.append('/configure service epipe "%s" service-mtu 9194' % mil2sdpinterasdict['bdrservice'])
            #bdrconfiglist.append('/configure service epipe "%s" endpoint "%s" revert-time 180' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrendpoint']))
            #bdrconfiglist.append('/configure service epipe "%s" endpoint "%s" standby-signaling master' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrendpoint']))

        elif 'VPLS' == mil2sdpinterasdict['bdrservicetype']:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR VPLS service and sdp configuration to ASBR node ###')
            bdrconfiglist.append('/configure service vpls "%s" admin-state enable' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" description "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrservice']))
            bdrconfiglist.append('/configure service vpls "%s" service-id %s' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrserviceid']))
            bdrconfiglist.append('/configure service vpls "%s" customer "%s"' % (mil2sdpinterasdict['bdrservice'], mil2sdpinterasdict['bdrcusid']))
            bdrconfiglist.append('/configure service vpls "%s" service-mtu 9194' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" mac-flush tldp send-on-failure true' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" fdb mac-move admin-state enable' % mil2sdpinterasdict['bdrservice'])


    elif lastbdrnode:
        for c in columnlist:
            if bdrl2sdpinterasws['%s2' % c].value != None:
                nodecolumnlist.append(bdrl2sdpinterasws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mil2sdpinterasdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrl2sdpinterasws['D1'] = existcolumn[-1]

    if 'EPIPE' == mil2sdpinterasdict['bdrservicetype']:
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s admin-state enable' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s description "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid'],mil2sdpinterasdict['bdrasbrsdpdesc']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s entropy-label' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s control-word true' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid']))
        #bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s endpoint name "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid'], mil2sdpinterasdict['asbrendpoint']))
        #if mil2sdpinterasdict['asbrsdpprecedence'] == 'primary':
        #    bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s endpoint precedence primary' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid']))
        #elif mil2sdpinterasdict['asbrsdpprecedence'] == 'backup':
        #    bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s endpoint precedence 3' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid']))
        for r in bdrconfiglist:
            remarklist.append('epipe')

    elif 'VPLS' == mil2sdpinterasdict['bdrservicetype']:
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s admin-state enable' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s description "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid'],mil2sdpinterasdict['bdrasbrsdpdesc']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s entropy-label' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s control-word true' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrasbrsdp'], mil2sdpinterasdict['bdrvcid']))
        for r in bdrconfiglist:
            remarklist.append('vpls')

    if bdrl2sdpinterasws['%s1' % chr(ord(column) - 1)].value == None:
        l2sdpwsrow = bdrl2sdpinterasws['B1'].value
    else:
        l2sdpwsrow = bdrl2sdpinterasws['%s1' % chr(ord(column) - 1)].value

    for line in range(len(bdrconfiglist)):
        bdrl2sdpinterasws['%s%s' % (chr(ord(column) - 1), l2sdpwsrow)] = remarklist[line]
        bdrl2sdpinterasws['%s%s' % (column, l2sdpwsrow)] = bdrconfiglist[line]
        l2sdpwsrow += 1

    lastbdrnode = mil2sdpinterasdict['bdrnode']
    bdrl2sdpinterasws['%s1' % chr(ord(column) - 1)] = l2sdpwsrow
    return lastbdrnode

def bdrl2evpninterasworksheet(serviceid,mil2evpninterasdict,lastbdrnode, bdrcfgwb):
    bdrl2evpninterasws = bdrcfgwb['l2inter-as-configuration']
    bdrconfiglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print(mil2evpninterasdict['bdrservice'])
    #print(lastbdrnode)
    #print(mil2evpninterasdict['bdrnode'])
    #print(bdrcfgwb['l2inter-as-configuration']['D1'].value)
    if lastbdrnode != mil2evpninterasdict['bdrnode']:
        for c in columnlist:
            if bdrl2evpninterasws['%s2'%c].value != None:
                nodecolumnlist.append(bdrl2evpninterasws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mil2evpninterasdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrl2evpninterasws['D1'] = existcolumn[-1]
                #print(bdrcfgwb['l2inter-as-configuration']['D1'].value)
        if mil2evpninterasdict['bdrnode'] not in nodecolumnlist:
            bdrl2evpninterasws['D1'] = existcolumn[-1]
            column = bdrl2evpninterasws['D1'].value
            column = chr(ord(column) + 3)
            bdrl2evpninterasws['D1'] = column
            bdrl2evpninterasws['%s2' % column] = mil2evpninterasdict['bdrnode']
            bdrl2evpninterasws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrl2evpninterasws.column_dimensions['%s' % column].width = 120.0

    elif lastbdrnode:
        for c in columnlist:
            if bdrl2evpninterasws['%s2' % c].value != None:
                nodecolumnlist.append(bdrl2evpninterasws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mil2evpninterasdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrl2evpninterasws['D1'] = existcolumn[-1]


    bdrconfiglist.append('')
    bdrconfiglist.append('### BDR EPIPE-EVPN ethernet-segment configuration ###')
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" admin-state enable' % mil2evpninterasdict['bdrpeethseg'])
    if mil2evpninterasdict['bdrpevlan'] == None:
        bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" type none' % mil2evpninterasdict['bdrpeethseg'])
    else:
        bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" type virtual' % mil2evpninterasdict['bdrpeethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" esi 0x%s' % (mil2evpninterasdict['bdrpeethseg'],mil2evpninterasdict['bdrpeesi']))
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" multi-homing-mode single-active' % mil2evpninterasdict['bdrpeethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election es-activation-timer 0' % mil2evpninterasdict['bdrpeethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election service-carving-mode manual' % mil2evpninterasdict['bdrpeethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election manual preference value %s' % (mil2evpninterasdict['bdrpeethseg'],mil2evpninterasdict['bdrpepreference']))
    if mil2evpninterasdict['bdrpevlan'] == None:
        if 'lag' in mil2evpninterasdict['bdrpeport']:
            bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association lag %s' % (mil2evpninterasdict['bdrpeethseg'],mil2evpninterasdict['bdrpeport'].split('-')[-1]))
        else:
            bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association port %s' % (mil2evpninterasdict['bdrpeethseg'],mil2evpninterasdict['bdrpeport']))
    else:
        if 'lag' in mil2evpninterasdict['bdrpeport']:
            bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association lag %s virtual-ranges dot1q q-tag %s end %s' % (mil2evpninterasdict['bdrpeethseg'],mil2evpninterasdict['bdrpeport'].split('-')[-1],mil2evpninterasdict['bdrpevlan'],mil2evpninterasdict['bdrpevlan']))
        else:
            bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association port %s virtual-ranges dot1q q-tag %s end %s' % (mil2evpninterasdict['bdrpeethseg'],mil2evpninterasdict['bdrpeport'],mil2evpninterasdict['bdrpevlan'],mil2evpninterasdict['bdrpevlan']))

    bdrconfiglist.append('')
    bdrconfiglist.append('### BDR EPIPE-EVPN service configuration ###')
    bdrconfiglist.append('/configure service epipe "%s" admin-state enable' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" description "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrservice']))
    bdrconfiglist.append('/configure service epipe "%s" service-id %s' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrserviceid']))
    bdrconfiglist.append('/configure service epipe "%s" customer "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrcustid']))
    bdrconfiglist.append('/configure service epipe "%s" service-mtu 9194' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" { bgp 1 }' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn evi %s' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrevi']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn local-ac name "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrpeacname']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn local-ac eth-tag %s' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrpeethtag']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn remote-ac name "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrasbracname']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn remote-ac eth-tag %s' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrasbrethtag']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 admin-state enable' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 entropy-label true' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 control-word true' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 auto-bind-tunnel resolution filter' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 auto-bind-tunnel resolution-filter rsvp true' % mil2evpninterasdict['bdrservice'])


    for i in bdrconfiglist:
        if 'ethernet-segment' in i:
            remarklist.append('ether-segment')
        else:
            remarklist.append('epipe-evpn')

    if bdrl2evpninterasws['%s1' % chr(ord(column) - 1)].value == None:
        l2evpnwsrow = bdrl2evpninterasws['B1'].value
    else:
        l2evpnwsrow = bdrl2evpninterasws['%s1' % chr(ord(column) - 1)].value

    for line in range(len(bdrconfiglist)):
        bdrl2evpninterasws['%s%s' % (chr(ord(column) - 1), l2evpnwsrow)] = remarklist[line]
        bdrl2evpninterasws['%s%s' % (column, l2evpnwsrow)] = bdrconfiglist[line]
        l2evpnwsrow += 1

    lastbdrnode = mil2evpninterasdict['bdrnode']
    bdrl2evpninterasws['%s1' % chr(ord(column) - 1)] = l2evpnwsrow
    return lastbdrnode

def bdrl2sdpasbrworksheet(serviceid, mil2sdpinterasdict,lastasbrnode, bdrcfgwb):
    bdrl2sdpinterasws = bdrcfgwb['l2inter-as-configuration']
    bdrconfiglist = []
    remarklist = []
    if lastasbrnode != mil2sdpinterasdict['bdrasbrnode']:
        column = bdrl2sdpinterasws['D1'].value
        column = chr(ord(column) + 3)
        bdrl2sdpinterasws['D1'] = column
        bdrl2sdpinterasws['%s2' % chr(ord(column) - 1)] = 'ASBR'
        bdrl2sdpinterasws['%s2' % column] = mil2sdpinterasdict['bdrasbrnode']
        bdrl2sdpinterasws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
        bdrl2sdpinterasws.column_dimensions['%s' % column].width = 120.0

        bdrconfiglist.append('')
        ('### BDR eth-cfm configuration of ASBR node ###')
        bdrconfiglist.append('/configure eth-cfm domain "%s" level 5' % mil2sdpinterasdict['asbrmdindex'])
        bdrconfiglist.append('/configure eth-cfm domain "%s" name "%s"' % (mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmdname']))
        bdrconfiglist.append('/configure eth-cfm domain "%s" association "%s" string "%s"' % (mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmdname']))
        bdrconfiglist.append('/configure eth-cfm domain "%s" association "%s" ccm-interval 1s' % (mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex']))
        bdrconfiglist.append('/configure eth-cfm domain "%s" association "%s" bridge-identifier "%s" vlan %s' % (mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrbridgeid'],mil2sdpinterasdict['asbrbridgevlan']))
        bdrconfiglist.append('/configure eth-cfm domain "%s" association "%s" remote-mep %s' % (mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbriptnmepid']))

        if 'EPIPE' == mil2sdpinterasdict['bdrservicetype']:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR EPIPE service and sdp configuration of ASBR node ###')
            bdrconfiglist.append('/configure service epipe "%s" admin-state enable' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service epipe "%s" description "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrservice']))
            bdrconfiglist.append('/configure service epipe "%s" service-id %s' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrserviceid']))
            bdrconfiglist.append('/configure service epipe "%s" customer "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrcusid']))
            bdrconfiglist.append('/configure service epipe "%s" service-mtu 9194' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service epipe "%s" sap %s:%s admin-state enable' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan']))
            bdrconfiglist.append('/configure service epipe "%s" sap %s:%s ingress qos sap-ingress policy-name "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['bdrcusid']))
            bdrconfiglist.append('/configure service epipe "%s" sap %s:%s egress qos sap-egress policy-name "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['bdrcusid']))
            bdrconfiglist.append('/configure service epipe "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s admin-state enable' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmepid']))
            bdrconfiglist.append('/configure service epipe "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s low-priority-defect all-def' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmepid']))
            bdrconfiglist.append('/configure service epipe "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s fault-propagation use-if-status-tlv' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmepid']))
            bdrconfiglist.append('/configure service epipe "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s ccm true' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmepid']))


        elif 'VPLS' == mil2sdpinterasdict['bdrservicetype']:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR VPLS service and sap configuration of ASBR node ###')
            bdrconfiglist.append('/configure service vpls "%s" admin-state enable' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" description "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrservice']))
            bdrconfiglist.append('/configure service vpls "%s" service-id %s' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrserviceid']))
            bdrconfiglist.append('/configure service vpls "%s" customer "%s"' % (mil2sdpinterasdict['bdrservice'], mil2sdpinterasdict['bdrcusid']))
            bdrconfiglist.append('/configure service vpls "%s" service-mtu 9194' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" mac-flush tldp send-on-failure true' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" fdb mac-move admin-state enable' % mil2sdpinterasdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" sap %s:%s admin-state disable' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan']))
            bdrconfiglist.append('/configure service vpls "%s" sap %s:%s ingress qos sap-ingress policy-name "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['bdrcusid']))
            bdrconfiglist.append('/configure service vpls "%s" sap %s:%s egress qos sap-egress policy-name "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['bdrcusid']))
            bdrconfiglist.append('/configure service vpls "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s admin-state enable' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmepid']))
            bdrconfiglist.append('/configure service vpls "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s low-priority-defect all-def' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmepid']))
            bdrconfiglist.append('/configure service vpls "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s fault-propagation use-if-status-tlv' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmepid']))
            bdrconfiglist.append('/configure service vpls "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s ccm true' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['asbrinteraslag'],mil2sdpinterasdict['asbrbridgevlan'],mil2sdpinterasdict['asbrmdindex'],mil2sdpinterasdict['asbrmaindex'],mil2sdpinterasdict['asbrmepid']))

    elif lastasbrnode:
        column = bdrl2sdpinterasws['D1'].value

    if 'EPIPE' == mil2sdpinterasdict['bdrservicetype']:
        bdrconfiglist.append('### BDR EPIPE sdp configuration of ASBR node ###')
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s admin-state enable' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrpesdp'],mil2sdpinterasdict['bdrvcid']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s description "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrpesdp'],mil2sdpinterasdict['bdrvcid'],mil2sdpinterasdict['bdrpesdpdesc']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s entropy-label' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrpesdp'],mil2sdpinterasdict['bdrvcid']))
        bdrconfiglist.append('/configure service epipe "%s" spoke-sdp %s:%s control-word true' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrpesdp'],mil2sdpinterasdict['bdrvcid']))
        for i in bdrconfiglist:
            remarklist.append('epipe')

    elif 'VPLS' == mil2sdpinterasdict['bdrservicetype']:
        bdrconfiglist.append('### BDR VPLS sdp configuration of ASBR node ###')
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s admin-state enable' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrpesdp'],mil2sdpinterasdict['bdrvcid']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s description "%s"' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrpesdp'],mil2sdpinterasdict['bdrvcid'],mil2sdpinterasdict['bdrpesdpdesc']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s entropy-label' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrpesdp'],mil2sdpinterasdict['bdrvcid']))
        bdrconfiglist.append('/configure service vpls "%s" mesh-sdp %s:%s control-word true' % (mil2sdpinterasdict['bdrservice'],mil2sdpinterasdict['bdrpesdp'],mil2sdpinterasdict['bdrvcid']))
        for i in bdrconfiglist:
            remarklist.append('vpls')

    if bdrl2sdpinterasws['%s1' % chr(ord(column) - 1)].value == None:
        l2sdpwsrow = bdrl2sdpinterasws['J1'].value
    else:
        l2sdpwsrow = bdrl2sdpinterasws['%s1' % chr(ord(column) - 1)].value

    for line in range(len(bdrconfiglist)):
        bdrl2sdpinterasws['%s%s' % (chr(ord(column) - 1), l2sdpwsrow)] = remarklist[line]
        bdrl2sdpinterasws['%s%s' % (column, l2sdpwsrow)] = bdrconfiglist[line]
        l2sdpwsrow += 1

    lastasbrnode = mil2sdpinterasdict['bdrasbrnode']
    bdrl2sdpinterasws['%s1' % chr(ord(column) - 1)] = l2sdpwsrow
    return lastasbrnode

def bdrl2evpnasbrworksheet(serviceid, mil2evpninterasdict,lastasbrnode, bdrcfgwb):
    bdrl2evpninterasws = bdrcfgwb['l2inter-as-configuration']
    bdrconfiglist = []
    remarklist = []
    if lastasbrnode != mil2evpninterasdict['asbrnode']:
        column = bdrl2evpninterasws['D1'].value
        column = chr(ord(column) + 3)
        bdrl2evpninterasws['D1'] = column
        bdrl2evpninterasws['%s2' % chr(ord(column) - 1)] = 'ASBR'
        bdrl2evpninterasws['%s2' % column] = mil2evpninterasdict['asbrnode']
        bdrl2evpninterasws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
        bdrl2evpninterasws.column_dimensions['%s' % column].width = 120.0

    elif lastasbrnode:
        column = bdrl2evpninterasws['D1'].value


    bdrconfiglist.append('')
    ('### BDR eth-cfm configuration of ASBR node ###')
    bdrconfiglist.append('/configure eth-cfm domain "%s" level 5' % mil2evpninterasdict['asbrmdindex'])
    bdrconfiglist.append('/configure eth-cfm domain "%s" name "%s"' % (mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmdname']))
    bdrconfiglist.append('/configure eth-cfm domain "%s" association "%s" string "%s"' % (mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmaindex'],mil2evpninterasdict['asbrmdname']))
    bdrconfiglist.append('/configure eth-cfm domain "%s" association "%s" ccm-interval 1s' % (mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmaindex']))
    bdrconfiglist.append('/configure eth-cfm domain "%s" association "%s" bridge-identifier "%s" vlan %s' % (mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmaindex'],mil2evpninterasdict['asbrbridgeid'],mil2evpninterasdict['asbrbridgevlan']))
    bdrconfiglist.append('/configure eth-cfm domain "%s" association "%s" remote-mep %s' % (mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmaindex'],mil2evpninterasdict['asbriptnmep']))

    bdrconfiglist.append('')
    bdrconfiglist.append('### BDR EPIPE-EVPN ethernet-segment configuration ###')
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" admin-state enable' % mil2evpninterasdict['asbrethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" type virtual' % mil2evpninterasdict['asbrethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" esi 0x%s' % (mil2evpninterasdict['asbrethseg'],mil2evpninterasdict['asbresi']))
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" multi-homing-mode single-active' % mil2evpninterasdict['asbrethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election es-activation-timer 0' % mil2evpninterasdict['asbrethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election service-carving-mode manual' % mil2evpninterasdict['asbrethseg'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election manual preference value %s' % (mil2evpninterasdict['asbrethseg'],mil2evpninterasdict['asbrpreference']))
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association lag %s virtual-ranges dot1q q-tag %s end %s' % (mil2evpninterasdict['asbrethseg'],mil2evpninterasdict['asbrinteraslag'].split('-')[-1],mil2evpninterasdict['asbrbridgevlan'],mil2evpninterasdict['asbrbridgevlan']))

    bdrconfiglist.append('')
    bdrconfiglist.append('### BDR EPIPE-EVPN service configuration ###')
    bdrconfiglist.append('/configure service epipe "%s" admin-state enable' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" description "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrservice']))
    bdrconfiglist.append('/configure service epipe "%s" service-id %s' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrserviceid']))
    bdrconfiglist.append('/configure service epipe "%s" customer "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrcustid']))
    bdrconfiglist.append('/configure service epipe "%s" service-mtu 9194' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" { bgp 1 }' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn evi %s' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrevi']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn local-ac name "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrasbracname']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn local-ac eth-tag %s' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrasbrethtag']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn remote-ac name "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrpeacname']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn remote-ac eth-tag %s' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['bdrpeethtag']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 admin-state enable' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 entropy-label true' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 control-word true' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 auto-bind-tunnel resolution filter' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 auto-bind-tunnel resolution-filter rsvp true' % mil2evpninterasdict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" sap %s:%s admin-state disable' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['asbrinteraslag'],mil2evpninterasdict['asbrbridgevlan']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s:%s ingress qos sap-ingress policy-name "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['asbrinteraslag'],mil2evpninterasdict['asbrbridgevlan'],mil2evpninterasdict['bdrcustid']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s:%s egress qos sap-egress policy-name "%s"' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['asbrinteraslag'],mil2evpninterasdict['asbrbridgevlan'],mil2evpninterasdict['bdrcustid']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s admin-state enable' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['asbrinteraslag'],mil2evpninterasdict['asbrbridgevlan'],mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmaindex'],mil2evpninterasdict['asbrlocalmep']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s low-priority-defect all-def' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['asbrinteraslag'],mil2evpninterasdict['asbrbridgevlan'],mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmaindex'],mil2evpninterasdict['asbrlocalmep']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s fault-propagation use-if-status-tlv' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['asbrinteraslag'],mil2evpninterasdict['asbrbridgevlan'],mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmaindex'],mil2evpninterasdict['asbrlocalmep']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s:%s eth-cfm mep md-admin-name "%s" ma-admin-name "%s" mep-id %s ccm true' % (mil2evpninterasdict['bdrservice'],mil2evpninterasdict['asbrinteraslag'],mil2evpninterasdict['asbrbridgevlan'],mil2evpninterasdict['asbrmdindex'],mil2evpninterasdict['asbrmaindex'],mil2evpninterasdict['asbrlocalmep']))

    for i in bdrconfiglist:
        if 'ethernet-segment' in i:
            remarklist.append('ether-segment')
        elif 'eth-cfm domain' in i:
            remarklist.append('eth-cfm')
        else:
            remarklist.append('epipe-evpn')

    if bdrl2evpninterasws['%s1'%(chr(ord(column) - 1))].value == None:
        l2evpnwsrow = bdrl2evpninterasws['B1'].value
    else:
        l2evpnwsrow = bdrl2evpninterasws['%s1' % (chr(ord(column) - 1))].value

    for line in range(len(bdrconfiglist)):
        bdrl2evpninterasws['%s%s' % (chr(ord(column) - 1), l2evpnwsrow)] = remarklist[line]
        bdrl2evpninterasws['%s%s' % (column, l2evpnwsrow)] = bdrconfiglist[line]
        l2evpnwsrow += 1

    lastasbrnode = mil2evpninterasdict['asbrnode']
    bdrl2evpninterasws['%s1'%(chr(ord(column) - 1))] = l2evpnwsrow
    return lastasbrnode

def bdrpolicyworksheet(policypath, prefixpath, policytype,bdrpolicydict,bdrnode,bdrservice,bdrpolicyname):
    policycfg = []
    policypath = policypath.strip('.')
    pol = open(migratepath + misumtime + policypath, "r")
    pollines = pol.readlines()
    pol.close()

    policycfg.append('#######################################')
    policycfg.append('# BDR %s policy configuration' % policytype)
    policycfg.append('#######################################')

    if prefixpath == None:
        prefixpath = ''
    else:
        prefixpath = prefixpath.strip('.')
        pre = open(migratepath + misumtime + prefixpath, "r")
        prelines = pre.readlines()
        pre.close()
        for pr in range(len(prelines)):
            prelines[pr] = re.sub(r"^\s+", "", prelines[pr])  # remove space from beginning
            prelines[pr] = prelines[pr].rstrip('\n')  # remove newline('\n') from end of line
            prelines[pr] = re.sub(r"\s+$", "", prelines[pr])  # remove space from ending
            #print(prelines[pr])
            if 'prefix-list' in prelines[pr]:
                if '%s_%s_%s_existing' %(bdrnode,bdrservice,bdrpolicyname) in bdrpolicydict:
                    for existingprefixlist in bdrpolicydict['%s_%s_%s_existing' %(bdrnode,bdrservice,bdrpolicyname)]:
                        if existingprefixlist != None:
                            if prelines[pr].split()[3] == '"%s"' % existingprefixlist:
                                #print(existingprefixlist)
                                bdrprefixlist = bdrpolicydict['%s_%s_%s_%s' %(bdrnode,bdrservice,bdrpolicyname,existingprefixlist)]
                                #print(bdrprefixlist)
                                prelines[pr] = prelines[pr].replace('prefix-list "%s" prefix' % existingprefixlist,'prefix-list "%s" prefix' % bdrprefixlist)
            policycfg.append(prelines[pr])

    for a in range(len(pollines)):
        pollines[a] = re.sub(r"^\s+", "", pollines[a])  # remove space from beginning
        pollines[a] = pollines[a].rstrip('\n')  # remove newline('\n') from end of line
        pollines[a] = re.sub(r"\s+$", "", pollines[a])  # remove space from ending
        if 'named-entry' in pollines[a]:
            #print(pollines[a])
            if '%s_%s_%s_entry-existing' %(bdrnode,bdrservice,bdrpolicyname) in bdrpolicydict:
                for existingentry in bdrpolicydict['%s_%s_%s_entry-existing' %(bdrnode,bdrservice,bdrpolicyname)]:
                    if existingentry != None:
                        #print(existingprefixlist)
                        bdrentry = bdrpolicydict['%s_%s_%s_%s_entry' %(bdrnode,bdrservice,bdrpolicyname,existingentry)]
                        #print(bdrprefixlist)
                        pollines[a] = pollines[a].replace('named-entry "%s"' % existingentry,'named-entry "%s"' % bdrentry)
        if 'prefix-list' in pollines[a]:
            #print(pollines[a])
            if '%s_%s_%s_existing' %(bdrnode,bdrservice,bdrpolicyname) in bdrpolicydict:
                for existingprefixlist in bdrpolicydict['%s_%s_%s_existing' %(bdrnode,bdrservice,bdrpolicyname)]:
                    if existingprefixlist != None:
                        #print(existingprefixlist)
                        bdrprefixlist = bdrpolicydict['%s_%s_%s_%s' %(bdrnode,bdrservice,bdrpolicyname,existingprefixlist)]
                        #print(bdrprefixlist)
                        pollines[a] = pollines[a].replace('prefix-list ["%s"]' % existingprefixlist,'prefix-list ["%s"]' % bdrprefixlist)
            #print(pollines[a])

        policycfg.append(pollines[a])

    return policycfg


def bdrglobalvprnworksheet(serviceid, miglobalvprndict, lastbdrnode, bdrcfgwb,bdrpolicydict):
    bdrglobalvprnws = bdrcfgwb['bdr-configuration']
    bdrpolicyws = bdrcfgwb['bdr-policy']
    bdrfilterws = bdrcfgwb['bdr-filter']
    bdrconfiglist = []
    bdrpolicycfglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print('column %s' %bdrglobalvprnws['A1'].value)
    if lastbdrnode != miglobalvprndict['bdrnode']:
        for c in columnlist:
            if bdrglobalvprnws['%s2'%c].value != None:
                nodecolumnlist.append(bdrglobalvprnws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if miglobalvprndict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrglobalvprnws['D1'] = existcolumn[-1]
        if miglobalvprndict['bdrnode'] not in nodecolumnlist:
            bdrglobalvprnws['D1'] = existcolumn[-1]
            column = bdrglobalvprnws['D1'].value
            column = chr(ord(column) + 3)
            bdrglobalvprnws['D1'] = column
            bdrglobalvprnws['%s2' % column] = miglobalvprndict['bdrnode']
            bdrpolicyws['%s2' % column] = miglobalvprndict['bdrnode']
            bdrfilterws['%s2' % column] = miglobalvprndict['bdrnode']
            bdrglobalvprnws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrglobalvprnws.column_dimensions['%s' % column].width = 120.0
            bdrpolicyws.column_dimensions['%s' % chr(ord(column) - 1)].width = 17.0
            bdrpolicyws.column_dimensions['%s' % column].width = 120.0
            bdrfilterws.column_dimensions['%s' % chr(ord(column) - 1)].width = 17.0
            bdrfilterws.column_dimensions['%s' % column].width = 120.0

    elif lastbdrnode:
        for c in columnlist:
            if bdrglobalvprnws['%s2' % c].value != None:
                nodecolumnlist.append(bdrglobalvprnws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if miglobalvprndict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrglobalvprnws['D1'] = existcolumn[-1]

    bdrconfiglist.append('')
    #print(miglobalvprndict['bdrservice'])
    bdrconfiglist.append('######################################')
    bdrconfiglist.append('### BDR VPRN service configuration ###')
    bdrconfiglist.append('######################################')
    bdrconfiglist.append('/configure service vprn "%s" admin-state enable' % miglobalvprndict['bdrservice'])
    bdrconfiglist.append('/configure service vprn "%s" service-id %s' % (miglobalvprndict['bdrservice'],miglobalvprndict['bdrserviceid']))
    bdrconfiglist.append('/configure service vprn "%s" customer "%s"' % (miglobalvprndict['bdrservice'],miglobalvprndict['bdrcustid']))
    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service vprn "%s" autonomous-system 65051' % miglobalvprndict['bdrservice'])
    bdrconfiglist.append('/configure service vprn "%s" ecmp 16' % miglobalvprndict['bdrservice'])
    bdrconfiglist.append('/configure service vprn "%s" route-distinguisher "%s"' % (miglobalvprndict['bdrservice'],miglobalvprndict['bdrrd']))
    bdrconfiglist.append('')
    if miglobalvprndict['bdrvrfimppolname'] != None:
        imppolicycfg = bdrpolicyworksheet(miglobalvprndict['bdrvrfimppol'],miglobalvprndict['bdrvrfimpprefix'],'vrf-import',bdrpolicydict,miglobalvprndict['bdrnode'],miglobalvprndict['bdrservice'],miglobalvprndict['bdrvrfimppolname'])
        bdrpolicycfglist.extend(imppolicycfg)
        bdrconfiglist.append('/configure service vprn "%s" vrf-import policy "%s"' % (miglobalvprndict['bdrservice'],miglobalvprndict['bdrvrfimppolname']))
    if miglobalvprndict['bdrvrfexppolname'] != None:
        exppolicycfg = bdrpolicyworksheet(miglobalvprndict['bdrvrfexppol'],miglobalvprndict['bdrvrfexpprefix'],'vrf-export',bdrpolicydict,miglobalvprndict['bdrnode'],miglobalvprndict['bdrservice'],miglobalvprndict['bdrvrfexppolname'])
        bdrpolicycfglist.extend(exppolicycfg)
        bdrconfiglist.append('/configure service vprn "%s" vrf-export policy "%s"' % (miglobalvprndict['bdrservice'],miglobalvprndict['bdrvrfexppolname']))
    if miglobalvprndict['bdrvrfagg'] != None:
        bdrconfiglist.append('')
        miglobalvprndict['bdrvrfagg'] = miglobalvprndict['bdrvrfagg'].strip('.')
        f = open(migratepath + misumtime + miglobalvprndict['bdrvrfagg'] , "r")
        lines = f.readlines()
        f.close()
        for a in range(len(lines)):
            lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
            lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
            lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
            #print(lines[a])
            if lines[a] not in bdrconfiglist:
                bdrconfiglist.append(lines[a])
    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service vprn "%s" auto-bind-tunnel resolution filter' % miglobalvprndict['bdrservice'])
    bdrconfiglist.append('/configure service vprn "%s" auto-bind-tunnel resolution-filter rsvp true' % miglobalvprndict['bdrservice'])
    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service vprn "%s" bgp-vpn-backup ipv4 true' % miglobalvprndict['bdrservice'])
    bdrconfiglist.append('/configure service vprn "%s" bgp-vpn-backup ipv6 true' % miglobalvprndict['bdrservice'])

    for i in bdrconfiglist:
        remarklist.append('vprn')

    if bdrglobalvprnws['%s1'%(chr(ord(column) - 1))].value == None:
        globalvprnwsrow = bdrglobalvprnws['B1'].value
    else:
        globalvprnwsrow = bdrglobalvprnws['%s1' % (chr(ord(column) - 1))].value

    if bdrpolicyws['%s1'%(chr(ord(column) - 1))].value == None:
        bdrpolicyrow = bdrpolicyws['B1'].value
    else:
        bdrpolicyrow = bdrpolicyws['%s1' % (chr(ord(column) - 1))].value

    #print(l3prefix)
    for line in range(len(bdrconfiglist)):
        bdrglobalvprnws['%s%s' % (chr(ord(column) - 1), globalvprnwsrow)] = remarklist[line]
        bdrglobalvprnws['%s%s' % (column, globalvprnwsrow)] = bdrconfiglist[line]
        globalvprnwsrow += 1
    for line in range(len(bdrpolicycfglist)):
        bdrpolicyws['%s%s' % (chr(ord(column) - 1), bdrpolicyrow)] = 'vrf imp/exp policy'
        bdrpolicyws['%s%s' % (column, bdrpolicyrow)] = bdrpolicycfglist[line]
        bdrpolicyrow += 1

    lastbdrnode = miglobalvprndict['bdrnode']
    bdrglobalvprnws['%s1'%(chr(ord(column) - 1))] = globalvprnwsrow
    bdrpolicyws['%s1'%(chr(ord(column) - 1))] = bdrpolicyrow
    return lastbdrnode


def bdrvprnintworksheet(serviceid, mivprnintdict, lastbdrnode, bdrcfgwb,bdrpolicydict,l3prefix):
    bdrvprnintws = bdrcfgwb['bdr-configuration']
    bdrpolicyws = bdrcfgwb['bdr-policy']
    bdrconfiglist = []
    bdrpolicycfglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    mapsheet = openpyxl.load_workbook(inputpath + 'nodemapping.xlsx')['VPRN-service']
    #print('column %s' %bdrvprnintws['A1'].value)
    if lastbdrnode != mivprnintdict['bdrnode']:
        for c in columnlist:
            if bdrvprnintws['%s2'%c].value != None:
                nodecolumnlist.append(bdrvprnintws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mivprnintdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrvprnintws['D1'] = existcolumn[-1]
        if mivprnintdict['bdrnode'] not in nodecolumnlist:
            bdrvprnintws['D1'] = existcolumn[-1]
            column = bdrvprnintws['D1'].value
            column = chr(ord(column) + 3)
            bdrvprnintws['D1'] = column
            bdrvprnintws['%s2' % column] = mivprnintdict['bdrnode']
            bdrpolicyws['%s2' % column] = mivprnintdict['bdrnode']
            bdrfilterws['%s2' % column] = mivprnintdict['bdrnode']
            bdrvprnintws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrvprnintws.column_dimensions['%s' % column].width = 120.0
            bdrpolicyws.column_dimensions['%s' % chr(ord(column) - 1)].width = 17.0
            bdrpolicyws.column_dimensions['%s' % column].width = 120.0
            bdrfilterws.column_dimensions['%s' % chr(ord(column) - 1)].width = 17.0
            bdrfilterws.column_dimensions['%s' % column].width = 120.0


    elif lastbdrnode:
        for c in columnlist:
            if bdrvprnintws['%s2' % c].value != None:
                nodecolumnlist.append(bdrvprnintws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mivprnintdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrvprnintws['D1'] = existcolumn[-1]

    #print(mivprnintdict)
    networkaddr = IPNetwork(mivprnintdict['bdrpriip']).network
    bdrpriip = mivprnintdict['bdrpriip'].split('/')[0]
    bdrprilength = mivprnintdict['bdrpriip'].split('/')[-1]
    l3prefix.append(str(networkaddr) + '/' + bdrprilength)
    if mivprnintdict['bdrsecip'] != None:
        networkaddr = IPNetwork(mivprnintdict['bdrsecip']).network
        bdrsecip = mivprnintdict['bdrsecip'].split('/')[0]
        bdrseclength = mivprnintdict['bdrsecip'].split('/')[-1]
        l3prefix.append(str(networkaddr) + '/' + bdrseclength)

    bdrconfiglist.append('')
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('# BDR VPRN Interface %s configuration' % mivprnintdict['bdrintname'])
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('/configure service vprn "%s" interface "%s" admin-state disable'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname']))
    bdrconfiglist.append('/configure service vprn "%s" interface "%s" description "%s"'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintdesc']))
    bdrconfiglist.append('#/configure service vprn "%s" interface "%s" hold-time %s down seconds 600'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam']))
    bdrconfiglist.append('#/configure service vprn "%s" interface "%s" hold-time %s down init-only true'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam']))
    bdrconfiglist.append('')
    if mivprnintdict['bdrservice'] == 'L3_Gn':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" load-balancing teid-load-balancing true'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname']))
    if mivprnintdict['bdrintfam'] == 'ipv4':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s primary address %s'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],bdrpriip))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s primary prefix-length %s'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],bdrprilength))
        if mivprnintdict['bdrsecip'] != None:
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s secondary %s prefix-length %s'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],bdrsecip,bdrseclength))
    elif mivprnintdict['bdrintfam'] == 'ipv6':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s prefix-length %s' % (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],bdrpriip,bdrprilength))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s primary-preference 1'% (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],bdrpriip))
        if mivprnintdict['bdrvrrpid'] != None and mivprnintdict['bdrintfam'] == 'ipv6':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s duplicate-address-detection false' %(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],bdrpriip))
            if 'PE01' in mivprnintdict['bdrnode']:
                bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s link-local-address address fe80::2' %(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam']))
            elif 'PE02' in mivprnintdict['bdrnode']:
                bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s link-local-address address fe80::3' %(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam']))
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s link-local-address duplicate-address-detection false' %(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam']))
        if mivprnintdict['bdrsecip'] != None:
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s prefix-length %s'% (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],bdrsecip,bdrseclength))
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s primary-preference 10'% (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],bdrsecip))

    if mivprnintdict['bdrvrrpid'] != None:
        bdrvrrpinterval = ''
        if mivprnintdict['bdrvrrpint'] != None and mivprnintdict['bdrintfam'] == 'ipv4':
            bdrvrrpinterval = int(mivprnintdict['bdrvrrpint'])//100
        elif mivprnintdict['bdrvrrpint'] != None and mivprnintdict['bdrintfam'] == 'ipv6':
            bdrvrrpinterval = int(mivprnintdict['bdrvrrpint'])//10
        bdrconfiglist.append('')
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s admin-state enable'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrvrrpid']))
        if mivprnintdict['bdrvrrpid'] != None and mivprnintdict['bdrintfam'] == 'ipv6':
            grouphex = str(hex(int(mivprnintdict['bdrvrrpid']))).strip('0x')
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s backup [%s fe80::200:5eff:fe00:2%s]'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrvrrpid'],mivprnintdict['bdrvrrpvip'],grouphex))
        else:
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s backup [%s]'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrvrrpid'],mivprnintdict['bdrvrrpvip']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s ping-reply true'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrvrrpid']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s traceroute-reply true'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrvrrpid']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s standby-forwarding true'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrvrrpid']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s priority %s'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrvrrpid'],mivprnintdict['bdrvrrppri']))
        if mivprnintdict['bdrvrrpid'] != None and mivprnintdict['bdrintfam'] == 'ipv6':
            bdrconfiglist.append('/configure service vprn "%s" ipv6 router-advertisement interface "%s" admin-state enable' %(mivprnintdict['bdrservice'],mivprnintdict['bdrintname']))
            bdrconfiglist.append('/configure service vprn "%s" ipv6 router-advertisement interface "%s" use-virtual-mac true' %(mivprnintdict['bdrservice'],mivprnintdict['bdrintname']))

    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrsap']))
    if mivprnintdict['bdrsapqos'] != None:
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s ingress qos sap-ingress policy-name "%s"'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrsap'],mivprnintdict['bdrsapqos']))
    if mivprnintdict['bdrsapinfilter'] != None and mivprnintdict['bdrintfam'] == 'ipv4':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s ingress filter ip "%s"'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrsap'],mivprnintdict['bdrsapinfilter']))
    elif mivprnintdict['bdrsapinfilter'] != None and mivprnintdict['bdrintfam'] == 'ipv6':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s ingress filter ipv6 "%s"'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrsap'],mivprnintdict['bdrsapinfilter']))

    if mivprnintdict['bdrsapqos'] != None:
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s egress qos sap-egress policy-name "%s"'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrsap'],mivprnintdict['bdrsapqos']))
    if mivprnintdict['bdrsapegfilter'] != None and mivprnintdict['bdrintfam'] == 'ipv4':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s egress filter ip "%s"'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrsap'],mivprnintdict['bdrsapegfilter']))
    elif mivprnintdict['bdrsapegfilter'] != None and mivprnintdict['bdrintfam'] == 'ipv6':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s egress filter ipv6 "%s"'%(mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrsap'],mivprnintdict['bdrsapegfilter']))

    if mivprnintdict['bdrstatic'] != None:
        bdrconfiglist.append('')
        bdrconfiglist.append('######################################################')
        bdrconfiglist.append('# BDR PE-CE Static Routing configuration of interface %s' % mivprnintdict['bdrintname'])
        staticpath = mivprnintdict['bdrstatic'].strip('.')
        sta = open(migratepath + misumtime + staticpath, "r")
        staticline = sta.readlines()
        sta.close()
        for st in range(len(staticline)):
            staticline[st] = re.sub(r"^\s+", "", staticline[st])  # remove space from beginning
            staticline[st] = staticline[st].rstrip('\n')  # remove newline('\n') from end of line
            staticline[st] = re.sub(r"\s+$", "", staticline[st])  # remove space from ending
            bdrconfiglist.append(staticline[st])
            if 'static-routes route' in staticline[st]:
                l3prefix.append(staticline[st].split()[6])

    if mivprnintdict['bdrbgpneighbor'] != None:
        bdrconfiglist.append('')
        bdrconfiglist.append('######################################################')
        bdrconfiglist.append('# BDR PE-CE BGP Routing configuration of interface %s' % mivprnintdict['bdrintname'])
        if mivprnintdict['bdrbgpbfdint'] != None:

            bdrconfiglist.append('')
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s bfd admin-state enable' % (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam']))
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s bfd transmit-interval %s' % (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrbgpbfdint']))
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s bfd receive %s' % (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrbgpbfdint']))
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s bfd multiplier %s' % (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam'],mivprnintdict['bdrbgpbfdmul']))
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s bfd type cpm-np' % (mivprnintdict['bdrservice'],mivprnintdict['bdrintname'],mivprnintdict['bdrintfam']))
        bdrconfiglist.append('')
        bdrconfiglist.append('/configure service vprn "%s" bgp min-route-advertisement 1' % mivprnintdict['bdrservice'])
        bdrconfiglist.append('/configure service vprn "%s" bgp rapid-withdrawal true' % mivprnintdict['bdrservice'])
        bdrconfiglist.append('/configure service vprn "%s" bgp peer-ip-tracking true' % mivprnintdict['bdrservice'])
        bdrconfiglist.append('/configure service vprn "%s" bgp split-horizon true' % mivprnintdict['bdrservice'])
        bdrconfiglist.append('/configure service vprn "%s" bgp ebgp-default-reject-policy import true' % mivprnintdict['bdrservice'])
        bdrconfiglist.append('/configure service vprn "%s" bgp ebgp-default-reject-policy export true' % mivprnintdict['bdrservice'])
        bdrconfiglist.append('')
        bdrconfiglist.append('/configure service vprn "%s" bgp group "%s" type external' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpgroup']))
        bdrconfiglist.append('/configure service vprn "%s" bgp group "%s" peer-as %s' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpgroup'],mivprnintdict['bdrbgppeeras']))
        bdrconfiglist.append('/configure service vprn "%s" bgp group "%s" family %s true' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpgroup'],mivprnintdict['bdrintfam']))

        if mivprnintdict['bdrbgpimppol'] != None:
            imppolicycfg = bdrpolicyworksheet(mivprnintdict['bdrbgpimppol'],mivprnintdict['bdrbgpimpprefix'],'bgp-pe-ce-import',bdrpolicydict,mivprnintdict['bdrnode'],mivprnintdict['bdrservice'],mivprnintdict['bdrbgpimppolname'])
            bdrpolicycfglist.extend(imppolicycfg)
            bdrconfiglist.append('/configure service vprn "%s" bgp group "%s" import policy ["%s"]' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpgroup'],mivprnintdict['bdrbgpimppolname']))
        if mivprnintdict['bdrbgpexppol'] != None:
            exppolicycfg = bdrpolicyworksheet(mivprnintdict['bdrbgpexppol'],mivprnintdict['bdrbgpexpprefix'],'bgp-pe-ce-export',bdrpolicydict,mivprnintdict['bdrnode'],mivprnintdict['bdrservice'],mivprnintdict['bdrbgpexppolname'])
            bdrpolicycfglist.extend(exppolicycfg)
            bdrconfiglist.append('/configure service vprn "%s" bgp group "%s" export policy ["%s"]' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpgroup'],mivprnintdict['bdrbgpexppolname']))

        bdrconfiglist.append('')
        if mivprnintdict['bdrbgpneighbordesc'] != None:
            bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" description "%s"' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor'],mivprnintdict['bdrbgpneighbordesc']))
        if mivprnintdict['bdrbgpbfdint'] != None:
            bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" bfd-liveness true' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor']))
        if mivprnintdict['bdrbgpfeature'] != None:
            if 'as-override' in mivprnintdict['bdrbgpfeature']:
                bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" as-override true' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor']))
            if 'remove-private' in mivprnintdict['bdrbgpfeature']:
                bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" remove-private' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor']))
            if 'authentication-key' in mivprnintdict['bdrbgpfeature']:
                bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" authentication-key XX:PLEASE-REPLACE-THIS-KEY:XX' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor']))
            if 'prefix-limit-maximum' in mivprnintdict['bdrbgpfeature'] and 'prefix-limit-teardown' not in mivprnintdict['bdrbgpfeature']:
                prefixmax = mivprnintdict['bdrbgpfeature'].split('prefix-limit-maximum')[-1].split()[0]
                bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" prefix-limit %s maximum %s' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor'],mivprnintdict['bdrintfam'],prefixmax))
                bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" prefix-limit %s log-only true' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor'],mivprnintdict['bdrintfam']))
                bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" prefix-limit %s threshold 100' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor'],mivprnintdict['bdrintfam']))
            if 'prefix-limit-maximum' in mivprnintdict['bdrbgpfeature'] and 'prefix-limit-teardown' in mivprnintdict['bdrbgpfeature']:
                prefixmax = mivprnintdict['bdrbgpfeature'].split('prefix-limit-maximum')[-1].split()[0]
                prefixth = mivprnintdict['bdrbgpfeature'].split('prefix-limit-teardown')[-1].split()[0]
                bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" prefix-limit %s maximum %s' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor'],mivprnintdict['bdrintfam'],prefixmax))
                bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" prefix-limit %s threshold %s' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor'],mivprnintdict['bdrintfam'],prefixth))
        bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" group "%s"' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor'],mivprnintdict['bdrbgpgroup']))
        bdrconfiglist.append('/configure service vprn "%s" bgp neighbor "%s" admin-state disable' % (mivprnintdict['bdrservice'],mivprnintdict['bdrbgpneighbor']))
        bdrconfiglist.append('')

        for n in range(3, mapsheet.max_row + 1):
            #print(mapsheet['A%s' % n].value)
            if mivprnintdict['bdrservice'] == mapsheet['A%s' % n].value:
                if mivprnintdict['bdrintfam'] == 'ipv4':
                    bdrconfiglist.append('/configure filter match-list ip-prefix-list "BGP-Peers" apply-path bgp-peers %s group ".*"' % mapsheet['B%s' % n].value)
                    bdrconfiglist.append('/configure filter match-list ip-prefix-list "BGP-Peers" apply-path bgp-peers %s neighbor ".*"' % mapsheet['B%s' % n].value)
                    bdrconfiglist.append('/configure filter match-list ip-prefix-list "BGP-Peers" apply-path bgp-peers %s router-instance "%s"' % (mapsheet['B%s' % n].value,mivprnintdict['bdrservice']))
                elif mivprnintdict['bdrintfam'] == 'ipv6':
                    bdrconfiglist.append('/configure filter match-list ipv6-prefix-list "BGP-Peers" apply-path bgp-peers %s group ".*"' % mapsheet['B%s' % n].value)
                    bdrconfiglist.append('/configure filter match-list ipv6-prefix-list "BGP-Peers" apply-path bgp-peers %s neighbor ".*"' % mapsheet['B%s' % n].value)
                    bdrconfiglist.append('/configure filter match-list ipv6-prefix-list "BGP-Peers" apply-path bgp-peers %s router-instance "%s"' % (mapsheet['B%s' % n].value,mivprnintdict['bdrservice']))

    for i in bdrconfiglist:
        remarklist.append('vprn-int')

    if bdrvprnintws['%s1'%(chr(ord(column) - 1))].value == None:
        vprnintwsrow = bdrvprnintws['B1'].value
    else:
        vprnintwsrow = bdrvprnintws['%s1' % (chr(ord(column) - 1))].value

    if bdrpolicyws['%s1'%(chr(ord(column) - 1))].value == None:
        bdrpolicyrow = bdrpolicyws['B1'].value
    else:
        bdrpolicyrow = bdrpolicyws['%s1' % (chr(ord(column) - 1))].value


    for line in range(len(bdrconfiglist)):
        bdrvprnintws['%s%s' % (chr(ord(column) - 1), vprnintwsrow)] = remarklist[line]
        bdrvprnintws['%s%s' % (column, vprnintwsrow)] = bdrconfiglist[line]
        vprnintwsrow += 1
    for line in range(len(bdrpolicycfglist)):
        bdrpolicyws['%s%s' % (chr(ord(column) - 1), bdrpolicyrow)] = 'bgp pe-ce policy'
        bdrpolicyws['%s%s' % (column, bdrpolicyrow)] = bdrpolicycfglist[line]
        bdrpolicyrow += 1

    lastbdrnode = mivprnintdict['bdrnode']
    bdrvprnintws['%s1'%(chr(ord(column) - 1))] = vprnintwsrow
    bdrpolicyws['%s1'%(chr(ord(column) - 1))] = bdrpolicyrow
    return lastbdrnode, l3prefix

def bdrrvplsintworksheet(serviceid, mirvplsintdict, lastbdrnode, bdrcfgwb,l3prefix):
    bdrrvplsintws = bdrcfgwb['bdr-configuration']
    bdrfilterws = bdrcfgwb['bdr-filter']
    bdrconfiglist = []
    bdrpolicycfglist = []
    bdrfiltercfglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print('column %s' %bdrrvplsintws['A1'].value)
    if lastbdrnode != mirvplsintdict['bdrnode']:
        for c in columnlist:
            if bdrrvplsintws['%s2'%c].value != None:
                nodecolumnlist.append(bdrrvplsintws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mirvplsintdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrrvplsintws['D1'] = existcolumn[-1]
        if mirvplsintdict['bdrnode'] not in nodecolumnlist:
            bdrrvplsintws['D1'] = existcolumn[-1]
            column = bdrrvplsintws['D1'].value
            column = chr(ord(column) + 3)
            bdrrvplsintws['D1'] = column
            bdrrvplsintws['%s2' % column] = mirvplsintdict['bdrnode']
            bdrrvplsintws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrrvplsintws.column_dimensions['%s' % column].width = 120.0

    elif lastbdrnode:
        for c in columnlist:
            if bdrrvplsintws['%s2'%c].value != None:
                nodecolumnlist.append(bdrrvplsintws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mirvplsintdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrrvplsintws['D1'] = existcolumn[-1]

    #print(mirvplsintdict)
    networkaddr = IPNetwork(mirvplsintdict['bdrpriip']).network
    bdrpriip = mirvplsintdict['bdrpriip'].split('/')[0]
    bdrprilength = mirvplsintdict['bdrpriip'].split('/')[-1]
    l3prefix.append(str(networkaddr)+'/'+bdrprilength)
    if mirvplsintdict['bdrsecip'] != None:
        networkaddr = IPNetwork(mirvplsintdict['bdrsecip']).network
        bdrsecip = mirvplsintdict['bdrsecip'].split('/')[0]
        bdrseclength = mirvplsintdict['bdrsecip'].split('/')[-1]
        l3prefix.append(str(networkaddr)+'/'+bdrseclength)


    bdrconfiglist.append('')
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('# BDR VPRN Interface %s configuration' % mirvplsintdict['bdrintname'])
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('/configure service vprn "%s" interface "%s" admin-state disable'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname']))
    bdrconfiglist.append('/configure service vprn "%s" interface "%s" description "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintdesc']))
    bdrconfiglist.append('#/configure service vprn "%s" interface "%s" hold-time %s down seconds 600'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam']))
    bdrconfiglist.append('#/configure service vprn "%s" interface "%s" hold-time %s down init-only true'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam']))
    bdrconfiglist.append('')
    if mirvplsintdict['bdrservice'] == 'L3_Gn':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" load-balancing teid-load-balancing true'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname']))
    if mirvplsintdict['bdrintfam'] == 'ipv4':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s primary address %s'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],bdrpriip))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s primary prefix-length %s'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],bdrprilength))
        if mirvplsintdict['bdrsecip'] != None:
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s secondary %s prefix-length %s'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],bdrsecip,bdrseclength))
    elif mirvplsintdict['bdrintfam'] == 'ipv6':
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s prefix-length %s' % (mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],bdrpriip,bdrprilength))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s primary-preference 1'% (mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],bdrpriip))
        if mirvplsintdict['bdrsecip'] != None:
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s prefix-length %s'% (mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],bdrsecip,bdrseclength))
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s address %s primary-preference 10'% (mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],bdrsecip))

    if mirvplsintdict['bdrvrrpid'] != None:
        bdrvrrpinterval = ''
        if mirvplsintdict['bdrvrrpint'] != None and mirvplsintdict['bdrintfam'] == 'ipv4':
            bdrvrrpinterval = int(mirvplsintdict['bdrvrrpint'])//100
        elif mirvplsintdict['bdrvrrpint'] != None and mirvplsintdict['bdrintfam'] == 'ipv6':
            bdrvrrpinterval = int(mirvplsintdict['bdrvrrpint'])//10
        bdrconfiglist.append('')
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s admin-state enable'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],mirvplsintdict['bdrvrrpid']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s backup [%s]'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],mirvplsintdict['bdrvrrpid'],mirvplsintdict['bdrvrrpvip']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s ping-reply true'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],mirvplsintdict['bdrvrrpid']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s traceroute-reply true'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],mirvplsintdict['bdrvrrpid']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s standby-forwarding true'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],mirvplsintdict['bdrvrrpid']))
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" %s vrrp %s priority %s'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrintfam'],mirvplsintdict['bdrvrrpid'],mirvplsintdict['bdrvrrppri']))

    if 'SAP:$' not in mirvplsintdict['bdrrvpls']:
        bdrconfiglist.append('')
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" vpls "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrrvpls']))
        if mirvplsintdict['bdrrvplsinfilter'] != None and mirvplsintdict['bdrintfam'] == 'ipv4':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" vpls "%s" ingress routed-override-filter ip "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrrvpls'],mirvplsintdict['bdrrvplsinfilter']))
        elif mirvplsintdict['bdrrvplsinfilter'] != None and mirvplsintdict['bdrintfam'] == 'ipv6':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" vpls "%s" ingress routed-override-filter ipv6 "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrrvpls'],mirvplsintdict['bdrrvplsinfilter']))

        if mirvplsintdict['bdrrvplsegfilter'] != None and mirvplsintdict['bdrintfam'] == 'ipv4':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" vpls "%s" egress routed-override-filter ip "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrrvpls'],mirvplsintdict['bdrrvplsegfilter']))
        elif mirvplsintdict['bdrrvplsegfilter'] != None and mirvplsintdict['bdrintfam'] == 'ipv6':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" vpls "%s" egress routed-override-filter ipv6 "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],mirvplsintdict['bdrrvpls'],mirvplsintdict['bdrrvplsegfilter']))
    else:
        bdrconfiglist.append('')
        sapid = mirvplsintdict['bdrrvpls'].split(':$')[1]
        sapqos = mirvplsintdict['bdrrvpls'].split(':$')[3]
        bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],sapid))
        if sapqos:
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s ingress qos sap-ingress policy-name "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],sapid,sapqos))
        if mirvplsintdict['bdrrvplsinfilter'] != None and mirvplsintdict['bdrintfam'] == 'ipv4':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s ingress filter ip "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],sapid,mirvplsintdict['bdrrvplsinfilter']))
        elif mirvplsintdict['bdrrvplsinfilter'] != None and mirvplsintdict['bdrintfam'] == 'ipv6':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s ingress filter ipv6 "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],sapid,mirvplsintdict['bdrrvplsinfilter']))
        if sapqos:
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s egress qos sap-egress policy-name "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],sapid,sapqos))
        if mirvplsintdict['bdrrvplsegfilter'] != None and mirvplsintdict['bdrintfam'] == 'ipv4':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s egress filter ip "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],sapid,mirvplsintdict['bdrrvplsegfilter']))
        elif mirvplsintdict['bdrrvplsegfilter'] != None and mirvplsintdict['bdrintfam'] == 'ipv6':
            bdrconfiglist.append('/configure service vprn "%s" interface "%s" sap %s egress filter ipv6 "%s"'%(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname'],sapid,mirvplsintdict['bdrrvplsegfilter']))

    if mirvplsintdict['bdrvrrpid'] != None and mirvplsintdict['bdrintfam'] == 'ipv6':
        bdrconfiglist.append('/configure service vprn "%s" ipv6 router-advertisement interface "%s" admin-state enable' %(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname']))
        bdrconfiglist.append('/configure service vprn "%s" ipv6 router-advertisement interface "%s" use-virtual-mac true' %(mirvplsintdict['bdrservice'],mirvplsintdict['bdrintname']))

    if mirvplsintdict['bdrstatic'] != None:
        bdrconfiglist.append('')
        bdrconfiglist.append('######################################################')
        bdrconfiglist.append('# BDR PE-CE Static Routing configuration of interface %s' % mirvplsintdict['bdrintname'])
        staticpath = mirvplsintdict['bdrstatic'].strip('.')
        sta = open(migratepath + misumtime + staticpath, "r")
        staticline = sta.readlines()
        sta.close()
        for st in range(len(staticline)):
            staticline[st] = re.sub(r"^\s+", "", staticline[st])  # remove space from beginning
            staticline[st] = staticline[st].rstrip('\n')  # remove newline('\n') from end of line
            staticline[st] = re.sub(r"\s+$", "", staticline[st])  # remove space from ending
            #print(staticline[st])
            bdrconfiglist.append(staticline[st])
            if 'static-routes route' in staticline[st]:
                l3prefix.append(staticline[st].split()[6])

    if mirvplsintdict['bdrrvplsinfilterpol'] != None:
        bdrfiltercfglist.append('')
        mirvplsintdict['bdrrvplsinfilterpol'] = mirvplsintdict['bdrrvplsinfilterpol'].strip('.')
        f = open(migratepath + misumtime + mirvplsintdict['bdrrvplsinfilterpol'] , "r")
        lines = f.readlines()
        f.close()
        for a in range(len(lines)):
            lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
            lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
            lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
            #print(lines[a])
            if 'prefix-list' in lines[a]:
                if '%s_prefix' % mirvplsintdict['bdrnode'] in bdrfilterdict:
                    for existingprefixlist in bdrfilterdict['%s_prefix' % mirvplsintdict['bdrnode']]:
                        if existingprefixlist != None:
                            #print(existingprefixlist)
                            bdrprefixlist = bdrfilterdict['%s_pr_%s' % (mirvplsintdict['bdrnode'], existingprefixlist)]
                            #print(bdrprefixlist)
                            lines[a] = lines[a].replace(existingprefixlist,bdrprefixlist)
            if 'port-list' in lines[a]:
                if '%s_port' % mirvplsintdict['bdrnode'] in bdrfilterdict:
                    for existingportlist in bdrfilterdict['%s_port' % mirvplsintdict['bdrnode']]:
                        if existingportlist != None:
                            #print(existingportlist)
                            bdrportlist = bdrfilterdict['%s_po_%s' % (mirvplsintdict['bdrnode'], existingportlist)]
                            #print(bdrportlist)
                            lines[a] = lines[a].replace(existingportlist, bdrportlist)
            bdrfiltercfglist.append(lines[a])

    if mirvplsintdict['bdrrvplsegfilterpol'] != None:
        bdrfiltercfglist.append('')
        mirvplsintdict['bdrrvplsegfilterpol'] = mirvplsintdict['bdrrvplsegfilterpol'].strip('.')
        f = open(migratepath + misumtime + mirvplsintdict['bdrrvplsegfilterpol'] , "r")
        lines = f.readlines()
        f.close()
        for a in range(len(lines)):
            lines[a] = re.sub(r"^\s+", "", lines[a])  # remove space from beginning
            lines[a] = lines[a].rstrip('\n')  # remove newline('\n') from end of line
            lines[a] = re.sub(r"\s+$", "", lines[a])  # remove space from ending
            #print(lines[a])
            if 'prefix-list' in lines[a]:
                if '%s_prefix' % mirvplsintdict['bdrnode'] in bdrfilterdict:
                    for existingprefixlist in bdrfilterdict['%s_prefix' % mirvplsintdict['bdrnode']]:
                        if existingprefixlist != None:
                            #print(existingprefixlist)
                            bdrprefixlist = bdrfilterdict['%s_pr_%s' % (mirvplsintdict['bdrnode'], existingprefixlist)]
                            #print(bdrprefixlist)
                            lines[a] = lines[a].replace(existingprefixlist,bdrprefixlist)
            if 'port-list' in lines[a]:
                if '%s_port' % mirvplsintdict['bdrnode'] in bdrfilterdict:
                    for existingportlist in bdrfilterdict['%s_port' % mirvplsintdict['bdrnode']]:
                        if existingportlist != None:
                            #print(existingportlist)
                            bdrportlist = bdrfilterdict['%s_po_%s' % (mirvplsintdict['bdrnode'], existingportlist)]
                            #print(bdrportlist)
                            lines[a] = lines[a].replace(existingportlist, bdrportlist)

            bdrfiltercfglist.append(lines[a])

    for i in bdrconfiglist:
        remarklist.append('rvpls-int')

    if bdrrvplsintws['%s1'%(chr(ord(column) - 1))].value == None:
        rvplsintrow = bdrrvplsintws['B1'].value
    else:
        rvplsintrow = bdrrvplsintws['%s1' % (chr(ord(column) - 1))].value

    if bdrfilterws['%s1'%(chr(ord(column) - 1))].value == None:
        bdrfilterrow = bdrfilterws['B1'].value
    else:
        bdrfilterrow = bdrfilterws['%s1' % (chr(ord(column) - 1))].value

    for line in range(len(bdrconfiglist)):
        bdrrvplsintws['%s%s' % (chr(ord(column) - 1), rvplsintrow)] = remarklist[line]
        bdrrvplsintws['%s%s' % (column, rvplsintrow)] = bdrconfiglist[line]
        rvplsintrow += 1

    for line in range(len(bdrfiltercfglist)):
        bdrfilterws['%s%s' % (column, bdrfilterrow)] = bdrfiltercfglist[line]
        bdrfilterrow += 1

    lastbdrnode = mirvplsintdict['bdrnode']
    bdrrvplsintws['%s1'%(chr(ord(column) - 1))] = rvplsintrow
    bdrfilterws['%s1' % (chr(ord(column) - 1))] = bdrfilterrow

    return lastbdrnode, l3prefix

def createqosscheduler(sapqos,rate,sch_name):
    qoscfg = []
    sch_rate = ''
    if 'G' in rate or 'g' in rate:
        sch_rate = int(float(rate[:-1]) * 1000000)
    elif 'M' in rate or 'm' in rate:
        sch_rate = int(float(rate[:-1]) * 1000)


    if sapqos == 'best-effort':
        qoscfg.append('')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" description "sap-ingress policy for best-effort traffic"')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" policy-id 101')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" default-fc "be"')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" { queue 1 }')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 1 percent-rate pir 100.0')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 1 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 1 cbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 1 mbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 1 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 11 multipoint true')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 11 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 11 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 11 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 12 multipoint true')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 12 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 12 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" queue 12 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" fc "be" queue 1')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" fc "be" multicast-queue 11')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" fc "be" broadcast-queue 12')
        qoscfg.append('/configure qos sap-ingress "best-effort_scheduler" fc "be" unknown-queue 12')
    elif sapqos == 'corporate':
        qoscfg.append('')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" description "sap-ingress policy for corporate traffic"')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" policy-id 501')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" default-fc "h2"')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" default-priority high')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" { queue 5 }')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 5 percent-rate pir 100.0')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 5 percent-rate cir 100.0')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 5 cbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 5 mbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 5 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 11 multipoint true')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 11 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 11 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 11 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 12 multipoint true')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 12 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 12 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" queue 12 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" fc "h2" queue 5')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" fc "h2" multicast-queue 11')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" fc "h2" broadcast-queue 12')
        qoscfg.append('/configure qos sap-ingress "corporate_scheduler" fc "h2" unknown-queue 12')
    elif sapqos == 'internal':
        qoscfg.append('')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" description "sap-ingress policy for internal traffic"')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" policy-id 301')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" default-fc "af"')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" default-priority high')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" { queue 3 }')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 3 percent-rate pir 100.0')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 3 percent-rate cir 100.0')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 3 cbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 3 mbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 3 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 11 multipoint true')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 11 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 11 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 11 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 12 multipoint true')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 12 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 12 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" queue 12 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" fc "af" queue 3')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" fc "af" multicast-queue 11')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" fc "af" broadcast-queue 12')
        qoscfg.append('/configure qos sap-ingress "internal_scheduler" fc "af" unknown-queue 12')
    elif sapqos == 'mobile':
        qoscfg.append('')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" description "sap-ingress policy for mobile traffic"')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" policy-id 701')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" default-fc "h1"')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" default-priority high')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" { queue 7 }')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 7 percent-rate pir 100.0')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 7 percent-rate cir 100.0')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 7 cbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 7 mbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 7 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 11 multipoint true')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 11 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 11 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 11 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 12 multipoint true')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 12 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 12 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" queue 12 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" fc "h1" queue 7')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" fc "h1" multicast-queue 11')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" fc "h1" broadcast-queue 12')
        qoscfg.append('/configure qos sap-ingress "mobile_scheduler" fc "h1" unknown-queue 12')
    elif sapqos == 'network-control':
        qoscfg.append('')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" description "sap-ingress policy for network-control traffic"')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" policy-id 801')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" default-fc "nc"')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" default-priority high')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" { queue 8 }')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 8 percent-rate pir 100.0')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 8 percent-rate cir 100.0')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 8 cbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 8 mbs 150 megabytes')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 8 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 11 multipoint true')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 11 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 11 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 11 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 12 multipoint true')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 12 percent-rate pir 1.0')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 12 percent-rate cir 0.0')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" queue 12 scheduler-parent scheduler-name root')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" fc "nc" queue 8')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" fc "nc" multicast-queue 11')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" fc "nc" broadcast-queue 12')
        qoscfg.append('/configure qos sap-ingress "network-control_scheduler" fc "nc" unknown-queue 12')

    qoscfg.append('')
    qoscfg.append('/configure qos scheduler-policy "%s" tier 1 scheduler "root" rate pir %s' % (sch_name,sch_rate))
    qoscfg.append('/configure qos scheduler-policy "%s" tier 1 scheduler "root" rate cir %s' % (sch_name,sch_rate))
    qoscfg.append('')

    return qoscfg

def bdrepipeinterasworksheet(serviceid, miepipeinterasdict, lastbdrnode, bdrcfgwb):
    bdrepipeinterasws = bdrcfgwb['l2inter-as-configuration']
    bdrconfiglist = []
    bdrpolicycfglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print('column %s' %bdrepipeinterasws['A1'].value)
    if lastbdrnode != miepipeinterasdict['bdrnode']:
        for c in columnlist:
            if bdrepipeinterasws['%s2'%c].value != None:
                nodecolumnlist.append(bdrepipeinterasws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if miepipeinterasdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrepipeinterasws['D1'] = existcolumn[-1]
        if miepipeinterasdict['bdrnode'] not in nodecolumnlist:
            bdrepipeinterasws['D1'] = existcolumn[-1]
            column = bdrepipeinterasws['D1'].value
            column = chr(ord(column) + 3)
            bdrepipeinterasws['D1'] = column
            bdrepipeinterasws['%s2' % column] = miepipeinterasdict['bdrnode']
            bdrepipeinterasws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrepipeinterasws.column_dimensions['%s' % column].width = 120.0

    elif lastbdrnode:
        for c in columnlist:
            if bdrepipeinterasws['%s2' % c].value != None:
                nodecolumnlist.append(bdrepipeinterasws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if miepipeinterasdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrepipeinterasws['D1'] = existcolumn[-1]

    if miepipeinterasdict['bdrsapinrate'] != None:
        bdrqosinpol = miepipeinterasdict['bdrsapqos']+'_scheduler'
        bdrschinpol = 'scheduler_'+miepipeinterasdict['bdrsapinrate']
    else:
        bdrqosinpol = miepipeinterasdict['bdrsapqos']

    bdrqosegpol = miepipeinterasdict['bdrsapqos']

    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('# BDR EPIPE SAP configuration')
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service epipe "%s" sap %s admin-state disable' % (miepipeinterasdict['bdrservice'],miepipeinterasdict['bdrsap']))
    if miepipeinterasdict['bdrsapdesc'] != None:
        bdrconfiglist.append('/configure service epipe "%s" sap %s description "%s"' % (miepipeinterasdict['bdrservice'],miepipeinterasdict['bdrsap'],miepipeinterasdict['bdrsapdesc']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s ingress qos sap-ingress policy-name "%s"' % (miepipeinterasdict['bdrservice'],miepipeinterasdict['bdrsap'],bdrqosinpol))
    if miepipeinterasdict['bdrsapinrate'] != None:
        bdrconfiglist.append('/configure service epipe "%s" sap %s ingress qos scheduler-policy policy-name "%s"' % (miepipeinterasdict['bdrservice'],miepipeinterasdict['bdrsap'],bdrschinpol))
    if miepipeinterasdict['bdrsapinfilter'] != None:
        bdrconfiglist.append('/configure service epipe "%s" sap %s ingress filter ip "%s"' % (miepipeinterasdict['bdrservice'],miepipeinterasdict['bdrsap'],miepipeinterasdict['bdrsapinfilter']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s egress qos sap-egress policy-name "%s"' % (miepipeinterasdict['bdrservice'],miepipeinterasdict['bdrsap'],bdrqosegpol))
    if miepipeinterasdict['bdrsapegfilter'] != None:
        bdrconfiglist.append('/configure service epipe "%s" sap %s egress filter ip "%s"' % (miepipeinterasdict['bdrservice'],miepipeinterasdict['bdrsap'],miepipeinterasdict['bdrsapegfilter']))

    if miepipeinterasdict['bdrsapinrate'] != None:
        schcfg = createqosscheduler(miepipeinterasdict['bdrsapqos'], miepipeinterasdict['bdrsapinrate'],bdrschinpol)
        bdrconfiglist.extend(schcfg)

    for i in bdrconfiglist:
        remarklist.append('epipe-sap')

    if bdrepipeinterasws['%s1'%(chr(ord(column) - 1))].value == None:
        epipeinteraswsrow = bdrepipeinterasws['B1'].value
    else:
        epipeinteraswsrow = bdrepipeinterasws['%s1' % (chr(ord(column) - 1))].value

    for line in range(len(bdrconfiglist)):
        bdrepipeinterasws['%s%s' % (chr(ord(column) - 1), epipeinteraswsrow)] = remarklist[line]
        bdrepipeinterasws['%s%s' % (column, epipeinteraswsrow)] = bdrconfiglist[line]
        epipeinteraswsrow += 1

    lastbdrnode = miepipeinterasdict['bdrnode']
    bdrepipeinterasws['%s1'%(chr(ord(column) - 1))] = epipeinteraswsrow

    return lastbdrnode

def bdrepipesapworksheet(serviceid, miepipesapdict, lastbdrnode, bdrcfgwb, misum):
    bdrepipesapws = bdrcfgwb['bdr-configuration']
    bdrconfiglist = []
    bdrpolicycfglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print('column %s' %bdrepipesapws['A1'].value)
    if lastbdrnode != miepipesapdict['bdrnode']:
        for c in columnlist:
            if bdrepipesapws['%s2'%c].value != None:
                nodecolumnlist.append(bdrepipesapws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if miepipesapdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrepipesapws['D1'] = existcolumn[-1]
        if miepipesapdict['bdrnode'] not in nodecolumnlist:
            bdrepipesapws['D1'] = existcolumn[-1]
            column = bdrepipesapws['D1'].value
            column = chr(ord(column) + 3)
            bdrepipesapws['D1'] = column
            bdrepipesapws['%s2' % column] = miepipesapdict['bdrnode']
            bdrepipesapws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrepipesapws.column_dimensions['%s' % column].width = 120.0

        sdpserviceidlist = []
        for l2sdpline in range(3, misum['l2service-sdp'].max_row + 1):
            if misum['l2service-sdp']['D%s' % l2sdpline].value != None:
                sdpserviceidlist.append(int(misum['l2service-sdp']['D%s' % l2sdpline].value))

        if int(miepipesapdict['bdrserviceid']) not in sdpserviceidlist:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR EPIPE service configuration ###')
            bdrconfiglist.append('/configure service epipe "%s" admin-state enable' % miepipesapdict['bdrservice'])
            bdrconfiglist.append('/configure service epipe "%s" description "%s"' % (miepipesapdict['bdrservice'], miepipesapdict['bdrservice']))
            bdrconfiglist.append('/configure service epipe "%s" service-id %s' % (miepipesapdict['bdrservice'], miepipesapdict['bdrserviceid']))
            bdrconfiglist.append('/configure service epipe "%s" customer "%s"' % (miepipesapdict['bdrservice'], miepipesapdict['bdrsapqos']))
            bdrconfiglist.append('/configure service epipe "%s" service-mtu 9194' % miepipesapdict['bdrservice'])

    elif lastbdrnode:
        for c in columnlist:
            if bdrepipesapws['%s2' % c].value != None:
                nodecolumnlist.append(bdrepipesapws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if miepipesapdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrepipesapws['D1'] = existcolumn[-1]

    if miepipesapdict['bdrsapinrate'] != None:
        bdrqosinpol = miepipesapdict['bdrsapqos']+'_scheduler'
        bdrschinpol = 'scheduler_'+miepipesapdict['bdrsapinrate']
    else:
        bdrqosinpol = miepipesapdict['bdrsapqos']

    bdrqosegpol = miepipesapdict['bdrsapqos']

    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('# BDR EPIPE SAP configuration')
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service epipe "%s" sap %s admin-state disable' % (miepipesapdict['bdrservice'],miepipesapdict['bdrsap']))
    if miepipesapdict['bdrsapdesc'] != None:
        bdrconfiglist.append('/configure service epipe "%s" sap %s description "%s"' % (miepipesapdict['bdrservice'],miepipesapdict['bdrsap'],miepipesapdict['bdrsapdesc']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s ingress qos sap-ingress policy-name "%s"' % (miepipesapdict['bdrservice'],miepipesapdict['bdrsap'],bdrqosinpol))
    if miepipesapdict['bdrsapinrate'] != None:
        bdrconfiglist.append('/configure service epipe "%s" sap %s ingress qos scheduler-policy policy-name "%s"' % (miepipesapdict['bdrservice'],miepipesapdict['bdrsap'],bdrschinpol))
    if miepipesapdict['bdrsapinfilter'] != None:
        bdrconfiglist.append('/configure service epipe "%s" sap %s ingress filter ip "%s"' % (miepipesapdict['bdrservice'],miepipesapdict['bdrsap'],miepipesapdict['bdrsapinfilter']))
    bdrconfiglist.append('/configure service epipe "%s" sap %s egress qos sap-egress policy-name "%s"' % (miepipesapdict['bdrservice'],miepipesapdict['bdrsap'],bdrqosegpol))
    if miepipesapdict['bdrsapegfilter'] != None:
        bdrconfiglist.append('/configure service epipe "%s" sap %s egress filter ip "%s"' % (miepipesapdict['bdrservice'],miepipesapdict['bdrsap'],miepipesapdict['bdrsapegfilter']))

    if miepipesapdict['bdrsapinrate'] != None:
        schcfg = createqosscheduler(miepipesapdict['bdrsapqos'], miepipesapdict['bdrsapinrate'],bdrschinpol)
        bdrconfiglist.extend(schcfg)

    for i in bdrconfiglist:
        remarklist.append('epipe-sap')

    if bdrepipesapws['%s1'%(chr(ord(column) - 1))].value == None:
        epipesapwsrow = bdrepipesapws['B1'].value
    else:
        epipesapwsrow = bdrepipesapws['%s1' % (chr(ord(column) - 1))].value

    for line in range(len(bdrconfiglist)):
        bdrepipesapws['%s%s' % (chr(ord(column) - 1), epipesapwsrow)] = remarklist[line]
        bdrepipesapws['%s%s' % (column, epipesapwsrow)] = bdrconfiglist[line]
        epipesapwsrow += 1

    lastbdrnode = miepipesapdict['bdrnode']
    bdrepipesapws['%s1'%(chr(ord(column) - 1))] = epipesapwsrow

    return lastbdrnode

def bdrvplssapworksheet(serviceid, mivplssapdict, lastbdrnode, bdrcfgwb, misum):
    bdrvplssapws = bdrcfgwb['bdr-configuration']
    bdrconfiglist = []
    bdrpolicycfglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print('column %s' %bdrvplssapws['A1'].value)
    if lastbdrnode != mivplssapdict['bdrnode']:
        for c in columnlist:
            if bdrvplssapws['%s2'%c].value != None:
                nodecolumnlist.append(bdrvplssapws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mivplssapdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrvplssapws['D1'] = existcolumn[-1]
        if mivplssapdict['bdrnode'] not in nodecolumnlist:
            bdrvplssapws['D1'] = existcolumn[-1]
            column = bdrvplssapws['D1'].value
            column = chr(ord(column) + 3)
            bdrvplssapws['D1'] = column
            bdrvplssapws['%s2' % column] = mivplssapdict['bdrnode']
            bdrvplssapws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrvplssapws.column_dimensions['%s' % column].width = 120.0

        sdpserviceidlist = []
        for l2sdpline in range(3,misum['l2service-sdp'].max_row + 1):
            if misum['l2service-sdp']['D%s' % l2sdpline].value != None:
                sdpserviceidlist.append(int(misum['l2service-sdp']['D%s' % l2sdpline].value))

        if int(mivplssapdict['bdrserviceid']) not in sdpserviceidlist:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR VPLS service configuration ###')
            bdrconfiglist.append('/configure service vpls "%s" admin-state enable' % mivplssapdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" description "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrservice']))
            bdrconfiglist.append('/configure service vpls "%s" service-id %s' % (mivplssapdict['bdrservice'],mivplssapdict['bdrserviceid']))
            bdrconfiglist.append('/configure service vpls "%s" customer "%s"' % (mivplssapdict['bdrservice'], mivplssapdict['bdrsapqos']))
            bdrconfiglist.append('/configure service vpls "%s" service-mtu 9194' % mivplssapdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" mac-flush tldp send-on-failure true' % mivplssapdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" fdb mac-move admin-state enable' % mivplssapdict['bdrservice'])

    elif lastbdrnode:
        for c in columnlist:
            if bdrvplssapws['%s2' % c].value != None:
                nodecolumnlist.append(bdrvplssapws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mivplssapdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrvplssapws['D1'] = existcolumn[-1]

    bdrqosinpol = mivplssapdict['bdrsapqos']
    bdrqosegpol = mivplssapdict['bdrsapqos']

    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('# BDR VPLS SAP configuration')
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service vpls "%s" sap %s admin-state disable' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap']))
    if mivplssapdict['bdrsapdesc'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s description "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],mivplssapdict['bdrsapdesc']))
    bdrconfiglist.append('/configure service vpls "%s" sap %s ingress qos sap-ingress policy-name "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],bdrqosinpol))
    if mivplssapdict['bdrsapinfilter'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s ingress filter ip "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],mivplssapdict['bdrsapinfilter']))
    bdrconfiglist.append('/configure service vpls "%s" sap %s egress qos sap-egress policy-name "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],bdrqosegpol))
    if mivplssapdict['bdrsapegfilter'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s egress filter ip "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],mivplssapdict['bdrsapegfilter']))

    for i in bdrconfiglist:
        remarklist.append('vpls-sap')

    if bdrvplssapws['%s1'%(chr(ord(column) - 1))].value == None:
        vplssapwsrow = bdrvplssapws['B1'].value
    else:
        vplssapwsrow = bdrvplssapws['%s1' % (chr(ord(column) - 1))].value

    for line in range(len(bdrconfiglist)):
        bdrvplssapws['%s%s' % (chr(ord(column) - 1), vplssapwsrow)] = remarklist[line]
        bdrvplssapws['%s%s' % (column, vplssapwsrow)] = bdrconfiglist[line]
        vplssapwsrow += 1

    lastbdrnode = mivplssapdict['bdrnode']
    bdrvplssapws['%s1'%(chr(ord(column) - 1))] = vplssapwsrow
    return lastbdrnode

def bdrvplsinterasworksheet(serviceid, mivplssapdict, lastbdrnode, bdrcfgwb):
    bdrvplssapws = bdrcfgwb['l2inter-as-configuration']
    bdrconfiglist = []
    bdrpolicycfglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print('column %s' %bdrvplssapws['A1'].value)
    if lastbdrnode != mivplssapdict['bdrnode']:
        for c in columnlist:
            if bdrvplssapws['%s2'%c].value != None:
                nodecolumnlist.append(bdrvplssapws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mivplssapdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrvplssapws['D1'] = existcolumn[-1]
        if mivplssapdict['bdrnode'] not in nodecolumnlist:
            bdrvplssapws['D1'] = existcolumn[-1]
            column = bdrvplssapws['D1'].value
            column = chr(ord(column) + 3)
            bdrvplssapws['D1'] = column
            bdrvplssapws['%s2' % column] = mivplssapdict['bdrnode']
            bdrvplssapws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrvplssapws.column_dimensions['%s' % column].width = 120.0

    elif lastbdrnode:
        for c in columnlist:
            if bdrvplssapws['%s2' % c].value != None:
                nodecolumnlist.append(bdrvplssapws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mivplssapdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrvplssapws['D1'] = existcolumn[-1]

    bdrqosinpol = mivplssapdict['bdrsapqos']
    bdrqosegpol = mivplssapdict['bdrsapqos']

    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('# BDR VPLS SAP configuration')
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service vpls "%s" sap %s admin-state disable' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap']))
    if mivplssapdict['bdrsapdesc'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s description "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],mivplssapdict['bdrsapdesc']))
    bdrconfiglist.append('/configure service vpls "%s" sap %s ingress qos sap-ingress policy-name "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],bdrqosinpol))
    if mivplssapdict['bdrsapinfilter'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s ingress filter ip "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],mivplssapdict['bdrsapinfilter']))
    bdrconfiglist.append('/configure service vpls "%s" sap %s egress qos sap-egress policy-name "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],bdrqosegpol))
    if mivplssapdict['bdrsapegfilter'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s egress filter ip "%s"' % (mivplssapdict['bdrservice'],mivplssapdict['bdrsap'],mivplssapdict['bdrsapegfilter']))

    for i in bdrconfiglist:
        remarklist.append('vpls-sap')

    if bdrvplssapws['%s1'%(chr(ord(column) - 1))].value == None:
        vplssapwsrow = bdrvplssapws['B1'].value
    else:
        vplssapwsrow = bdrvplssapws['%s1' % (chr(ord(column) - 1))].value

    for line in range(len(bdrconfiglist)):
        bdrvplssapws['%s%s' % (chr(ord(column) - 1), vplssapwsrow)] = remarklist[line]
        bdrvplssapws['%s%s' % (column, vplssapwsrow)] = bdrconfiglist[line]
        vplssapwsrow += 1

    lastbdrnode = mivplssapdict['bdrnode']
    bdrvplssapws['%s1'%(chr(ord(column) - 1))] = vplssapwsrow
    return lastbdrnode

def bdrrvplssapworksheet(serviceid, mirvplssapdict, lastbdrnode, bdrcfgwb,misum):
    bdrrvplssapws = bdrcfgwb['bdr-configuration']
    bdrconfiglist = []
    bdrpolicycfglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    #print('column %s' %bdrrvplssapws['A1'].value)
    if lastbdrnode != mirvplssapdict['bdrnode']:
        for c in columnlist:
            if bdrrvplssapws['%s2'%c].value != None:
                nodecolumnlist.append(bdrrvplssapws['%s2'%c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mirvplssapdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrrvplssapws['D1'] = existcolumn[-1]
        if mirvplssapdict['bdrnode'] not in nodecolumnlist:
            bdrrvplssapws['D1'] = existcolumn[-1]
            column = bdrrvplssapws['D1'].value
            column = chr(ord(column) + 3)
            bdrrvplssapws['D1'] = column
            bdrrvplssapws['%s2' % column] = mirvplssapdict['bdrnode']
            bdrrvplssapws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrrvplssapws.column_dimensions['%s' % column].width = 120.0

        sdpserviceidlist = []
        for l2sdpline in range(3,misum['l2service-sdp'].max_row + 1):
            if misum['l2service-sdp']['D%s' % l2sdpline].value != None:
                sdpserviceidlist.append(int(misum['l2service-sdp']['D%s' % l2sdpline].value))

        if int(mirvplssapdict['bdrserviceid']) not in sdpserviceidlist:
            bdrconfiglist.append('')
            bdrconfiglist.append('### BDR VPLS service configuration  ###')
            bdrconfiglist.append('/configure service vpls "%s" admin-state enable' % mirvplssapdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" description "%s"' % (mirvplssapdict['bdrservice'],mirvplssapdict['bdrservice']))
            bdrconfiglist.append('/configure service vpls "%s" service-id %s' % (mirvplssapdict['bdrservice'],mirvplssapdict['bdrserviceid']))
            bdrconfiglist.append('/configure service vpls "%s" customer "%s"' % (mirvplssapdict['bdrservice'], mirvplssapdict['bdrsapqos']))
            bdrconfiglist.append('/configure service vpls "%s" service-mtu 9194' % mirvplssapdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" mac-flush tldp send-on-failure true' % mirvplssapdict['bdrservice'])
            bdrconfiglist.append('/configure service vpls "%s" fdb mac-move admin-state enable' % mirvplssapdict['bdrservice'])
            if 'rVPLS' in mirvplssapdict['bdrservice']:
                bdrconfiglist.append('/configure service vpls "%s" routed-vpls' % mirvplssapdict['bdrservice'])

    elif lastbdrnode:
        for c in columnlist:
            if bdrrvplssapws['%s2' % c].value != None:
                nodecolumnlist.append(bdrrvplssapws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mirvplssapdict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrrvplssapws['D1'] = existcolumn[-1]



    bdrqosinpol = mirvplssapdict['bdrsapqos']
    bdrqosegpol = mirvplssapdict['bdrsapqos']

    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('# BDR rVPLS SAP configuration')
    bdrconfiglist.append('######################################################')
    bdrconfiglist.append('')
    bdrconfiglist.append('/configure service vpls "%s" sap %s admin-state disable' % (mirvplssapdict['bdrservice'],mirvplssapdict['bdrsap']))
    if mirvplssapdict['bdrsapdesc'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s description "%s"' % (mirvplssapdict['bdrservice'],mirvplssapdict['bdrsap'],mirvplssapdict['bdrsapdesc']))
    bdrconfiglist.append('/configure service vpls "%s" sap %s ingress qos sap-ingress policy-name "%s"' % (mirvplssapdict['bdrservice'],mirvplssapdict['bdrsap'],bdrqosinpol))
    if mirvplssapdict['bdrsapinfilter'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s ingress filter ip "%s"' % (mirvplssapdict['bdrservice'],mirvplssapdict['bdrsap'],mirvplssapdict['bdrsapinfilter']))
    bdrconfiglist.append('/configure service vpls "%s" sap %s egress qos sap-egress policy-name "%s"' % (mirvplssapdict['bdrservice'],mirvplssapdict['bdrsap'],bdrqosegpol))
    if mirvplssapdict['bdrsapegfilter'] != None:
        bdrconfiglist.append('/configure service vpls "%s" sap %s egress filter ip "%s"' % (mirvplssapdict['bdrservice'],mirvplssapdict['bdrsap'],mirvplssapdict['bdrsapegfilter']))

    for i in bdrconfiglist:
        remarklist.append('rvpls-sap')

    if bdrrvplssapws['%s1'%(chr(ord(column) - 1))].value == None:
        rvplssapwsrow = bdrrvplssapws['B1'].value
    else:
        rvplssapwsrow = bdrrvplssapws['%s1' % (chr(ord(column) - 1))].value

    for line in range(len(bdrconfiglist)):
        bdrrvplssapws['%s%s' % (chr(ord(column) - 1), rvplssapwsrow)] = remarklist[line]
        bdrrvplssapws['%s%s' % (column, rvplssapwsrow)] = bdrconfiglist[line]
        rvplssapwsrow += 1

    lastbdrnode = mirvplssapdict['bdrnode']
    bdrrvplssapws['%s1'%(chr(ord(column) - 1))] = rvplssapwsrow
    return lastbdrnode


def creatervplsintdict(rvplsintws, row):
    rvplsintdict = {}
    rvplsintdict['bdrnode'] = rvplsintws['A%s' % row].value # 'nokia-node'
    rvplsintdict['bdrsystem'] = rvplsintws['B%s' % row].value # 'system-ip'
    rvplsintdict['bdrservice'] = rvplsintws['C%s' % row].value # 'service-name'
    rvplsintdict['bdrserviceid'] = rvplsintws['D%s' % row].value # 'service-id'
    rvplsintdict['bdrintname'] = rvplsintws['E%s' % row].value # 'interface-name'
    rvplsintdict['bdrintdesc'] = rvplsintws['F%s' % row].value # 'interface description'
    rvplsintdict['bdrrvpls'] = rvplsintws['G%s' % row].value # 'vpls'
    rvplsintdict['bdrrvplsinfilter'] = rvplsintws['H%s' % row].value # 'vpls-ingress-filter'
    rvplsintdict['bdrrvplsinfilterpol'] = rvplsintws['I%s' % row].value # 'rvpls-ingress-filter-policy'
    rvplsintdict['bdrrvplsegfilter'] = rvplsintws['J%s' % row].value # 'vpls-egress-filter'
    rvplsintdict['bdrrvplsegfilterpol'] = rvplsintws['K%s' % row].value # 'rvpls-egress-filter-policy'
    rvplsintdict['bdrpriip'] = rvplsintws['L%s' % row].value # 'ip/ipv6 address'
    rvplsintdict['bdrsecip'] = rvplsintws['M%s' % row].value # 'secondary ip/ipv6'
    rvplsintdict['bdrintfam'] = rvplsintws['N%s' % row].value # 'family'
    rvplsintdict['bdrvrrpid'] = rvplsintws['O%s' % row].value # 'vrrp id'
    rvplsintdict['bdrvrrpvip'] = rvplsintws['P%s' % row].value # 'vrrp vip'
    rvplsintdict['bdrvrrppri'] = rvplsintws['Q%s' % row].value # 'vrrp priority'
    rvplsintdict['bdrvrrpint'] = rvplsintws['R%s' % row].value # 'vrrp interval'
    rvplsintdict['bdrstatic'] = rvplsintws['S%s' % row].value # 'static-route'
    rvplsintdict['iptnnode'] = rvplsintws['T%s' % row].value # 'iptn-node'
    rvplsintdict['iptnservice'] = rvplsintws['U%s' % row].value # 'iptn-service'
    rvplsintdict['iptnint'] = rvplsintws['V%s' % row].value # 'iptn-interface'
    return rvplsintdict

def creatervplssapdict(rvplssapws, row):
    rvplssapdict = {}
    rvplssapdict['bdrnode'] = rvplssapws['A%s' % row].value  # 'nokia-node'
    rvplssapdict['bdrsystem'] = rvplssapws['B%s' % row].value  # 'system-ip'
    rvplssapdict['bdrservice'] = rvplssapws['C%s' % row].value  # 'service-name'
    rvplssapdict['bdrserviceid'] = rvplssapws['D%s' % row].value  # 'service-id'
    rvplssapdict['bdrsap'] = rvplssapws['E%s' % row].value  # 'sap'
    rvplssapdict['bdrsapdesc'] = rvplssapws['F%s' % row].value  # 'sap-description'
    rvplssapdict['bdrsapinfilter'] = rvplssapws['G%s' % row].value  # 'sap-ingress-filter'
    rvplssapdict['bdrsapegfilter'] = rvplssapws['H%s' % row].value  # 'sap-egress-filter'
    rvplssapdict['bdrsapqos'] = rvplssapws['I%s' % row].value  # 'sap-in/egress qos'
    rvplssapdict['iptnnode'] = rvplssapws['J%s' % row].value  # 'iptn-node'
    rvplssapdict['iptnservice'] = rvplssapws['K%s' % row].value  # 'iptn-service'
    rvplssapdict['iptnint'] = rvplssapws['L%s' % row].value  # 'iptn-interface'
    return rvplssapdict

def createvplssapdict(vplssapws, row):
    vplssapdict = {}
    vplssapdict['bdrnode'] = vplssapws['A%s' % row].value  # 'nokia-node'
    vplssapdict['bdrsystem'] = vplssapws['B%s' % row].value  # 'system-ip'
    vplssapdict['bdrservice'] = vplssapws['C%s' % row].value  # 'service-name'
    vplssapdict['bdrserviceid'] = vplssapws['D%s' % row].value  # 'service-id'
    vplssapdict['bdrsap'] = vplssapws['E%s' % row].value  # 'sap'
    vplssapdict['bdrsapdesc'] = vplssapws['F%s' % row].value  # 'sap-description'
    vplssapdict['bdrsapinfilter'] = vplssapws['G%s' % row].value  # 'sap-ingress-filter'
    vplssapdict['bdrsapegfilter'] = vplssapws['H%s' % row].value  # 'sap-egress-filter'
    vplssapdict['bdrsapqos'] = vplssapws['I%s' % row].value  # 'sap-in/egress qos'
    vplssapdict['iptnnode'] = vplssapws['J%s' % row].value  # 'iptn-node'
    vplssapdict['iptnservice'] = vplssapws['K%s' % row].value  # 'iptn-service'
    vplssapdict['iptnint'] = vplssapws['L%s' % row].value  # 'iptn-interface'
    return vplssapdict

def createepipesapdict(epipesapws, row):
    epipesapdict = {}
    epipesapdict['bdrnode'] = epipesapws['A%s' % row].value  # 'nokia-node'
    epipesapdict['bdrsystem'] = epipesapws['B%s' % row].value  # 'system-ip'
    epipesapdict['bdrservice'] = epipesapws['C%s' % row].value  # 'service-name'
    epipesapdict['bdrserviceid'] = epipesapws['D%s' % row].value  # 'service-id'
    epipesapdict['bdrsap'] = epipesapws['E%s' % row].value  # 'sap'
    epipesapdict['bdrsapdesc'] = epipesapws['F%s' % row].value  # 'sap-description'
    epipesapdict['bdrsapinfilter'] = epipesapws['G%s' % row].value  # 'sap-ingress-filter'
    epipesapdict['bdrsapegfilter'] = epipesapws['H%s' % row].value  # 'sap-egress-filter'
    epipesapdict['bdrsapqos'] = epipesapws['I%s' % row].value  # 'sap-in/egress qos'
    epipesapdict['bdrsapinrate'] = epipesapws['J%s' % row].value  # 'ingress-rate'
    epipesapdict['bdrsapegrate'] = epipesapws['K%s' % row].value  # 'egress-rate'
    epipesapdict['iptnnode'] = epipesapws['L%s' % row].value  # 'iptn-node'
    epipesapdict['iptnservice'] = epipesapws['M%s' % row].value  # 'iptn-service'
    epipesapdict['iptnint'] = epipesapws['N%s' % row].value  # 'iptn-interface'
    return epipesapdict

def createvprnintdict(vprnintws,row):
    vprnintdict = {}
    vprnintdict['bdrnode'] = vprnintws['A%s' % row ].value   #'nokia-node'
    vprnintdict['bdrsystem'] = vprnintws['B%s' % row ].value   #'system-ip'
    vprnintdict['bdrservice'] = vprnintws['C%s' % row ].value   #'service-name'
    vprnintdict['bdrserviceid'] = vprnintws['D%s' % row ].value   #'service-id'
    vprnintdict['bdrintname'] = vprnintws['E%s' % row ].value   #'interface-name'
    vprnintdict['bdrintdesc'] = vprnintws['F%s' % row ].value   #'interface description'
    vprnintdict['bdrsap'] = vprnintws['G%s' % row ].value   #'sap'
    vprnintdict['bdrsapinfilter'] = vprnintws['H%s' % row ].value   #'sap-ingress-filter'
    vprnintdict['bdrsapegfilter'] = vprnintws['I%s' % row ].value   #'sap-egress-filter'
    vprnintdict['bdrsapqos'] = vprnintws['J%s' % row ].value   #'sap-in/egress qos'
    vprnintdict['bdrpriip'] = vprnintws['K%s' % row ].value   #'ip/ipv6 address'
    vprnintdict['bdrsecip'] = vprnintws['L%s' % row ].value   #'secondary ip/ipv6'
    vprnintdict['bdrintfam'] = vprnintws['M%s' % row ].value   #'family'
    vprnintdict['bdrvrrpid'] = vprnintws['N%s' % row ].value   #'vrrp id'
    vprnintdict['bdrvrrpvip'] = vprnintws['O%s' % row ].value   #'vrrp vip'
    vprnintdict['bdrvrrppri'] = vprnintws['P%s' % row ].value   #'vrrp priority'
    vprnintdict['bdrvrrpint'] = vprnintws['Q%s' % row ].value   #'vrrp interval'
    vprnintdict['bdrstatic'] = vprnintws['R%s' % row ].value   #'static-route'
    vprnintdict['bdrbgpfeature'] = vprnintws['S%s' % row ].value   #'bgp'
    vprnintdict['bdrbgpgroup'] = vprnintws['T%s' % row ].value   #'group'
    vprnintdict['bdrbgppeeras'] = vprnintws['U%s' % row ].value   #'peer-as'
    vprnintdict['bdrbgpneighbor'] = vprnintws['V%s' % row ].value   #'neighbor'
    vprnintdict['bdrbgpneighbordesc'] = vprnintws['W%s' % row ].value   #'neighbor-description'
    vprnintdict['bdrbgpimppolname'] = vprnintws['X%s' % row ].value   #'bgp-imp-policy-name'
    vprnintdict['bdrbgpimppol'] = vprnintws['Y%s' % row ].value   #'bgp-imp-policy'
    vprnintdict['bdrbgpimpprefix'] = vprnintws['Z%s' % row ].value   #'bgp-imp-policy-prefix'
    vprnintdict['bdrbgpimpcommu'] = vprnintws['AA%s' % row].value  #'bgp-imp-policy-commu'
    vprnintdict['bdrbgpexppolname'] = vprnintws['AB%s' % row].value  #'bgp-exp-policy-name'
    vprnintdict['bdrbgpexppol'] = vprnintws['AC%s' % row].value  #'bgp-exp-policy'
    vprnintdict['bdrbgpexpprefix'] = vprnintws['AD%s' % row].value  #'bgp-exp-policy-prefix'
    vprnintdict['bdrbgpexpcommu'] = vprnintws['AE%s' % row].value  #'bgp-exp-policy-commu'
    vprnintdict['bdrbgpbfdint'] = vprnintws['AF%s' % row].value  #'bgp-bfd-interval'
    vprnintdict['bdrbgpbfdmul'] = vprnintws['AG%s' % row].value  #'bgp-bfd-multiply'
    vprnintdict['iptnnode'] = vprnintws['AH%s' % row].value  #'iptn-node'
    vprnintdict['iptnservice'] = vprnintws['AI%s' % row].value  #'iptn-service'
    vprnintdict['iptnint'] = vprnintws['AJ%s' % row].value  #'iptn-interface'
    return vprnintdict

def createglobalvprndict(vprnws,row):
    globalvprndict = {}
    globalvprndict['bdrnode'] = vprnws['A%s' % row].value #'nokia-node'
    globalvprndict['bdrsystem'] = vprnws['B%s' % row].value #'system-ip'
    globalvprndict['bdrservice'] = vprnws['C%s' % row].value #'service-name'
    globalvprndict['bdrserviceid'] = vprnws['D%s' % row].value #'service-id'
    globalvprndict['bdrservicestate'] = vprnws['E%s' % row].value #'admin-state'
    globalvprndict['bdrrd'] = vprnws['F%s' % row].value #'rd'
    globalvprndict['bdrvrfimppolname'] = vprnws['G%s' % row].value #'vrf-import-policy-name'
    globalvprndict['bdrvrfimppol'] = vprnws['H%s' % row].value #'vrf-import-policy'
    globalvprndict['bdrvrfimpprefix'] = vprnws['I%s' % row].value #'vrf-import-prefix'
    globalvprndict['bdrvrfimpcommu'] = vprnws['J%s' % row].value #'vrf-import-commu'
    globalvprndict['bdrvrfexppolname'] = vprnws['K%s' % row].value #'vrf-export-policy-name'
    globalvprndict['bdrvrfexppol'] = vprnws['L%s' % row].value #'vrf-export-policy'
    globalvprndict['bdrvrfexpprefix'] = vprnws['M%s' % row].value #'vrf-export-prefix'
    globalvprndict['bdrvrfexpcommu'] = vprnws['N%s' % row].value #'vrf-export-commu'
    globalvprndict['bdrvrftarget'] = vprnws['O%s' % row].value #'vrf-target'
    globalvprndict['bdrvrfagg'] = vprnws['P%s' % row].value #'aggregate-route'
    globalvprndict['bdrcustid'] = vprnws['Q%s' % row].value #'customer-id'
    globalvprndict['iptnnode'] = vprnws['R%s' % row].value #'iptn-node'
    globalvprndict['iptnservice'] = vprnws['S%s' % row].value #'iptn-service'
    return globalvprndict

def createevpndict(l2evpnws, row):
    evpndict = {}
    evpndict['bdrnode'] = l2evpnws['A%s' % row].value #'nokia-node'
    evpndict['bdrsystem'] = l2evpnws['B%s' % row].value #'system-ip'
    evpndict['bdrservice'] = l2evpnws['C%s' % row].value #'service-name'
    evpndict['bdrserviceid'] = l2evpnws['D%s' % row].value #'service-id'
    evpndict['bdrethersegment'] = l2evpnws['E%s' % row].value #'ethernet-segment'
    evpndict['bdresi'] = l2evpnws['F%s' % row].value #'esi'
    evpndict['bdrpreference'] = l2evpnws['G%s' % row].value #'preference'
    evpndict['bdrport'] = l2evpnws['H%s' % row].value #'port/lag'
    evpndict['bdrportvlan'] = l2evpnws['I%s' % row].value #'vlan'
    evpndict['bdrlocalac'] = l2evpnws['J%s' % row].value #'local-ac-name'
    evpndict['bdrlocalethtag'] = l2evpnws['K%s' % row].value #'local-eth-tag'
    evpndict['bdrremoteac'] = l2evpnws['L%s' % row].value #'remote-ac-name'
    evpndict['bdrremoteethtag'] = l2evpnws['M%s' % row].value #'remote-eth-tag'
    evpndict['bdrevi'] = l2evpnws['N%s' % row].value #'evi'
    evpndict['bdrcustid'] = l2evpnws['O%s' % row].value #'customer-id'
    evpndict['bdrentropy'] = l2evpnws['P%s' % row].value #'entropy'
    return evpndict


def bdrl2evpnworksheet(serviceid, mievpndict, lastbdrnode, bdrcfgwb):
    bdrl2evpnws = bdrcfgwb['bdr-configuration']
    bdrconfiglist = []
    remarklist = []
    nodecolumnlist = []
    existcolumn = []
    # print('column %s' %bdrl2evpnws['A1'].value)
    if lastbdrnode != mievpndict['bdrnode']:
        for c in columnlist:
            if bdrl2evpnws['%s2' % c].value != None:
                nodecolumnlist.append(bdrl2evpnws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mievpndict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrl2evpnws['D1'] = existcolumn[-1]
        if mievpndict['bdrnode'] not in nodecolumnlist:
            bdrl2evpnws['D1'] = existcolumn[-1]
            column = bdrl2evpnws['D1'].value
            column = chr(ord(column) + 3)
            bdrl2evpnws['D1'] = column
            bdrl2evpnws['%s2' % column] = mievpndict['bdrnode']
            bdrl2evpnws.column_dimensions['%s' % chr(ord(column) - 1)].width = 12.0
            bdrl2evpnws.column_dimensions['%s' % column].width = 120.0
    elif lastbdrnode:
        for c in columnlist:
            if bdrl2evpnws['%s2' % c].value != None:
                nodecolumnlist.append(bdrl2evpnws['%s2' % c].value)
                existcolumn.append(c)
        for n in range(len(nodecolumnlist)):
            if mievpndict['bdrnode'] == nodecolumnlist[n]:
                column = existcolumn[n]
                bdrl2evpnws['D1'] = existcolumn[-1]

    bdrconfiglist.append('')
    bdrconfiglist.append('### BDR EPIPE-EVPN ethernet-segment configuration ###')
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" admin-state enable' % mievpndict['bdrethersegment'])
    if mievpndict['bdrportvlan'] == None:
        bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" type none' % mievpndict['bdrethersegment'])
    else:
        bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" type virtual' % mievpndict['bdrethersegment'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" esi 0x%s' % (mievpndict['bdrethersegment'],mievpndict['bdresi']))
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" multi-homing-mode single-active' % mievpndict['bdrethersegment'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election es-activation-timer 0' % mievpndict['bdrethersegment'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election service-carving-mode manual' % mievpndict['bdrethersegment'])
    bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" df-election manual preference value %s' % (mievpndict['bdrethersegment'],mievpndict['bdrpreference']))
    if mievpndict['bdrportvlan'] == None:
        if 'lag' in mievpndict['bdrport']:
            bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association lag %s' % (mievpndict['bdrethersegment'],mievpndict['bdrport'].split('-')[-1]))
        else:
            bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association port %s' % (mievpndict['bdrethersegment'],mievpndict['bdrport']))
    else:
        if 'lag' in mievpndict['bdrport']:
            bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association lag %s virtual-ranges dot1q q-tag %s end %s' % (mievpndict['bdrethersegment'],mievpndict['bdrport'].split('-')[-1],mievpndict['bdrportvlan'],mievpndict['bdrportvlan']))
        else:
            bdrconfiglist.append('/configure service system bgp evpn ethernet-segment "%s" association port %s virtual-ranges dot1q q-tag %s end %s' % (mievpndict['bdrethersegment'],mievpndict['bdrport'],mievpndict['bdrportvlan'],mievpndict['bdrportvlan']))

    bdrconfiglist.append('')
    bdrconfiglist.append('### BDR EPIPE-EVPN service configuration ###')
    bdrconfiglist.append('/configure service epipe "%s" admin-state enable' % mievpndict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" description "%s"' % (mievpndict['bdrservice'],mievpndict['bdrservice']))
    bdrconfiglist.append('/configure service epipe "%s" service-id %s' % (mievpndict['bdrservice'],mievpndict['bdrserviceid']))
    bdrconfiglist.append('/configure service epipe "%s" customer "%s"' % (mievpndict['bdrservice'],mievpndict['bdrcustid']))
    bdrconfiglist.append('/configure service epipe "%s" service-mtu 9194' % mievpndict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" { bgp 1 }' % mievpndict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn evi %s' % (mievpndict['bdrservice'],mievpndict['bdrevi']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn local-ac name "%s"' % (mievpndict['bdrservice'],mievpndict['bdrlocalac']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn local-ac eth-tag %s' % (mievpndict['bdrservice'],mievpndict['bdrlocalethtag']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn remote-ac name "%s"' % (mievpndict['bdrservice'],mievpndict['bdrremoteac']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn remote-ac eth-tag %s' % (mievpndict['bdrservice'],mievpndict['bdrremoteethtag']))
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 admin-state enable' % mievpndict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 entropy-label true' % mievpndict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 control-word true' % mievpndict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 auto-bind-tunnel resolution filter' % mievpndict['bdrservice'])
    bdrconfiglist.append('/configure service epipe "%s" bgp-evpn mpls 1 auto-bind-tunnel resolution-filter rsvp true' % mievpndict['bdrservice'])


    for i in bdrconfiglist:
        if 'ethernet-segment' in i:
            remarklist.append('ether-segment')
        else:
            remarklist.append('epipe-evpn')


    if bdrl2evpnws['%s1' % chr(ord(column) - 1)].value == None:
        l2evpnwsrow = bdrl2evpnws['B1'].value
    else:
        l2evpnwsrow = bdrl2evpnws['%s1' % chr(ord(column) - 1)].value

    for line in range(len(bdrconfiglist)):
        bdrl2evpnws['%s%s' % (chr(ord(column) - 1), l2evpnwsrow)] = remarklist[line]
        bdrl2evpnws['%s%s' % (column, l2evpnwsrow)] = bdrconfiglist[line]
        l2evpnwsrow += 1

    lastbdrnode = mievpndict['bdrnode']
    bdrl2evpnws['%s1' % chr(ord(column) - 1)] = l2evpnwsrow
    return lastbdrnode



def main():
    e = open(extractpath+'latest.txt', "r")
    extime = e.readlines()[0][19:32]
    listserviceid, miiptndict = createlistservice()
    bdrfilterdict = {}
    bdrpolicydict = {}
    for filterrow in range(3, misum['filter-naming'].max_row + 1):
        bdrfilterdict.setdefault('%s_prefix' %misum['filter-naming']['A%s' % filterrow].value, []).append(misum['filter-naming']['H%s' % filterrow].value)
        bdrfilterdict['%s_pr_%s' % (misum['filter-naming']['A%s' % filterrow].value,misum['filter-naming']['H%s' % filterrow].value)] = misum['filter-naming']['I%s' % filterrow].value
        bdrfilterdict.setdefault('%s_port' %misum['filter-naming']['A%s' % filterrow].value, []).append(misum['filter-naming']['K%s' % filterrow].value)
        bdrfilterdict['%s_po_%s' % (misum['filter-naming']['A%s' % filterrow].value,misum['filter-naming']['K%s' % filterrow].value)] = misum['filter-naming']['L%s' % filterrow].value

    for policyrow in range(3, misum['policy-naming'].max_row + 1):
        bdrpolicydict.setdefault('%s_%s_%s_existing' % (misum['policy-naming']['A%s' % policyrow].value,misum['policy-naming']['C%s' % policyrow].value,misum['policy-naming']['E%s' % policyrow].value), []).append(misum['policy-naming']['J%s' % policyrow].value)
        bdrpolicydict['%s_%s_%s_%s' % (misum['policy-naming']['A%s' % policyrow].value,misum['policy-naming']['C%s' % policyrow].value,misum['policy-naming']['E%s' % policyrow].value,misum['policy-naming']['J%s' % policyrow].value)] = misum['policy-naming']['K%s' % policyrow].value
        bdrpolicydict.setdefault('%s_%s_%s_entry-existing' % (misum['policy-naming']['A%s' % policyrow].value,misum['policy-naming']['C%s' % policyrow].value,misum['policy-naming']['E%s' % policyrow].value), []).append(misum['policy-naming']['G%s' % policyrow].value)
        bdrpolicydict['%s_%s_%s_%s_entry' % (misum['policy-naming']['A%s' % policyrow].value,misum['policy-naming']['C%s' % policyrow].value,misum['policy-naming']['E%s' % policyrow].value,misum['policy-naming']['G%s' % policyrow].value)] = misum['policy-naming']['H%s' % policyrow].value
    #print(bdrfilterdict)
    for serviceid in listserviceid:
        lastbdrnode = ''
        portcheck = []
        l3prefix = []
        bdrcfgwb = createbdrconfigfile()
        createiptncfgworksheet(serviceid, extime, miiptndict,bdrcfgwb)
        for misumline in range(3,misum['port-lag'].max_row + 1):
            if misum['port-lag']['E%s' % misumline].value != None:
                if serviceid == misum['port-lag']['E%s' % misumline].value:
                    miportdict = createportdict(misum['port-lag'],misumline)
                    portcheck, lastbdrnode = bdrportlagconfigworksheet(serviceid,miportdict,lastbdrnode, bdrcfgwb,portcheck,bdrfilterdict)
                    servicename = miportdict['bdrservicename']
        lastbdrnode = ''
        for l2sdpline in range(3,misum['l2service-sdp'].max_row + 1):
            if misum['l2service-sdp']['D%s' % l2sdpline].value != None:
                if serviceid == misum['l2service-sdp']['D%s' % l2sdpline].value:
                    mil2sdpdict = createl2sdpdict(misum['l2service-sdp'], l2sdpline)
                    lastbdrnode = bdrl2sdpworksheet(serviceid, mil2sdpdict, lastbdrnode, bdrcfgwb)
        lastbdrnode = ''
        for l2evpnline in range(3,misum['l2-evpn'].max_row + 1):
            if misum['l2-evpn']['D%s' %l2evpnline].value != None:
                if serviceid == misum['l2-evpn']['D%s' %l2evpnline].value:
                    mievpndict = createevpndict(misum['l2-evpn'], l2evpnline)
                    lastbdrnode = bdrl2evpnworksheet(serviceid, mievpndict, lastbdrnode, bdrcfgwb)
        lastbdrnode = ''
        lastasbr1 = ''
        asbr1line = []
        asbr2line = []
        asbrline = []
        for l2sdpinterasline in range(3, misum['l2inter-as'].max_row + 1):
            if misum['l2inter-as']['D%s' %l2sdpinterasline].value != None:
                if serviceid == misum['l2inter-as']['D%s' %l2sdpinterasline].value:
                    if not lastasbr1:
                        lastasbr1 = misum['l2inter-as']['G%s' % l2sdpinterasline].value
                        asbr1line.append(l2sdpinterasline)
                    else:
                        if misum['l2inter-as']['G%s' % l2sdpinterasline].value == lastasbr1:
                            asbr1line.append(l2sdpinterasline)
                        else:
                            asbr2line.append(l2sdpinterasline)
                    mil2sdpinterasdict = createl2sdpinterasdict(misum['l2inter-as'], l2sdpinterasline)
                    lastbdrnode = bdrl2sdpinterasworksheet(serviceid, mil2sdpinterasdict, lastbdrnode, bdrcfgwb)
        asbrline.append(asbr1line)
        asbrline.append(asbr2line)
        lastasbrnode = ''
        for asbr in asbrline:
            for l2sdpinterasasbrline in asbr:
                mil2sdpinterasdict = createl2sdpinterasdict(misum['l2inter-as'], l2sdpinterasasbrline)
                lastasbrnode = bdrl2sdpasbrworksheet(serviceid, mil2sdpinterasdict,lastasbrnode, bdrcfgwb)
        lastbdrnode = ''
        lastasbr1 = ''
        evpnpeservicelist = []
        asbr1line = []
        asbr2line = []
        asbrline = []
        for l2evpninterasline in range(3, misum['l2-evpn-inter-as'].max_row + 1):
            if misum['l2-evpn-inter-as']['D%s' %l2evpninterasline].value != None:
                if serviceid == misum['l2-evpn-inter-as']['D%s' %l2evpninterasline].value:
                    if not lastasbr1:
                        lastasbr1 = misum['l2-evpn-inter-as']['O%s' % l2evpninterasline].value
                        asbr1line.append(l2evpninterasline)
                    else:
                        if misum['l2-evpn-inter-as']['O%s' % l2evpninterasline].value == lastasbr1:
                            asbr1line.append(l2evpninterasline)
                        else:
                            asbr2line.append(l2evpninterasline)
                    evpnpeservice = '%s%s%s' %(misum['l2-evpn-inter-as']['A%s' % l2evpninterasline].value, misum['l2-evpn-inter-as']['C%s' % l2evpninterasline].value, misum['l2-evpn-inter-as']['D%s' % l2evpninterasline].value)
                    if evpnpeservice not in evpnpeservicelist:
                        evpnpeservicelist.append(evpnpeservice)
                        mil2evpninterasdict = createl2evpninterasdict(misum['l2-evpn-inter-as'], l2evpninterasline)
                        lastbdrnode = bdrl2evpninterasworksheet(serviceid, mil2evpninterasdict, lastbdrnode, bdrcfgwb)
        asbrline.append(asbr1line)
        asbrline.append(asbr2line)
        lastasbrnode = ''
        evpnasbrservicelist = []
        for asbr in asbrline:
            for l2evpninterasasbrline in asbr:
                evpnasbrservice = '%s%s%s' %(misum['l2-evpn-inter-as']['O%s' % l2evpninterasasbrline].value, misum['l2-evpn-inter-as']['C%s' % l2evpninterasasbrline].value, misum['l2-evpn-inter-as']['D%s' % l2evpninterasasbrline].value)
                if evpnasbrservice not in evpnasbrservicelist:
                    evpnasbrservicelist.append(evpnasbrservice)
                    mil2evpninterasdict = createl2evpninterasdict(misum['l2-evpn-inter-as'], l2evpninterasasbrline)
                    lastasbrnode = bdrl2evpnasbrworksheet(serviceid, mil2evpninterasdict,lastasbrnode, bdrcfgwb)


        lastbdrnode = ''
        lastbdrvprnintnode = ''
        vprncheck = []
        for vprnintline in range(3, misum['vprn-interface'].max_row + 1):
            if misum['vprn-interface']['D%s' % vprnintline].value != None:
                if serviceid == int(misum['vprn-interface']['D%s' % vprnintline].value):
                    mivprnintdict = createvprnintdict(misum['vprn-interface'],vprnintline)
                    #print(mivprnintdict)
                    lastbdrvprnintnode, l3prefix = bdrvprnintworksheet(serviceid, mivprnintdict, lastbdrvprnintnode, bdrcfgwb,bdrpolicydict,l3prefix)
                    for globalvprnline in range(3, misum['global-vprn'].max_row + 1):
                        if misum['global-vprn']['D%s' % globalvprnline].value != None:
                            if serviceid == int(misum['global-vprn']['D%s' % globalvprnline].value):
                                miglobalvprndict = createglobalvprndict(misum['global-vprn'],globalvprnline)
                                if mivprnintdict['bdrnode'] == miglobalvprndict['bdrnode']:
                                    #print(mivprnintdict['bdrnode'])
                                    vprnservice = miglobalvprndict['bdrnode']+str(serviceid)
                                    if vprnservice not in vprncheck:
                                        vprncheck.append(vprnservice)
                                        lastbdrnode = bdrglobalvprnworksheet(serviceid, miglobalvprndict, lastbdrnode, bdrcfgwb,bdrpolicydict)


        lastbdrnode = ''
        lastinterasbdrnode = ''
        for epipesapline in range(3, misum['epipe-sap'].max_row + 1):
            if misum['epipe-sap']['D%s' % epipesapline].value != None:
                if serviceid == int(misum['epipe-sap']['D%s' % epipesapline].value):
                    miepipesapdict = createepipesapdict(misum['epipe-sap'], epipesapline)
                    lastbdrnode = bdrepipesapworksheet(serviceid, miepipesapdict, lastbdrnode, bdrcfgwb, misum)
                    lastinterasbdrnode = bdrepipeinterasworksheet(serviceid, miepipesapdict, lastinterasbdrnode, bdrcfgwb)

        lastbdrnode = ''
        lastinterasbdrnode = ''
        for vplssapline in range(3, misum['vpls-sap'].max_row + 1):
            if misum['vpls-sap']['D%s' % vplssapline].value != None:
                if serviceid == int(misum['vpls-sap']['D%s' % vplssapline].value):
                    mivplssapdict = createvplssapdict(misum['vpls-sap'], vplssapline)
                    lastbdrnode = bdrvplssapworksheet(serviceid, mivplssapdict, lastbdrnode, bdrcfgwb, misum)
                    lastinterasbdrnode = bdrvplsinterasworksheet(serviceid, mivplssapdict, lastinterasbdrnode, bdrcfgwb)

        lastbdrnode = ''
        lastinterasbdrnode = ''
        for rvplssapline in range(3, misum['rvpls-sap'].max_row + 1):
            if misum['rvpls-sap']['D%s' % rvplssapline].value != None:
                if serviceid == int(misum['rvpls-sap']['D%s' % rvplssapline].value):
                    mirvplssapdict = creatervplssapdict(misum['rvpls-sap'], rvplssapline)
                    if 'rVPLS' not in misum['rvpls-sap']['C%s' % rvplssapline].value:
                        lastbdrnode = bdrrvplssapworksheet(serviceid, mirvplssapdict, lastbdrnode, bdrcfgwb,misum)
                    for rvplsintline in range(3, misum['rvpls-interface'].max_row + 1):
                        if misum['rvpls-sap']['A%s' % rvplssapline].value == misum['rvpls-interface']['A%s' % rvplsintline].value and misum['rvpls-sap']['C%s' % rvplssapline].value == misum['rvpls-interface']['G%s' % rvplsintline].value:
                            if mirvplssapdict['bdrnode'] == misum['rvpls-interface']['A%s' % rvplsintline].value:
                                lastbdrnode = bdrrvplssapworksheet(serviceid, mirvplssapdict, lastbdrnode, bdrcfgwb,misum)

        lastbdrnode = ''
        lastbdrrvplsintnode = ''
        rvplscheck = []
        for rvplssapline in range(3, misum['rvpls-sap'].max_row + 1):
            if misum['rvpls-sap']['D%s' % rvplssapline].value != None:
                if serviceid == int(misum['rvpls-sap']['D%s' % rvplssapline].value):
                    mirvplssapdict = creatervplssapdict(misum['rvpls-sap'], rvplssapline)
                    for rvplsintline in range(3, misum['rvpls-interface'].max_row + 1):
                        if 'SAP:$' in misum['rvpls-interface']['G%s' % rvplsintline].value:
                            sapid = misum['rvpls-interface']['G%s' % rvplsintline].value.split(':$')[1]
                            if mirvplssapdict['bdrnode'] == misum['rvpls-interface']['A%s' % rvplsintline].value and mirvplssapdict['bdrsap'] == sapid:
                                mirvplsintdict = creatervplsintdict(misum['rvpls-interface'],rvplsintline)
                                lastbdrrvplsintnode, l3prefix = bdrrvplsintworksheet(serviceid, mirvplsintdict, lastbdrrvplsintnode, bdrcfgwb,l3prefix)
                                for globalvprnline in range(3, misum['global-vprn'].max_row + 1):
                                    if mirvplsintdict['bdrservice'] == misum['global-vprn']['C%s' % globalvprnline].value:
                                        miglobalvprndict = createglobalvprndict(misum['global-vprn'],globalvprnline)
                                        if mirvplsintdict['bdrnode'] == miglobalvprndict['bdrnode']:
                                            rvplsservice = miglobalvprndict['bdrnode']+str(serviceid)
                                            if rvplsservice not in rvplscheck:
                                                rvplscheck.append(rvplsservice)
                                                lastbdrnode = bdrglobalvprnworksheet(serviceid, miglobalvprndict, lastbdrnode, bdrcfgwb,bdrpolicydict)
                        else:
                            if mirvplssapdict['bdrnode'] == misum['rvpls-interface']['A%s' % rvplsintline].value and mirvplssapdict['bdrservice'] == misum['rvpls-interface']['G%s' % rvplsintline].value:
                                mirvplsintdict = creatervplsintdict(misum['rvpls-interface'],rvplsintline)
                                lastbdrrvplsintnode, l3prefix = bdrrvplsintworksheet(serviceid, mirvplsintdict, lastbdrrvplsintnode, bdrcfgwb,l3prefix)
                                for globalvprnline in range(3, misum['global-vprn'].max_row + 1):
                                    if mirvplsintdict['bdrservice'] == misum['global-vprn']['C%s' % globalvprnline].value:
                                        miglobalvprndict = createglobalvprndict(misum['global-vprn'],globalvprnline)
                                        if mirvplsintdict['bdrnode'] == miglobalvprndict['bdrnode']:
                                            rvplsservice = miglobalvprndict['bdrnode']+str(serviceid)
                                            if rvplsservice not in rvplscheck:
                                                rvplscheck.append(rvplsservice)
                                                lastbdrnode = bdrglobalvprnworksheet(serviceid, miglobalvprndict, lastbdrnode, bdrcfgwb,bdrpolicydict)


        l3prefixlist = []
        addonl3 = []
        maxrow = bdrcfgwb['bdr-policy'].max_row + 1
        maxcolumn = get_column_letter(bdrcfgwb['bdr-policy'].max_column)
        advrr = open(migratepath + time + '\\' + 'show_adv_rr.txt', "a", encoding="utf-8")
        prefixcheck = []
        for prefix in l3prefix:
            if ':' in prefix:
                addonl3.append(prefix.replace('/','0/'))
        l3prefix.extend(addonl3)
        for prefix in l3prefix:
            if prefix not in prefixcheck:
                prefixcheck.append(prefix)
                if ':' in prefix:
                    if '0/' not in prefix:
                        advrr.write("/show router bgp neighbor 10.129.147.1 advertised-routes vpn-ipv6 | match '%s' post-lines 2" % prefix + '\n')
                else:
                    advrr.write("/show router bgp neighbor 10.129.147.1 advertised-routes vpn-ipv4 | match '%s' post-lines 2" % prefix + '\n')
            for allcolumn in bdrcfgwb['bdr-policy']['B3:%s%s' % (maxcolumn,maxrow)]:
                for bdrpolicycfg in allcolumn:
                    if bdrpolicycfg.value != None:
                        if prefix in bdrpolicycfg.value:
                            bdrpolicycfg.font = Font(color='FFFF0000')
                            if 'policy-options prefix-list' in bdrpolicycfg.value:
                                if bdrpolicycfg.value.split()[3] not in l3prefixlist:
                                    l3prefixlist.append(bdrpolicycfg.value.split()[3])
        for pfl in l3prefixlist:
            for allcolumn in bdrcfgwb['bdr-policy']['B3:%s%s' % (maxcolumn, maxrow)]:
                for bdrpolicycfg in allcolumn:
                    if bdrpolicycfg.value != None:
                        if pfl in bdrpolicycfg.value:
                            bdrpolicycfg.font = Font(color='FFFF0000')
        if miiptndict['%s_type'% serviceid] == 'VPRN' or miiptndict['%s_type'% serviceid] == 'Bridge-Domain' or miiptndict['%s_type'% serviceid] == 'Virtual-Switch':
            bdrcfgwb.remove(bdrcfgwb['l2inter-as-configuration'])
        if miiptndict['%s_type'% serviceid] == 'EPIPE':
            bdrcfgwb.remove(bdrcfgwb['bdr-policy'])
        bdrcfgwb.save(migratepath + time + '\\' + '%s_%s_configuration.xlsx' %(serviceid,servicename))

if __name__ == "__main__":
    main()





